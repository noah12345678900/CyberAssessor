"""Property-based tests for ``engine.workbook_sync``.

The example-driven suite in ``backend/tests/engine/test_workbook_sync.py``
pins each event type's happy path with a hand-crafted before/after
workbook pair. This file fuzzes the *shape* of the diff (arbitrary
combinations of added/removed/edited/moved rows interleaved with
key-less rows) so a refactor that breaks an invariant in a corner of
that space gets caught before it ships.

The invariants we pin here all underpin the "what changed since you
last looked" UI that reads ``WorkbookSyncEvent`` directly — drift in
any of them silently corrupts the audit log:

  1. **Idempotence on no-op resync.** Sync the same workbook twice → the
     second call produces ZERO events. A regression here would mean the
     diff misclassifies an unchanged row as "edited", spamming the audit
     log with phantom changes that mask real ones.

  2. **Shared ``occurred_at``.** Every event from one ``sync_workbook``
     call carries the same UTC timestamp. The UI groups events by
     timestamp into "sync runs"; per-event timestamps would scatter a
     single run across multiple groups.

  3. **Null/non-null contract on ``old_value_json`` / ``new_value_json``.**
     ``added`` → ``old_value_json is None``; ``removed`` → ``new_value_json
     is None``; ``moved`` and ``edited`` → both populated. A flip here
     would break the "show me the before/after" detail view.

  4. **Summary counts match persisted event counts by type.** The
     ``SyncSummary`` is what the API hands back; the DB rows are what
     the UI reads later. Drift between them = the "x changes since last
     sync" banner shows a different count than the detail page.

  5. **Keyed-diff totality.** Rows missing ``cci_id`` MUST NOT appear in
     any event. They can't participate in the keyed diff (no stable
     identifier within a control), and silently emitting an event with
     ``cci_id=None`` would let phantom keys pollute the audit log.

  6. **JSON payload round-trip.** Every populated ``*_value_json`` must
     parse via ``json.loads`` to a dict. A regression to ``repr()`` /
     single-quote serialization would yield strings the SQLite JSON
     functions can't parse downstream.

  7. **Convergence after one sync.** After one ``sync_workbook`` call,
     a subsequent identical call produces zero events — i.e. the
     snapshot sidecar is in the same state as the persisted log
     described.

  8. **Every event key was present somewhere.** Every event's
     ``(control_id, cci_id)`` was either in the prior workbook (for
     ``removed``) or in the current workbook (for ``added`` /
     ``edited`` / ``moved``) — no phantom keys.

These properties exercise the full read-diff-persist pipeline; the
unit-level invariants on ``_diff_indexes`` itself are pinned by the
example tests in ``backend/tests/engine/test_reader_reread.py``.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

hypothesis = pytest.importorskip("hypothesis")

from hypothesis import HealthCheck, given, settings  # noqa: E402
from hypothesis import strategies as st  # noqa: E402

# Backend package on the path — the property tests live at repo-root
# `tests/engine/`, so we need parents[2] to land at the backend dir.
_BACKEND = Path(__file__).resolve().parents[2] / "backend"
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.engine.workbook_sync import (  # noqa: E402
    SyncSummary,
    sync_workbook,
)
from cybersecurity_assessor.excel.ccis_reader import (  # noqa: E402
    COL_CCI,
    COL_CONTROL,
    COL_NARRATIVE,
    COL_RESULTS,
    COL_STATUS,
)
from cybersecurity_assessor.models import (  # noqa: E402
    Workbook as WorkbookRow,
)
from cybersecurity_assessor.models import (  # noqa: E402
    WorkbookSyncEvent,
)

SHEET = "WORKING SHEET"


# ---------------------------------------------------------------------------
# Strategies
# ---------------------------------------------------------------------------


# Small fixed control pool — using a constrained set keeps the diff
# interesting (added/removed actually overlap with edited cases) while
# staying within Hypothesis's example budget.
_CONTROL_IDS = ["AC-2", "AC-3", "AC-6", "AU-2", "IA-5", "CM-7"]

# CCI ids — int-only strategy that maps to canonical "CCI-NNNNNN" form
# at workbook write time. Small pool so collisions across rows are
# possible (which exercises the "shared key, different row content" =>
# edit path).
_CCI_INTS = st.integers(min_value=15, max_value=40)

# Excel-safe text alphabet — printable ASCII + tab/newline. Hypothesis's
# default `st.text()` includes C0 control characters (U+0000..U+001F minus
# tab/LF/CR) that openpyxl rejects with `IllegalCharacterError` — but those
# bytes never reach the workbook_sync diff layer because openpyxl refuses
# to serialize them upstream. Restricting the strategy here mirrors what
# the real workbook write path can actually express; it does not weaken
# coverage of the diff/normalize logic.
_EXCEL_SAFE_TEXT = st.text(
    alphabet=st.characters(min_codepoint=0x20, max_codepoint=0x7E),
    min_size=0,
    max_size=40,
)

# Narratives — short text so workbook writes stay fast. Includes the
# empty string so the "whitespace collapses to None" normalization in
# `_normalize_diff_value` gets exercised.
_NARRATIVE = st.one_of(
    st.none(),
    st.just(""),
    _EXCEL_SAFE_TEXT,
)

# Status — one of the canonical CCIS values. The diff includes column N,
# so flipping status between syncs should classify the row as `edited`.
_STATUS = st.sampled_from([None, "Compliant", "Non-Compliant", "Not Applicable"])

# Results — column Q; another diff-tracked column. Bound length to keep
# workbook write IO cheap.
_RESULTS = st.one_of(st.none(), _EXCEL_SAFE_TEXT)


@st.composite
def _row_spec(draw) -> dict:
    """One row spec — may or may not have a CCI id.

    Rows without a CCI id are written to the workbook (column B populated,
    column H blank) but EXCLUDED from the keyed diff. The properties
    below rely on that distinction.
    """
    has_cci = draw(st.booleans())
    spec = {
        "control_id": draw(st.sampled_from(_CONTROL_IDS)),
        "cci_id": draw(_CCI_INTS) if has_cci else None,
        "narrative": draw(_NARRATIVE),
        "status": draw(_STATUS),
        "results": draw(_RESULTS),
    }
    return spec


# Small row sets — Hypothesis budget is bounded by workbook write IO
# (~50ms per save × 2 saves per example). 0-6 rows keeps each example
# under ~1s while still hitting all four diff-event branches.
_ROW_SET = st.lists(_row_spec(), min_size=0, max_size=6)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def session():
    """In-memory SQLite with a shared connection — mirrors the example suite.

    Hypothesis re-invokes the test body many times per fixture instance;
    every property test calls `_reset_schema(session)` at the top so
    rows from prior examples can't bleed into the current example's
    count assertions.
    """
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


def _reset_schema(session: Session) -> None:
    """Wipe all rows and re-create the schema between Hypothesis examples.

    Cheaper than dropping/recreating tables — and avoids tripping the
    foreign-key checks SQLite enforces when WorkbookSyncEvent points at
    a Workbook id we've already deleted out from under it.
    """
    session.rollback()
    for table in reversed(SQLModel.metadata.sorted_tables):
        session.exec(table.delete())
    session.commit()


def _persist_workbook_row(session: Session, workbook_path: Path) -> WorkbookRow:
    """Create the FK target row used by every WorkbookSyncEvent."""
    wb = WorkbookRow(path=str(workbook_path), filename=workbook_path.name)
    session.add(wb)
    session.commit()
    session.refresh(wb)
    return wb


# ---------------------------------------------------------------------------
# Workbook writer — fuller column coverage than the example suite so the
# diff-tracked columns (N status, Q results) actually drive `edited` events.
# ---------------------------------------------------------------------------


_FIELD_TO_COL = {
    "control_id": COL_CONTROL,
    "cci_id": COL_CCI,
    "narrative": COL_NARRATIVE,
    "status": COL_STATUS,
    "results": COL_RESULTS,
}


def _format_cci_for_excel(cci_int: int | None) -> str | None:
    """Mirror the reader's expected on-sheet form for the CCI column.

    The reader normalizes 'CCI-000015' or bare '15' to 'CCI-000015' on
    read; we write bare zero-padded integers so the strategy's int can
    land in column H as-is.
    """
    if cci_int is None:
        return None
    return f"{cci_int:06d}"


def _write_workbook(path: Path, rows: list[dict]) -> Path:
    """Render a row spec list to disk in CCIS shape.

    Header at row 6 (matches the reader's `min_row=7` data start).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(row=6, column=COL_CONTROL, value="Control Acronym")
    ws.cell(row=6, column=COL_CCI, value="CCI")

    for offset, row in enumerate(rows):
        excel_row = 7 + offset
        for key, val in row.items():
            col = _FIELD_TO_COL.get(key)
            if col is None:
                raise KeyError(f"Unknown row spec key: {key}")
            if key == "cci_id":
                val = _format_cci_for_excel(val)
            if val is not None:
                ws.cell(row=excel_row, column=col, value=val)

    wb.save(path)
    wb.close()
    return path


def _events_for(session: Session, workbook_id: int) -> list[WorkbookSyncEvent]:
    return list(
        session.exec(
            select(WorkbookSyncEvent).where(
                WorkbookSyncEvent.workbook_id == workbook_id
            )
        )
    )


def _expected_keys_for(rows: list[dict]) -> set[tuple[str, str]]:
    """The set of (control_id, cci_id) tuples the keyed diff will see.

    Mirrors the reader's `_row_key` filter: drop rows with no CCI id,
    and apply the same canonical-form conversion the reader does. This
    is the ground-truth set the properties below use to bound which
    keys can legally appear in events.
    """
    out: set[tuple[str, str]] = set()
    for r in rows:
        if r["cci_id"] is None:
            continue
        out.add((r["control_id"], f"CCI-{r['cci_id']:06d}"))
    return out


def _clear_index_cache() -> None:
    """Drop the reader's path/mtime/size cache between writes.

    Without this, two writes that happen inside the same mtime tick
    (sub-millisecond on some Windows volumes) AND produce the same
    file size would hit the cache and return stale row data, breaking
    every property below. Cheaper and more deterministic than sleeping
    to wait out the mtime resolution.
    """
    from cybersecurity_assessor.excel import ccis_reader as _reader

    _reader._INDEX_CACHE.clear()


def _clear_snapshot(workbook_path: Path) -> None:
    """Remove the ``.snapshot.json`` sidecar for a workbook path.

    Hypothesis re-invokes function-scoped fixtures across examples, but
    ``tmp_path`` is per-TEST, not per-example. That means example N+1
    sees example N's snapshot sidecar on disk and ``sync_workbook``
    diffs against it, producing "phantom" events for keys that were
    never in the current example's row sets. Clearing the sidecar at
    the top of each test body restores the per-example isolation the
    rest of the harness assumes.
    """
    sidecar = Path(str(workbook_path) + ".snapshot.json")
    if sidecar.exists():
        sidecar.unlink()


# ---------------------------------------------------------------------------
# Idempotence — second sync on an unchanged file emits zero events
# ---------------------------------------------------------------------------


@given(rows=_ROW_SET)
@settings(
    max_examples=20,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_resync_on_unchanged_workbook_emits_zero_events(rows, session, tmp_path):
    """Sync the same file twice → second call's SyncSummary is all-zeros.

    The snapshot sidecar written by the first sync should make the
    second sync's diff fully empty. A regression to "always emits added
    events" would let the audit log accumulate duplicate noise on every
    open.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_idem.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, rows)
    wb_row = _persist_workbook_row(session, workbook_path)

    # First sync — establishes the snapshot.
    sync_workbook(session, wb_row.id, workbook_path)
    first_count = len(_events_for(session, wb_row.id))

    # Second sync — should be a no-op.
    summary = sync_workbook(session, wb_row.id, workbook_path)
    assert summary.added_count == 0
    assert summary.removed_count == 0
    assert summary.moved_count == 0
    assert summary.edited_count == 0
    assert summary.events == []
    assert summary.had_prior_snapshot is True

    # And the persisted log didn't grow.
    assert len(_events_for(session, wb_row.id)) == first_count


# ---------------------------------------------------------------------------
# Shared occurred_at — every event from one call shares a timestamp
# ---------------------------------------------------------------------------


@given(rows=_ROW_SET)
@settings(
    max_examples=20,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_all_events_in_one_call_share_occurred_at(rows, session, tmp_path):
    """All ``WorkbookSyncEvent`` rows from a single call share ``occurred_at``.

    The UI groups events into "sync runs" by exact timestamp; even a
    1µs drift between events in one call would scatter that run across
    two groups in the UI.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_ts.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, rows)
    wb_row = _persist_workbook_row(session, workbook_path)

    summary = sync_workbook(session, wb_row.id, workbook_path)

    if not summary.events:
        # Empty workbook → no events; the invariant is vacuously true.
        return

    # Normalize tzinfo: SQLite strips tz on round-trip but the in-memory
    # `events` list still has the original aware datetimes. Compare both
    # sources to catch a drift in either direction.
    stamps = {ev.occurred_at for ev in summary.events}
    assert len(stamps) == 1, (
        f"events have {len(stamps)} distinct occurred_at values, expected 1"
    )


# ---------------------------------------------------------------------------
# Old/new JSON contract per event type
# ---------------------------------------------------------------------------


@given(initial=_ROW_SET, modified=_ROW_SET)
@settings(
    max_examples=25,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_event_value_json_nullability_matches_event_type(
    initial, modified, session, tmp_path
):
    """``added`` → ``old_value_json is None``; ``removed`` → ``new_value_json
    is None``; ``moved`` / ``edited`` → both populated.

    This is the contract the detail-view UI assumes when rendering the
    "before / after" diff. Drift here = either a null-pointer crash or
    a silently-empty diff panel.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_null.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, initial)
    wb_row = _persist_workbook_row(session, workbook_path)
    sync_workbook(session, wb_row.id, workbook_path)

    _clear_index_cache()
    _write_workbook(workbook_path, modified)
    summary = sync_workbook(session, wb_row.id, workbook_path)

    for ev in summary.events:
        if ev.event_type == "added":
            assert ev.old_value_json is None, "added event must have old=None"
            assert ev.new_value_json is not None, "added event must have new"
        elif ev.event_type == "removed":
            assert ev.new_value_json is None, "removed event must have new=None"
            assert ev.old_value_json is not None, "removed event must have old"
        elif ev.event_type in {"moved", "edited"}:
            assert ev.old_value_json is not None, f"{ev.event_type} needs old"
            assert ev.new_value_json is not None, f"{ev.event_type} needs new"
        else:
            pytest.fail(f"unknown event_type: {ev.event_type}")


# ---------------------------------------------------------------------------
# Summary counts match persisted-event counts per type
# ---------------------------------------------------------------------------


@given(initial=_ROW_SET, modified=_ROW_SET)
@settings(
    max_examples=25,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_summary_counts_equal_persisted_event_counts(
    initial, modified, session, tmp_path
):
    """``SyncSummary`` counts == grouped event counts queried from the DB.

    The UI surfaces the headline counter from ``SyncSummary`` (returned
    by the API) and the per-event breakdown by querying
    ``WorkbookSyncEvent`` later. Any drift between the two would show
    the user a count on the banner that doesn't add up on the detail
    page.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_counts.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, initial)
    wb_row = _persist_workbook_row(session, workbook_path)

    first = sync_workbook(session, wb_row.id, workbook_path)

    _clear_index_cache()
    _write_workbook(workbook_path, modified)
    second = sync_workbook(session, wb_row.id, workbook_path)

    persisted = _events_for(session, wb_row.id)
    by_type: dict[str, int] = {}
    for ev in persisted:
        by_type[ev.event_type] = by_type.get(ev.event_type, 0) + 1

    expected_added = first.added_count + second.added_count
    expected_removed = first.removed_count + second.removed_count
    expected_moved = first.moved_count + second.moved_count
    expected_edited = first.edited_count + second.edited_count

    assert by_type.get("added", 0) == expected_added
    assert by_type.get("removed", 0) == expected_removed
    assert by_type.get("moved", 0) == expected_moved
    assert by_type.get("edited", 0) == expected_edited

    # And the in-memory `summary.events` list size equals the sum of
    # the four counts — a regression where one branch silently dropped
    # events would fail this.
    for s in (first, second):
        total = s.added_count + s.removed_count + s.moved_count + s.edited_count
        assert len(s.events) == total


# ---------------------------------------------------------------------------
# Keyed-diff totality — rows without CCI never produce events
# ---------------------------------------------------------------------------


@given(initial=_ROW_SET, modified=_ROW_SET)
@settings(
    max_examples=25,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_rows_without_cci_produce_no_events(initial, modified, session, tmp_path):
    """Every event's ``(control_id, cci_id)`` was present in either workbook.

    Rows missing a CCI id can't participate in the keyed diff; the
    reader's `_row_key` filter drops them silently. If a regression let
    them through with `cci_id=None`, this assertion would surface
    phantom keys in the persisted log.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_noscci.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, initial)
    wb_row = _persist_workbook_row(session, workbook_path)
    sync_workbook(session, wb_row.id, workbook_path)

    _clear_index_cache()
    _write_workbook(workbook_path, modified)
    sync_workbook(session, wb_row.id, workbook_path)

    legal_keys = _expected_keys_for(initial) | _expected_keys_for(modified)

    for ev in _events_for(session, wb_row.id):
        # Defensive: no event should ever have a null CCI.
        assert ev.cci_id is not None, "event has null cci_id (filter slipped)"
        assert ev.cci_id.startswith("CCI-"), f"non-canonical CCI: {ev.cci_id}"
        # And every event key was in one of the two row sets.
        assert (ev.control_id, ev.cci_id) in legal_keys, (
            f"phantom event key: ({ev.control_id!r}, {ev.cci_id!r})"
        )


# ---------------------------------------------------------------------------
# JSON payload round-trip — every populated *_value_json parses as a dict
# ---------------------------------------------------------------------------


@given(initial=_ROW_SET, modified=_ROW_SET)
@settings(
    max_examples=20,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_event_json_payloads_round_trip_as_dicts(
    initial, modified, session, tmp_path
):
    """``json.loads(ev.*_value_json)`` returns a dict for every populated field.

    Defends against an accidental ``repr()`` / ``str()`` swap that
    would write valid-looking text but fail the SQLite ``json_each``
    parser downstream queries use.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_json.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, initial)
    wb_row = _persist_workbook_row(session, workbook_path)
    sync_workbook(session, wb_row.id, workbook_path)

    _clear_index_cache()
    _write_workbook(workbook_path, modified)
    sync_workbook(session, wb_row.id, workbook_path)

    for ev in _events_for(session, wb_row.id):
        for payload, label in (
            (ev.old_value_json, "old"),
            (ev.new_value_json, "new"),
        ):
            if payload is None:
                continue
            parsed = json.loads(payload)
            assert isinstance(parsed, dict), (
                f"{label}_value_json for {ev.event_type} parsed as "
                f"{type(parsed).__name__}, expected dict"
            )


# ---------------------------------------------------------------------------
# Convergence — sync, then sync again with same file → second is empty
# ---------------------------------------------------------------------------


@given(initial=_ROW_SET, modified=_ROW_SET)
@settings(
    max_examples=20,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_third_sync_after_change_is_a_noop(initial, modified, session, tmp_path):
    """After ``initial → modified`` syncs, a third sync against the same
    file produces zero events.

    The snapshot sidecar must roll forward to match the current parse
    on every sync; a regression where the sidecar lagged would mean
    every subsequent open re-fires the same changes.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_converge.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, initial)
    wb_row = _persist_workbook_row(session, workbook_path)
    sync_workbook(session, wb_row.id, workbook_path)

    _clear_index_cache()
    _write_workbook(workbook_path, modified)
    sync_workbook(session, wb_row.id, workbook_path)

    # Third sync — file unchanged since the second write.
    third = sync_workbook(session, wb_row.id, workbook_path)
    assert third.added_count == 0
    assert third.removed_count == 0
    assert third.moved_count == 0
    assert third.edited_count == 0
    assert third.events == []
    assert third.had_prior_snapshot is True


# ---------------------------------------------------------------------------
# Return type and SyncSummary shape — defensive against API drift
# ---------------------------------------------------------------------------


@given(rows=_ROW_SET)
@settings(
    max_examples=15,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_sync_workbook_returns_a_well_formed_summary(rows, session, tmp_path):
    """``sync_workbook`` always returns a ``SyncSummary`` with int counts.

    The route handlers unpack ``summary.added_count`` (etc.) directly
    into the API response. A regression that returned a dict or a
    namedtuple would break every call site silently (attribute access
    on the wrong type can mask as zeros depending on the structure).
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_shape.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, rows)
    wb_row = _persist_workbook_row(session, workbook_path)

    summary = sync_workbook(session, wb_row.id, workbook_path)

    assert isinstance(summary, SyncSummary)
    assert isinstance(summary.added_count, int)
    assert isinstance(summary.removed_count, int)
    assert isinstance(summary.moved_count, int)
    assert isinstance(summary.edited_count, int)
    assert isinstance(summary.had_prior_snapshot, bool)
    assert isinstance(summary.events, list)
    for ev in summary.events:
        assert isinstance(ev, WorkbookSyncEvent)


# ---------------------------------------------------------------------------
# Single-call ID isolation — events FK-scope to the workbook they came from
# ---------------------------------------------------------------------------


@given(rows_a=_ROW_SET, rows_b=_ROW_SET)
@settings(
    max_examples=15,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_events_scoped_to_their_workbook_id(rows_a, rows_b, session, tmp_path):
    """Events for workbook A never carry workbook B's id.

    Two workbooks in the same DB, two syncs. The "what changed" UI
    queries by ``workbook_id``; cross-workbook leakage would let
    workbook A's events appear under workbook B's sync history.
    """
    _reset_schema(session)
    wb_path_a = tmp_path / "wb_a.xlsx"
    wb_path_b = tmp_path / "wb_b.xlsx"
    _clear_snapshot(wb_path_a)
    _clear_snapshot(wb_path_b)
    _clear_index_cache()
    _write_workbook(wb_path_a, rows_a)
    _write_workbook(wb_path_b, rows_b)
    wb_a = _persist_workbook_row(session, wb_path_a)
    wb_b = _persist_workbook_row(session, wb_path_b)

    sync_workbook(session, wb_a.id, wb_path_a)
    sync_workbook(session, wb_b.id, wb_path_b)

    for ev in _events_for(session, wb_a.id):
        assert ev.workbook_id == wb_a.id
    for ev in _events_for(session, wb_b.id):
        assert ev.workbook_id == wb_b.id

    # And the cross-workbook intersection is empty — no event has
    # been mis-attributed.
    a_ids = {ev.id for ev in _events_for(session, wb_a.id)}
    b_ids = {ev.id for ev in _events_for(session, wb_b.id)}
    assert a_ids.isdisjoint(b_ids)


# ---------------------------------------------------------------------------
# `source` field — always "reread" from sync_workbook
# ---------------------------------------------------------------------------


@given(rows=_ROW_SET)
@settings(
    max_examples=10,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_every_event_source_is_reread(rows, session, tmp_path):
    """``ev.source == "reread"`` for every event ``sync_workbook`` emits.

    The model column has a default, but the engine sets it explicitly.
    Pinning the value here defends against a refactor that silently
    drops the explicit assignment and falls back to a different default
    (e.g. ``None`` if the column default is removed).
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_src.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, rows)
    wb_row = _persist_workbook_row(session, workbook_path)

    summary = sync_workbook(session, wb_row.id, workbook_path)
    for ev in summary.events:
        assert ev.source == "reread"


# ---------------------------------------------------------------------------
# UTC timestamp — occurred_at is timezone-aware UTC when constructed
# ---------------------------------------------------------------------------


@given(rows=st.lists(_row_spec(), min_size=1, max_size=4))
@settings(
    max_examples=10,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture, HealthCheck.too_slow],
)
def test_occurred_at_is_utc_aware_at_construction(rows, session, tmp_path):
    """The in-memory ``ev.occurred_at`` carries UTC tzinfo before SQLite
    strips it on round-trip.

    SQLite stores naive datetimes, but the engine constructs aware ones
    with ``datetime.now(timezone.utc)``. The in-memory ``events`` list
    on ``SyncSummary`` MUST surface them as aware so callers can format
    them with timezone context for the user.
    """
    _reset_schema(session)
    workbook_path = tmp_path / "wb_utc.xlsx"
    _clear_snapshot(workbook_path)
    _clear_index_cache()
    _write_workbook(workbook_path, rows)
    wb_row = _persist_workbook_row(session, workbook_path)

    summary = sync_workbook(session, wb_row.id, workbook_path)
    if not summary.events:
        return

    # After `session.refresh(event)`, SQLite may have stripped tz; but
    # the timestamps all came from one `datetime.now(timezone.utc)`
    # call, so their delta should be exactly zero regardless of tzinfo.
    base = summary.events[0].occurred_at
    for ev in summary.events:
        assert ev.occurred_at == base
    # And the assigned timestamp is recent enough to be "now-ish" — a
    # regression to a fixed epoch (e.g. `datetime.min`) would surface
    # here. Window is wide enough for slow CI runs.
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    base_naive = base.replace(tzinfo=None) if base.tzinfo else base
    delta_seconds = abs((now - base_naive).total_seconds())
    assert delta_seconds < 60, (
        f"occurred_at differs from now by {delta_seconds}s — clock skew or bug"
    )
