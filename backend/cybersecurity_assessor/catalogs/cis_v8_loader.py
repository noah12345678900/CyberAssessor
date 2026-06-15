"""CIS Controls v8 catalog loader (license-aware).

WHY THIS LOADER IS DIFFERENT FROM THE NIST LOADERS
--------------------------------------------------
The NIST OSCAL catalogs are U.S. Government public-domain content, so
``oscal_loader.py`` is free to download and bundle the real control text.

The CIS Controls v8 (and their Safeguards) are **copyrighted by the Center
for Internet Security (CIS)**. We may NOT bundle, fabricate, paraphrase, or
otherwise ship the real Safeguard text. Instead, an organization that has
lawfully obtained the CIS Controls supplies its own export, and this loader
reads that user-supplied file. With no path (or in offline mode) the loader
refuses to run and tells the user to supply their licensed export — it never
invents content to fill the gap.

THE LICENSED-IMPORT CONTRACT
----------------------------
``load_cis_v8_catalog`` accepts a user-supplied file ``path`` pointing at
either:

  * a ``.csv`` file (Excel-friendly — what assessors usually have),
  * a ``.json`` file (a list of objects), or
  * a ``.xlsx`` file — specifically the native CIS Controls workbook that CIS
    distributes to licensees ("CIS_Controls_Version_8.x.xlsx"). Its
    ``Controls vN`` sheet carries the columns ``CIS Control | CIS Safeguard |
    Asset Class | Security Function | Title | Description | IG1 | IG2 | IG3``
    and interleaves 18 parent-control rows (blank Safeguard) with the
    Safeguard rows. The reader normalizes BOTH into catalog rows so every
    Control AND every Safeguard is ingested: a parent row keys on its CIS
    Control number (e.g. ``"1"``); a Safeguard row keys on its Safeguard id
    (e.g. ``"1.1"``) with ``family`` = the parent control number.

Accepted column / field names are matched case-insensitively and are liberal
on input, strict on output:

  REQUIRED
    - id    : one of ``id`` | ``control_id`` | ``ref``
              (CIS Safeguard ids look like ``1.1`` or ``18.5`` — stored as given)
    - title : one of ``title`` | ``name``
    - text  : one of ``text`` | ``requirement`` | ``statement``
  OPTIONAL
    - family/category : one of ``family`` | ``category`` | ``theme`` | ``function``
                        (the parent CIS Control number or its name). Stored on
                        ``Control.family``. When ABSENT, ``family`` is derived
                        from the Safeguard id by splitting on ``"."`` and taking
                        the first segment (Safeguard ``1.1`` -> family ``"1"``).

A missing REQUIRED column raises a clear ``ValueError`` naming the field.

Reloads are idempotent: the Framework is upserted by ``(name, version)`` and
each Control by ``(framework_id, control_id)``.
"""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from sqlmodel import Session, select

from ..models import Control, Framework

# --- Framework identity (NOT copyrighted — these are bibliographic facts) ---
FRAMEWORK_NAME = "CIS Controls"
FRAMEWORK_VERSION = "v8"
FRAMEWORK_ID = "CIS-v8"

# Header aliases — lower-cased on lookup. Order = preference.
_ID_KEYS = ("id", "control_id", "ref")
_TITLE_KEYS = ("title", "name")
_TEXT_KEYS = ("text", "requirement", "statement")
_FAMILY_KEYS = ("family", "category", "theme", "function")

_LICENSE_MSG = (
    "CIS Controls v8 (and their Safeguards) are copyrighted by the Center for "
    "Internet Security and cannot be bundled or downloaded by this "
    "application. To load this framework you must supply a licensed export "
    "your organization already owns: pass path=<file> pointing at a CSV or "
    "JSON file of the Safeguards (columns: id/control_id/ref, title/name, "
    "text/requirement/statement, and optionally "
    "family/category/theme/function). Offline mode cannot satisfy this "
    "requirement — there is no public copy to fall back on."
)


def _pick(row: dict[str, Any], keys: tuple[str, ...]) -> str | None:
    """Return the first present, non-empty value among ``keys`` (case-insensitive)."""
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        if key in lowered:
            val = lowered[key]
            if val is None:
                continue
            text = str(val).strip()
            if text:
                return text
    return None


def _present_header(headers: set[str], keys: tuple[str, ...]) -> bool:
    """True if any alias in ``keys`` appears among ``headers`` (case-insensitive)."""
    lowered = {h.strip().lower() for h in headers}
    return any(k in lowered for k in keys)


def _family_from_safeguard_id(control_id: str) -> str:
    """Safeguard '1.1' -> parent CIS Control '1'.

    Splits on the first '.' and returns the leading segment. Ids with no '.'
    (a bare CIS Control number) map to themselves.
    """
    return control_id.split(".", 1)[0].strip()


def _leading_control_number(value: Any) -> str:
    """Extract the leading CIS Control integer from a cell value.

    The native CIS workbook decorates parent-control cells with a trailing
    footnote/marker glyph (e.g. ``"1\ufffd"``), while Safeguard rows carry a
    clean integer. Both normalize to the bare control number (``"1"``). Returns
    ``""`` if the cell has no leading digits.
    """
    m = re.match(r"\s*(\d+)", str(value))
    return m.group(1) if m else ""


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    """Read the native CIS Controls workbook into normalized row dicts.

    Locates the Safeguards sheet (the one whose header row contains a
    "CIS Safeguard" column, falling back to a sheet named like ``Controls``),
    maps columns by header name (case-insensitive), and emits one normalized
    dict per data row with keys ``id``/``title``/``text``/``family`` so the
    shared upsert path treats it identically to a CSV/JSON export.

    Both row kinds are emitted:
      * parent control row (blank Safeguard) -> id = control number, family = same
      * safeguard row                        -> id = safeguard id, family = control number
    """
    # openpyxl is a hard dependency of the sidecar; import lazily so the rest
    # of the loader stays usable even in a stripped environment.
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = None
        for cand in wb.worksheets:
            # Prefer a sheet that actually has the Safeguard header.
            for row in cand.iter_rows(min_row=1, max_row=1, values_only=True):
                hdr = {str(c).strip().lower() for c in row if c is not None}
                if "cis safeguard" in hdr:
                    ws = cand
                    break
            if ws is not None:
                break
        if ws is None:
            for cand in wb.worksheets:
                if cand.title.strip().lower().startswith("controls"):
                    ws = cand
                    break
        if ws is None:
            raise ValueError(
                "CIS Controls workbook has no recognizable Safeguards sheet "
                "(expected a 'Controls vN' sheet with a 'CIS Safeguard' column)."
            )

        # Header → column-index map (case-insensitive, first match wins).
        header_map: dict[str, int] = {}
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for idx, cell in enumerate(row):
                if cell is None:
                    continue
                key = str(cell).strip().lower()
                if key and key not in header_map:
                    header_map[key] = idx
            break

        col_control = header_map.get("cis control")
        col_safeguard = header_map.get("cis safeguard")
        col_title = header_map.get("title")
        col_desc = header_map.get("description")

        def _at(row: tuple, idx: int | None):
            """Return the openpyxl cell at ``idx`` (or None if out of range)."""
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        def _cell_str(cell) -> str:
            if cell is None or cell.value is None:
                return ""
            return str(cell.value).strip()

        def _safeguard_id(cell) -> str:
            """Recover the CIS Safeguard id, preserving a trailing zero.

            openpyxl reads a numeric safeguard cell like ``3.10`` as the float
            ``3.1`` — Excel does not store the trailing zero in the cell value —
            which would COLLIDE with the genuine Safeguard ``3.1`` and silently
            drop one of the two on upsert (keyed by control_id). The workbook
            keeps the distinction in the cell's NUMBER FORMAT: the ``.10``
            safeguards carry a two-decimal format (``0.00`` / ``#,##0.00``)
            while every other safeguard is ``General``. We use that signal to
            restore the lost zero so ``3.1`` and ``3.10`` stay distinct.
            """
            if cell is None or cell.value is None:
                return ""
            val = cell.value
            if isinstance(val, str):
                return val.strip()
            if isinstance(val, (int, float)):
                fmt = cell.number_format or ""
                if "0.00" in fmt:
                    return f"{val:.2f}"  # 3.1 -> '3.10'
                # General: strip float noise/trailing zeros (3.1 -> '3.1').
                return f"{val:.10f}".rstrip("0").rstrip(".")
            return str(val).strip()

        rows: list[dict[str, Any]] = []
        first = True
        for row in ws.iter_rows():
            if first:
                first = False  # skip header row
                continue
            control_num = _leading_control_number(_cell_str(_at(row, col_control)))
            safeguard = _safeguard_id(_at(row, col_safeguard))
            title = _cell_str(_at(row, col_title))
            text = _cell_str(_at(row, col_desc))

            if safeguard:
                rid = safeguard
                family = control_num or _family_from_safeguard_id(safeguard)
            elif control_num:
                rid = control_num
                family = control_num
            else:
                continue  # neither id present — not a control/safeguard row

            rows.append(
                {"id": rid, "title": title, "text": text, "family": family}
            )
        return rows
    finally:
        wb.close()


def _read_rows(path: Path) -> list[dict[str, Any]]:
    """Read a user-supplied CSV, JSON, or native CIS XLSX into row dicts.

    CSV: first line is the header. JSON: a list of objects (each object is a
    row). XLSX: the native CIS Controls workbook (see :func:`_read_xlsx_rows`).
    All are read as UTF-8. Required-column validation happens against the union
    of headers so an empty data file still names the missing field.
    """
    if not path.exists():
        raise FileNotFoundError(f"Licensed CIS Controls v8 export not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = _read_xlsx_rows(path)
        if not rows:
            raise ValueError(
                "CIS Controls workbook parsed to zero rows — the Safeguards "
                "sheet may be empty or in an unexpected layout."
            )
        return rows
    if suffix == ".json":
        raw = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(raw, list):
            raise ValueError(
                "CIS Controls v8 JSON export must be a list of safeguard objects."
            )
        rows = [dict(r) for r in raw if isinstance(r, dict)]
        headers: set[str] = set()
        for r in rows:
            headers.update(str(k) for k in r.keys())
    elif suffix == ".csv":
        with path.open("r", encoding="utf-8", newline="") as fh:
            reader = csv.DictReader(fh)
            headers = set(reader.fieldnames or [])
            rows = [dict(r) for r in reader]
    else:
        raise ValueError(
            f"Unsupported licensed-export format {suffix!r}; supply a .csv or .json file."
        )

    # Strict required-column validation, liberal alias matching.
    if not _present_header(headers, _ID_KEYS):
        raise ValueError(
            "Licensed CIS Controls v8 export is missing the required 'id' column "
            f"(accepted: {', '.join(_ID_KEYS)})."
        )
    if not _present_header(headers, _TITLE_KEYS):
        raise ValueError(
            "Licensed CIS Controls v8 export is missing the required 'title' column "
            f"(accepted: {', '.join(_TITLE_KEYS)})."
        )
    if not _present_header(headers, _TEXT_KEYS):
        raise ValueError(
            "Licensed CIS Controls v8 export is missing the required 'text' column "
            f"(accepted: {', '.join(_TEXT_KEYS)})."
        )
    return rows


def load_cis_v8_catalog(
    session: Session,
    *,
    path: str | Path | None = None,
    offline: bool = False,
) -> Framework:
    """Idempotently load CIS Controls v8 Safeguards from a user-supplied export.

    Args:
        session: an active SQLModel Session.
        path: REQUIRED — a local CSV or JSON file containing the licensed CIS
            Controls v8 Safeguards the organization already owns. There is no
            download fallback because the content is copyrighted.
        offline: accepted for signature parity with the NIST loaders. Because
            there is never a network source for this framework, ``offline=True``
            is treated the same as ``path=None``: it raises with the licensing
            guidance.

    Returns:
        The Framework row (created or updated).

    Raises:
        ValueError: if ``path`` is None or ``offline`` is True (licensing
            guard), or if a required column is missing.
        FileNotFoundError: if the supplied path does not exist.
    """
    if path is None or offline:
        raise ValueError(_LICENSE_MSG)

    file_path = Path(path)
    rows = _read_rows(file_path)

    # --- Framework upsert by (name, version) -------------------------------
    framework = session.exec(
        select(Framework).where(
            Framework.name == FRAMEWORK_NAME,
            Framework.version == FRAMEWORK_VERSION,
        )
    ).first()
    if framework is None:
        framework = Framework(
            name=FRAMEWORK_NAME,
            version=FRAMEWORK_VERSION,
            framework_id=FRAMEWORK_ID,
            oscal_uri=None,
            parent_framework_id=None,
            # enabled left untouched — defaults True.
        )
        session.add(framework)
        session.commit()
        session.refresh(framework)
    elif framework.framework_id is None:
        framework.framework_id = FRAMEWORK_ID
        session.add(framework)
        session.commit()
        session.refresh(framework)

    # --- Control upsert by (framework_id, control_id) ----------------------
    existing = {
        c.control_id: c
        for c in session.exec(
            select(Control).where(Control.framework_id == framework.id)
        ).all()
    }

    for row in rows:
        control_id = _pick(row, _ID_KEYS)
        if not control_id:
            # Row with no id — skip rather than persist an unkeyed safeguard.
            continue
        title = _pick(row, _TITLE_KEYS) or ""
        statement = _pick(row, _TEXT_KEYS)
        family = _pick(row, _FAMILY_KEYS)
        if not family:
            family = _family_from_safeguard_id(control_id)

        existing_row = existing.get(control_id)
        if existing_row is None:
            existing_row = Control(
                framework_id=framework.id,  # type: ignore[arg-type]
                control_id=control_id,
                title=title,
                family=family,
                statement=statement,
            )
            session.add(existing_row)
            existing[control_id] = existing_row
        else:
            existing_row.title = title
            existing_row.family = family
            existing_row.statement = statement
            session.add(existing_row)

    session.commit()
    session.refresh(framework)
    return framework
