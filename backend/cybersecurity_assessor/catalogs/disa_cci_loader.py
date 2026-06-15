"""CCI catalog loader.

Background — why this file exists
---------------------------------
DoD eMASS / CCIS workbooks are organized around **Control Correlation
Identifiers (CCIs)**: roughly ~3500 assessment-objective rows, each one
decomposing a NIST 800-53 control into a single testable statement
(``CCI-000001`` through ``CCI-00xxxx``). The workbook rows are the unit
the assessor scores -- not the parent control. Without the CCI catalog
loaded, controls in this app show ``0 CCIs`` and the Control Detail
page has nothing to assess against.

For ~15 years the canonical source was DISA's ``U_CCI_List.xml``,
distributed as a zip from ``https://public.cyber.mil/stigs/cci/``.
**That is no longer the case.** DISA discontinued the standalone CCI
list distribution. Current authoritative sources, in priority order:

1. **NIST CSRC mapping spreadsheet** -- the same CCI data NIST now
   curates as ``stig-mapping-to-nist-800-53.xlsx`` at
   ``https://csrc.nist.gov/csrc/media/projects/forum/documents/``.
   Public, no CAC, no zip step. **This is the recommended source.**
2. **STIG exports** -- per-product CCI references emitted by STIG
   Viewer (``.ckl`` / ``.cklb``). Use these if you only need the CCIs
   relevant to one product; they will not cover the full catalog.
3. **Archived ``U_CCI_List.xml``** -- still valid input if you have a
   copy from before DISA pulled the page. The XML schema didn't
   change; old archives are stable.

This loader accepts (1) and (3) transparently -- it sniffs the file
extension (``.xlsx`` vs ``.xml``) and dispatches to the right parser.
Both parsers normalize to the same internal ``_CciItem`` shape, so the
upsert path that writes ``Objective`` rows is shared.

(2) -- STIG exports -- is **not** handled here; that ingest belongs in
the STIG evidence pipeline (``cybersecurity_assessor.excel.stig_*``),
not in the catalog layer.

What this gets us
-----------------
Without this loader, ``Objective`` rows are populated lazily from
whatever CCIs the opened CCIS workbook happens to contain (~319 for
a typical Example System baseline). That works for the loaded system but
leaves the catalog incomplete: cross-system reuse, framework
switching, and the rev4-vs-rev5 crosswalk all want the full ~3500-row
CCI set indexed ahead of time. Loading the full catalog up front also
means a freshly-installed app has the right shape from first launch
instead of "grow into the truth as workbooks are opened."

Both this loader and the workbook reader upsert by
``(control_id_fk, objective_id)``, so they cooperate without conflict:
this fills in canonical CCI text + deprecation status;
``populate_objectives`` adds workbook-only enrichment
(``implementation_guidance``, ``assessment_procedures`` from cols J/K).

What this skips
---------------
- CCIs marked ``deprecated``: still upserted but tagged
  ``source="CCI-deprecated"`` so the UI can hide them. Old workbooks
  may still reference them, so we don't drop the rows entirely.
- CCIs whose only references are non-NIST policy docs (the source
  also catalogs DoDI 8500.01 mappings). We only attach to ``Control``
  rows we recognize for the requested framework.
- When multiple NIST references exist for a CCI (rev3 / rev4 / rev5
  historical), we keep the highest revision. ``_pick_best_nist_index``
  enforces this for both parser paths.
"""

from __future__ import annotations

import logging
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

from sqlalchemy import select as sa_select
from sqlmodel import Session, select

log = logging.getLogger(__name__)

from ..models import (
    Assessment,
    BaselineObjective,
    Control,
    Crosswalk,
    EvidenceTag,
    Framework,
    Objective,
    PoamObjective,
    RequirementMap,
)

# DISA's XML uses a default namespace; ElementTree needs it spelled out.
_NS = {"cci": "http://iase.disa.mil/cci"}

# NIST CSRC xlsx column headers (row 2 of the workbook -- row 1 is blank).
# Hard-coded against the published spec so a column reorder breaks loudly
# at parse time instead of silently wiring "definition" into "creator".
_XLSX_SHEET = "U_CCI_List"
_XLSX_HEADER_ROW = 2
_XLSX_COL_CCI = "CCI"
_XLSX_COL_DEFINITION = "/cci_items/cci_item/definition"
_XLSX_COL_INDEX = "index"   # control reference, e.g. "AC-2 (1)" or "AC-1 b 1"
_XLSX_COL_TITLE = "title"   # publication, e.g. "NIST SP 800-53 Revision 4"
_XLSX_COL_STATUS = "status"


@dataclass
class DisaCciLoadResult:
    framework_id: int
    cci_items_in_xml: int  # name kept for API stability; counts items from xlsx too
    cci_items_attached: int  # matched a Control row
    cci_items_unmatched: int  # referenced an 800-53 control not in the catalog
    cci_items_no_nist_ref: int  # only referenced non-NIST policy docs
    cci_items_deprecated: int
    objectives_created: int
    objectives_updated: int
    # CCIs that HAVE a NIST 800-53 reference but only for a revision other
    # than the target. Distinct from cci_items_no_nist_ref (which means
    # "no 800-53 ref at all" — only DoDI / non-NIST policy refs).
    # Zero when no target_revision filter is in effect.
    cci_items_below_revision: int = 0
    # Stub cleanup counters — see the post-loop subtractive delete in
    # ``load_disa_cci_catalog`` for why both are needed.
    objectives_removed: int = 0
    objectives_preserved_with_refs: int = 0


@dataclass
class _CciItem:
    cci_id: str  # "CCI-000001"
    status: str  # "draft" | "published" | "deprecated"
    definition: str
    # Raw (title, index) reference pairs. Picking the right one is deferred
    # to load time so the target-revision filter (which is framework-dependent,
    # not parser-dependent) can be applied with full ref context.
    refs: list[tuple[str, str]]


def _ccis_ref_to_oscal_control_id(raw: str) -> str:
    """'AC-2 (1)' / 'AC-2(1)' -> 'ac-2.1' (matches OSCAL canonical form).

    Mirrors :func:`cybersecurity_assessor.excel.ccis_reader._ccis_to_oscal_control_id`
    -- duplicated here on purpose so this loader has no excel/ dependency
    (the catalog layer should be importable without openpyxl available).

    Note: the NIST CSRC xlsx ``index`` column carries fine-grained refs
    like ``"AC-1 a"`` or ``"AC-1 b 1"`` for the item-letter callouts.
    We strip those down to the base control id; the item-letter detail
    lives in the CCI ``definition`` text itself, not in the FK.
    """
    s = raw.strip().lower()
    # Drop any trailing item-letter / sub-letter callouts after the base
    # control id (e.g. "ac-1 a" -> "ac-1", "ac-1 b 1" -> "ac-1"). Catch
    # this BEFORE the enhancement-paren collapse so "ac-2 (1) a" still
    # becomes "ac-2.1" rather than just "ac-2".
    #
    # Do NOT use `\b` to anchor the tail of this match: `)` and EOL are
    # both non-word characters, so there's no word boundary after `(1)`,
    # the optional group never fires, and "ac-2 (1)" collapses to "ac-2".
    # Plain greedy match + no anchor is correct here -- regex stops where
    # the capture ends and the trailing " a" / " b 1" is discarded.
    s = re.sub(r"\s+", " ", s).strip()
    m = re.match(r"^([a-z]+-\d+(?:\s*\(\d+\))?)", s)
    if m:
        s = m.group(1)
    s = re.sub(r"\s+", "", s)  # "ac-2 (1)" -> "ac-2(1)"
    s = re.sub(r"\((\d+)\)", r".\1", s)  # "ac-2(1)" -> "ac-2.1"
    return s


def _rev_from_title(title: str) -> int:
    """Pull the revision number out of an 800-53 reference title.

    "NIST SP 800-53 Revision 4" -> 4
    "NIST SP 800-53"             -> 0  (untagged baseline; lose to any rev)
    "NIST SP 800-53A Revision 1" -> -1 (assessment guide, not the catalog)

    Returning -1 for 800-53A is how we exclude assessment-procedure refs
    from "best 800-53 ref" picking without scattering ``"800-53A" in title``
    guards through the rest of the file.
    """
    t = title.strip()
    if "800-53A" in t or "800-53a" in t:
        return -1
    if "800-53" not in t:
        return -1
    m = re.search(r"Revision\s+(\d+)", t, flags=re.I)
    return int(m.group(1)) if m else 0


def _pick_best_nist_index(
    refs: list[tuple[str, str]],
    *,
    target_revision: int | None = None,
) -> str | None:
    """From ``[(title, index), ...]`` pick the 800-53 control reference.

    Behavior:
      * ``target_revision=None`` (default) — pick the highest-rev 800-53
        reference present (legacy behavior, framework-agnostic).
      * ``target_revision=N`` — require an EXACT match on Revision N.
        Older refs (rev=0 unrevisioned Rev3-era, rev=3) and newer refs
        (Rev5+) are ignored. This is how we filter Rev3-only CCIs out
        of a Rev4 framework: their only 800-53 refs are unrevisioned
        ``"NIST SP 800-53"`` (rev=0) + ``"NIST SP 800-53A"`` (rev=-1),
        neither of which matches ``target_revision=4``.

    Returns the raw ``index`` string (e.g. ``"AC-2 (1)"``), or ``None`` if
    no qualifying NIST 800-53 reference is present. Caller is responsible
    for normalizing through :func:`_ccis_ref_to_oscal_control_id`.
    """
    best: tuple[int, str] | None = None
    for title, index in refs:
        if not index:
            continue
        rev = _rev_from_title(title)
        if rev < 0:
            continue  # 800-53A or unrelated policy
        if target_revision is not None:
            if rev != target_revision:
                continue
            return index  # exact match — no need to keep scanning
        if best is None or rev > best[0]:
            best = (rev, index)
    return best[1] if best else None


# ---------------------------------------------------------------------------
# XML parser -- handles archived U_CCI_List.xml
# ---------------------------------------------------------------------------


def _parse_cci_xml(xml_path: Path) -> list[_CciItem]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    items: list[_CciItem] = []
    for item in root.findall("cci:cci_items/cci:cci_item", _NS):
        cci_id = (item.get("id") or "").strip()
        if not cci_id:
            continue
        status_el = item.find("cci:status", _NS)
        status = (status_el.text or "").strip().lower() if status_el is not None else ""
        def_el = item.find("cci:definition", _NS)
        definition = (def_el.text or "").strip() if def_el is not None else ""

        refs: list[tuple[str, str]] = []
        for ref in item.findall("cci:references/cci:reference", _NS):
            refs.append(
                ((ref.get("title") or "").strip(), (ref.get("index") or "").strip())
            )

        items.append(
            _CciItem(
                cci_id=cci_id,
                status=status or "unknown",
                definition=definition,
                refs=refs,
            )
        )
    return items


# ---------------------------------------------------------------------------
# xlsx parser -- handles the NIST CSRC stig-mapping-to-nist-800-53.xlsx,
# which is the active source as of June 2026 now that DISA pulled the
# standalone XML from public.cyber.mil.
# ---------------------------------------------------------------------------


def _parse_cci_xlsx(xlsx_path: Path) -> list[_CciItem]:
    """Parse the NIST CSRC CCI-to-800-53 mapping spreadsheet.

    Shape we expect:
      - Single sheet named ``U_CCI_List`` (or first sheet if mis-named)
      - Row 1 blank, row 2 column headers, row 3+ data
      - One row per (CCI, reference) -- so a single CCI appears multiple
        times when it has refs for rev3 / rev4 / rev5 / 800-53A
      - Columns we care about: ``CCI``, ``/cci_items/cci_item/definition``,
        ``index``, ``title``, ``status``

    We import openpyxl lazily so the catalog layer stays importable even
    if openpyxl isn't installed (e.g. a future trimmed-down deployment).
    """
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as e:  # pragma: no cover - openpyxl is a runtime dep
        raise RuntimeError(
            "openpyxl is required to load the NIST CSRC CCI xlsx. "
            "Install it with `uv add openpyxl` in the backend project."
        ) from e

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[_XLSX_SHEET] if _XLSX_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    # Skip leading blanks until we hit the header row.
    header: tuple | None = None
    seen_rows = 0
    for r in rows:
        seen_rows += 1
        if any(cell is not None and str(cell).strip() for cell in r):
            header = r
            break
    if header is None:
        raise ValueError("xlsx appears empty -- no header row found")

    # Build column-name -> column-index map; fail loud if the schema shifted.
    col_idx: dict[str, int] = {}
    for i, name in enumerate(header):
        if name is None:
            continue
        col_idx[str(name).strip()] = i

    missing = [
        c
        for c in (_XLSX_COL_CCI, _XLSX_COL_DEFINITION, _XLSX_COL_INDEX, _XLSX_COL_TITLE)
        if c not in col_idx
    ]
    if missing:
        raise ValueError(
            "xlsx is missing expected columns: "
            f"{missing}. Saw columns: {sorted(col_idx)}. "
            "If NIST changed the spreadsheet shape, update _XLSX_COL_* constants."
        )

    # Group rows by CCI -- the xlsx has one row per reference, so the same
    # CCI can repeat 3-5 times.
    grouped: dict[str, dict] = {}
    for r in rows:
        cci_raw = r[col_idx[_XLSX_COL_CCI]]
        if cci_raw is None:
            continue
        cci_id = str(cci_raw).strip()
        if not cci_id or not cci_id.upper().startswith("CCI-"):
            continue

        bucket = grouped.setdefault(
            cci_id,
            {
                "status": "unknown",
                "definition": "",
                "refs": [],  # list[(title, index)]
            },
        )

        # Status / definition are the same across all rows for one CCI;
        # take the first non-empty value we see (read-only mode preserves
        # row order).
        status_raw = r[col_idx[_XLSX_COL_STATUS]] if _XLSX_COL_STATUS in col_idx else None
        if status_raw and bucket["status"] == "unknown":
            bucket["status"] = str(status_raw).strip().lower()

        def_raw = r[col_idx[_XLSX_COL_DEFINITION]]
        if def_raw and not bucket["definition"]:
            bucket["definition"] = str(def_raw).strip()

        title_raw = r[col_idx[_XLSX_COL_TITLE]]
        index_raw = r[col_idx[_XLSX_COL_INDEX]]
        title = str(title_raw).strip() if title_raw else ""
        index = str(index_raw).strip() if index_raw else ""
        if title or index:
            bucket["refs"].append((title, index))

    items: list[_CciItem] = []
    for cci_id, b in grouped.items():
        items.append(
            _CciItem(
                cci_id=cci_id,
                status=b["status"] or "unknown",
                definition=b["definition"],
                refs=b["refs"],
            )
        )
    return items


# ---------------------------------------------------------------------------
# Format dispatcher + public API
# ---------------------------------------------------------------------------


def _parse_cci_source(path: Path) -> list[_CciItem]:
    """Pick parser by extension. Raises ``ValueError`` for unknown formats."""
    suffix = path.suffix.lower()
    if suffix in (".xml",):
        return _parse_cci_xml(path)
    if suffix in (".xlsx",):
        return _parse_cci_xlsx(path)
    raise ValueError(
        f"Unrecognized CCI source extension {suffix!r} (path={path}). "
        "Expected .xlsx (NIST CSRC stig-mapping-to-nist-800-53.xlsx) "
        "or .xml (archived DISA U_CCI_List.xml)."
    )


def load_disa_cci_catalog(
    session: Session,
    *,
    source_path: str | Path | None = None,
    framework_id: int,
    xml_path: str | Path | None = None,  # legacy kwarg name -- pre-xlsx-support
    target_revision: int | None = None,
) -> DisaCciLoadResult:
    """Upsert every NIST-referencing CCI from the source into ``Objective``.

    Args:
        session: active SQLModel session.
        source_path: local path to either the NIST CSRC
            ``stig-mapping-to-nist-800-53.xlsx`` (preferred) or an
            archived DISA ``U_CCI_List.xml``. The format is sniffed by
            extension.
        framework_id: which loaded Framework to attach CCIs to. We don't
            auto-pick because rev4 vs rev5 catalogs can coexist; the
            caller decides which one this CCI list maps to.
        xml_path: deprecated alias for ``source_path`` (kept so callers
            from before xlsx support don't break). Ignored if
            ``source_path`` is also passed.
        target_revision: when set, only import CCIs whose 800-53 reference
            EXACTLY matches this revision. Filters out Rev3-only legacy
            CCIs (whose only refs are unrevisioned ``"NIST SP 800-53"``)
            from a Rev4 framework load. ``None`` = legacy behavior:
            accept whatever the highest-rev ref is. The route layer
            derives this from ``Framework.version`` ("Rev 4" -> 4).

    Returns:
        Counts the UI can show after import.
    """
    framework = session.get(Framework, framework_id)
    if framework is None:
        raise ValueError(f"Framework id={framework_id} does not exist")

    chosen = source_path or xml_path
    if chosen is None:
        raise ValueError("source_path is required")
    items = _parse_cci_source(Path(chosen))

    # CCIs are revision-AGNOSTIC identifiers -- there is no such thing as a
    # "Rev 5 CCI". A given CCI carries whatever 800-53 control references the
    # source file happens to encode, and the canonical NIST CSRC
    # ``stig-mapping-to-nist-800-53.xlsx`` only carries **Revision 4** index
    # refs (plus unrevisioned + 800-53A). The ``target_revision`` EXACT-match
    # filter exists for one narrow job: keep Rev3-only legacy stubs OUT of a
    # Rev4 load when a newer ref set is also present. It must NEVER be the
    # reason a load matches zero CCIs.
    #
    # Regression this guards against: loading the Rev4-only CSRC file into a
    # Rev 5 framework derived ``target_revision=5``; no item has an exact Rev5
    # ref, so every CCI was filtered out, ``keep_keys`` came back empty, and
    # the subtractive cleanup then tried to delete the framework's ENTIRE
    # objective set (including workbook CCIs with live baseline/assessment FKs)
    # -> FK IntegrityError. The fix: if the source contains no refs at the
    # requested revision, the filter is inapplicable -- fall back to legacy
    # highest-revision matching (the behavior that worked before).
    if target_revision is not None:
        available_revs = {
            _rev_from_title(title)
            for it in items
            for title, index in it.refs
            if index and _rev_from_title(title) >= 0
        }
        if target_revision not in available_revs:
            log.warning(
                "CCI source %s has no Revision %s control references "
                "(available revisions: %s). CCIs are revision-agnostic, so "
                "falling back to highest-revision matching instead of "
                "excluding the entire catalog.",
                Path(chosen).name,
                target_revision,
                sorted(available_revs) or "none",
            )
            target_revision = None

    # Build control_id -> Control.id lookup for this framework only.
    controls: dict[str, int] = {}
    for c in session.exec(
        select(Control).where(Control.framework_id == framework_id)
    ).all():
        if c.id is not None:
            controls[c.control_id] = c.id

    # Existing Objectives in this framework, keyed by (control_id_fk, objective_id).
    existing: dict[tuple[int, str], Objective] = {}
    for o in session.exec(
        select(Objective).where(
            Objective.control_id_fk.in_(list(controls.values()))  # type: ignore[attr-defined]
        )
    ).all():
        existing[(o.control_id_fk, o.objective_id)] = o

    attached = 0
    unmatched = 0
    no_nist = 0
    below_revision = 0
    deprecated = 0
    created = 0
    updated = 0
    # Every (control_pk, cci_id) that this run legitimately wrote — both
    # newly-created and refreshed-existing. Anything in ``existing`` with
    # source CCI/CCI-deprecated that's NOT in this set is a stub from a
    # prior import where the filter was looser (or the CCI was Rev3-only
    # and shouldn't be on this framework). The post-loop cleanup uses
    # this set to decide what to delete.
    keep_keys: set[tuple[int, str]] = set()

    # NOTE: do NOT branch on ``status == "draft"`` here. In the NIST CSRC
    # ``stig-mapping-to-nist-800-53.xlsx`` (the active source as of 2026),
    # ~3,480 of ~3,633 CCIs are tagged ``draft`` — essentially the entire
    # active catalog including CCI-000015 etc. that eMASS exports without
    # issue. ``status`` in this file is effectively a default value and
    # does not signal supersession. A brief earlier branch that skipped
    # drafts dropped ~96% of the catalog; reverted. The actual signal
    # for "in eMASS workbook vs. legacy" is whether the CCI has a
    # ``NIST SP 800-53 Revision N`` reference — pre-Rev4 CCIs only carry
    # the unrevisioned ``NIST SP 800-53`` title and end up as legitimate
    # catalog-only rows. We import them all; the UI shows which are not
    # in the user's workbook based on workbook-side bookkeeping.
    for item in items:
        if item.status == "deprecated":
            deprecated += 1
            source_tag = "CCI-deprecated"
        else:
            source_tag = "CCI"

        # Pick the 800-53 ref, respecting the framework-revision filter
        # if the caller set one. Bucket failures into:
        #   below_revision — CCI has an 800-53 ref but only for the wrong
        #     revision (e.g. Rev3-only when loading into Rev4). Distinct
        #     from no_nist so the UI can show "filtered N Rev<x> stubs"
        #     vs. "N CCIs reference only DoDI policy".
        #   no_nist       — CCI has no 800-53 ref at all under any revision
        #     (only DoDI 8500.01 or other non-NIST policy refs).
        nist_raw = _pick_best_nist_index(item.refs, target_revision=target_revision)
        if nist_raw is None:
            if (
                target_revision is not None
                and _pick_best_nist_index(item.refs) is not None
            ):
                below_revision += 1
            else:
                no_nist += 1
            continue
        nist_canon = _ccis_ref_to_oscal_control_id(nist_raw)

        control_pk = controls.get(nist_canon)
        if control_pk is None:
            unmatched += 1
            continue

        attached += 1
        key = (control_pk, item.cci_id)
        keep_keys.add(key)
        obj = existing.get(key)
        if obj is None:
            session.add(
                Objective(
                    control_id_fk=control_pk,
                    objective_id=item.cci_id,
                    source=source_tag,
                    text=item.definition,
                )
            )
            created += 1
        else:
            # Don't clobber workbook-enriched fields (implementation_guidance,
            # assessment_procedures) -- only refresh the canonical CCI text
            # and the source tag (in case it flipped to deprecated).
            obj.text = item.definition or obj.text
            obj.source = source_tag
            session.add(obj)
            updated += 1

    session.commit()

    # ---- Subtractive cleanup of stale stubs ---------------------------------
    # Why this is needed even though we already filter at pick-time:
    # ``target_revision`` prevents NEW Rev3-only stubs from being inserted on
    # this import, but the loader is otherwise pure-upsert and will never
    # touch rows from PRIOR imports. Concrete case: 6 CCIs on AC-2
    # (CCI-000009 / 000013 / 000014 / 000237 / 001354 / 001355) were
    # imported by an earlier permissive run, the user's workbook doesn't
    # reference them, and re-importing did nothing because the upsert path
    # never visits them. This block walks every existing CCI Objective in
    # the framework, deletes the ones this run did NOT touch — unless they
    # have outgoing FK references (Assessment/Baseline/Crosswalk/
    # RequirementMap/EvidenceTag/PoamObjective), in which case we leave them
    # in place to avoid breaking in-flight assessments. All Objective FKs
    # are ON DELETE NO ACTION in SQLite, so we MUST check before deleting.
    removed = 0
    preserved_with_refs = 0
    candidates = [
        (pk, obj_id, obj)
        for (pk, obj_id), obj in existing.items()
        if (pk, obj_id) not in keep_keys
        and (obj.source or "").startswith("CCI")
    ]
    # Safety guard: a load that legitimately wrote ZERO objectives must never
    # trigger the subtractive cleanup. An empty ``keep_keys`` means "this run
    # matched nothing" -- treating every existing objective as a stale stub and
    # deleting the whole framework is never the intent, and is pure data-loss
    # risk (it's exactly what the Rev5/Rev4-source mismatch used to do before
    # the revision fallback above). With the fallback in place ``keep_keys`` is
    # populated for the real-world case; this guard covers a genuinely empty or
    # malformed source.
    if candidates and not keep_keys:
        log.error(
            "CCI cleanup skipped: load matched 0 CCIs but %d existing "
            "objectives would have been deleted. Refusing to wipe the "
            "framework catalog from an empty/non-matching source.",
            len(candidates),
        )
        candidates = []

    if candidates:
        candidate_ids = [obj.id for _, _, obj in candidates if obj.id is not None]
        # Bulk-fetch the FK-referencing tables once instead of per-candidate
        # round-trips; with ~3500 CCIs the per-row N+1 would dominate.
        referenced: set[int] = set()
        if candidate_ids:
            for col in (
                Assessment.objective_id,
                BaselineObjective.objective_id,
                Crosswalk.from_objective_id,
                Crosswalk.to_objective_id,
                RequirementMap.objective_id,
                EvidenceTag.objective_id,
                PoamObjective.objective_id,
            ):
                rows = session.exec(
                    sa_select(col).where(col.in_(candidate_ids)).distinct()
                ).all()
                # CRITICAL: ``session.exec(sa_select(single_col))`` yields
                # SQLAlchemy ``Row`` objects, and in SQLAlchemy 2.0 ``Row`` is
                # NOT a ``tuple`` subclass. The old ``isinstance(r, tuple)``
                # guard was therefore always False, so whole ``Row`` objects
                # got stuffed into ``referenced`` instead of bare ints. The
                # later ``obj.id in referenced`` then compared an int against
                # ``Row(id,)`` -- never equal -- so EVERY FK-referenced
                # objective slipped past the preserve check and was deleted,
                # raising the FK IntegrityError on commit. Index [0] unwraps a
                # Row (and a plain tuple); fall back to the value only if it is
                # already a bare scalar.
                for r in rows:
                    referenced.add(r[0] if not isinstance(r, int) else r)
        for _, _, obj in candidates:
            if obj.id is not None and obj.id in referenced:
                preserved_with_refs += 1
                continue
            session.delete(obj)
            removed += 1
        if removed:
            session.commit()

    return DisaCciLoadResult(
        framework_id=framework_id,
        cci_items_in_xml=len(items),
        cci_items_attached=attached,
        cci_items_unmatched=unmatched,
        cci_items_no_nist_ref=no_nist,
        cci_items_deprecated=deprecated,
        objectives_created=created,
        objectives_updated=updated,
        cci_items_below_revision=below_revision,
        objectives_removed=removed,
        objectives_preserved_with_refs=preserved_with_refs,
    )
