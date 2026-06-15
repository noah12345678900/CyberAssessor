"""Program-specific control overlay loader.

A program overlay is a workbook tab that maps program-numbered requirements
(e.g. "SDA-AC-01", "ENCLAVE-AU-04") to one or more NIST 800-53 CCIs. Each
overlay becomes a ``RequirementSource`` row with N ``RequirementMap`` rows —
the assessor UI can show, per CCI, which program requirements drive that
CCI's status, and the rules engine can flag Rule #8a/#8b inheritance.

Two overlay shapes are auto-detected:

* **CCI-grain** (e.g. SDA Enterprise Services Controls) — sheet has an
  explicit CCI column; each row enumerates the CCIs the program req maps to.
* **Control-grain** (e.g. T1TL Ground Security Controls) — no CCI column;
  the NIST control IDs are embedded in the shall-statement prose as
  ``Associated CNSSI 1253 [Control Tag:] AC-2(13)`` (sometimes a comma-
  separated list). Each parsed control is fanned out to every CCI
  ``Objective`` underneath it in the target framework, so the resulting
  ``RequirementMap`` rows are the same shape as the CCI-grain path. The
  user's mental model — "SDA req → control → CCI" — holds either way.

Why a separate loader (not part of OSCAL)
-----------------------------------------
The same Control/Objective model will eventually host 800-171, but
program overlays only apply to specific assessments. Keeping them in their
own loader means ``RequirementSource`` rows can stack — one per program —
without conflating them with the canonical framework catalog.

Dependency on Objective rows
----------------------------
``RequirementMap`` rows point at ``Objective`` rows by CCI. CCIs for
800-53 are not in the NIST OSCAL catalog (DoD/DISA publishes them); they
get loaded by the DISA CCI loader (or by the CCIS workbook reader). Until
those run, this loader will create the ``RequirementSource`` row and emit
``unmapped_ccis`` / ``unmapped_control_ids`` in the return summary so the
UI can surface "these reqs reference catalog rows we haven't seen yet"
without erroring.

Synthetic ``Baseline`` materialization
--------------------------------------
After writing maps, the loader also upserts a ``Baseline`` row (keyed by
``framework_id`` + ``source_name``, ``source_type=PROGRAM_CONTROLS``)
covering every ``Control`` whose CCIs were referenced. This makes program
overlays first-class citizens of the ``WorkbookOverlay`` surface — the
Workbooks page "Overlays" column and the ManageOverlaysDialog read that
table, so without the synthetic baseline an SDA load succeeds but never
shows up in the UI. On *first* creation, the loader also auto-attaches
the baseline to every workbook whose ``framework_id`` matches; reloads
skip auto-attach so user detach choices stick.
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlmodel import Session, select

from ..models import (
    Baseline,
    BaselineControl,
    BaselineSourceType,
    Control,
    Framework,
    IngestReport,
    Objective,
    RequirementMap,
    RequirementSource,
)

# Match CCI references in any reasonable form: "CCI-000015", "CCI 15", "cci-15".
# We normalize all of these to the canonical zero-padded "CCI-000015" form.
_CCI_RE = re.compile(r"CCI[-\s]?(\d{1,7})", re.IGNORECASE)

# Header keywords we accept for each logical column, in PRIORITY ORDER.
# The first alias that resolves to a column with non-empty data wins
# (see ``_resolve_logical_col``). Order matters when two candidate columns
# both exist in a sheet:
#
# * SDA Enterprise Services Controls has BOTH a "Security Control" col
#   (holding the parent NIST id like "AC-10") AND a "Threshold" col (the
#   shall statement). We want "Threshold" for the shall text, so it's
#   listed first; "security control" is the last-resort fallback for
#   T1TL-style sheets where the shall sits under "Security Control" and
#   "Threshold"/"Objective" columns exist as headers but are empty.
# * "control no" / "control number" / "no." cover SDA's program-req-number
#   column (which is the SDA-assigned ID, NOT the parent NIST control).
_HEADER_ALIASES = {
    "req_number": [
        "req number", "requirement number", "control no", "control number",
        "req #", "requirement #", "no.", "id", "req",
    ],
    "requirement_text": [
        "threshold", "shall statement", "requirement",
        "control text", "description", "objective",
        "security control",
    ],
    "cci_refs": ["cci references", "cci ref", "ccis", "nist cci", "cci"],
    # Dedicated "Security Control" column that holds the parent NIST
    # control ID as structured data (e.g. T1TL Ground Security Controls
    # — col carries literal "AC-10" or "AC-10, AC-11" with no anchor
    # prose). Aliases the SAME header as the requirement_text fallback;
    # when both resolve to the same column the loader uses the cell as
    # display text AND parses it via the direct (no-anchor) extractor,
    # which is exactly the T1TL shape. SDA Enterprise Services keeps its
    # Threshold col winning for requirement_text, so this resolves
    # alongside it without changing existing behaviour.
    "control_id": ["security control"],
}

# How many rows from the top of the sheet we'll scan looking for the header.
_HEADER_SCAN_DEPTH = 10

# Loader code version. Bumped any time the row-walk logic changes in a way
# that could produce different RequirementMap rows from the same workbook
# (header resolution, forward-fill rules, extractor regexes). Persisted on
# each IngestReport so replay can tell whether a historical load was done
# by a known-buggy or known-good loader.
#
# History:
#   1 — initial release (no forward-fill; unmerged cell-block continuation
#       rows produced "(unnumbered)" sentinels).
#   2 — border-gated forward-fill of req_number across unmerged tall col-A
#       cell blocks; rows-forward-filled / rows-unnumbered counters; per-
#       row action log. Persisted via IngestReport (alembic 0005).
LOADER_VERSION = "program_controls_loader@2"

# How many data rows below the header we'll probe to decide whether a
# header-aliased column is actually populated. Some overlay sheets carry
# header rows for columns that are then left blank for the whole sheet
# (e.g. T1TL Ground has "Threshold"/"Objective" headers but no values).
_DATA_PROBE_ROWS = 5

# Match a single NIST control ID, with optional parenthesised enhancement:
# "AC-2", "AC-2(13)", "IA-5(7)", "PM-12".
_CONTROL_ID_RE = re.compile(r"[A-Z]{2}-\d+(?:\(\d+\))?")

# Anchor for control-grain overlays: only pull control IDs that the overlay
# author explicitly tagged as the canonical mapping. Without the anchor we'd
# pick up control-shaped tokens in narrative prose. Observed phrasings:
#   "Associated CNSSI 1253 Control Tag: AC-2"
#   "Associated CNSSI 1253 Control Tag: AC-2(13)"
#   "Associated CNSSI 1253 CM-7(8)"                  (no "Control Tag:")
#   "Associated CNSSI 1253 Control Tag: PM-12, AT-2(2), IR-4(7)"
_CONTROL_PROSE_ANCHOR_RE = re.compile(
    r"Associated\s+CNSSI\s+1253(?:\s+Control\s+Tag)?\s*:?\s*"
    r"((?:[A-Z]{2}-\d+(?:\(\d+\))?[\s,]*)+)",
    re.IGNORECASE,
)


def _normalize_cci(raw: str) -> str:
    """'CCI 15' / 'cci-15' / 'CCI-000015' -> 'CCI-000015'."""
    m = _CCI_RE.search(raw)
    if not m:
        return raw.strip()
    return f"CCI-{int(m.group(1)):06d}"


_BARE_CCI_CELL_RE = re.compile(r"^[\d\s,;|/.\-]+$")
_BARE_NUMBER_RE = re.compile(r"\d{1,7}")


def _extract_ccis(cell_value: Any) -> list[str]:
    """Pull every CCI reference out of a cell, deduped, in stable order.

    Two-pass extraction so we handle both prefixed and bare-number CCI cells
    without misreading numbers that appear in prose:

    1. Strict pass — matches ``CCI-000123`` / ``CCI 15`` / ``cci-15``
       anywhere in the cell. This is what most overlay sheets use, and it
       safely ignores stray digits in narrative text.
    2. Bare-number fallback — only fires when the cell is *purely*
       structured (digits + whitespace/separators only, no letters). This
       catches the SDA Enterprise Services Controls overlay where the CCI
       column holds newline-joined bare numbers like
       ``"000054\\n000055\\n002252"`` with the ``CCI-`` prefix stripped.
    """
    if cell_value is None:
        return []
    text = str(cell_value)
    seen: set[str] = set()
    out: list[str] = []

    for m in _CCI_RE.finditer(text):
        cci = f"CCI-{int(m.group(1)):06d}"
        if cci not in seen:
            seen.add(cci)
            out.append(cci)
    if out:
        return out

    stripped = text.strip()
    if stripped and _BARE_CCI_CELL_RE.fullmatch(stripped):
        for m in _BARE_NUMBER_RE.finditer(stripped):
            cci = f"CCI-{int(m.group(0)):06d}"
            if cci not in seen:
                seen.add(cci)
                out.append(cci)
    return out


def _normalize_control_id(raw: str) -> str:
    """Convert prose-form control IDs to the OSCAL catalog form.

    Prose:   ``AC-2``, ``AC-2(13)``, ``IA-5(7)``
    Catalog: ``ac-2``, ``ac-2.13``,  ``ia-5.7``  (lowercase, dot-enhancement)

    Without this normalization the control_objectives_lookup miss-rate is 100%
    against framework_id=2 (r5), since the OSCAL loader stores ids lowercased
    with a dot, not parens — verified by GETting /api/catalog/frameworks/2/controls.
    """
    s = raw.strip().lower()
    # AC-2(13) → AC-2.13; tolerate stray whitespace inside the parens
    return re.sub(r"\(\s*(\d+)\s*\)", r".\1", s)


def _extract_control_ids(cell_value: Any) -> list[str]:
    """Pull NIST control IDs out of shall-text prose.

    Only IDs that appear inside an ``Associated CNSSI 1253 ...`` anchor are
    returned — narrative mentions of controls elsewhere in the text are
    ignored. Returns deduped, **catalog-normalized** IDs in encounter order;
    preserves enhancements (``AC-2(13)`` → ``ac-2.13``, distinct from ``ac-2``).
    """
    if cell_value is None:
        return []
    text = str(cell_value)
    seen: set[str] = set()
    out: list[str] = []
    for anchor in _CONTROL_PROSE_ANCHOR_RE.finditer(text):
        for m in _CONTROL_ID_RE.finditer(anchor.group(1)):
            cid = _normalize_control_id(m.group(0))
            if cid not in seen:
                seen.add(cid)
                out.append(cid)
    return out


def _extract_control_ids_direct(cell_value: Any) -> list[str]:
    """Pull NIST control IDs from a dedicated control-id column.

    No prose anchor required — the column is by convention a list of
    control IDs as structured data (e.g. T1TL Ground Security Controls
    "Security Control" col holding "AC-10" or "AC-10, AC-11"). Safe
    because the caller routes only header-aliased columns to this path,
    so we're not scraping narrative prose. Returns deduped,
    **catalog-normalized** IDs in encounter order; preserves
    enhancements (``AC-2(13)`` → ``ac-2.13``, distinct from ``ac-2``).
    """
    if cell_value is None:
        return []
    text = str(cell_value)
    seen: set[str] = set()
    out: list[str] = []
    for m in _CONTROL_ID_RE.finditer(text):
        cid = _normalize_control_id(m.group(0))
        if cid not in seen:
            seen.add(cid)
            out.append(cid)
    return out


def _resolve_logical_col(
    sheet, header_row: int, headers: dict[int, str], aliases: list[str]
) -> int | None:
    """Return the 1-based column index for the first alias (priority order)
    that matches a header AND has non-empty data in the first
    ``_DATA_PROBE_ROWS`` rows below the header. ``None`` if no candidate
    column carries data."""
    for alias in aliases:
        for col_idx, header in headers.items():
            if alias not in header:
                continue
            for r in range(header_row + 1, header_row + 1 + _DATA_PROBE_ROWS):
                value = sheet.cell(row=r, column=col_idx).value
                if value is not None and str(value).strip():
                    return col_idx
            # Header matched but column is empty; try the next match for
            # this alias, then fall through to lower-priority aliases.
    return None


def _column_has_data(sheet, col_idx: int, start_row: int) -> bool:
    """True if column ``col_idx`` is non-empty in the first ``_DATA_PROBE_ROWS``
    rows below ``start_row``."""
    for r in range(start_row + 1, start_row + 1 + _DATA_PROBE_ROWS):
        value = sheet.cell(row=r, column=col_idx).value
        if value is not None and str(value).strip() and str(value).strip() != "-":
            return True
    return False


def _find_header_row(sheet) -> tuple[int, dict[str, int]]:
    """Locate the header row and return (row_index, column-name -> 1-based col index).

    Requires at least ONE of (``cci_refs``, ``requirement_text``) to resolve
    to a populated column — those are the two paths the data loop can use.

    Fallback: if ``req_number`` doesn't match any header alias but column 1
    carries data (e.g. T1TL Ground Security Controls, where col A header is
    a dash but rows hold integer row numbers), treat column 1 as the program
    requirement number column. The value is used verbatim — no prefix
    synthesis.
    """
    for row_idx in range(1, _HEADER_SCAN_DEPTH + 1):
        headers = {
            col_idx: str(cell.value).strip().lower() if cell.value else ""
            for col_idx, cell in enumerate(sheet[row_idx], start=1)
        }
        if not any(headers.values()):
            continue
        mapping: dict[str, int] = {}
        for logical, aliases in _HEADER_ALIASES.items():
            hit = _resolve_logical_col(sheet, row_idx, headers, aliases)
            if hit is not None:
                mapping[logical] = hit
        if "cci_refs" in mapping or "requirement_text" in mapping:
            # Column-A fallback for sheets whose req-number column doesn't
            # match any header alias. Only fires when (a) we didn't already
            # find a req_number col by header match, (b) column 1 isn't
            # already claimed by another logical column, and (c) column 1
            # has non-empty data within the data-probe window.
            if (
                "req_number" not in mapping
                and 1 not in mapping.values()
                and _column_has_data(sheet, 1, row_idx)
            ):
                mapping["req_number"] = 1
            return row_idx, mapping
    raise ValueError(
        "Could not locate a header row with either a CCI column or a "
        f"requirement-text column within the first {_HEADER_SCAN_DEPTH} rows "
        "of the overlay sheet"
    )


def _has_top_border(cell: Any) -> bool:
    """True if ``cell`` has a visible top border in the workbook XML.

    Used to distinguish two visually-different sources of a blank col-A cell
    in eMASS/T1TL overlay workbooks:

    * **Continuation row of an unmerged tall cell block** — col A renders
      visually across many sub-bullet rows but is unmerged in the XML, so
      openpyxl sees the first row carrying the value and the rest carrying
      ``None``. Excel encodes the visual continuity by drawing NO top border
      on the continuation rows; the parent's top border is the only one.
      Returns ``False`` for these → safe to forward-fill.

    * **Genuine workbook gap** — col A is blank because the row really has
      no mapping (orphan shall-statement, comment row, etc.). Excel draws
      the standard cell border. Returns ``True`` for these → we MUST NOT
      forward-fill; the row stays as a ``(unnumbered)`` sentinel so the
      operator sees it in the ingest report and can decide.

    Verified against T1TL Ground Security Controls: the 9 known
    continuation rows (sheet positions 461-472 under col A=460/AU-2,
    526/528 under 484-485/CP-9, 530 under 486/CP-10, 564 under 511/AC-17,
    610/612/614/616 under 543-546/SI-3) all report ``top_border_style=None``;
    the parent numbered rows above them report ``style='thin'``.

    Requires the workbook to be opened WITHOUT ``read_only=True`` — read-
    only mode strips border data from the parsed cells.
    """
    try:
        return bool(
            cell is not None
            and cell.border
            and cell.border.top
            and cell.border.top.style
        )
    except AttributeError:
        return False


def _resolve_sheet(wb, sheet_name: str):
    """Return the named sheet. We require an explicit name — overlay workbooks
    vary too much across programs for fuzzy matching to be safe."""
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )
    return wb[sheet_name]


def load_program_controls(
    session: Session,
    *,
    source_name: str,
    workbook_path: str | Path,
    framework_id: int,
    sheet_name: str,
) -> RequirementSource:
    """Idempotently load a program-controls overlay tab.

    Args:
        session: an active SQLModel Session.
        source_name: human-readable label for this overlay
            (e.g. "SDA Enterprise Services Controls", "Example System Program Overlay").
            Used as the upsert key together with ``framework_id``.
        workbook_path: path to the overlay workbook.
        framework_id: the Framework (typically 800-53r5) these reqs map to.
        sheet_name: explicit sheet name within the workbook.

    Returns:
        The RequirementSource row (created or updated). Transient attributes
        ``_rows_seen``, ``_maps_written``, ``_unmapped_ccis`` are attached so
        the caller can return a summary without a follow-up query.
    """
    wb_path = Path(workbook_path)
    if not wb_path.exists():
        raise FileNotFoundError(f"Program controls workbook not found: {wb_path}")

    framework = session.get(Framework, framework_id)
    if framework is None:
        raise ValueError(f"Framework id={framework_id} does not exist")

    # ``read_only=True`` strips cell border metadata, and the row walk needs
    # cell.border.top.style to distinguish unmerged-tall-cell continuation
    # rows (forward-fill safe) from genuine workbook gaps (must stay
    # unnumbered). The memory hit is bounded — overlay sheets top out in the
    # low thousands of rows.
    wb = load_workbook(wb_path, data_only=True)
    sheet = _resolve_sheet(wb, sheet_name)
    header_row, col_map = _find_header_row(sheet)

    # --- RequirementSource upsert (one row per framework + path) -----------
    # Keyed on (framework_id, path) — NOT (framework_id, name) — so two
    # distinct files become two distinct rows even when the user happens to
    # pass the same source_name (e.g. re-using "SDA Controls" as the label
    # for a Threshold catalog and a Objectives catalog). Re-uploading the
    # *same* file path still upserts in place (the "reload to refresh"
    # UX). The matching Baseline upsert below uses the same key for
    # symmetry — see the second upsert block.
    workbook_path_str = str(wb_path)
    source = session.exec(
        select(RequirementSource).where(
            RequirementSource.framework_id == framework_id,
            RequirementSource.path == workbook_path_str,
        )
    ).first()
    if source is None:
        source = RequirementSource(
            framework_id=framework_id,
            name=source_name,
            path=workbook_path_str,
        )
        session.add(source)
        session.commit()
        session.refresh(source)
    else:
        # Keep name in sync so the UI label reflects the latest load.
        source.name = source_name
        session.add(source)
        session.commit()

    # --- Build CCI -> Objective.id and Control -> [CCI obj ids] lookups ----
    # Critical: scope to the target framework. The same CCI string (e.g.
    # "CCI-000001") exists in both 800-53r4 and r5 catalogs as distinct
    # Objective rows; without the framework filter, dict-comp last-write-wins
    # would silently route maps to whichever framework iterated last, and the
    # UI join (Control -> Objective -> RequirementMap) would find nothing
    # because the map's objective_id belongs to the other framework's rows.
    # Reported as the "everything blank in the CSV after a successful overlay
    # load" bug.
    #
    # One query feeds three lookups (~3k rows for r5 — cheap).
    # ``objective_to_control_pk`` lets us derive which Control PKs the
    # synthetic Baseline should cover: any control with at least one CCI
    # that ended up on a RequirementMap row belongs in the baseline.
    objective_lookup: dict[str, int] = {}
    control_objectives_lookup: dict[str, list[int]] = {}
    objective_to_control_pk: dict[int, int] = {}
    for obj, ctrl in session.exec(
        select(Objective, Control)
        .join(Control, Objective.control_id_fk == Control.id)  # type: ignore[arg-type]
        .where(Objective.source == "CCI")
        .where(Control.framework_id == framework_id)
    ).all():
        if obj.id is None or ctrl.id is None:
            continue
        objective_lookup[obj.objective_id] = obj.id
        control_objectives_lookup.setdefault(ctrl.control_id, []).append(obj.id)
        objective_to_control_pk[obj.id] = ctrl.id

    # --- Wipe + rewrite maps for this source (idempotent reruns) -----------
    existing_maps = session.exec(
        select(RequirementMap).where(
            RequirementMap.requirement_source_id == source.id  # type: ignore[arg-type]
        )
    ).all()
    for m in existing_maps:
        session.delete(m)
    session.commit()

    # --- Walk data rows ----------------------------------------------------
    req_col = col_map.get("req_number")
    text_col = col_map.get("requirement_text")
    cci_col = col_map.get("cci_refs")
    # Dedicated control-id column (T1TL "Security Control" — structured
    # data, not prose). May share its physical column with text_col
    # when the sheet has no separate Threshold/Objective col, in which
    # case the same cell feeds both display text AND direct control-id
    # extraction.
    control_id_col = col_map.get("control_id")

    rows_seen = 0
    maps_written = 0
    rows_forward_filled = 0
    rows_unnumbered = 0
    unmapped_ccis: set[str] = set()
    unmapped_control_ids: set[str] = set()
    # Distinct Control PKs touched by any RequirementMap row. Drives the
    # synthetic Baseline's BaselineControl rows so the Workbooks page
    # "Overlays" column has something to render under this baseline.
    baseline_control_pks: set[int] = set()
    # Per-row structured action log persisted to IngestReport.actions.
    # Bounded by row count of the source sheet (low thousands max). Keep
    # entries small and human-readable — a 3PAO reads this JSON directly
    # via the bundle export.
    actions: list[dict] = []
    # Sticky parent req_number across the walk, used to forward-fill
    # continuation rows of unmerged tall col-A cell blocks. Reset only
    # on a row whose col A carries a real value.
    last_req_number = ""

    # ``values_only=False`` so we have Cell objects to inspect border
    # metadata on the req_number column. We re-derive the value tuple
    # locally to keep the rest of the existing loop body unchanged.
    for cells in sheet.iter_rows(min_row=header_row + 1):
        row = tuple(c.value for c in cells)
        if not row or all(cell is None for cell in row):
            continue
        rows_seen += 1
        row_idx = cells[0].row  # openpyxl row index for the action log

        raw_req = str(row[req_col - 1]).strip() if req_col and row[req_col - 1] else ""
        req_text = str(row[text_col - 1]).strip() if text_col and row[text_col - 1] else ""

        # Border-gated forward-fill. See ``_has_top_border`` docstring for
        # why the top-border signal is the safe boundary between "tall
        # unmerged cell continuation" and "real workbook gap".
        if raw_req:
            req_number = raw_req
            last_req_number = raw_req
        else:
            req_col_cell = cells[req_col - 1] if req_col else None
            if last_req_number and not _has_top_border(req_col_cell):
                req_number = last_req_number
                rows_forward_filled += 1
                actions.append(
                    {
                        "row": row_idx,
                        "action": "forward_fill",
                        "from_value": last_req_number,
                    }
                )
            else:
                req_number = ""
                rows_unnumbered += 1
                actions.append(
                    {
                        "row": row_idx,
                        "action": "unnumbered_block_start",
                        "reason": (
                            "top_border_present"
                            if _has_top_border(req_col_cell)
                            else "no_prior_value"
                        ),
                    }
                )

        # Resolve this row's target CCI Objective ids via two paths:
        #
        # 1. Explicit CCI column (SDA-style).
        # 2. Control IDs parsed from the shall text (T1TL-style), fanned
        #    out to every CCI Objective under each control.
        #
        # Path 2 only runs when path 1 produced nothing for this row, so a
        # mixed sheet that fills CCIs on some rows and uses prose tags on
        # others is handled row-by-row. Either path lands on Objective(CCI)
        # rows, so RequirementMap shape stays uniform.
        target_objective_ids: list[int] = []

        if cci_col is not None:
            for cci in _extract_ccis(row[cci_col - 1]):
                obj_id = objective_lookup.get(cci)
                if obj_id is None:
                    unmapped_ccis.add(cci)
                else:
                    target_objective_ids.append(obj_id)

        if not target_objective_ids and req_text:
            for control_id in _extract_control_ids(req_text):
                obj_ids = control_objectives_lookup.get(control_id)
                if not obj_ids:
                    unmapped_control_ids.add(control_id)
                    continue
                target_objective_ids.extend(obj_ids)

        # Path 3: dedicated "Security Control" column carrying the parent
        # NIST id as structured data (T1TL Ground Security Controls).
        # Only fires when paths 1 & 2 are empty so SDA Enterprise
        # Services Controls (which has both Threshold prose AND a
        # Security Control id col) keeps using its explicit CCI col.
        # No anchor required — the column is by convention a list of
        # control ids, so we scan the cell wholesale via the direct
        # extractor.
        if not target_objective_ids and control_id_col is not None:
            for control_id in _extract_control_ids_direct(row[control_id_col - 1]):
                obj_ids = control_objectives_lookup.get(control_id)
                if not obj_ids:
                    unmapped_control_ids.add(control_id)
                    continue
                target_objective_ids.extend(obj_ids)

        if not target_objective_ids:
            continue

        # Dedupe (an obj id can show up twice if e.g. control-grain fans out
        # AC-2 and AC-2(13) and both share some CCIs).
        for obj_id in dict.fromkeys(target_objective_ids):
            session.add(
                RequirementMap(
                    requirement_source_id=source.id,  # type: ignore[arg-type]
                    objective_id=obj_id,
                    requirement_number=req_number or "(unnumbered)",
                    requirement_text=req_text,
                )
            )
            maps_written += 1
            ctrl_pk = objective_to_control_pk.get(obj_id)
            if ctrl_pk is not None:
                baseline_control_pks.add(ctrl_pk)

    session.commit()
    session.refresh(source)

    # --- Synthetic Baseline materialization --------------------------------
    # The RequirementSource/RequirementMap rows are the source of truth, but
    # the UI's "Overlays" surface (Workbooks page + ManageOverlaysDialog)
    # reads WorkbookOverlay → Baseline. Without a Baseline row, program
    # overlays exist in a parallel namespace and never render. We mirror
    # the source as a synthetic Baseline + BaselineControl set so program
    # reqs flow through the same UI plumbing as FedRAMP/CRM overlays.
    baseline_was_created = False
    # Mirror the RequirementSource key: identify the synthetic Baseline by
    # (framework_id, source_ref=path, PROGRAM_CONTROLS) rather than by name,
    # so two distinct files with the same human label stay distinct.
    baseline = session.exec(
        select(Baseline).where(
            Baseline.framework_id == framework_id,
            Baseline.source_ref == workbook_path_str,
            Baseline.source_type == BaselineSourceType.PROGRAM_CONTROLS,
        )
    ).first()
    if baseline is None:
        baseline = Baseline(
            framework_id=framework_id,
            name=source_name,
            source_type=BaselineSourceType.PROGRAM_CONTROLS,
            source_ref=workbook_path_str,
        )
        session.add(baseline)
        session.commit()
        session.refresh(baseline)
        baseline_was_created = True
    else:
        baseline.name = source_name
        baseline.refreshed_at = datetime.now(timezone.utc)
        session.add(baseline)
        session.commit()

    # Wipe + rewrite BaselineControl rows for idempotent reloads.
    existing_bcs = session.exec(
        select(BaselineControl).where(
            BaselineControl.baseline_id == baseline.id  # type: ignore[arg-type]
        )
    ).all()
    for bc in existing_bcs:
        session.delete(bc)
    session.commit()

    tailoring_reason = f"Mapped by program overlay {source_name}"
    for ctrl_pk in sorted(baseline_control_pks):
        session.add(
            BaselineControl(
                baseline_id=baseline.id,  # type: ignore[arg-type]
                control_id=ctrl_pk,
                in_scope=True,
                tailoring_reason=tailoring_reason,
            )
        )
    session.commit()

    # No auto-attach. Loading a PSC overlay is a pure catalog operation:
    # the synthetic Baseline and per-CCI RequirementMap rows land in the
    # global catalog, and the user explicitly attaches it to whichever
    # workbook(s) need it via the Manage Overlays dialog. This keeps load
    # operations side-effect-free at the workbook layer — the previous
    # "bulk auto-attach on first creation" pass was implicit workbook
    # context that confused reload semantics (first load attached to all
    # framework-matching workbooks; reload attached to nothing).
    auto_attached_workbook_ids: list[int] = []

    # --- IngestReport audit row -------------------------------------------
    # One row per load. The structural decisions the loader just made
    # (forward-fills, surviving unnumbered rows, unresolved CCIs / control
    # ids) live here instead of evaporating with the HTTP response. The
    # bundle exporter joins RequirementSource → IngestReport to render the
    # audit trail; the UI can show "AU-2 row 472 was forward-filled from
    # AU-2 row 460" by reading actions[]. Cascades on RequirementSource
    # delete so re-imports don't accumulate orphan audit rows.
    sorted_unmapped_ccis = sorted(unmapped_ccis)
    sorted_unmapped_control_ids = sorted(unmapped_control_ids)
    ingest_report = IngestReport(
        requirement_source_id=source.id,
        framework_id=framework_id,
        source_path=workbook_path_str,
        sheet_name=sheet_name,
        loader_version=LOADER_VERSION,
        rows_seen=rows_seen,
        maps_written=maps_written,
        rows_forward_filled=rows_forward_filled,
        rows_unnumbered=rows_unnumbered,
        unmapped_ccis=sorted_unmapped_ccis,
        unmapped_control_ids=sorted_unmapped_control_ids,
        actions=actions,
    )
    session.add(ingest_report)
    session.commit()
    session.refresh(ingest_report)

    # Transient — see docstring.
    source.__dict__["_rows_seen"] = rows_seen
    source.__dict__["_maps_written"] = maps_written
    source.__dict__["_rows_forward_filled"] = rows_forward_filled
    source.__dict__["_rows_unnumbered"] = rows_unnumbered
    source.__dict__["_actions"] = actions
    source.__dict__["_loader_version"] = LOADER_VERSION
    source.__dict__["_sheet_name"] = sheet_name
    source.__dict__["_ingest_report_id"] = ingest_report.id
    source.__dict__["_unmapped_ccis"] = sorted_unmapped_ccis
    source.__dict__["_unmapped_control_ids"] = sorted_unmapped_control_ids
    # Synthetic-baseline materialization summary (so callers can render a
    # toast like "SDA overlay: attached to 3 workbooks, 412 controls").
    source.__dict__["_baseline_id"] = baseline.id
    source.__dict__["_baseline_was_created"] = baseline_was_created
    source.__dict__["_baseline_controls_written"] = len(baseline_control_pks)
    source.__dict__["_auto_attached_workbook_ids"] = auto_attached_workbook_ids

    return source
