"""Read an eMASS RMF POAM workbook back into the DB.

Round-trip pair to ``exporter.py``. Use openpyxl here — we're only reading,
so we don't need Excel COM, and openpyxl handles closed workbooks faster
than xlwings.

Merge policy:
  - Rows with an ``emass_poam_id`` (column A) that already exists in the DB
    for this workbook → UPDATE that Poam (preserves FK links to objectives).
  - Rows whose ID is missing or starts with ``DRAFT-`` → treated as new POAMs
    and INSERTED.
  - Milestone rows (rows that share a POAM ID with a previous row in the
    same import) → append PoamMilestone rows to the existing Poam.

We deliberately do NOT delete POAMs that are absent from the imported file —
the assessor's DB is the source of truth for in-flight work; eMASS exports
may have been filtered.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from sqlmodel import Session, select

from ..models import Poam, PoamMilestone, PoamStatus, RiskLevel
from .template import COLS, DATA_START_ROW, SHEET_NAME


def _parse_date(v) -> datetime | None:
    """eMASS date columns may come back as datetime or as 'DD-Mmm-YYYY' strings."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt)
        except ValueError:
            continue
    return None


def _parse_enum(enum_cls, v):
    """Best-effort enum coerce. Returns None if value not in enum."""
    if v is None or v == "":
        return None
    s = str(v).strip()
    for member in enum_cls:
        if member.value.lower() == s.lower():
            return member
    return None


def _col(row_cells, key: str):
    """Look up a cell in the row by logical COLS key. row_cells is 1-based."""
    letter = COLS[key].letter
    # Convert Excel letter to 1-based column index.
    idx = 0
    for ch in letter:
        idx = idx * 26 + (ord(ch.upper()) - ord("A") + 1)
    return row_cells[idx - 1].value if idx - 1 < len(row_cells) else None


def import_poams(
    workbook_id: int,
    poam_file_path: str | Path,
    s: Session,
) -> dict:
    """Read a POAM workbook and merge its rows into this workbook's POAMs.

    Returns ``{poams_created, poams_updated, milestones_created, skipped}``.
    """
    import openpyxl

    path = Path(poam_file_path)
    if not path.exists():
        raise FileNotFoundError(f"POAM file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet {SHEET_NAME!r} not found in {path.name}. "
            f"Available: {wb.sheetnames}"
        )
    sh = wb[SHEET_NAME]

    # Index existing POAMs for this workbook by emass id for O(1) merge lookup.
    existing = {
        p.emass_poam_id: p
        for p in s.exec(
            select(Poam).where(Poam.workbook_id == workbook_id)
        ).all()
        if p.emass_poam_id
    }

    created = 0
    updated = 0
    milestones_created = 0
    skipped = 0

    # Track the "current" POAM across milestone-continuation rows.
    current: Poam | None = None
    current_poam_id_cell: str | None = None

    for row in sh.iter_rows(min_row=DATA_START_ROW):
        row_cells = list(row)
        poam_id_cell = _col(row_cells, "id")
        if poam_id_cell is None or str(poam_id_cell).strip() == "":
            current = None
            current_poam_id_cell = None
            continue

        poam_id_str = str(poam_id_cell).strip()

        # Same POAM ID as previous row → milestone continuation.
        if (
            current is not None
            and current_poam_id_cell == poam_id_str
            and _col(row_cells, "vulnerability_description") in (None, "")
        ):
            m_desc = _col(row_cells, "milestone_description")
            if m_desc:
                s.add(
                    PoamMilestone(
                        poam_id=current.id,
                        description=str(m_desc),
                        scheduled_date=_parse_date(_col(row_cells, "milestone_scheduled_date")),
                        completion_date=_parse_date(_col(row_cells, "milestone_completion_date")),
                        changes_history=_col(row_cells, "milestone_status_comments"),
                    )
                )
                milestones_created += 1
            continue

        # New POAM row.
        is_draft = poam_id_str.startswith("DRAFT-")
        target = None if is_draft else existing.get(poam_id_str)

        vuln = _col(row_cells, "vulnerability_description")
        if not vuln:
            # No description and no existing match → not enough to act on.
            skipped += 1
            continue

        if target is None:
            target = Poam(
                workbook_id=workbook_id,
                control_cluster=str(_col(row_cells, "controls_aps") or "").split(",")[0].strip()
                or "unknown",
                vulnerability_description=str(vuln),
                emass_poam_id=None if is_draft else poam_id_str,
            )
            s.add(target)
            created += 1
        else:
            target.vulnerability_description = str(vuln)
            updated += 1

        target.security_control_number = _col(row_cells, "controls_aps") or target.security_control_number
        target.office_org = _col(row_cells, "office_org") or target.office_org
        target.resources_required = _col(row_cells, "resources") or target.resources_required
        target.mitigations = _col(row_cells, "mitigations") or target.mitigations
        target.comments = _col(row_cells, "comments") or target.comments

        status = _parse_enum(PoamStatus, _col(row_cells, "status"))
        if status:
            target.status = status

        sched = _parse_date(_col(row_cells, "scheduled_completion_date"))
        if sched:
            target.scheduled_completion_date = sched
        actual = _parse_date(_col(row_cells, "completion_date"))
        if actual:
            target.actual_completion_date = actual

        for fld in ("raw_severity", "likelihood", "impact", "relevance_of_threat", "residual_risk"):
            lvl = _parse_enum(RiskLevel, _col(row_cells, fld))
            if lvl:
                setattr(target, fld, lvl)

        target.updated_at = datetime.now(timezone.utc)
        s.flush()

        # First milestone (if present on the POAM row itself).
        m_desc = _col(row_cells, "milestone_description")
        if m_desc:
            s.add(
                PoamMilestone(
                    poam_id=target.id,
                    description=str(m_desc),
                    scheduled_date=_parse_date(_col(row_cells, "milestone_scheduled_date")),
                    completion_date=_parse_date(_col(row_cells, "milestone_completion_date")),
                    changes_history=_col(row_cells, "milestone_status_comments"),
                )
            )
            milestones_created += 1

        current = target
        current_poam_id_cell = poam_id_str

    s.commit()
    wb.close()

    return {
        "poams_created": created,
        "poams_updated": updated,
        "milestones_created": milestones_created,
        "skipped": skipped,
    }
