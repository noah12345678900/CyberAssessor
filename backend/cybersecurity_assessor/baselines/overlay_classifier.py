"""Overlay-kind classifier — sniff an xlsx and decide CRM / PSC / OTHER.

The Settings → Catalogs page used to surface two separate upload buttons
(CRM via :mod:`baselines.crm_xlsx`, PSC via
:mod:`catalogs.program_controls_loader`) because the two loaders consume
disjoint header vocabularies. To the assessor, both are just "overlay
spreadsheets the program threw at me." This module unifies the front
door: one classifier reads the header rows once, returns a kind, and the
new ``POST /api/catalog/overlays/import`` route dispatches to the right
loader without the user having to pick.

Decision rules
--------------
For each visible sheet, in this order:

* **CRM** — a control-id-shaped header column AND a responsibility-shaped
  header column both appear on the same sheet. Vocab drawn from
  :mod:`baselines.crm_xlsx` (``_CONTROL_ID_HEADERS``,
  ``_RESPONSIBILITY_HEADERS_*``). CRM is the more specific shape (the
  responsibility column is the unmistakable signal), so we check it first
  to avoid misrouting a CRM sheet that happens to carry a "control text"
  column to the PSC loader.
* **PSC** — a threshold / shall / requirement / objective / control-text
  column appears on the sheet. Vocab drawn from
  :mod:`catalogs.program_controls_loader._HEADER_ALIASES["requirement_text"]`.
  PSC overlays are control-grain or shall-grain — they carry NO CCI
  column. CCIs are derived later by joining the resolved control IDs
  (from the ``Associated CNSSI 1253 Control Tag:`` anchor in the shall
  prose) against the global DISA CCI ↔ 800-53 mapping.

Aggregation:

* First sheet that matches CRM or PSC wins. The per-sheet rule already
  encodes priority (CRM > PSC within a sheet), so cross-sheet aggregation
  is just "stop at the first hit." First CRM-shaped sheet → CRM; else
  first PSC-shaped sheet → PSC.
* Otherwise → ``OTHER`` (inert; gets a Baseline row so it's visible/
  attachable but no resolver runs against it during assessment).

Read-only — never mutates the file. Safe to call on the original upload
path before the loader copies it into the catalog cache.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from enum import Enum
from pathlib import Path

from openpyxl import load_workbook

# How many rows from the top of each sheet we'll scan looking for a
# header. Matches ``catalogs.program_controls_loader._HEADER_SCAN_DEPTH``
# so the classifier sees the same rows the PSC loader would parse.
_HEADER_SCAN_DEPTH = 10

# Strip everything but ASCII letters + digits so "Control ID" / "control_id"
# / "CONTROL-ID" / "control #" all collapse to "controlid". Mirrors the
# normalization the CRM and PSC loaders do header-by-header internally,
# but applied as a single regex here for speed.
_HEADER_NORMALIZE_RE = re.compile(r"[^a-z0-9]+")


def _norm(s: str) -> str:
    return _HEADER_NORMALIZE_RE.sub("", s.lower())


# CRM vocab. Derived from baselines/crm_xlsx.py:79-137.
#
# Control-id column tokens — any of these in normalized form means
# "this column probably holds AC-2 / AC-2(1) / etc."
_CRM_CONTROL_ID_TOKENS: frozenset[str] = frozenset(
    {
        _norm("control id"),
        _norm("control identifier"),
        _norm("control"),
        _norm("control number"),
        _norm("control #"),
        _norm("control no"),
    }
)

# Responsibility column tokens — Customer / Provider / Hybrid / Inherited
# / NA. The CRM loader has dual cloud/on-prem variants; the classifier
# just needs to see *any* responsibility column to confirm CRM shape.
_CRM_RESPONSIBILITY_TOKENS: frozenset[str] = frozenset(
    {
        _norm("responsibility"),
        _norm("control responsibility"),
        _norm("implementation status"),
        _norm("csp responsibility"),
        _norm("responsibility assignment"),
        _norm("cloud responsibility"),
        _norm("cloud control responsibility"),
        _norm("on-prem responsibility"),
        _norm("onprem responsibility"),
        _norm("on prem responsibility"),
        _norm("on-premise responsibility"),
        _norm("on-premises responsibility"),
        _norm("on premises responsibility"),
        _norm("on-prem control responsibility"),
    }
)

# PSC vocab. Derived from catalogs/program_controls_loader.py:91-102.
#
# Threshold / shall-statement column tokens — the requirement text the
# PSC loader treats as the authoritative "shall." Presence of this
# column alone is enough to call a sheet PSC-shaped; PSC overlays do
# NOT carry a CCI column (CCIs come from the global DISA mapping at
# resolve time, not from the overlay).
_PSC_REQUIREMENT_TOKENS: frozenset[str] = frozenset(
    {
        _norm("threshold"),
        _norm("shall statement"),
        _norm("shall"),
        _norm("requirement"),
        _norm("control text"),
        _norm("objective"),
        # "security control" is the last-resort fallback in the PSC loader
        # for T1TL-style sheets; we include it here too so we don't miss
        # a file the loader would happily parse.
        _norm("security control"),
    }
)


class OverlayKind(str, Enum):
    """What an overlay xlsx looks like, by header shape."""

    CRM = "crm"
    PSC = "psc"
    OTHER = "other"


def _classify_sheet(header_tokens: set[str]) -> OverlayKind | None:
    """Classify a single sheet by its (normalized) header tokens.

    Returns ``None`` when neither vocab matches — caller aggregates
    across sheets.

    CRM is checked first because the responsibility column is the
    unmistakable signal; a CRM sheet that happens to carry a "control
    text" column would otherwise hit the (broader) PSC requirement vocab
    and be misrouted to the PSC loader.
    """
    has_crm_id = bool(header_tokens & _CRM_CONTROL_ID_TOKENS)
    has_crm_resp = bool(header_tokens & _CRM_RESPONSIBILITY_TOKENS)
    has_psc_req = bool(header_tokens & _PSC_REQUIREMENT_TOKENS)

    if has_crm_id and has_crm_resp:
        return OverlayKind.CRM
    if has_psc_req:
        return OverlayKind.PSC
    return None


def sniff_overlay_kind(xlsx_path: Path) -> OverlayKind:
    """Open ``xlsx_path`` read-only and return the inferred kind.

    Aggregation rule: first sheet that matches wins — see module
    docstring.

    Raises:
      FileNotFoundError: if ``xlsx_path`` does not exist.
      ValueError: if ``xlsx_path`` is not a parseable xlsx (openpyxl
        will raise its own ``InvalidFileException``; we let it bubble).
    """
    return classify_overlay(xlsx_path).kind


@dataclass
class OverlayClassification:
    """What ``classify_overlay`` returns.

    ``kind`` is the aggregated decision. ``sheet_name`` is the first sheet
    that matched ``kind``'s vocabulary (None for ``OTHER`` — no sheet
    matched). The unified import route needs the sheet name because
    :func:`catalogs.program_controls_loader.load_program_controls` takes
    ``sheet_name`` as a required argument — it intentionally won't fuzzy-
    match because overlay tabs vary too wildly across programs.
    """

    kind: OverlayKind
    sheet_name: str | None = None


@dataclass
class OverlaySheetCandidate:
    """One sheet's per-sheet classification — used by the sheet-picker UI.

    Unlike :class:`OverlayClassification` (which aggregates to "first match
    wins" across the whole workbook), this row reports what *this* sheet
    looks like in isolation. The Settings → Import overlay card uses the
    list to populate a "Sheet" dropdown so the user can target a specific
    tab — the T1TL workbook ships with both Ground and SV PSC-shaped
    sheets, and without a per-sheet view there's no way to pick SV (the
    aggregate classifier always returns the first matched sheet, Ground).

    ``kind`` is ``None`` when the sheet's headers don't match any known
    vocabulary — those sheets still appear in the dropdown so the user can
    force them via ``kind_hint`` if they know what's there, but they're
    flagged as non-candidates in the UI.
    """

    name: str
    kind: OverlayKind | None


def classify_overlay_sheets(xlsx_path: Path) -> list[OverlaySheetCandidate]:
    """Open ``xlsx_path`` read-only and classify every sheet independently.

    Walks ``wb.sheetnames`` in workbook order and applies the same
    per-sheet rule :func:`classify_overlay` uses, but without the
    "first match wins" aggregation — every sheet is reported, including
    sheets whose headers match no known vocabulary (``kind=None``).

    This is the data behind the Settings → Import overlay sheet picker.
    Pair the result with :func:`classify_overlay` (the auto-pick) so the
    UI can render "Auto-pick (Ground Security Controls)" as the default
    option followed by every sheet labeled with its candidate kind.

    Raises:
      FileNotFoundError: if ``xlsx_path`` does not exist.
      ValueError: if ``xlsx_path`` is not a parseable xlsx.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"overlay file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out: list[OverlaySheetCandidate] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tokens: set[str] = set()
            for row_idx, row in enumerate(
                ws.iter_rows(values_only=True), start=1
            ):
                if row_idx > _HEADER_SCAN_DEPTH:
                    break
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if not text:
                        continue
                    tokens.add(_norm(text))
            out.append(
                OverlaySheetCandidate(
                    name=sheet_name,
                    kind=_classify_sheet(tokens),
                )
            )
        return out
    finally:
        wb.close()


def classify_overlay(xlsx_path: Path) -> OverlayClassification:
    """Open ``xlsx_path`` read-only and return the kind + matched sheet.

    Same aggregation rules as :func:`sniff_overlay_kind` (first matched
    sheet wins; CRM has priority over PSC within a single sheet), plus
    the sheet name that triggered the match — needed by the PSC loader,
    which requires an explicit ``sheet_name`` argument.

    Raises:
      FileNotFoundError: if ``xlsx_path`` does not exist.
      ValueError: if ``xlsx_path`` is not a parseable xlsx.
    """
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"overlay file not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Collect normalized header tokens from the top N rows. Both
            # loaders scan for a header row instead of trusting row 1, so
            # the classifier does the same — many CRM/PSC sheets have a
            # title block above the actual header.
            tokens: set[str] = set()
            for row_idx, row in enumerate(
                ws.iter_rows(values_only=True), start=1
            ):
                if row_idx > _HEADER_SCAN_DEPTH:
                    break
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if not text:
                        continue
                    tokens.add(_norm(text))

            kind = _classify_sheet(tokens)
            if kind is not None:
                # First matched sheet wins. Per-sheet priority (CRM
                # before PSC) is already encoded in _classify_sheet, so
                # cross-sheet aggregation is just "stop at the first
                # hit."
                return OverlayClassification(
                    kind=kind, sheet_name=sheet_name
                )

        return OverlayClassification(kind=OverlayKind.OTHER, sheet_name=None)
    finally:
        wb.close()
