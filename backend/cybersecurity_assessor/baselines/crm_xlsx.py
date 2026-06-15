"""CRM (Customer Responsibility Matrix) overlay loader.

Ingests a FedRAMP-style xlsx where each row carries:
  - Control ID  (AC-2, AC-2(1), etc. -- normalized to OSCAL catalog form)
  - Responsibility  ("Customer" / "Provider" / "Hybrid" / "Inherited" /
                     "Not Applicable" -- normalized to lowercase enum vals)
  - Customer Responsibility  (narrative text; optional)

Output: a Baseline (source_type=CRM) with one BaselineControl per row
carrying responsibility + responsibility_narrative. Reuses the CCIS
adapter's control-ID normalizer chain (``_normalize_control`` ->
``_ccis_to_oscal_control_id``) so "AC-02 (1)" and "AC-2(1)" both
resolve to the catalog's canonical "ac-2.1".

Idempotent: re-loading the same CRM xlsx upserts the Baseline on
(source_type, source_ref) and BaselineControl rows on
(baseline_id, control_id). Rows that disappear from a re-uploaded CRM
are removed from the baseline so the overlay stays truthful; callers
who want to preserve an old inheritance trail for audit should upload
the new CRM as a fresh Baseline (different source_ref) rather than
overwriting in place.

Default-customer on absence: per the overlay-default-local rule, a
control with no CRM row gets no BaselineControl here at all -- the
kernel's CrmContext.lookup() returns None and the assessor falls
through to its normal full-assessment path. The CRM cannot opt
controls *out* of assessment by silence; only explicit Provider /
Inherited / NA rows short-circuit.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from sqlmodel import Session, select

# Reuse the CCIS adapter's control-ID normalizer chain -- single source
# of truth for "AC-02 (1)" -> "ac-2.1" canonicalization. Both helpers
# are module-private but stable; the CCIS loader imports them the same
# way (see baselines/ccis_workbook.py).
from ..excel.ccis_reader import _ccis_to_oscal_control_id, _normalize_control
from ..models import (
    Baseline,
    BaselineControl,
    BaselineSourceType,
    Control,
    Framework,
)
from .base import BaselineApplyResult

# Responsibility-value normalization. Keys are lowercased, stripped
# header strings as they appear in FedRAMP CRM templates (varies by
# revision). Values are the four-state enum we store on
# BaselineControl.responsibility. Unknown values are counted under
# controls_unknown so the user sees how much of the CRM we couldn't
# parse instead of silently dropping rows.
_RESPONSIBILITY_MAP: dict[str, str] = {
    "customer": "customer",
    "customer responsibility": "customer",
    # "Customer Configured" is its own state: the customer must actively
    # configure the control (vs plain "customer responsibility"). Kept
    # distinct from "customer" so the chip and narrative can reflect the
    # configuration burden. Like "customer", it is NOT short-circuited —
    # it still gets a full assessment.
    "customer configured": "customer_configured",
    "customer-configured": "customer_configured",
    "configured by customer": "customer_configured",
    "provider": "provider",
    "service provider": "provider",
    "csp": "provider",
    "provider system": "provider",
    "hybrid": "hybrid",
    "shared": "hybrid",
    "shared responsibility": "hybrid",
    "inherited": "inherited",
    "fully inherited": "inherited",
    "not applicable": "not_applicable",
    "n/a": "not_applicable",
    "na": "not_applicable",
}

# Column-header sniffing. FedRAMP templates have varied wording across
# revisions (Rev4 vs Rev5 vs 20x), so we match a small synonym set per
# column and pick the first that hits. Case + whitespace insensitive.
_CONTROL_ID_HEADERS = frozenset(
    {
        "control id",
        "control identifier",
        "control",
        "control number",
        "control #",
        "control no",
    }
)
_RESPONSIBILITY_HEADERS_CLOUD = frozenset(
    {
        # Generic / CSP-provided CRM templates -- always cloud-scope today.
        "responsibility",
        "control responsibility",
        "implementation status",
        "csp responsibility",
        "responsibility assignment",
        # Explicit cloud-scope synonyms for dual-column CRMs.
        "cloud responsibility",
        "cloud control responsibility",
    }
)
_RESPONSIBILITY_HEADERS_ONPREM = frozenset(
    {
        "on-prem responsibility",
        "onprem responsibility",
        "on prem responsibility",
        "on-premise responsibility",
        "on-premises responsibility",
        "on premises responsibility",
        "on_prem_responsibility",
        "on-prem control responsibility",
    }
)
_NARRATIVE_HEADERS_CLOUD = frozenset(
    {
        "customer responsibility",
        "customer responsibility narrative",
        "customer implementation",
        "customer description",
        "customer narrative",
        "customer responsibilities",
        # Explicit cloud-scope synonyms for dual-column CRMs.
        "cloud customer responsibility",
        "cloud narrative",
    }
)
_NARRATIVE_HEADERS_ONPREM = frozenset(
    {
        "on-prem narrative",
        "onprem narrative",
        "on-prem customer responsibility",
        "on-premises customer responsibility",
        "on-prem responsibility narrative",
        "on-prem description",
        "on-prem implementation",
    }
)


class CrmXlsxBaselineSource:
    """BaselineSource adapter for FedRAMP-style CRM xlsx files."""

    source_type = BaselineSourceType.CRM

    def __init__(
        self,
        workbook_path: str | Path,
        *,
        name: str | None = None,
        system_id: int | None = None,
        scope_label: str | None = None,
    ) -> None:
        self.workbook_path = Path(workbook_path)
        self.name = name or f"CRM: {self.workbook_path.stem}"
        self.system_id = system_id
        # v0.2 multi-implementation: each CRM upload represents one
        # implementation slice (e.g. ``"AWS GovCloud"``). The route
        # layer enforces non-null + canonical normalization; we just
        # pass it through to the Baseline row.
        self.scope_label = scope_label

    # ------------------------------------------------------------------
    # Column-header sniffing
    # ------------------------------------------------------------------
    @staticmethod
    def _locate_columns(
        header: tuple,
    ) -> tuple[int | None, int | None, int | None, int | None, int | None]:
        """Return (control_id, resp_cloud, resp_onprem, narr_cloud, narr_onprem).

        First match wins per column. Only ``control_id_col`` and at
        least one of the two responsibility cols are required (the
        caller raises). Narrative cols are always optional and the
        on-prem pair is optional too — a legacy single-column CRM
        (AWS GovCloud template) only carries the cloud pair, so the
        on-prem cols come back None and stay None on the model.
        """
        col_ctrl: int | None = None
        col_resp_cloud: int | None = None
        col_resp_onprem: int | None = None
        col_narr_cloud: int | None = None
        col_narr_onprem: int | None = None
        for i, cell in enumerate(header):
            if cell is None:
                continue
            key = str(cell).strip().lower()
            if not key:
                continue
            if col_ctrl is None and key in _CONTROL_ID_HEADERS:
                col_ctrl = i
                continue
            # On-prem checked before cloud so a column literally titled
            # "On-Prem Responsibility" never accidentally hits the
            # broader "responsibility" alias in the cloud set.
            if col_resp_onprem is None and key in _RESPONSIBILITY_HEADERS_ONPREM:
                col_resp_onprem = i
                continue
            if col_resp_cloud is None and key in _RESPONSIBILITY_HEADERS_CLOUD:
                col_resp_cloud = i
                continue
            if col_narr_onprem is None and key in _NARRATIVE_HEADERS_ONPREM:
                col_narr_onprem = i
                continue
            if col_narr_cloud is None and key in _NARRATIVE_HEADERS_CLOUD:
                col_narr_cloud = i
                continue
        return col_ctrl, col_resp_cloud, col_resp_onprem, col_narr_cloud, col_narr_onprem

    # ------------------------------------------------------------------
    # Upsert helpers
    # ------------------------------------------------------------------
    def _upsert_baseline(self, session: Session, framework_id: int) -> Baseline:
        source_ref = str(self.workbook_path)
        # Upsert key includes scope_label so two CRMs targeting the same
        # workbook path under different implementation labels coexist.
        # Replace-by-label semantics (different path, same label) live one
        # layer up in routes/catalog.py — by the time _upsert_baseline runs,
        # the prior label-collision Baseline has already been deleted.
        baseline = session.exec(
            select(Baseline).where(
                Baseline.source_type == self.source_type,
                Baseline.source_ref == source_ref,
                Baseline.scope_label == self.scope_label,
            )
        ).first()
        if baseline is None:
            baseline = Baseline(
                framework_id=framework_id,
                system_id=self.system_id,
                name=self.name,
                source_type=self.source_type,
                source_ref=source_ref,
                scope_label=self.scope_label,
            )
            session.add(baseline)
            session.commit()
            session.refresh(baseline)
        else:
            baseline.framework_id = framework_id
            if self.system_id is not None:
                baseline.system_id = self.system_id
            # scope_label is part of the lookup key, so it's already
            # correct on the existing row — no re-assignment needed.
            baseline.refreshed_at = datetime.now(timezone.utc)
            session.add(baseline)
            session.commit()
            session.refresh(baseline)
        return baseline

    # ------------------------------------------------------------------
    # Apply
    # ------------------------------------------------------------------
    def apply(self, session: Session, *, framework_id: int) -> BaselineApplyResult:
        framework = session.get(Framework, framework_id)
        if framework is None:
            raise ValueError(f"Framework id={framework_id} does not exist")

        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as exc:  # pragma: no cover - dep is in pyproject
            raise RuntimeError(
                "openpyxl is required for CRM xlsx loading"
            ) from exc

        wb = load_workbook(self.workbook_path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                raise ValueError("CRM xlsx has no header row")
            (
                col_ctrl,
                col_resp_cloud,
                col_resp_onprem,
                col_narr_cloud,
                col_narr_onprem,
            ) = self._locate_columns(header)
            if col_ctrl is None or (col_resp_cloud is None and col_resp_onprem is None):
                raise ValueError(
                    "CRM xlsx must have Control ID and at least one "
                    "Responsibility column (cloud or on-prem); "
                    f"got headers: {[str(c) for c in header if c is not None]}"
                )

            baseline = self._upsert_baseline(session, framework_id)

            # Catalog lookup: string control_id -> Control.id (PK).
            # BaselineControl.control_id is the FK to Control.id, NOT the
            # human string, so we need this map to write rows.
            control_pk_by_id: dict[str, int] = {
                c.control_id: c.id
                for c in session.exec(
                    select(Control).where(Control.framework_id == framework_id)
                ).all()
                if c.id is not None
            }

            existing_baseline_ctls = {
                bc.control_id: bc
                for bc in session.exec(
                    select(BaselineControl).where(
                        BaselineControl.baseline_id == baseline.id
                    )
                ).all()
            }

            seen_control_pks: set[int] = set()
            controls_in_scope = 0  # CRM rows we wrote (any responsibility)
            controls_unknown_ids: set[str] = set()
            unknown_responsibility = 0

            def _cell(row: tuple, col: int | None):
                if col is None or col >= len(row):
                    return None
                return row[col]

            for row in rows_iter:
                ctrl_raw = _cell(row, col_ctrl)
                resp_cloud_raw = _cell(row, col_resp_cloud)
                resp_onprem_raw = _cell(row, col_resp_onprem)
                narr_cloud_raw = _cell(row, col_narr_cloud)
                narr_onprem_raw = _cell(row, col_narr_onprem)

                if ctrl_raw in (None, ""):
                    continue
                # A row must carry at least one of the two scope verdicts
                # to be meaningful; otherwise it's a blank/spacer row.
                if (
                    resp_cloud_raw in (None, "")
                    and resp_onprem_raw in (None, "")
                ):
                    continue

                normalized = _normalize_control(str(ctrl_raw))
                if not normalized:
                    continue
                oscal_ctl_id = _ccis_to_oscal_control_id(normalized)

                # Normalize each scope independently. An unknown value
                # in ONE scope counts toward unknown_responsibility but
                # the other scope's value (if recognized) is still kept.
                responsibility_cloud: str | None = None
                if resp_cloud_raw not in (None, ""):
                    responsibility_cloud = _RESPONSIBILITY_MAP.get(
                        str(resp_cloud_raw).strip().lower()
                    )
                    if responsibility_cloud is None:
                        unknown_responsibility += 1

                responsibility_onprem: str | None = None
                if resp_onprem_raw not in (None, ""):
                    responsibility_onprem = _RESPONSIBILITY_MAP.get(
                        str(resp_onprem_raw).strip().lower()
                    )
                    if responsibility_onprem is None:
                        unknown_responsibility += 1

                # If both scopes were specified but both unrecognized,
                # there's nothing to write — skip the row entirely so
                # the kernel falls through to its full-assessment path.
                if responsibility_cloud is None and responsibility_onprem is None:
                    continue

                ctl_pk = control_pk_by_id.get(oscal_ctl_id)
                if ctl_pk is None:
                    # CRM references a control the loaded catalog doesn't
                    # have (wrong rev, typo, withdrawn). Surface in the
                    # response so the user can fix the CRM rather than
                    # silently dropping inheritance signal.
                    controls_unknown_ids.add(oscal_ctl_id)
                    continue

                narrative_cloud = (
                    str(narr_cloud_raw).strip() if narr_cloud_raw else None
                )
                narrative_onprem = (
                    str(narr_onprem_raw).strip() if narr_onprem_raw else None
                )
                seen_control_pks.add(ctl_pk)

                bc = existing_baseline_ctls.get(ctl_pk)
                if bc is None:
                    # CRM baselines don't carry tailoring (the FedRAMP
                    # profile baseline does that). in_scope=True is a
                    # vacuous default here -- the kernel reads
                    # responsibility, not in_scope, from CRM overlays.
                    session.add(
                        BaselineControl(
                            baseline_id=baseline.id,  # type: ignore[arg-type]
                            control_id=ctl_pk,
                            in_scope=True,
                            responsibility=responsibility_cloud,
                            responsibility_narrative=narrative_cloud,
                            responsibility_onprem=responsibility_onprem,
                            responsibility_onprem_narrative=narrative_onprem,
                        )
                    )
                else:
                    bc.responsibility = responsibility_cloud
                    bc.responsibility_narrative = narrative_cloud
                    bc.responsibility_onprem = responsibility_onprem
                    bc.responsibility_onprem_narrative = narrative_onprem
                    session.add(bc)
                controls_in_scope += 1

            # Prune rows the new CRM no longer references. Keeps the
            # overlay truthful when the user re-uploads a tightened CRM.
            # Audit-preservation is handled by uploading the new CRM as
            # a distinct Baseline (different source_ref), not by hanging
            # on to stale rows here.
            for ctl_pk, bc in existing_baseline_ctls.items():
                if ctl_pk not in seen_control_pks:
                    session.delete(bc)

            session.commit()
        finally:
            wb.close()

        return BaselineApplyResult(
            baseline=baseline,
            controls_in_scope=controls_in_scope,
            controls_out_of_scope=0,
            controls_unknown=len(controls_unknown_ids),
            notes={
                "loader": "crm_xlsx",
                "path": str(self.workbook_path),
                "unknown_control_ids": sorted(controls_unknown_ids),
                "unknown_responsibility_rows": unknown_responsibility,
            },
        )
