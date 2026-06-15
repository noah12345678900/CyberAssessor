"""CCIS workbook reader (openpyxl, read-only).

Column layout is taken verbatim from the nist-assessor plugin's
``ccis-workbook-guide.md`` (the eMASS Export schema) — do not invent a
different mapping. WORKING SHEET data rows start at row 7; row 6 is
headers; rows 1-5 are system metadata.

    Col A  Required for assessment?       "YES" or blank
    Col B  Control Acronym                "AC-2(1)"
    Col C  Control Information            full control text + supplemental
    Col D  Control Implementation Status  "Planned" / "Implemented"
    Col E  Security Control Designation   "Hybrid" / "Common" / "System-Specific"
    Col F  Implementation Narrative       free text
    Col G  AP Acronym                     "AC-2.1"   (== deterministic objective id)
    Col H  CCI                            "002110"   (often bare, no "CCI-" prefix)
    Col I  CCI Definition                 what's required
    Col J  Implementation Guidance        how to implement + evidence
    Col K  Assessment Procedures          how to verify
    Col L  Inherited                      "Local" / "DoW Enterprise" / system name
    Col M  Remote Inheritance Instance    specific system for remote inheritance
    Col N  Compliance Status              "Compliant" / "Non-Compliant" / "Not Applicable"
    Col O  Date Tested                    ISO 8601
    Col P  Tested By                      assessor name
    Col Q  Test Results                   facts-only narrative
    Col R  PREVIOUS Compliance Status
    Col S  PREVIOUS Date Tested
    Col T  PREVIOUS Tested By
    Col U  PREVIOUS Test Results

The reader is read-only by design. The find-evidence and review-assessment
flows both need to parse the workbook without touching the database, and
the assessor flow uses the same parser then hands rows to
``populate_objectives`` for DB upsert.
"""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from ..models import Control, Framework, Objective

# Sheet name candidates in priority order.
_WORKING_SHEET_NAMES = ["WORKING SHEET", "Working Sheet", "Working sheet"]

# Assignment Values tab carries ODP (Organization-Defined Parameter)
# values. Column layout drifts between eMASS template revs, so the
# reader detects headers rather than hard-coding column indexes. See
# memory/project_odp_architecture.md for why ODPs are stored separately
# from the Control statement.
_ASSIGNMENT_VALUES_SHEET_NAMES = ["Assignment Values", "Assignment values"]

# Data starts at row 7; stop after this many consecutive blank control-id
# cells (the eMASS template pads with thousands of empty rows).
_EMPTY_ROW_RUN_LIMIT = 5

# Safety cap (plugin's B7:B325 implies ~319 CCIs; we go higher to be safe).
_MAX_DATA_ROW = 500

# Match "CCI-000015" / "CCI 15" — REQUIRES the CCI prefix.
_CCI_PREFIXED_RE = re.compile(r"CCI[-\s]?(\d{1,7})", re.IGNORECASE)
# Match bare numbers (used for the dedicated CCI column H which often
# holds "002110" without a prefix).
_CCI_BARE_RE = re.compile(r"\b(\d{1,7})\b")
_CONTROL_RE = re.compile(r"^[A-Z]{2}-\d{1,2}(?:\(\d{1,2}\))?$")

# Columns (1-based, matches openpyxl).
COL_REQUIRED = 1  # A
COL_CONTROL = 2  # B
COL_IMPL_STATUS = 4  # D
COL_DESIGNATION = 5  # E
COL_NARRATIVE = 6  # F
COL_AP_ACRONYM = 7  # G
COL_CCI = 8  # H
COL_DEFINITION = 9  # I
COL_GUIDANCE = 10  # J
COL_PROCEDURES = 11  # K
COL_INHERITED = 12  # L
COL_REMOTE_INHERITANCE = 13  # M
COL_STATUS = 14  # N
COL_DATE_TESTED = 15  # O
COL_TESTER = 16  # P
COL_RESULTS = 17  # Q
COL_PREV_STATUS = 18  # R
COL_PREV_DATE = 19  # S
COL_PREV_TESTER = 20  # T
COL_PREV_RESULTS = 21  # U


# ---------------------------------------------------------------------------
# Parsed-row dataclass
# ---------------------------------------------------------------------------


@dataclass
class CcisRow:
    """One CCI row from the WORKING SHEET.

    ``excel_row`` is the absolute 1-based row index — write paths use it
    to address the same row via xlwings without re-finding it.
    """

    excel_row: int
    required: bool  # col A == "YES"
    control_id: str  # col B, "AC-2(1)"
    ap_acronym: str | None  # col G, "AC-2.1"
    cci_id: str | None  # col H, canonical "CCI-002110"
    implementation_status: str | None  # col D
    designation: str | None  # col E
    narrative: str | None  # col F
    definition: str | None  # col I
    guidance: str | None  # col J
    procedures: str | None  # col K
    inherited: str | None  # col L — "Local" / "DoW Enterprise" / system
    remote_inheritance: str | None  # col M
    # Current assessment (writable cells)
    status: str | None  # col N
    date_tested: datetime | None  # col O
    tester: str | None  # col P
    results: str | None  # col Q
    # Previous assessment (read-only)
    previous_status: str | None
    previous_date: datetime | None
    previous_tester: str | None
    previous_results: str | None
    raw: dict[str, Any] = field(default_factory=dict)


@dataclass
class AssignmentValueRow:
    """One ODP value row from the Assignment Values tab.

    Mapped 1:1 to a future :class:`OdpAssignment` row. ``odp_id`` is
    preserved verbatim from the workbook (``{$37$}`` for Rev 4,
    ``ac-02_odp.03`` for Rev 5) so the render-time tokenizer in
    :func:`controls.odp_render.resolve_odps` can match either syntax
    without normalizing here.
    """

    excel_row: int
    control_id: str  # canonicalized to upper-case "AC-2(1)"
    odp_id: str
    value: str
    assigned_from: str


@dataclass
class CcisIndex:
    """Result of parsing a CCIS workbook."""

    workbook_path: Path
    sheet_name: str
    rows: list[CcisRow]

    def by_control(self) -> dict[str, list[CcisRow]]:
        """Group rows by control ID (one control -> N CCIs)."""
        out: dict[str, list[CcisRow]] = {}
        for r in self.rows:
            out.setdefault(r.control_id, []).append(r)
        return out

    def by_cci(self) -> dict[str, CcisRow]:
        """Index by canonical CCI id."""
        return {r.cci_id: r for r in self.rows if r.cci_id}

    def by_ap(self) -> dict[str, CcisRow]:
        """Index by AP acronym (e.g. 'AC-2.1' — matches OSCAL objective ids)."""
        return {r.ap_acronym: r for r in self.rows if r.ap_acronym}


# ---------------------------------------------------------------------------
# Cell coercion helpers
# ---------------------------------------------------------------------------


def _resolve_sheet(wb) -> Worksheet:
    for name in _WORKING_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        if "working" in name.lower():
            return wb[name]
    raise ValueError(f"No WORKING SHEET found. Available sheets: {wb.sheetnames}")


def _normalize_cci_cell(raw: Any) -> str | None:
    """Col H may hold 'CCI-000015' or just '15' / '000015'. Both -> 'CCI-000015'."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Prefer prefixed match if present
    m = _CCI_PREFIXED_RE.search(s)
    if m:
        return f"CCI-{int(m.group(1)):06d}"
    # Fall back to bare integer (only acceptable in the dedicated column)
    m = _CCI_BARE_RE.search(s)
    if m:
        return f"CCI-{int(m.group(1)):06d}"
    return None


def _normalize_control(raw: Any) -> str | None:
    if raw is None:
        return None
    val = str(raw).strip().upper().replace(" ", "")
    if not val:
        return None
    return val if _CONTROL_RE.match(val) else val


def _coerce_text(raw: Any) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s or None


def _coerce_bool_yes(raw: Any) -> bool:
    if raw is None:
        return False
    return str(raw).strip().upper() == "YES"


def _coerce_date(raw: Any) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(str(raw).strip(), fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Public API: parse
# ---------------------------------------------------------------------------


_MAX_COL = COL_PREV_RESULTS  # furthest column we read (U = 21)


def _read_index_uncached(path: Path) -> CcisIndex:
    """The actual parse. Split out so the cached wrapper stays thin."""
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = _resolve_sheet(wb)

    rows: list[CcisRow] = []
    empty_run = 0

    # iter_rows in read-only mode streams the worksheet in a single pass.
    # Calling sheet.cell(row=R, column=C) instead is O(rows) per call because
    # openpyxl rescans XML each time — that pattern was costing ~123s on a
    # 319-row workbook. iter_rows drops it to ~1s.
    for row_idx, values in enumerate(
        sheet.iter_rows(min_row=7, max_row=_MAX_DATA_ROW, max_col=_MAX_COL, values_only=True),
        start=7,
    ):
        # Pad short rows so column indices below are safe.
        if len(values) < _MAX_COL:
            values = values + (None,) * (_MAX_COL - len(values))

        control_raw = values[COL_CONTROL - 1]
        control_id = _normalize_control(control_raw)
        if not control_id:
            empty_run += 1
            if empty_run >= _EMPTY_ROW_RUN_LIMIT:
                break
            continue
        empty_run = 0

        rows.append(
            CcisRow(
                excel_row=row_idx,
                required=_coerce_bool_yes(values[COL_REQUIRED - 1]),
                control_id=control_id,
                ap_acronym=_coerce_text(values[COL_AP_ACRONYM - 1]),
                cci_id=_normalize_cci_cell(values[COL_CCI - 1]),
                implementation_status=_coerce_text(values[COL_IMPL_STATUS - 1]),
                designation=_coerce_text(values[COL_DESIGNATION - 1]),
                narrative=_coerce_text(values[COL_NARRATIVE - 1]),
                definition=_coerce_text(values[COL_DEFINITION - 1]),
                guidance=_coerce_text(values[COL_GUIDANCE - 1]),
                procedures=_coerce_text(values[COL_PROCEDURES - 1]),
                inherited=_coerce_text(values[COL_INHERITED - 1]),
                remote_inheritance=_coerce_text(values[COL_REMOTE_INHERITANCE - 1]),
                status=_coerce_text(values[COL_STATUS - 1]),
                date_tested=_coerce_date(values[COL_DATE_TESTED - 1]),
                tester=_coerce_text(values[COL_TESTER - 1]),
                results=_coerce_text(values[COL_RESULTS - 1]),
                previous_status=_coerce_text(values[COL_PREV_STATUS - 1]),
                previous_date=_coerce_date(values[COL_PREV_DATE - 1]),
                previous_tester=_coerce_text(values[COL_PREV_TESTER - 1]),
                previous_results=_coerce_text(values[COL_PREV_RESULTS - 1]),
            )
        )

    sheet_title = sheet.title
    wb.close()
    return CcisIndex(workbook_path=path, sheet_name=sheet_title, rows=rows)


# Cache by (resolved-path, mtime_ns, size). The size catches editors that touch
# mtime without rewriting; mtime catches edits that preserve size. Bounded by
# the number of distinct workbooks the sidecar sees in one session — fine.
_INDEX_CACHE: dict[tuple[str, int, int], CcisIndex] = {}


def read_workbook_index(workbook_path: str | Path) -> CcisIndex:
    """Parse a CCIS workbook into a structured index. No DB writes.

    Idempotent and read-only. Safe to call from find-evidence,
    review-assessment, or any UI screen that needs the latest workbook
    snapshot without committing to the catalog.

    Results are cached by (path, mtime, size) within the process — repeated
    calls (e.g. PDF + SAR back-to-back) reuse the parse.
    """
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"CCIS workbook not found: {path}")

    stat = path.stat()
    key = (str(path.resolve()), stat.st_mtime_ns, stat.st_size)
    cached = _INDEX_CACHE.get(key)
    if cached is not None:
        return cached

    index = _read_index_uncached(path)
    _INDEX_CACHE[key] = index
    return index


# ---------------------------------------------------------------------------
# Public API: Assignment Values (ODPs)
# ---------------------------------------------------------------------------


def _resolve_assignment_values_sheet(wb) -> Worksheet | None:
    """Find the Assignment Values tab. None (not raise) if absent —
    older/lighter workbooks may not include the tab."""
    for name in _ASSIGNMENT_VALUES_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        if "assignment" in name.lower() and "value" in name.lower():
            return wb[name]
    return None


# Header matching for the four columns we care about.
#
# The eMASS export packs FIVE columns containing "control" ("Control
# Acronym", "Control Name", "Control Set", "Control Description (Base)",
# "Control Description (with IDs)") and FOUR columns containing "value"
# ("Assignment Value ID", "Assignment Value Question Text", "Assignment
# Value Description", "Assignment Value"). Substring matching is too
# coarse — we must prefer the SPECIFIC labels first, then fall back to
# generic tokens for less-strict eMASS revs.
#
# Order matters in _match_header: more specific phrases checked first.

# Exact (case-insensitive) header → field mapping. Wins over fuzzy match.
_AV_EXACT_HEADERS: dict[str, str] = {
    "control acronym": "control_id",
    "assignment value id": "odp_id",
    "assignment value": "value",
    "assigned from": "assigned_from",
    # The parameterized control statement enumerates EVERY ODP slot for
    # the control with its {$N$} (Rev 4) or _odp.NN (Rev 5) token,
    # including slots that have no assigned value. This is the
    # authoritative slot ordering for the OSCAL positional bridge —
    # without it, the bridge would see only value-bearing rows and
    # mis-align controls whose workbook intentionally leaves some
    # slots unassigned (e.g. AC-2 in Example System, where {$36$} and
    # {$38$} are never filled in).
    "parameterized control": "parameterized_statement",
    "control with assignments": "parameterized_statement",
}


def _match_header(cell_text: str) -> str | None:
    """Return the canonical field name this header cell maps to, or None.

    Priority:
      1. Exact match against ``_AV_EXACT_HEADERS`` (covers current eMASS
         template). Handles the "Assignment Value ID" vs "Assignment
         Value" collision and the five "Control *" columns.
      2. Fuzzy substring fallback for older/drifted templates. Each
         check excludes tokens that would collide with another field.
    """
    s = cell_text.strip().lower()
    if not s:
        return None

    # 1. Exact-match the known canonical labels.
    if s in _AV_EXACT_HEADERS:
        return _AV_EXACT_HEADERS[s]

    # 2. Fuzzy fallback (older templates). Parameterized-statement
    # detection runs FIRST so its tokens win over the generic "parameter"
    # → odp_id rule below (the column header literally contains both
    # "parameterized" and "control" and we don't want it routed to
    # control_id either).
    if "parameterized" in s or "with assignments" in s or "with ids" in s:
        return "parameterized_statement"
    # odp_id checked before value because some templates use
    # "parameter value id" / "ODP value".
    if any(tok in s for tok in ("parameter", "odp", "variable")):
        return "odp_id"
    if "assignment id" in s or "assignment value id" in s:
        return "odp_id"
    if any(tok in s for tok in ("source", "assigned", "from", "origin", "overlay")):
        return "assigned_from"
    # Only treat a "control" column as control_id when it also says
    # "acronym" — otherwise it's "Control Name" / "Control Set" /
    # "Control Description", which we don't want.
    if "control" in s and "acronym" in s:
        return "control_id"
    # "value" is the loosest token; only accept if no other qualifier
    # turns it into a non-value column.
    if "value" in s and not any(
        qual in s for qual in ("id", "question", "description")
    ):
        return "value"
    return None


# Rev 4 ODP id token format. Workbook stores the bare integer (37);
# template carries "{$37$}". Normalize at read time so the stored key
# matches what odp_render.resolve_odps() looks up.
_REV4_INT_RE = re.compile(r"^\d+$")


# Slot tokenizer for the parameterized control statement column. Matches
# the SAME Rev 4 ``{$N$}`` and Rev 5 ``ac-XX_odp.NN`` forms that
# controls/odp_render.py's _ODP_PATTERN handles — they must stay in sync
# so a slot seen at ingest is a slot the renderer can later look up.
# We intentionally exclude the OSCAL wrapper form here because the
# parameterized statement column is the eMASS authoring view (Rev 4
# bare-integer tokens), not the OSCAL prose form.
_SLOT_PATTERN = re.compile(
    r"(?P<rev4>\{\$\d+\$\})"  # {$37$}
    r"|"
    r"(?P<rev5>[a-z]{2}-\d{1,2}(?:\(\d+\))?_odp(?:\.\d+)?)",  # ac-02_odp.03
    re.IGNORECASE,
)


def _extract_slot_order(statement: str | None) -> list[str]:
    """Return the slot tokens in first-occurrence order.

    Dedup is required because the parameterized statement may reference
    the same slot twice (eMASS sometimes echoes ``{$39$}`` in the
    Supplemental Guidance block as well as the main statement). The
    canonical OSCAL positional-bridge alignment is one entry per
    distinct slot, in the order it first appears.
    """
    if not statement:
        return []
    seen: set[str] = set()
    order: list[str] = []
    for m in _SLOT_PATTERN.finditer(statement):
        tok = m.group(0)
        if tok not in seen:
            seen.add(tok)
            order.append(tok)
    return order


def _normalize_odp_id(raw: Any) -> str | None:
    """Coerce an Assignment Value ID cell into a render-comparable token.

    The eMASS workbook stores Rev 4 ODP ids as bare integers (``37``)
    even though the control statement template carries them as
    ``{$37$}``. Wrap integers in ``{$...$}`` so the stored key equals
    the placeholder string. Rev 5 ids (``ac-02_odp.03``) are already
    canonical — pass through unchanged.
    """
    if raw is None:
        return None
    # openpyxl returns int for numeric cells; coerce to string first.
    if isinstance(raw, int):
        return f"{{${raw}$}}"
    s = str(raw).strip()
    if not s:
        return None
    # String-form integer ("37" or "37.0") — also Rev 4.
    if _REV4_INT_RE.match(s):
        return f"{{${int(s)}$}}"
    if s.endswith(".0") and _REV4_INT_RE.match(s[:-2]):
        return f"{{${int(float(s))}$}}"
    # Already a Rev 4 token, or a Rev 5 token like "ac-02_odp.03".
    return s


def read_assignment_values(
    workbook_path: str | Path,
) -> tuple[list[AssignmentValueRow], dict[str, list[str]]]:
    """Parse the Assignment Values tab.

    Returns a 2-tuple:

    * ``rows`` — deduped list of value-bearing ODP rows. One per
      ``(control_id, odp_id, assigned_from)`` triple.
    * ``slot_orders`` — ``{control_id: [odp_id, ...]}`` capturing EVERY
      ODP slot the control declares, in first-occurrence order, drawn
      from the parameterized control statement column. This is the
      authoritative source for the OSCAL positional bridge in
      :mod:`baselines.ccis_workbook` Step 6b because the eMASS workbook
      does NOT emit a row for slots that have no assigned value
      (sparse-workbook problem — e.g. AC-2 in Example System declares four
      slots but only fills two of them). Deriving slot order from
      ``rows`` alone would mis-count and trigger a count-mismatch
      abstain on every sparse control.

    Pure function — no DB, no session. Returns ``([], {})`` if the
    workbook has no Assignment Values tab (older eMASS exports omit
    it).

    The parser:
      * Scans the first 10 rows for a header row containing recognizable
        column titles, then maps physical column indexes from the match.
      * Reads data rows until ``_EMPTY_ROW_RUN_LIMIT`` blank
        ``control_id`` cells in a row, mirroring WORKING SHEET behavior.
      * Dedups within the parsed list on
        ``(control_id, odp_id, assigned_from)`` so re-imports of an
        unchanged workbook don't fire the OdpAuditLog diff path on every
        row.
      * Captures the parameterized statement column (if recognized) on
        the first row seen for each control; all rows for the same
        control share the same statement so first-seen-wins is exact.
        When the column is absent (older exports), ``slot_orders`` falls
        back to value-bearing slot order — same behavior as v0.1
        pre-Option-A, which is fine for non-sparse controls.

    Raises ``ValueError`` if the tab is present but no recognizable
    header row can be found — that means the workbook structure has
    drifted further than this reader knows how to handle, and silent
    skip would mask data loss.
    """
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"CCIS workbook not found: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = _resolve_assignment_values_sheet(wb)
        if sheet is None:
            return [], {}

        # Find header row: scan up to row 10, pick the row that maps the
        # most of our recognized fields. First-write-wins per field so
        # an early "Control Acronym" column doesn't get overwritten by a
        # later "Control Description" column. parameterized_statement is
        # optional and contributes to the score so we prefer a header
        # row that exposes it when available.
        header_map: dict[str, int] = {}
        header_row_idx: int | None = None
        best_score = 0
        for row_idx, values in enumerate(
            sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1
        ):
            candidate: dict[str, int] = {}
            for col_idx, raw in enumerate(values, start=1):
                if raw is None:
                    continue
                field_name = _match_header(str(raw))
                if field_name and field_name not in candidate:
                    candidate[field_name] = col_idx
            score = len(candidate)
            if score > best_score:
                best_score = score
                header_map = candidate
                header_row_idx = row_idx
            # Perfect match — stop early. 5 = the four required fields
            # plus parameterized_statement.
            if score == 5:
                break

        required_fields = {"control_id", "odp_id", "value", "assigned_from"}
        missing = required_fields - set(header_map.keys())
        if header_row_idx is None or missing:
            raise ValueError(
                f"Assignment Values tab in {path.name} has unrecognized "
                f"header layout (missing fields: {sorted(missing) or 'all'}). "
                "Update _AV_HEADER_TOKENS or _match_header to match."
            )

        param_col = header_map.get("parameterized_statement")

        data_start = header_row_idx + 1
        max_col = max(header_map.values())

        rows: list[AssignmentValueRow] = []
        seen_keys: set[tuple[str, str, str]] = set()
        # Slot ordering per control. Captured from the parameterized
        # statement column on the first row seen for each control (all
        # rows for the same control share the same statement). If the
        # column is absent or empty, the fallback below derives order
        # from value-bearing rows — fine for non-sparse controls.
        slot_orders: dict[str, list[str]] = {}
        # Fallback accumulator: per-control, the slots we actually saw
        # rows for, in first-occurrence order. Used only when the
        # parameterized statement column didn't yield a slot order.
        fallback_orders: dict[str, list[str]] = {}
        fallback_seen: dict[str, set[str]] = {}
        empty_run = 0

        for row_idx, values in enumerate(
            sheet.iter_rows(
                min_row=data_start,
                max_row=_MAX_DATA_ROW,
                max_col=max_col,
                values_only=True,
            ),
            start=data_start,
        ):
            if len(values) < max_col:
                values = values + (None,) * (max_col - len(values))

            control_raw = values[header_map["control_id"] - 1]
            control_id = _normalize_control(control_raw)
            if not control_id:
                empty_run += 1
                if empty_run >= _EMPTY_ROW_RUN_LIMIT:
                    break
                continue
            empty_run = 0

            # Capture the parameterized statement once per control. The
            # first row for a control wins; subsequent rows redundantly
            # carry the same statement. If extraction yields an empty
            # list (column was blank for this row), don't lock in the
            # empty list — try again on the next row of the same control.
            if param_col is not None and control_id not in slot_orders:
                stmt_raw = values[param_col - 1]
                stmt = _coerce_text(stmt_raw) if stmt_raw is not None else None
                order = _extract_slot_order(stmt)
                if order:
                    slot_orders[control_id] = order

            odp_raw = values[header_map["odp_id"] - 1]
            value_raw = values[header_map["value"] - 1]
            assigned_raw = values[header_map["assigned_from"] - 1]

            # ODP id needs Rev 4 integer → "{$N$}" wrapping; see
            # _normalize_odp_id docstring for why.
            odp_id = _normalize_odp_id(odp_raw)
            value = _coerce_text(value_raw)
            assigned_from = _coerce_text(assigned_raw)

            # ODP id required, value is NOT — empty-value rows preserve the
            # slot identity used by the OSCAL positional bridge in
            # baselines/ccis_workbook.py Step 6b. Without the empty rows,
            # a sparse workbook (e.g. AC-2: only {$37$} and {$39$} filled
            # of four declared slots) would cause Step 6b's count check
            # to mismatch the OSCAL param count and abstain on the whole
            # control. Render layer treats empty value as unresolved so
            # the assessor still sees a visible placeholder.
            if not odp_id:
                continue
            if value is None:
                value = ""
            if not assigned_from:
                assigned_from = "workbook"

            # Track fallback slot order regardless of whether the
            # parameterized column produced one — used only when that
            # column was missing/blank.
            ctrl_seen = fallback_seen.setdefault(control_id, set())
            if odp_id not in ctrl_seen:
                ctrl_seen.add(odp_id)
                fallback_orders.setdefault(control_id, []).append(odp_id)

            key = (control_id, odp_id, assigned_from)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            rows.append(
                AssignmentValueRow(
                    excel_row=row_idx,
                    control_id=control_id,
                    odp_id=odp_id,
                    value=value,
                    assigned_from=assigned_from,
                )
            )

        # Backfill any control whose parameterized statement was missing
        # or unparseable with the value-bearing fallback order.
        for ctrl_id, order in fallback_orders.items():
            slot_orders.setdefault(ctrl_id, order)

        return rows, slot_orders
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Public API: upsert into catalog
# ---------------------------------------------------------------------------


def _ccis_to_oscal_control_id(ccis_id: str) -> str:
    """'AC-2(1)' -> 'ac-2.1' (matches OSCAL canonical form).

    Translation is deterministic so storing one canonical form in the
    Control row is enough — workbook reads translate on lookup.
    """
    s = ccis_id.strip().lower()
    s = re.sub(r"\((\d+)\)", r".\1", s)
    return s


def read_workbook_summary(workbook_path: str | Path) -> dict[str, Any]:
    """Lightweight summary for the Workbooks UI.

    Returns counts and a status histogram without exposing every row —
    cheap enough to call on every open / refresh.
    """
    index = read_workbook_index(workbook_path)
    status_counts: dict[str, int] = {}
    required = 0
    with_cci = 0
    for row in index.rows:
        if row.required:
            required += 1
        if row.cci_id:
            with_cci += 1
        key = (row.status or "(unassessed)").strip() or "(unassessed)"
        status_counts[key] = status_counts.get(key, 0) + 1

    controls_seen = len({row.control_id for row in index.rows})
    return {
        "filename": Path(workbook_path).name,
        "sheet": index.sheet_name,
        "rows": len(index.rows),
        "controls": controls_seen,
        "required": required,
        "with_cci": with_cci,
        "status_counts": status_counts,
    }


def populate_objectives(
    session: Session, *, framework_id: int, index: CcisIndex
) -> dict[str, int]:
    """Upsert Objective rows from a parsed CCIS index.

    For each row with a CCI id, find the matching Control (by translated
    id) and create/update an Objective row populated with cols I/J/K.
    Controls the OSCAL loader didn't produce are counted under
    ``missing_controls`` — this happens when a program overlay adds
    controls outside the rev-5 baseline.

    Returns counts: created, updated, missing_controls, missing_cci.
    """
    framework = session.get(Framework, framework_id)
    if framework is None:
        raise ValueError(f"Framework id={framework_id} does not exist")

    control_lookup: dict[str, int] = {
        c.control_id: c.id  # type: ignore[misc]
        for c in session.exec(select(Control).where(Control.framework_id == framework_id)).all()
        if c.id is not None
    }

    created = 0
    updated = 0
    missing_controls = 0
    missing_cci = 0

    for row in index.rows:
        if not row.cci_id:
            missing_cci += 1
            continue
        oscal_id = _ccis_to_oscal_control_id(row.control_id)
        control_pk = control_lookup.get(oscal_id)
        if control_pk is None:
            missing_controls += 1
            continue

        existing = session.exec(
            select(Objective).where(
                Objective.control_id_fk == control_pk,
                Objective.objective_id == row.cci_id,
            )
        ).first()
        if existing is None:
            session.add(
                Objective(
                    control_id_fk=control_pk,
                    objective_id=row.cci_id,
                    source="CCI",
                    text=row.definition or "",
                    implementation_guidance=row.guidance,
                    assessment_procedures=row.procedures,
                )
            )
            created += 1
        else:
            existing.text = row.definition or existing.text
            existing.implementation_guidance = row.guidance
            existing.assessment_procedures = row.procedures
            session.add(existing)
            updated += 1

    session.commit()
    return {
        "created": created,
        "updated": updated,
        "missing_controls": missing_controls,
        "missing_cci": missing_cci,
    }


# ---------------------------------------------------------------------------
# Re-read mode: snapshot + diff
# ---------------------------------------------------------------------------

# Columns compared by the diff when classifying a row as "edited".
# F = narrative, I = CCI definition, J = guidance, K = procedures,
# N = compliance status, O = date tested, P = tester, Q = test results,
# U = previous test results (kept because reviewers care if it shifts).
_DIFF_COLUMNS: tuple[tuple[str, str], ...] = (
    ("F", "narrative"),
    ("I", "definition"),
    ("J", "guidance"),
    ("K", "procedures"),
    ("N", "status"),
    ("O", "date_tested"),
    ("P", "tester"),
    ("Q", "results"),
    ("U", "previous_results"),
)

# Sidecar file suffix. Kept next to the workbook so it travels with it.
_SNAPSHOT_SUFFIX = ".snapshot.json"


RowKey = tuple[str, str]
"""Composite key: (control_id, cci_id). Stable across re-orderings."""


@dataclass
class RereadDiff:
    """Structured diff between previous snapshot and current workbook.

    All entries are dicts (not CcisRow) so the result round-trips cleanly
    through JSON for the UI/IPC layer.
    """

    added: list[dict[str, Any]] = field(default_factory=list)
    removed: list[dict[str, Any]] = field(default_factory=list)
    moved: list[dict[str, Any]] = field(default_factory=list)
    edited: list[dict[str, Any]] = field(default_factory=list)

    @property
    def is_empty(self) -> bool:
        return not (self.added or self.removed or self.moved or self.edited)


@dataclass
class RereadResult:
    """Output of :func:`reread_workbook` — fresh rows + diff."""

    index: CcisIndex
    diff: RereadDiff
    had_prior_snapshot: bool


def _snapshot_path(workbook_path: str | Path) -> Path:
    return Path(str(workbook_path) + _SNAPSHOT_SUFFIX)


def _row_key(row_or_dict: CcisRow | dict[str, Any]) -> RowKey | None:
    """Return ``(control_id, cci_id)`` or None if either component is missing.

    Rows lacking a CCI id can't participate in the keyed diff because we
    have no stable identifier for them within a control.
    """
    if isinstance(row_or_dict, CcisRow):
        control = row_or_dict.control_id
        cci = row_or_dict.cci_id
    else:
        control = row_or_dict.get("control_id")
        cci = row_or_dict.get("cci_id")
    if not control or not cci:
        return None
    return (control, cci)


def _row_to_snapshot_dict(row: CcisRow) -> dict[str, Any]:
    """Serialize a CcisRow to a JSON-safe dict (datetimes -> ISO strings)."""
    d = asdict(row)
    # Drop the opaque raw bag — it's not stable and not needed for diffs.
    d.pop("raw", None)
    for k in ("date_tested", "previous_date"):
        v = d.get(k)
        if isinstance(v, datetime):
            d[k] = v.isoformat()
    return d


def _normalize_diff_value(value: Any) -> Any:
    """Reduce a cell to a canonical form for equality comparison.

    Whitespace-only strings collapse to None so a stray space doesn't
    register as an edit. Datetime objects compare against ISO strings
    written to the snapshot.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        return s or None
    return value


def _load_snapshot(path: Path) -> dict[RowKey, dict[str, Any]] | None:
    """Read the sidecar snapshot, or None if it doesn't exist / is invalid."""
    if not path.exists():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    rows = data.get("rows") if isinstance(data, dict) else None
    if not isinstance(rows, list):
        return None
    out: dict[RowKey, dict[str, Any]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = _row_key(row)
        if key is None:
            continue
        out[key] = row
    return out


def _write_snapshot(path: Path, index: CcisIndex) -> None:
    """Persist the current parse as the new baseline for next re-read."""
    payload = {
        "workbook_path": str(index.workbook_path),
        "sheet_name": index.sheet_name,
        "captured_at": datetime.now(timezone.utc).isoformat(),
        "rows": [_row_to_snapshot_dict(r) for r in index.rows],
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _diff_indexes(
    previous: dict[RowKey, dict[str, Any]],
    current: dict[RowKey, CcisRow],
) -> RereadDiff:
    """Compute the structured diff between the two keyed views."""
    diff = RereadDiff()

    prev_keys = set(previous.keys())
    curr_keys = set(current.keys())

    for key in sorted(curr_keys - prev_keys):
        row = current[key]
        diff.added.append({"key": list(key), "row": _row_to_snapshot_dict(row)})

    for key in sorted(prev_keys - curr_keys):
        diff.removed.append({"key": list(key), "row": previous[key]})

    for key in sorted(prev_keys & curr_keys):
        prev_row = previous[key]
        curr_row = current[key]
        prev_excel_row = prev_row.get("excel_row")
        if isinstance(prev_excel_row, int) and prev_excel_row != curr_row.excel_row:
            diff.moved.append(
                {
                    "key": list(key),
                    "old_row": prev_excel_row,
                    "new_row": curr_row.excel_row,
                }
            )

        changed: list[str] = []
        for col_letter, attr_name in _DIFF_COLUMNS:
            new_val = _normalize_diff_value(getattr(curr_row, attr_name))
            old_val = _normalize_diff_value(prev_row.get(attr_name))
            if new_val != old_val:
                changed.append(col_letter)
        if changed:
            diff.edited.append({"key": list(key), "changed_cols": changed})

    return diff


def reread_workbook(
    workbook_path: str | Path, *, update_snapshot: bool = True
) -> RereadResult:
    """Parse the workbook and diff it against the sidecar snapshot.

    On first invocation (no snapshot present) every row is reported as
    "added" — callers can detect this via ``had_prior_snapshot`` and
    suppress the noise if they want a true delta-only view.

    The snapshot is rewritten to reflect the current parse unless the
    caller passes ``update_snapshot=False`` (useful for read-only
    preview UIs).
    """
    index = read_workbook_index(workbook_path)
    snapshot_file = _snapshot_path(index.workbook_path)
    previous = _load_snapshot(snapshot_file)
    had_prior_snapshot = previous is not None

    current_keyed: dict[RowKey, CcisRow] = {}
    for row in index.rows:
        key = _row_key(row)
        if key is None:
            continue
        current_keyed[key] = row

    if previous is None:
        diff = RereadDiff(
            added=[
                {"key": list(k), "row": _row_to_snapshot_dict(v)}
                for k, v in sorted(current_keyed.items())
            ]
        )
    else:
        diff = _diff_indexes(previous, current_keyed)

    if update_snapshot:
        _write_snapshot(snapshot_file, index)

    return RereadResult(
        index=index,
        diff=diff,
        had_prior_snapshot=had_prior_snapshot,
    )
