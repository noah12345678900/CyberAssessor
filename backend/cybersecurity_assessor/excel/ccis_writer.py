"""CCIS workbook writer (headless .xlsx surgery).

Writes assessment results into the four assessor-writable cells per CCI row:

    Col N  Compliance Status   "Compliant" / "Non-Compliant" / "Not Applicable"
    Col O  Date Tested         datetime (rendered as ISO 8601 string)
    Col P  Tested By           assessor name
    Col Q  Test Results        facts-only narrative

Historically the write path drove a live Excel COM session through xlwings so
comments, named ranges, merged cells, conditional formatting, data validation,
and formula references would survive the write. openpyxl strips some of those
features silently on save, which is why xlwings was chosen originally.

This module now uses :mod:`xlsx_surgery` — a pure-Python ``.xlsx`` cell patcher
that opens the workbook as a zip of XML, surgically updates target ``<c>``
elements, and byte-copies every other zip part verbatim. That preserves the
exact same set of features xlwings did, without needing Excel installed on the
host. The ``safe_write`` backup-and-verify harness still wraps every write so a
corrupted or silently-dropped write is rolled back from a snapshot.
"""

from __future__ import annotations

import json
import shutil
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from ..models import ComplianceStatus
from .xlsx_surgery import CellValue, patch_cells as _patch_cells

# Sheet name candidates in priority order. Must match ccis_reader.
_WORKING_SHEET_NAMES = ["WORKING SHEET", "Working Sheet", "Working sheet"]

# Writable columns (1-based, matches A1 column letters via _col_letter).
COL_REQUIRED = 1  # A
COL_CONTROL = 2  # B
COL_AP_ACRONYM = 7  # G
COL_CCI = 8  # H
COL_DEFINITION = 9  # I
COL_GUIDANCE = 10  # J
COL_PROCEDURES = 11  # K
COL_STATUS = 14  # N
COL_DATE_TESTED = 15  # O
COL_TESTER = 16  # P
COL_RESULTS = 17  # Q

# First data row in WORKING SHEET (rows 1-5 are metadata, 6 is headers).
_FIRST_DATA_ROW = 7
# Upper bound matches ccis_reader._MAX_DATA_ROW.
_MAX_DATA_ROW = 500

# Date format written to col O. eMASS accepts ISO 8601.
_DATE_FMT = "%Y-%m-%d"


# ---------------------------------------------------------------------------
# Public dataclass: one write request
# ---------------------------------------------------------------------------


@dataclass
class CcisWrite:
    """One row's worth of assessor output, addressed by absolute Excel row.

    ``excel_row`` comes straight from the matching ``CcisRow.excel_row`` so
    callers never have to re-find the row. Any field set to ``None`` is left
    untouched in the workbook (allowing partial updates — e.g. only
    re-writing the narrative without touching status/date/tester).

    ``needs_review`` is the v0.2 precision-over-recall gate: when True the
    writer treats the row as abstained and writes NOTHING — status/date/
    tester/results are all left exactly as the user (or prior assessor)
    left them in the workbook. The skipped row is counted in the result's
    ``skipped_needs_review`` so callers (and curl bypass attempts) can see
    that the row was intentionally not written. This is a defense-in-depth
    layer beneath the UI's Apply-button hard-gate and the route's 409
    response — even if a stale client sneaks an apply through, the
    workbook never gets a triage-pending verdict baked in.
    """

    excel_row: int
    status: ComplianceStatus | str | None = None
    date_tested: datetime | None = None
    tester: str | None = None
    results: str | None = None
    needs_review: bool = False
    # v0.2 citation-hygiene fields (NOT an abstain). When ``rewrite_requested``
    # is True, the verdict still rides into the workbook — but a
    # "Cite refresh requested: 'legacy' → 'current'" footer is appended to the
    # Q-column results so the next narrative pass knows which doc cite to
    # swap. ``rewrite_requested_refs`` is the raw JSON-encoded
    # ``[[legacy, current], ...]`` list straight off Assessment; the writer
    # decodes it (and falls back to a generic footer when missing).
    rewrite_requested: bool = False
    rewrite_requested_refs: str | None = None


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _col_letter(col: int) -> str:
    """1-based column index → Excel letter. 1→A, 14→N, 17→Q, 27→AA."""
    letters = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return letters


def _resolve_sheet_name(path: Path) -> str:
    """Pick the actual sheet name in ``path`` matching our working-sheet ladder.

    Mirrors the resolution in :mod:`xlsx_surgery` but returns the *human* name
    (for use in the writer's result dict) rather than the zip path.
    """
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        names = wb.sheetnames
    finally:
        wb.close()
    for candidate in _WORKING_SHEET_NAMES:
        if candidate in names:
            return candidate
    for name in names:
        if "working" in name.lower():
            return name
    raise ValueError(
        f"No WORKING SHEET found in {path.name}. Sheets: {names}"
    )


def _coerce_status(value: ComplianceStatus | str | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, ComplianceStatus):
        return value.value
    s = str(value).strip()
    return s or None


def _coerce_date(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.strftime(_DATE_FMT)


def _format_cite_refresh_footer(refs_json: str | None) -> str | None:
    """Build the Q-column footer for a rewrite_requested row.

    ``refs_json`` is the raw JSON-encoded ``[[legacy, current], ...]`` list
    off ``Assessment.rewrite_requested_refs``. Returns a short footer the
    caller appends to the results cell — one line per pair, or a generic
    note when no pairs could be reconstructed (legacy rows where
    supersession couldn't recover the doc names).

    Mirrors the POAM generator's ``_render_cite_refresh_block`` in tone so
    reviewers see consistent language across both surfaces. Kept compact
    here because Excel cell wrap matters — POAM gets a markdown ``##``
    heading, the workbook cell gets a single-line lead-in.
    """
    if not refs_json or not refs_json.strip():
        # Empty, None, or whitespace-only — treat as "supersession couldn't
        # reconstruct pairs". Whitespace-only specifically defends against a
        # stale/malformed Assessment.rewrite_requested_refs row that would
        # otherwise fall through to json.loads, raise, and silently suppress
        # the footer entirely.
        return (
            "Cite refresh requested: legacy document reference detected; "
            "specific legacy/current pair could not be reconstructed. "
            "Re-run assess after updating the narrative to clear the flag."
        )
    try:
        decoded = json.loads(refs_json)
    except (TypeError, ValueError):
        return None
    if not isinstance(decoded, list):
        return None
    pairs: list[tuple[str, str]] = []
    for entry in decoded:
        if (
            isinstance(entry, (list, tuple))
            and len(entry) >= 2
            and entry[0]
            and entry[1]
        ):
            pairs.append((str(entry[0]), str(entry[1])))
    if not pairs:
        return None
    lines = ["Cite refresh requested (verdict still stands):"]
    for legacy, current in pairs:
        lines.append(f"- '{legacy}' \u2192 '{current}'")
    return "\n".join(lines)


def _build_row_cells(
    write: CcisWrite,
) -> tuple[dict[str, str], dict[str, str], dict[str, CellValue]]:
    """Compute the cell-level changes for a single ``CcisWrite``.

    Returns a tuple ``(display_changes, expected_full, patch_cells_map)``:
        - ``display_changes`` — summary for human-readable logs (Q-column
          values truncated to 80 chars).
        - ``expected_full`` — full text the safety harness compares
          against on read-back.
        - ``patch_cells_map`` — ``{A1: value}`` to feed into
          :func:`xlsx_surgery.patch_cells`.
    """
    changes: dict[str, str] = {}
    expected: dict[str, str] = {}
    cells: dict[str, CellValue] = {}

    # v0.2 precision-over-recall gate. needs_review rows are abstentions —
    # their proposed status isn't trusted yet, so we don't bake any of it
    # into the workbook. The caller counts these via skipped_needs_review.
    if write.needs_review:
        return changes, expected, cells

    row = write.excel_row

    status = _coerce_status(write.status)
    if status is not None:
        addr = f"{_col_letter(COL_STATUS)}{row}"
        cells[addr] = status
        changes[addr] = status
        expected[addr] = status

    date_str = _coerce_date(write.date_tested)
    if date_str is not None:
        addr = f"{_col_letter(COL_DATE_TESTED)}{row}"
        cells[addr] = date_str
        changes[addr] = date_str
        expected[addr] = date_str

    if write.tester is not None:
        addr = f"{_col_letter(COL_TESTER)}{row}"
        cells[addr] = write.tester
        changes[addr] = write.tester
        expected[addr] = write.tester

    if write.results is not None:
        results_text = write.results
        # v0.2 citation-hygiene: append a "Cite refresh requested" footer to
        # the Q-column text so the assessor's next pass sees which doc cites
        # to swap right in the workbook. Verdict still stands — this is NOT
        # an abstain, the row already passed the needs_review gate above.
        if write.rewrite_requested:
            footer = _format_cite_refresh_footer(write.rewrite_requested_refs)
            if footer:
                results_text = f"{results_text}\n\n{footer}"
        addr = f"{_col_letter(COL_RESULTS)}{row}"
        cells[addr] = results_text
        changes[addr] = (
            results_text if len(results_text) <= 80 else results_text[:77] + "..."
        )
        expected[addr] = results_text

    return changes, expected, cells


# ---------------------------------------------------------------------------
# Safety harness: backup-then-verify wrapper
# ---------------------------------------------------------------------------


# How many backups to retain per workbook. Anything older than the N most
# recent (by mtime) is deleted on every successful write so the assessor's
# Documents folder doesn't fill with .bak-* files from a busy session.
_BACKUP_RETENTION = 5

# Suffix prefix; the timestamp portion is compact ISO 8601 UTC with no
# colons so Windows treats it as a legal filename (`:` is illegal on NTFS).
_BACKUP_PREFIX = ".bak-"


class WorkbookWriteVerificationError(RuntimeError):
    """Raised when a post-write openpyxl read-back doesn't match what the
    writer claims to have written. The backup has already been restored by
    the time this is raised — the workbook on disk is back to its pre-write
    state.
    """


def _backup_path(workbook_path: Path) -> Path:
    # Compact ISO 8601 UTC — e.g. 20260603T142530Z. Filename-safe on
    # NTFS, sortable lexicographically, no separator chars.
    ts = datetime.now(tz=timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    return workbook_path.with_name(workbook_path.name + f"{_BACKUP_PREFIX}{ts}")


def _list_backups(workbook_path: Path) -> list[Path]:
    """All existing backups for ``workbook_path``, newest mtime last."""
    parent = workbook_path.parent
    prefix = workbook_path.name + _BACKUP_PREFIX
    backups = [p for p in parent.iterdir() if p.is_file() and p.name.startswith(prefix)]
    backups.sort(key=lambda p: p.stat().st_mtime)
    return backups


def _prune_backups(workbook_path: Path, *, keep: int = _BACKUP_RETENTION) -> None:
    backups = _list_backups(workbook_path)
    # Keep the ``keep`` newest entries (end of the sorted list).
    for stale in backups[:-keep] if keep > 0 else backups:
        try:
            stale.unlink()
        except OSError:
            # Best-effort cleanup — never let a leftover backup break the
            # write path. The assessor can sweep ``*.bak-*`` manually.
            pass


def _verify_writes(workbook_path: Path, expected: dict[str, str]) -> list[str]:
    """Read the workbook with openpyxl and return a list of cells whose
    actual value does not match the expected write. Empty list = all good.

    ``expected`` keys are A1 addresses; values are the stringified value
    the wrapper believes was written. We compare on stringified form
    because round-tripping may coerce the type (e.g. date strings stay
    strings unless a number format kicks in).
    """
    if not expected:
        return []
    # data_only=False so we read the raw cell values, not cached formula
    # results. read_only mode is fastest and doesn't lock the file.
    wb = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        # Pick the WORKING SHEET the same way as the writer.
        sheet = None
        names = wb.sheetnames
        for candidate in _WORKING_SHEET_NAMES:
            if candidate in names:
                sheet = wb[candidate]
                break
        if sheet is None:
            for name in names:
                if "working" in name.lower():
                    sheet = wb[name]
                    break
        if sheet is None:
            # Can't verify if we can't find the sheet. Treat as mismatch
            # so the caller restores the backup rather than silently
            # accepting an unverifiable write.
            return list(expected.keys())

        mismatches: list[str] = []
        for cell_addr, want in expected.items():
            got = sheet[cell_addr].value
            got_str = "" if got is None else str(got)
            want_str = "" if want is None else str(want)
            if got_str != want_str:
                mismatches.append(cell_addr)
        return mismatches
    finally:
        wb.close()


@contextmanager
def safe_write(workbook_path: Path):
    """Context manager that wraps a block of writes with a backup +
    post-write verification harness.

    Usage::

        with safe_write(path) as ctx:
            # ... do writes via xlsx_surgery.patch_cells ...
            ctx["expected"]["N7"] = "Compliant"   # record what was written
            # block exit: workbook is verified; backup pruned to 5

    The yielded context dict has two keys:
        - ``expected`` (dict[str, str]): callers append A1-addressed
          cell values they wrote so the harness can read them back.
        - ``backup`` (Path): the path of the snapshot taken before the
          block. The harness restores from it on verification failure.

    Raises:
        WorkbookWriteVerificationError if any cell in ``expected`` does
        not match its read-back value (after the writer has saved).
    """
    workbook_path = Path(workbook_path)
    backup = _backup_path(workbook_path)
    shutil.copy2(workbook_path, backup)

    ctx: dict[str, object] = {"expected": {}, "backup": backup}
    try:
        yield ctx
    except Exception as original_exc:
        # If the inner block exploded, the on-disk file may already have
        # been partially modified before the exception. Restore from
        # backup to be safe. If the rollback itself fails (e.g. the live
        # file is open in Excel — same lock that caused the write to fail
        # in the first place), DO NOT let the rollback exception mask the
        # real cause. Chain via __context__ so the original traceback is
        # still in the chain, and re-raise the original so route handlers
        # can map the actual failure type (PermissionError, ValueError…)
        # to the right HTTP status.
        try:
            shutil.copy2(backup, workbook_path)
        except Exception:
            pass
        try:
            backup.unlink()
        except OSError:
            pass
        raise original_exc

    expected: dict[str, str] = ctx["expected"]  # type: ignore[assignment]
    mismatches = _verify_writes(workbook_path, expected)
    if mismatches:
        # Roll back: copy backup over the corrupted live file, then drop
        # the backup so we don't leave a duplicate of the (now restored)
        # original behind.
        shutil.copy2(backup, workbook_path)
        try:
            backup.unlink()
        except OSError:
            pass
        raise WorkbookWriteVerificationError(
            f"Post-write verification failed for {workbook_path.name}. "
            f"Mismatched cells: {sorted(mismatches)}. Workbook restored "
            f"from backup."
        )

    # All good — prune so we keep at most _BACKUP_RETENTION snapshots.
    _prune_backups(workbook_path)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_assessment(
    workbook_path: str | Path,
    writes: Iterable[CcisWrite],
    *,
    save: bool = True,
    close: bool = False,
) -> dict[str, object]:
    """Write one or more CCI assessment results to a CCIS workbook.

    Args:
        workbook_path: Absolute path to the CCIS .xlsx file.
        writes: Iterable of ``CcisWrite`` records — one per CCI row.
        save: If True (default), commit the writes to disk. If False, the
            writes are computed but not applied (rarely useful; kept for
            API parity with the prior xlwings-based version).
        close: Deprecated no-op. The headless writer doesn't keep an Excel
            session open, so there's nothing to close. Accepted for API
            compatibility; ignored.

    Returns:
        Summary dict: ``{"rows_written": int, "cells_changed": int,
        "changes": {<cellA1>: <new_value_truncated>, ...}, "workbook":
        <name>, "sheet": <sheet_name>, "skipped_needs_review": int}``.

    Raises:
        FileNotFoundError: if the workbook does not exist.
        ValueError: if the WORKING SHEET cannot be found.
        WorkbookWriteVerificationError: if post-write read-back doesn't
            match what was written. Backup has been restored.
    """
    del close  # accepted-but-ignored; no live Excel session to close.
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"CCIS workbook not found: {path}")

    # Materialize the writes iterable up front so we can iterate inside
    # the safe_write block without re-entering a generator that may be
    # consumed.
    writes = list(writes)

    sheet_name = _resolve_sheet_name(path)

    with safe_write(path) as ctx:
        all_changes: dict[str, str] = {}
        all_expected: dict[str, str] = {}
        all_cells: dict[str, CellValue] = {}
        rows_written = 0
        skipped_needs_review = 0

        for write in writes:
            if write.needs_review:
                # Defensive skip — UI hard-gates Apply on needs_review and
                # the route returns 409, but a stale client or direct curl
                # could still construct a CcisWrite for an abstained row.
                # Count it but write nothing.
                skipped_needs_review += 1
                continue
            row_changes, row_expected, row_cells = _build_row_cells(write)
            if row_cells:
                rows_written += 1
                all_changes.update(row_changes)
                all_expected.update(row_expected)
                all_cells.update(row_cells)

        if save and all_cells:
            _patch_cells(path, sheet_name, all_cells)

        # Tell the safety harness what we expect to see on read-back.
        # Only register cells we actually saved; if save=False the
        # writes live only as a return-value summary and there's nothing
        # to verify against the on-disk file.
        if save:
            ctx["expected"].update(all_expected)  # type: ignore[union-attr]

        result = {
            "rows_written": rows_written,
            "cells_changed": len(all_changes),
            "changes": all_changes,
            "skipped_needs_review": skipped_needs_review,
            "workbook": path.name,
            "sheet": sheet_name,
        }

    return result


def write_single(
    workbook_path: str | Path,
    *,
    excel_row: int,
    status: ComplianceStatus | str | None = None,
    date_tested: datetime | None = None,
    tester: str | None = None,
    results: str | None = None,
    rewrite_requested: bool = False,
    rewrite_requested_refs: str | None = None,
    save: bool = True,
    close: bool = False,
) -> dict[str, object]:
    """Convenience wrapper: write a single CCI row.

    Mirrors the plugin's single-cell write pattern (one CCI at a time from
    the Control Detail screen). For batch writes from ``/assess-control``
    use ``write_assessment`` with multiple ``CcisWrite`` records.

    ``rewrite_requested`` / ``rewrite_requested_refs`` are the v0.2 citation
    -hygiene fields — when True, ``_build_row_cells`` appends a "Cite
    refresh requested" footer to the Q-column text. NOT an abstain; the
    verdict still stands.
    """
    return write_assessment(
        workbook_path,
        [
            CcisWrite(
                excel_row=excel_row,
                status=status,
                date_tested=date_tested,
                tester=tester,
                results=results,
                rewrite_requested=rewrite_requested,
                rewrite_requested_refs=rewrite_requested_refs,
            )
        ],
        save=save,
        close=close,
    )


# ---------------------------------------------------------------------------
# Row insertion: close a workbook completeness gap
# ---------------------------------------------------------------------------


def _normalize_control_for_match(value: object) -> str | None:
    """Mirror ccis_reader._normalize_control for col B comparisons."""
    if value is None:
        return None
    s = str(value).strip().upper().replace(" ", "")
    return s or None


def _canonical_cci(value: str) -> str:
    """Force a CCI id to canonical 'CCI-NNNNNN' form."""
    import re

    m = re.search(r"(\d{1,7})", str(value))
    if not m:
        raise ValueError(f"Cannot parse CCI id from {value!r}")
    return f"CCI-{int(m.group(1)):06d}"


def _scan_col_b(path: Path, sheet_name: str) -> list[tuple[int, str | None]]:
    """Return ``[(row, normalized_control), ...]`` for rows 7..MAX in col B.

    Read-only openpyxl is fast and doesn't take a file lock. Used by
    :func:`insert_cci_row` to locate the right insertion point.
    """
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        ws = wb[sheet_name]
        out: list[tuple[int, str | None]] = []
        # ws.iter_rows is the read-only path; we ask for col B only.
        for row_idx in range(_FIRST_DATA_ROW, _MAX_DATA_ROW + 1):
            cell = ws.cell(row=row_idx, column=COL_CONTROL)
            out.append((row_idx, _normalize_control_for_match(cell.value)))
        return out
    finally:
        wb.close()


def insert_cci_row(
    workbook_path: str | Path,
    *,
    control_id: str,
    cci_id: str,
    ap_acronym: str | None = None,
    definition: str | None = None,
    guidance: str | None = None,
    procedures: str | None = None,
    required: bool = True,
    save: bool = True,
) -> int:
    """Insert a new CCI row into the WORKING SHEET.

    Used when the framework catalog (DB) knows a CCI belongs to a control
    but the workbook is missing that row (eMASS Export variants
    occasionally omit catalog CCIs). The new row goes immediately after
    the LAST existing row for the same control so CCIs stay visually
    grouped. If the control has no rows yet, the row is appended after the
    last non-blank col B row.

    Uses :func:`xlsx_surgery.patch_cells` with ``insert_row_before`` so
    existing rows shift down and sqref ranges (mergeCells, dataValidation,
    conditionalFormatting) that span the insert point are extended — the
    new row inherits dropdowns and conditional formatting in the same way
    Excel's ``Rows.Insert()`` did under xlwings.

    Wrapped in ``safe_write`` so a backup is taken and verification rolls
    back on read-back mismatch.

    Returns the new ``excel_row`` (1-based).

    Raises:
        FileNotFoundError: if the workbook does not exist.
        ValueError: if the WORKING SHEET cannot be found, the CCI id
            cannot be parsed, or the workbook is already at the max
            data-row cap.
        WorkbookWriteVerificationError: if post-write read-back doesn't
            match what was written. Backup has been restored.
    """
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"CCIS workbook not found: {path}")

    control_norm = _normalize_control_for_match(control_id)
    if not control_norm:
        raise ValueError(f"control_id is empty or invalid: {control_id!r}")
    cci_canonical = _canonical_cci(cci_id)

    sheet_name = _resolve_sheet_name(path)

    with safe_write(path) as ctx:
        # Scan col B to find the insertion point. Track the last row that
        # matches our control, plus the last non-blank row overall as a
        # fallback when the control isn't in the workbook yet.
        last_match_row: int | None = None
        last_nonblank_row: int = _FIRST_DATA_ROW - 1
        for row_num, norm in _scan_col_b(path, sheet_name):
            if norm is None:
                # Don't break early — some templates leave occasional
                # blank rows between control groups. _MAX_DATA_ROW caps
                # the scan.
                continue
            last_nonblank_row = row_num
            if norm == control_norm:
                last_match_row = row_num

        if last_match_row is not None:
            insert_at = last_match_row + 1
        else:
            insert_at = last_nonblank_row + 1

        if insert_at > _MAX_DATA_ROW:
            raise ValueError(
                f"Workbook is full — cannot insert row beyond {_MAX_DATA_ROW}. "
                f"Last non-blank row: {last_nonblank_row}."
            )

        # Build the cells dict for the new row. Addresses use post-insert
        # row numbers — xlsx_surgery first bumps rows >= insert_at, then
        # applies these cells, so insert_at is the right index here.
        expected: dict[str, str] = {}
        cells: dict[str, CellValue] = {}

        required_val = "YES" if required else ""
        a_addr = f"{_col_letter(COL_REQUIRED)}{insert_at}"
        cells[a_addr] = required_val if required_val else None
        if required_val:
            expected[a_addr] = required_val

        b_addr = f"{_col_letter(COL_CONTROL)}{insert_at}"
        cells[b_addr] = control_id
        expected[b_addr] = control_id

        if ap_acronym is not None:
            g_addr = f"{_col_letter(COL_AP_ACRONYM)}{insert_at}"
            cells[g_addr] = ap_acronym
            expected[g_addr] = ap_acronym

        h_addr = f"{_col_letter(COL_CCI)}{insert_at}"
        cells[h_addr] = cci_canonical
        expected[h_addr] = cci_canonical

        if definition is not None:
            i_addr = f"{_col_letter(COL_DEFINITION)}{insert_at}"
            cells[i_addr] = definition
            expected[i_addr] = definition

        if guidance is not None:
            j_addr = f"{_col_letter(COL_GUIDANCE)}{insert_at}"
            cells[j_addr] = guidance
            expected[j_addr] = guidance

        if procedures is not None:
            k_addr = f"{_col_letter(COL_PROCEDURES)}{insert_at}"
            cells[k_addr] = procedures
            expected[k_addr] = procedures

        # Clear writable assessor cells in the freshly-inserted row so
        # we don't carry stale prior-assessor verdicts into the new CCI
        # row. (The bumped empty <row r="insert_at"/> already has nothing,
        # but we set them explicitly for traceability.)
        for col in (COL_STATUS, COL_DATE_TESTED, COL_TESTER, COL_RESULTS):
            cells[f"{_col_letter(col)}{insert_at}"] = None

        if save:
            _patch_cells(path, sheet_name, cells, insert_row_before=insert_at)
            ctx["expected"].update(expected)  # type: ignore[union-attr]

    return insert_at
