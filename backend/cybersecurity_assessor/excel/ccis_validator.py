"""CCIS workbook validator (openpyxl, read-only).

Catches the structural problems that would silently corrupt a re-import:
columns shifted out of position, duplicate ``(control_id, cci_id)`` keys
from a bad family-partition merge, orphan rows with a Status filled in but
no control identity, etc.

The contract with the user is three buckets:

  errors    Block re-import. The workbook is structurally broken or two
            rows claim the same identity — either silently overwrites real
            work on round-trip, so we refuse.
  warnings  Surface to the assessor but don't block. Things that *could*
            be a problem (unknown control id, formula in a value cell,
            stray text below the data area) but might also be intentional.
  tolerated Trailing fully-blank rows. eMASS templates pad with hundreds
            of empty rows; never count them as errors. We scan bottom-up
            from ``ws.max_row`` and stop at the last row with any data,
            so trailing nulls are invisible to the report.

Reuses the column constants and normalization helpers from ``ccis_reader``
so the two stay in lock-step. Any column-layout change here MUST be made
to ``ccis_reader`` in the same commit.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .ccis_reader import (
    COL_AP_ACRONYM,
    COL_CCI,
    COL_CONTROL,
    COL_DATE_TESTED,
    COL_REQUIRED,
    COL_RESULTS,
    COL_STATUS,
    COL_TESTER,
    _coerce_date,
    _coerce_text,
    _normalize_cci_cell,
    _normalize_control,
    _resolve_sheet,
)

# Expected header text at row 6 — used to detect "columns shifted" failures.
# Match is case-insensitive substring, so eMASS template tweaks like
# "Compliance Status (current)" still pass.
_EXPECTED_HEADERS: dict[int, str] = {
    COL_REQUIRED: "required",
    COL_CONTROL: "control",
    COL_AP_ACRONYM: "ap",
    COL_CCI: "cci",
    COL_STATUS: "compliance status",
    COL_DATE_TESTED: "date tested",
    COL_TESTER: "tested by",
    COL_RESULTS: "test results",
}

_HEADER_ROW = 6
_FIRST_DATA_ROW = 7

# Valid compliance statuses. Empty string is also valid (unassessed).
_VALID_STATUSES: frozenset[str] = frozenset(
    {"Compliant", "Non-Compliant", "Not Applicable"}
)

# Columns scanned by the bottom-up "last data row" sweep. We pick the
# columns an assessor would actually populate; the eMASS pre-fill columns
# (B through K) all stay set even on rows where the assessor hasn't done
# anything, so including them in the scan would defeat trailing-null
# tolerance. Limiting to writable + identity columns gives the right
# "stops at the last row a human touched" semantics.
_DATA_COLUMNS = (
    COL_CONTROL,
    COL_CCI,
    COL_STATUS,
    COL_DATE_TESTED,
    COL_TESTER,
    COL_RESULTS,
)


# ---------------------------------------------------------------------------
# Public dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ValidationIssue:
    """One problem with the workbook.

    ``excel_row`` is None for workbook-level problems (missing sheet, header
    row shifted, etc.); set for row-scoped problems so the UI can deep-link.
    ``cell`` is the A1 address (e.g. ``"N42"``) when the issue is cell-scoped.
    """

    code: str  # stable machine identifier, e.g. "duplicate_identity"
    message: str  # human-readable explanation
    excel_row: int | None = None
    cell: str | None = None
    control_id: str | None = None
    cci_id: str | None = None


@dataclass
class ValidationReport:
    """Full validator output for one workbook.

    ``valid`` is False if there are any errors; warnings alone do not flip
    it. Callers (UI + re-import path) gate on ``valid``: re-import refuses
    when False, but the UI still shows warnings on the import-preview
    screen so the assessor can decide whether to clean them up.
    """

    workbook_path: Path
    sheet_name: str | None
    header_row: int | None  # None if the header row could not be located
    first_data_row: int
    last_data_row: int  # 6 if the sheet is empty below the header
    data_row_count: int  # last_data_row - first_data_row + 1, or 0 if empty
    errors: list[ValidationIssue] = field(default_factory=list)
    warnings: list[ValidationIssue] = field(default_factory=list)

    @property
    def valid(self) -> bool:
        return not self.errors


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def validate_workbook(
    workbook_path: str | Path,
    *,
    known_control_ids: Iterable[str] | None = None,
) -> ValidationReport:
    """Validate a CCIS workbook end-to-end.

    Args:
        workbook_path: Absolute path to the .xlsx file.
        known_control_ids: Optional iterable of control IDs the catalog
            knows about (OSCAL canonical form, e.g. ``"ac-2.1"``). Any row
            whose normalized control is not in this set emits a *warning*
            (program overlays add controls outside rev5 — see
            ``ccis_reader._ccis_to_oscal_control_id``). Pass None to skip
            the check entirely (typical when the catalog is not loaded).

    Returns:
        ``ValidationReport`` — never raises for content-level problems;
        only raises for file-system / OS-level failures.

    Raises:
        FileNotFoundError: if the workbook does not exist on disk.
    """
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(f"CCIS workbook not found: {path}")

    # read_only=True: this function never calls wb.save(). Opening
    # read-write would let a crash mid-validation leave Excel holding a
    # write lock on the user's original CCIS workbook -- precisely the
    # footgun the working-copy split was meant to close.
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        try:
            sheet = _resolve_sheet(wb)
        except ValueError as exc:
            return ValidationReport(
                workbook_path=path,
                sheet_name=None,
                header_row=None,
                first_data_row=_FIRST_DATA_ROW,
                last_data_row=_HEADER_ROW,
                data_row_count=0,
                errors=[ValidationIssue(code="missing_working_sheet", message=str(exc))],
            )

        report = ValidationReport(
            workbook_path=path,
            sheet_name=sheet.title,
            header_row=_HEADER_ROW,
            first_data_row=_FIRST_DATA_ROW,
            last_data_row=_HEADER_ROW,
            data_row_count=0,
        )

        _check_headers(sheet, report)
        last_row = _last_data_row(sheet)
        report.last_data_row = last_row
        report.data_row_count = max(0, last_row - _FIRST_DATA_ROW + 1)

        if last_row >= _FIRST_DATA_ROW:
            known = (
                {str(c).strip().lower() for c in known_control_ids}
                if known_control_ids is not None
                else None
            )
            _check_rows(sheet, last_row, report, known)

        return report
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _check_headers(sheet: Worksheet, report: ValidationReport) -> None:
    """Verify the row-6 header text matches the expected column layout.

    A shift here means every downstream read/write would land on the wrong
    column — a hard error, not a warning.
    """
    for col, expected in _EXPECTED_HEADERS.items():
        raw = sheet.cell(row=_HEADER_ROW, column=col).value
        text = _coerce_text(raw) or ""
        if expected not in text.lower():
            report.errors.append(
                ValidationIssue(
                    code="header_mismatch",
                    message=(
                        f"Column {_col_letter(col)} header should contain "
                        f'"{expected}" but got "{text}". eMASS template may '
                        "have been re-saved with shifted columns."
                    ),
                    cell=f"{_col_letter(col)}{_HEADER_ROW}",
                )
            )


def _last_data_row(sheet: Worksheet) -> int:
    """Bottom-up scan: return the row index of the last row with data.

    Returns ``_HEADER_ROW`` if no data row has any value in the scanned
    columns — caller treats that as "empty workbook, no rows to validate".

    Trailing fully-blank rows are silently tolerated by virtue of being
    invisible to this scan: openpyxl's ``ws.max_row`` overcounts because
    of eMASS template padding, so we step downward until we find a row
    that a human actually touched.
    """
    max_row = sheet.max_row or _HEADER_ROW
    if max_row <= _HEADER_ROW:
        return _HEADER_ROW
    for row in range(max_row, _HEADER_ROW, -1):
        for col in _DATA_COLUMNS:
            if sheet.cell(row=row, column=col).value not in (None, ""):
                return row
    return _HEADER_ROW


def _check_rows(
    sheet: Worksheet,
    last_row: int,
    report: ValidationReport,
    known_control_ids: set[str] | None,
) -> None:
    """Per-row checks: identity, status, dates, duplicates, orphans, formulas."""
    seen_identity: dict[tuple[str, str], int] = {}

    for row_idx in range(_FIRST_DATA_ROW, last_row + 1):
        control_raw = sheet.cell(row=row_idx, column=COL_CONTROL).value
        cci_raw = sheet.cell(row=row_idx, column=COL_CCI).value
        status_raw = sheet.cell(row=row_idx, column=COL_STATUS).value
        date_raw = sheet.cell(row=row_idx, column=COL_DATE_TESTED).value
        results_raw = sheet.cell(row=row_idx, column=COL_RESULTS).value

        control_id = _normalize_control(control_raw)
        cci_id = _normalize_cci_cell(cci_raw)
        status = _coerce_text(status_raw)

        has_assessor_data = any(
            v not in (None, "") for v in (status_raw, date_raw, results_raw)
        )

        # --- Orphan check ------------------------------------------------
        if not control_id and not cci_id:
            if has_assessor_data:
                report.errors.append(
                    ValidationIssue(
                        code="orphan_assessment",
                        message=(
                            "Row has assessment data (status/date/results) but no "
                            "control_id or CCI — cannot key it to a catalog entry. "
                            "Re-import would lose this row."
                        ),
                        excel_row=row_idx,
                    )
                )
            else:
                # Stray text below the data area (e.g. assessor's TODO notes).
                # Warn, don't error — the bottom-up scan already includes it
                # because *something* was non-blank, but it isn't a CCI row.
                report.warnings.append(
                    ValidationIssue(
                        code="non_data_row",
                        message=(
                            "Row falls inside the data area but has no control "
                            "identity. Likely an assessor note — move it out of "
                            "the data range to silence this warning."
                        ),
                        excel_row=row_idx,
                    )
                )
            continue

        # --- Identity present but partial -------------------------------
        if control_id and not cci_id:
            # Some 800-53 rows legitimately have no CCI (control-level
            # rollup), so this is only a warning. The reader filters on
            # cci_id when populating Objectives.
            report.warnings.append(
                ValidationIssue(
                    code="missing_cci",
                    message=(
                        f"Row for control {control_id} has no CCI in column H. "
                        "Re-import will skip this row when populating Objectives."
                    ),
                    excel_row=row_idx,
                    control_id=control_id,
                )
            )

        # --- Duplicate identity (the load-bearing merge check) ----------
        if control_id and cci_id:
            key = (control_id, cci_id)
            prior = seen_identity.get(key)
            if prior is not None:
                report.errors.append(
                    ValidationIssue(
                        code="duplicate_identity",
                        message=(
                            f"Duplicate ({control_id}, {cci_id}) — also at row "
                            f"{prior}. Family-partition merge likely produced "
                            "two rows for the same CCI; one would silently "
                            "overwrite the other on re-import."
                        ),
                        excel_row=row_idx,
                        control_id=control_id,
                        cci_id=cci_id,
                    )
                )
            else:
                seen_identity[key] = row_idx

        # --- Status value check -----------------------------------------
        if status and status not in _VALID_STATUSES:
            report.errors.append(
                ValidationIssue(
                    code="invalid_status",
                    message=(
                        f'Column N status "{status}" is not one of '
                        f"{sorted(_VALID_STATUSES)}. eMASS will reject the "
                        "workbook on upload."
                    ),
                    excel_row=row_idx,
                    cell=f"N{row_idx}",
                    control_id=control_id,
                    cci_id=cci_id,
                )
            )

        # --- Date check (only required when a status was set) -----------
        if status and date_raw not in (None, ""):
            parsed = _coerce_date(date_raw)
            if parsed is None:
                report.errors.append(
                    ValidationIssue(
                        code="unparseable_date",
                        message=(
                            f'Column O date "{date_raw!r}" could not be parsed. '
                            "Use ISO 8601 (YYYY-MM-DD) or a real Excel date cell."
                        ),
                        excel_row=row_idx,
                        cell=f"O{row_idx}",
                        control_id=control_id,
                        cci_id=cci_id,
                    )
                )

        # --- Unknown control (only when caller passed a catalog) --------
        if known_control_ids is not None and control_id:
            from .ccis_reader import _ccis_to_oscal_control_id

            oscal = _ccis_to_oscal_control_id(control_id)
            if oscal not in known_control_ids:
                report.warnings.append(
                    ValidationIssue(
                        code="unknown_control",
                        message=(
                            f"Control {control_id} (OSCAL {oscal}) not in catalog. "
                            "Likely a program overlay (SDA, FedRAMP plus, etc.) — "
                            "verify it shouldn't be a typo."
                        ),
                        excel_row=row_idx,
                        control_id=control_id,
                    )
                )

        # --- Formula in writable cell -----------------------------------
        for col, letter in ((COL_STATUS, "N"), (COL_RESULTS, "Q"), (COL_TESTER, "P")):
            cell = sheet.cell(row=row_idx, column=col)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                report.warnings.append(
                    ValidationIssue(
                        code="formula_in_writable_cell",
                        message=(
                            f"Cell {letter}{row_idx} contains a formula. Re-import "
                            "reads the formula string, not its evaluated result — "
                            "replace with the literal value."
                        ),
                        excel_row=row_idx,
                        cell=f"{letter}{row_idx}",
                    )
                )


def _col_letter(col_index: int) -> str:
    """1 → 'A', 14 → 'N'. Sufficient for the columns we report on (<= 26)."""
    return chr(ord("A") + col_index - 1)
