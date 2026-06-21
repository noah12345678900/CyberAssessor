"""Controls export — eMASS-strict and working-view xlsx writers.

Two public entry points:

- :func:`export_controls_to_emass` copies a user-supplied ``enterprise
  services controls.xlsx`` template, inserts a Program-Specific Controls
  column right after ``Control Acronym``, and writes one row per
  in-scope control with a status rollup that surfaces CRM-inherited /
  hybrid splits explicitly per the user's requested format::

      Compliant: CCI-000196, CCI-000197 (inherited from AWS GovCloud)
      Non-Compliant: CCI-000198 (no documented sanctions procedure)

  Uses xlwings (Excel COM) for the same reason ``poam/exporter.py`` does:
  the template carries data validation, conditional formatting, and 29
  sibling tabs that openpyxl's write path strips. xlwings goes through
  the live app and preserves all of it.

- :func:`export_controls_working_view` emits a fresh xlsx (openpyxl, no
  template) that mirrors the Controls UI: one row per **objective**, all
  needs_review and abstain rows included, plus the same PSC column.
  Owned by the assessor for working/review, never an eMASS deliverable.

The eMASS export is **idempotent**. Re-running onto the same output
file detects the existing "Program-Specific Controls" column header and
re-writes in place instead of stacking a second PSC column.

Precision-over-recall gate: rows with ``Assessment.needs_review=True``
are excluded from the eMASS export (per ``feedback_precision_over_recall``)
but included in the working view with a Needs Review column so the
assessor can triage them.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from sqlmodel import Session, select

from ..excel.ccis_reader import _ccis_to_oscal_control_id
from ..models import (
    Assessment,
    Baseline,
    BaselineControl,
    BaselineObjective,
    ComplianceStatus,
    Control,
    Objective,
    RequirementMap,
    RequirementSource,
    Workbook,
    _utcnow,
)

# ---------------------------------------------------------------------------
# Public types
# ---------------------------------------------------------------------------


# Excel cell hard-cap. Each PSC line is also capped (see _PSC_LINE_MAX) so a
# single huge requirement_text can't blow the whole cell budget on row 1.
_EXCEL_CELL_MAX = 32_767
_PSC_LINE_MAX = 500

# Header text we look for to locate the Control Acronym column. eMASS
# template variants spell this differently across program copies; the
# matcher is case-insensitive and matches any of the candidates.
_CONTROL_ACRONYM_HEADERS: tuple[str, ...] = (
    "control acronym",
    "control number",
    "control id",
    "controls / aps",
)

# eMASS Controls tab — also matches the working-view templates we've seen.
_DEFAULT_SHEET_NAME = "Controls"

# Bucket display order for _rollup_status. Compliant first so the eMASS
# reviewer sees the positive case at the top; Needs Review last because
# (a) those rows are excluded from the eMASS export anyway and (b) when
# a working-view export DOES surface them, they're the most actionable
# triage item and should sit closest to the row's other context.
_BUCKET_ORDER: tuple[str, ...] = (
    "Compliant",
    "Non-Compliant",
    "Not Applicable",
    "Needs Review",
)


@dataclass(frozen=True)
class ProgramControlRow:
    """One RequirementMap row, denormalized for the PSC column formatter.

    Mirrors the shape ``routes/controls.py::list_program_controls_for_control``
    returns under ``rows[]``, but we keep it as a dataclass so the bulk
    fetch helper can hand it to the formatter without dict-juggling.
    """

    source_name: str
    requirement_number: str
    requirement_text: str
    objective_id: int


@dataclass(frozen=True)
class ObjectiveAssessment:
    """The (objective, assessment) pair the rollup helper bucket-sorts on."""

    objective_id: int
    objective_code: str  # e.g. "CCI-000196"
    status: ComplianceStatus
    narrative_q: str | None
    needs_review: bool
    inheritance_rule: str | None
    # CRM cloud-scope verdict + narrative (matches CSP-issued CRM templates
    # like AWS GovCloud / Azure / GCP). ``crm_responsibility_onprem`` carries
    # the separately-tracked on-prem verdict for mixed cloud + on-prem
    # systems; both may be set independently per the dual-scope CRM schema.
    crm_responsibility: str | None  # "customer" / "customer_configured" / "provider" / "hybrid" / "inherited" / "not_applicable" / None
    crm_narrative: str | None
    crm_responsibility_onprem: str | None = None
    crm_narrative_onprem: str | None = None
    # Dual-narrative fields written into the eMASS export's per-scope
    # columns alongside ``narrative_q`` (the canonical CCIS col-Q text).
    # Default None so existing test fixtures and callers that don't
    # populate them keep working — the working-copy exporter passes
    # explicit values when an Assessment row carries them.
    narrative_on_prem: str | None = None
    narrative_cloud: str | None = None


@dataclass(frozen=True)
class ControlExportResult:
    """Summary of one export run. Returned to the UI so the modal can show
    "42 rows written, 6 with PSC mappings, 0 skipped" and surface any
    template-shape warnings inline.
    """

    output_path: str
    rows_written: int
    controls_with_psc: int
    skipped: list[tuple[str, str]] = field(default_factory=list)
    template_warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Status rollup
# ---------------------------------------------------------------------------


def _rollup_status(objectives: list[ObjectiveAssessment]) -> str:
    """Return the multi-line Status cell text per user-specified format.

    Bucketing rules:
      - Objectives where ``crm_responsibility`` is ``inherited`` or
        ``provider`` are short-circuited into the ``Compliant`` bucket
        regardless of any stored Assessment status. Their reason carries
        "inherited from <source>" so the eMASS reviewer knows the verdict
        is overlay-driven, not investigation-driven.
      - ``crm_responsibility == "not_applicable"`` → Not Applicable
        bucket with "marked NA by CRM overlay".
      - ``needs_review=True`` rows go to the Needs Review bucket no
        matter what status the LLM proposed (precision-over-recall).
      - Everything else uses ``Assessment.status``.

    Output shape:
      - Empty input → "" (caller decides what to write into the cell).
      - Single bucket → just the status token (e.g. "Compliant") so the
        cell stays compatible with eMASS's single-value expectations
        when there's no ambiguity to surface.
      - Multiple buckets → one line per bucket in ``_BUCKET_ORDER``::

            Compliant: CCI-000196, CCI-000197 (inherited from AWS GovCloud)
            Non-Compliant: CCI-000198 (no documented sanctions procedure)
    """
    if not objectives:
        return ""

    # bucket_name -> list[(objective_code, reason)]
    buckets: dict[str, list[tuple[str, str]]] = {b: [] for b in _BUCKET_ORDER}

    for oa in objectives:
        bucket, reason = _classify(oa)
        buckets[bucket].append((oa.objective_code, reason))

    nonempty = [(b, items) for b, items in buckets.items() if items]
    if not nonempty:
        return ""

    if len(nonempty) == 1:
        # Single-bucket control → emit just the status token. eMASS's
        # status field is single-valued by design; no ambiguity to surface.
        return nonempty[0][0]

    lines: list[str] = []
    for bucket, items in nonempty:
        # Group by reason within a bucket so we collapse
        # "CCI-1, CCI-2, CCI-3 (inherited from AWS)" into one line per
        # distinct reason rather than three identical trailing parens.
        by_reason: dict[str, list[str]] = {}
        for code, reason in items:
            by_reason.setdefault(reason, []).append(code)
        for reason, codes in by_reason.items():
            cci_list = ", ".join(codes)
            if reason:
                lines.append(f"{bucket}: {cci_list} ({reason})")
            else:
                lines.append(f"{bucket}: {cci_list}")
    return "\n".join(lines)


def _classify(oa: ObjectiveAssessment) -> tuple[str, str]:
    """Return (bucket_name, reason) for one objective.

    CRM short-circuit takes precedence over Assessment status — the
    overlay decision is the ground truth for inherited / provider /
    not_applicable, per ``feedback_overlay_default_local``.

    Dual-scope semantics (cloud + on-prem): short-circuit only when EVERY
    specified scope is inheritable. A control that's ``inherited`` in the
    cloud but ``customer`` on-prem still requires real assessment —
    falling through to the Assessment status path keeps the on-prem
    verdict from being hidden behind the cloud short-circuit. Mirrors
    the gate in ``engine/assessor.py``.
    """
    crm = (oa.crm_responsibility or "").lower() or None
    crm_op = (oa.crm_responsibility_onprem or "").lower() or None

    short_circuit_set = {"inherited", "provider", "not_applicable"}
    specified = [r for r in (crm, crm_op) if r]
    all_short_circuit = bool(specified) and all(
        r in short_circuit_set for r in specified
    )

    if all_short_circuit:
        # When both scopes agree, use the single combined reason. When
        # scopes disagree (e.g. cloud=inherited, on-prem=provider) we
        # join the per-scope reasons so the eMASS reviewer sees both.
        cloud_reason = _scope_short_circuit_reason(crm, oa.crm_narrative, "cloud")
        onprem_reason = _scope_short_circuit_reason(
            crm_op, oa.crm_narrative_onprem, "on-prem"
        )
        scope_reasons = [r for r in (cloud_reason, onprem_reason) if r]
        # Bucket priority: any "inherited" → Compliant; else any
        # "provider" → Compliant; else all "not_applicable" → NA.
        if crm == "inherited" or crm_op == "inherited":
            return "Compliant", " | ".join(scope_reasons)
        if crm == "provider" or crm_op == "provider":
            return "Compliant", " | ".join(scope_reasons)
        return "Not Applicable", " | ".join(scope_reasons) or "marked NA by CRM overlay"

    if oa.needs_review:
        return "Needs Review", _short_reason(oa.narrative_q) or "needs reviewer check"

    if oa.inheritance_rule == "8a":
        return "Compliant", "Rule 8a auto-compliant"

    if oa.status == ComplianceStatus.COMPLIANT:
        return "Compliant", _short_reason(oa.narrative_q)
    if oa.status == ComplianceStatus.NON_COMPLIANT:
        return "Non-Compliant", _short_reason(oa.narrative_q) or "gap identified"
    if oa.status == ComplianceStatus.NOT_APPLICABLE:
        return "Not Applicable", _short_reason(oa.narrative_q)

    return "Needs Review", "unknown status"


def _scope_short_circuit_reason(
    resp: str | None,
    narrative: str | None,
    scope_label: str,
) -> str:
    """Format a per-scope short-circuit reason for the rollup cell.

    Examples (with scope_label="cloud"):
      - inherited + "AWS GovCloud — ..." → "cloud: inherited from AWS GovCloud"
      - provider  + None                 → "cloud: implemented by provider"
      - not_applicable                   → "cloud: marked NA by CRM overlay"

    Empty string when resp is None — caller filters those out so a
    single-scope CRM doesn't produce orphan " | " separators.
    """
    if not resp:
        return ""
    src = _crm_source_phrase(narrative)
    if resp == "inherited":
        return f"{scope_label}: inherited from {src}" if src else f"{scope_label}: inherited"
    if resp == "provider":
        return (
            f"{scope_label}: implemented by {src}" if src
            else f"{scope_label}: implemented by provider"
        )
    if resp == "not_applicable":
        return f"{scope_label}: marked NA by CRM overlay"
    return ""


def _crm_source_phrase(crm_narrative: str | None) -> str:
    """Extract a short source phrase from the CRM narrative (e.g. "AWS
    GovCloud") for the Status cell. We don't try to NLP-parse — just take
    the first ~50 chars of the first sentence, which is overwhelmingly how
    CRM narratives are written ("AWS GovCloud — inherited control...").
    """
    if not crm_narrative:
        return ""
    first = crm_narrative.split(".")[0].strip()
    if len(first) <= 50:
        return first
    return first[:47].rstrip() + "..."


def _short_reason(narrative_q: str | None) -> str:
    """First sentence of the narrative, truncated to ~80 chars."""
    if not narrative_q:
        return ""
    first = narrative_q.strip().split(".")[0].strip()
    if len(first) <= 80:
        return first
    return first[:77].rstrip() + "..."


# ---------------------------------------------------------------------------
# PSC column formatter
# ---------------------------------------------------------------------------


def _format_psc_column(rows: list[ProgramControlRow]) -> str:
    """Render the PSC cell text for one control.

    Already-grouped input (by source.name + requirement_number) — the
    bulk fetcher sorts. Output is one line per PSC row::

        SDA-127: <text>
        SDA-128: <text>
        T1TL-031: <text>

    Per-line cap of 500 chars; whole-cell cap of Excel's 32,767-char
    hard limit with a trailing ``...[N more truncated]`` marker so the
    operator knows something was dropped instead of silently losing data.
    """
    if not rows:
        return ""

    lines: list[str] = []
    used = 0
    truncated_count = 0
    for r in rows:
        text = r.requirement_text or ""
        if len(text) > _PSC_LINE_MAX:
            text = text[: _PSC_LINE_MAX - 3].rstrip() + "..."
        line = f"{r.requirement_number}: {text}"
        # +1 for the newline separator we'll add between lines.
        projected = used + len(line) + (1 if lines else 0)
        # Reserve ~32 chars for the trailing truncation marker.
        if projected > _EXCEL_CELL_MAX - 32:
            truncated_count = len(rows) - len(lines)
            break
        lines.append(line)
        used = projected

    if truncated_count > 0:
        lines.append(f"...[{truncated_count} more truncated]")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Bulk PSC fetch — avoids N+1 during export
# ---------------------------------------------------------------------------


def fetch_psc_rows_bulk(
    session: Session,
    framework_id: int | None,
    control_ids: list[int],
) -> dict[int, list[ProgramControlRow]]:
    """One query → ``{control_id: [ProgramControlRow, ...]}`` map.

    Single join across RequirementMap → RequirementSource → Objective,
    then bucketed in Python by ``Objective.control_id_fk``. Empty list
    for any control with no overlay rows so the caller can index without
    a ``get(..., [])`` defensive read.

    Sort order matches ``list_program_controls_for_control``: source name
    ascending, then requirement_number lexicographic. Stable across
    re-exports.
    """
    if not control_ids:
        return {}

    q = (
        select(RequirementMap, RequirementSource, Objective)
        .join(
            RequirementSource,
            RequirementSource.id == RequirementMap.requirement_source_id,
        )
        .join(Objective, Objective.id == RequirementMap.objective_id)
        .where(Objective.control_id_fk.in_(control_ids))  # type: ignore[attr-defined]
    )
    if framework_id is not None:
        q = q.where(RequirementSource.framework_id == framework_id)

    out: dict[int, list[ProgramControlRow]] = {cid: [] for cid in control_ids}
    for rm, src, obj in session.exec(q).all():
        out[obj.control_id_fk].append(
            ProgramControlRow(
                source_name=src.name,
                requirement_number=rm.requirement_number,
                requirement_text=rm.requirement_text,
                objective_id=obj.id,
            )
        )
    for cid in out:
        out[cid].sort(key=lambda r: (r.source_name, r.requirement_number))
    return out


# ---------------------------------------------------------------------------
# Internal data loaders
# ---------------------------------------------------------------------------


def _load_in_scope_controls(
    session: Session,
    baseline: Baseline,
    family_filter: str | None = None,
) -> list[tuple[Control, BaselineControl]]:
    """Return every in-scope ``(Control, BaselineControl)`` pair for the
    baseline. Mirrors the join shape in ``routes/controls.py`` so the
    in_scope semantics stay authoritative.
    """
    stmt = (
        select(Control, BaselineControl)
        .join(BaselineControl, BaselineControl.control_id == Control.id)
        .where(
            BaselineControl.baseline_id == baseline.id,
            BaselineControl.in_scope.is_(True),  # type: ignore[union-attr]
        )
    )
    if family_filter:
        stmt = stmt.where(Control.family == family_filter.upper())
    rows = list(session.exec(stmt).all())
    rows.sort(key=lambda pc: pc[0].control_id)
    return rows


def _load_objectives_for_control(
    session: Session,
    workbook_id: int,
    baseline: Baseline,
    control: Control,
    bc: BaselineControl,
) -> list[ObjectiveAssessment]:
    """Pull every objective for the control along with its latest
    Assessment for this workbook. Objectives without an assessment are
    still returned (status defaults to the CRM short-circuit or None)
    so the rollup helper doesn't silently drop them.
    """
    # Exclude soft-deleted CCIs from the per-control rollup so the
    # exporter reports the current workbook roster — not the historical
    # superset preserved for save-path source_row lookups.
    obj_rows = list(
        session.exec(
            select(Objective, BaselineObjective)
            .join(BaselineObjective, BaselineObjective.objective_id == Objective.id)
            .where(
                Objective.control_id_fk == control.id,
                BaselineObjective.baseline_id == baseline.id,
                BaselineObjective.is_deprecated.is_(False),  # type: ignore[union-attr]
            )
        ).all()
    )

    if not obj_rows:
        return []

    obj_ids = [o.id for o, _ in obj_rows]
    assessments_by_obj: dict[int, Assessment] = {}
    for a in session.exec(
        select(Assessment).where(
            Assessment.workbook_id == workbook_id,
            Assessment.objective_id.in_(obj_ids),  # type: ignore[attr-defined]
        )
    ).all():
        # Keep the most recent assessment per objective. Created_at sort
        # in Python — N is small (controls have ~5-20 CCIs).
        prior = assessments_by_obj.get(a.objective_id)
        if prior is None or (a.created_at or datetime.min) > (
            prior.created_at or datetime.min
        ):
            assessments_by_obj[a.objective_id] = a

    out: list[ObjectiveAssessment] = []
    for obj, _bo in obj_rows:
        a = assessments_by_obj.get(obj.id)
        out.append(
            ObjectiveAssessment(
                objective_id=obj.id,
                objective_code=obj.objective_id,
                status=a.status if a else ComplianceStatus.NOT_APPLICABLE,
                narrative_q=a.narrative_q if a else None,
                narrative_on_prem=a.narrative_on_prem if a else None,
                narrative_cloud=a.narrative_cloud if a else None,
                needs_review=bool(a.needs_review) if a else False,
                inheritance_rule=a.inheritance_rule if a else None,
                crm_responsibility=bc.responsibility,
                crm_narrative=bc.responsibility_narrative,
                crm_responsibility_onprem=bc.responsibility_onprem,
                crm_narrative_onprem=bc.responsibility_onprem_narrative,
            )
        )
    # Stable order — eMASS reviewers expect CCI ids ascending.
    out.sort(key=lambda oa: oa.objective_code)
    return out


# ---------------------------------------------------------------------------
# eMASS export (xlwings, template-preserving)
# ---------------------------------------------------------------------------


def export_controls_to_emass(
    *,
    session: Session,
    workbook_id: int,
    template_path: str | Path,
    output_path: str | Path,
    sheet_name: str = _DEFAULT_SHEET_NAME,
) -> ControlExportResult:
    """Copy ``template_path`` → ``output_path``, fill the Controls tab.

    HEADLESS (openpyxl, no Excel/COM). The previous implementation drove
    Excel via xlwings, which (a) FROZE on the second export — a sync route
    runs on FastAPI's threadpool with no COM init, so the 2nd ``xw.App()``
    in the long-lived sidecar hung — and (b) wrote controls top-down from
    row 2, OVERWRITING the template's pre-populated, per-acronym formula
    rows (``=N<row>`` narrative pulls, array-formula CCI/References/STIG
    columns), which corrupted the file on re-export. This rewrite matches
    the app's headless convention (POAM export, CCIS writer, working view).

    Behavior:
      1. Loads in-scope controls + objectives + CRM responsibility from
         the workbook's primary baseline.
      2. Matches each control to its EXISTING row in the template by
         Control Acronym (the template ships pre-populated with one row per
         control, each carrying its own formulas). We write ONLY the Status
         / Program-Specific-Controls cells into that row and never reposition
         rows — so the template's formula scaffolding stays intact and
         re-export is stable (same control → same row every time).
      3. NO SILENT SKIP: every in-scope control that has a template row is
         written. Controls whose CCIs are still ``needs_review`` get the
         "Needs Review" status bucket (via ``_rollup_status``) rather than
         being dropped — an in-scope control silently missing from a
         compliance deliverable is the failure mode RMF artifacts exist to
         prevent. Controls with no matching template row are reported in
         ``skipped`` (a template-roster mismatch the operator should know
         about), not dropped silently.
      4. Writes ONLY the Status column. The eMASS deliverable does not carry a
         Program-Specific Controls column (that lives in the working view),
         and the Implementation Narrative is pulled by the template's own
         ``=N<row>`` formula from the CCIS workbook — so we never overwrite it.
         If the template has no Status column we append one (reused, never
         duplicated, on re-export).
      5. Stamps ``Workbook.exported_at``.

    Raises plain ``ValueError`` / ``FileNotFoundError``; ``routes/controls.py``
    maps them to HTTP status codes.
    """
    import shutil

    import openpyxl

    wb = session.get(Workbook, workbook_id)
    if wb is None:
        raise ValueError(f"Workbook {workbook_id} not found")
    if wb.baseline_id is None:
        raise ValueError(
            "Workbook has no Baseline. Reopen with a Framework selected so the "
            "app can materialize the in-scope control set."
        )
    baseline = session.get(Baseline, wb.baseline_id)
    if baseline is None:
        raise ValueError(f"Workbook references missing Baseline {wb.baseline_id}")

    src = Path(template_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"Controls template not found: {src}")
    dst.parent.mkdir(parents=True, exist_ok=True)
    # Idempotent re-export: operator may point template_path at the previous
    # output to refresh in place. Skip the copy when src == dst — the
    # open-modify-save pass below handles the rewrite.
    if not (dst.exists() and src.resolve() == dst.resolve()):
        shutil.copyfile(src, dst)

    pairs = _load_in_scope_controls(session, baseline)

    rows_written = 0
    skipped: list[tuple[str, str]] = []
    warnings: list[str] = []

    # data_only=False (default): preserve the template's formulas (the
    # Implementation Narrative =N<row> pulls and the array-formula
    # CCI/References/STIG columns) so re-saving doesn't strip them. Excel
    # recalculates them when the operator opens the file.
    book = openpyxl.load_workbook(str(dst))
    if sheet_name not in book.sheetnames:
        raise ValueError(f"Template missing required sheet '{sheet_name}'")
    sh = book[sheet_name]

    header_row, acronym_col = _find_control_acronym_column_opx(sh, warnings)

    # Build acronym -> row index from the template's pre-populated rows.
    # Key on the OSCAL-normalized form (ac-2.1) so 'AC-2(1)' in the template
    # matches Control.control_id regardless of the template's delimiter style.
    row_by_control: dict[str, int] = {}
    for r in range(header_row + 1, sh.max_row + 1):
        raw = sh.cell(r, acronym_col).value
        if raw is None:
            continue
        key = _ccis_to_oscal_control_id(str(raw))
        # First occurrence wins (templates list each control once); ignore
        # any accidental duplicates rather than letting a later blank win.
        row_by_control.setdefault(key, r)

    status_col = _find_header_col_opx(
        sh, header_row, ("status", "compliance status", "implementation status")
    )
    if status_col is None:
        # This template variant has no Status column (it's a formula-driven
        # view). Append one so the rollup has a home — without it the export
        # would silently write nothing meaningful, the original complaint.
        # _find_header_col_opx scans past max_column, so a re-export finds and
        # reuses this appended column instead of appending a duplicate.
        status_col = sh.max_column + 1
        sh.cell(header_row, status_col).value = "Implementation Status"
        warnings.append(
            "Template had no Status column; appended 'Implementation Status' "
            "as the last column for the rollup."
        )

    for control, bc in pairs:
        objectives = _load_objectives_for_control(
            session, workbook_id, baseline, control, bc
        )
        key = _ccis_to_oscal_control_id(control.control_id)
        target_row = row_by_control.get(key)
        if target_row is None:
            # The control is in scope but the template has no row for it —
            # a roster mismatch (e.g. template predates a control's addition).
            # Report it so the operator can reconcile; do NOT append a bare
            # row, which would have no formulas and look broken in eMASS.
            skipped.append(
                (control.control_id, "no matching row in template roster")
            )
            continue

        # NO SILENT SKIP: needs_review controls are written with the
        # "Needs Review" status bucket (via _rollup_status) so every in-scope
        # control appears in the deliverable and the gap is visible.
        status_text = _rollup_status(objectives)
        sh.cell(target_row, status_col).value = status_text
        rows_written += 1

    book.save(str(dst))

    wb.exported_at = _utcnow()
    session.add(wb)
    session.commit()

    return ControlExportResult(
        output_path=str(dst),
        rows_written=rows_written,
        # PSC column intentionally not written to the eMASS deliverable.
        controls_with_psc=0,
        skipped=skipped,
        template_warnings=warnings,
    )


# ---------------------------------------------------------------------------
# eMASS template helpers — openpyxl (headless) column discovery
# ---------------------------------------------------------------------------


def _find_control_acronym_column_opx(sh, warnings: list[str]) -> tuple[int, int]:
    """openpyxl twin of the acronym-column finder.

    Scans rows 1-5 for the header row containing a Control Acronym column.
    Returns ``(header_row, col_1based)``. Raises ``ValueError`` if no
    recognizable header is found.
    """
    for header_row in range(1, 6):
        col = _find_header_col_opx(sh, header_row, _CONTROL_ACRONYM_HEADERS)
        if col:
            if header_row != 1:
                warnings.append(
                    f"Header row detected at row {header_row} (not row 1)."
                )
            return header_row, col
    raise ValueError(
        "Template missing a Control Acronym / Control Number / Controls / APs "
        "column in the first 5 rows. Cannot determine where to write controls."
    )


def _find_header_col_opx(
    sh, header_row: int, candidates: tuple[str, ...]
) -> int | None:
    """openpyxl twin of ``_find_header_col``. Case/whitespace-insensitive
    header lookup across candidate phrases.

    Scans the full used width (``max_column``, +8 slack for columns we may
    have appended on a prior export — e.g. a previously-appended status or
    PSC column. A fixed cap would miss those on re-export and append a
    DUPLICATE column each run.)
    """
    candidates_norm = {_norm_header(c) for c in candidates}
    for col in range(1, sh.max_column + 9):
        v = sh.cell(header_row, col).value
        if not v:
            continue
        if _norm_header(str(v)) in candidates_norm:
            return col
    return None


def _norm_header(s: str) -> str:
    """Collapse whitespace + lowercase for tolerant header matching.

    ``'Control \\nAcronym'`` and ``'COMMON  CONTROL'`` (double space) both
    appear in real templates; ``str.split() + ' '.join()`` collapses any
    run of whitespace — including embedded newlines — to a single space.
    """
    return " ".join(str(s).split()).lower()


# ---------------------------------------------------------------------------
# Working-view export (openpyxl, no template)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ControlsFilterState:
    """Mirrors the Controls list page filter state. None = no filter."""

    family: str | None = None
    status: str | None = None  # "Compliant" / "Non-Compliant" / ...
    search: str | None = None  # case-insensitive substring on control_id/title


def export_controls_working_view(
    *,
    session: Session,
    workbook_id: int,
    output_path: str | Path,
    filter_state: ControlsFilterState | None = None,
) -> ControlExportResult:
    """Emit a fresh xlsx mirroring the Controls UI: one row per OBJECTIVE.

    Includes needs_review rows (with the flag surfaced as a column) so the
    assessor can review everything they see on the page in one workbook.
    """
    from openpyxl import Workbook as PyXlWorkbook

    wb = session.get(Workbook, workbook_id)
    if wb is None:
        raise ValueError(f"Workbook {workbook_id} not found")
    if wb.baseline_id is None:
        raise ValueError("Workbook has no Baseline.")
    baseline = session.get(Baseline, wb.baseline_id)
    if baseline is None:
        raise ValueError(f"Missing Baseline {wb.baseline_id}")

    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)

    fs = filter_state or ControlsFilterState()
    pairs = _load_in_scope_controls(session, baseline, family_filter=fs.family)

    if fs.search:
        needle = fs.search.lower()
        pairs = [
            (c, bc) for c, bc in pairs
            if needle in (c.control_id or "").lower()
            or needle in (c.title or "").lower()
        ]

    control_ids = [c.id for c, _ in pairs]
    psc_map = fetch_psc_rows_bulk(session, baseline.framework_id, control_ids)

    py_wb = PyXlWorkbook()
    ws = py_wb.active
    ws.title = "Controls (Working View)"
    headers = [
        "Control",
        "Title",
        "Family",
        "Program-Specific Controls",
        "CCI",
        "Status",
        "Needs Review",
        "Narrative",
        "Narrative (On-Prem)",
        "Narrative (Cloud)",
        "Inheritance Rule",
        "Confidence",
        "CRM Responsibility (Cloud)",
        "CRM Responsibility (On-Prem)",
    ]
    ws.append(headers)

    rows_written = 0
    controls_with_psc = 0
    skipped: list[tuple[str, str]] = []

    for control, bc in pairs:
        objectives = _load_objectives_for_control(
            session, workbook_id, baseline, control, bc
        )
        psc_rows = psc_map.get(control.id, [])
        psc_text = _format_psc_column(psc_rows)
        if psc_text:
            controls_with_psc += 1

        # Per-objective row (working view shows the full breakdown). If
        # the optional status filter is set, drop rows that don't match.
        for oa in objectives:
            row_status = oa.status.value if oa.status else ""
            if fs.status and row_status != fs.status:
                continue
            ws.append([
                control.control_id,
                control.title,
                control.family,
                psc_text,
                oa.objective_code,
                row_status,
                "Yes" if oa.needs_review else "",
                oa.narrative_q or "",
                oa.narrative_on_prem or "",
                oa.narrative_cloud or "",
                oa.inheritance_rule or "",
                "",  # confidence — not loaded in this slice
                bc.responsibility or "",
                bc.responsibility_onprem or "",
            ])
            rows_written += 1

    py_wb.save(str(dst))

    return ControlExportResult(
        output_path=str(dst),
        rows_written=rows_written,
        controls_with_psc=controls_with_psc,
        skipped=skipped,
        template_warnings=[],
    )
