"""XLSX extractor (openpyxl — already a project dependency).

Concatenates every non-empty cell, sheet by sheet, comma-separated by
row. We deliberately do not preserve formulas — they don't help the
LLM cite the document and only inflate the cache key. Skips the CCIS
workbook itself by sniffing for the "WORKING SHEET" tab and refusing
to ingest it (it's the assessment target, not evidence).
"""

from __future__ import annotations

import re
from pathlib import PurePosixPath
from typing import BinaryIO

from ...models import EvidenceKind
from .base import ExtractedDoc, ExtractorError, ExtractorSkip, register, resolve_doc_number

_CCIS_MARKER_SHEETS = {"WORKING SHEET", "working sheet"}

# Canonical CCI token (kept identical to evidence.tagger._CCI_RE and
# extractors._stig_common._CCI_RE). Used to VALUE-validate cells in a
# detected CCI column so a mistyped header or a stray column never injects
# non-CCI strings into the ungated 0.95 Tier-2 path.
_CCI_RE = re.compile(r"CCI-\d{6}", re.IGNORECASE)

# Header strings (case + whitespace insensitive) that mark a dedicated CCI
# column in a generic evidence workbook (a compliance matrix, SCAP results
# export, RTM, etc.). Deliberately tight: a non-STIG xlsx that carries an
# explicit CCI column is a high-precision structured signal, so its values
# route to the SAME ungated 0.95 Tier-2 branch the STIG/Nessus extractors
# feed (evidence.tagger step 2). We do NOT include "ap acronym" (that names
# the AP, not the CCI) — and every collected cell is still CCI-####-validated,
# so a wrong-column match yields zero refs rather than garbage tags.
_CCI_HEADER_HEADERS = frozenset(
    {
        "cci",
        "ccis",
        "cci id",
        "cci-id",
        "cci_id",
        "cci ref",
        "cci-ref",
        "cci_ref",
        "cci refs",
        "cci reference",
        "cci references",
        "cci number",
        "cci no",
        "cci #",
    }
)

# How many leading rows of EACH sheet to inspect when hunting for the header
# row that classifies a workbook (HW / SW / account / asset). The header is NOT
# always sheet 0 / row 0: real exports lead with a Cover or Table-of-Contents
# sheet, and even a data sheet often has a title banner + blank rows before the
# column headers. The O&I "Roles & User Accounts" workbook zero-tagged for
# exactly this reason — its first sheet is a cover page, so a sheet0/row0-only
# probe saw nothing and Tier 4 never fired (BUG B re-fix, 2026-06-10). 25 rows
# comfortably clears typical banner/legend preamble without scanning whole
# sheets.
_HEADER_SCAN_ROWS = 25

# Header strings (case + whitespace insensitive) that mark a hostname column
# in an HW/SW xlsx. Kept in sync with asset_crosscheck.HOSTNAME_HEADERS so a
# user-flagged asset list resolves the same way whether the cache hit or the
# re-parse path is taken. First matching column on the first sheet wins.
_HOSTNAME_HEADERS = frozenset(
    {
        "hostname",
        "host",
        "host name",
        "fqdn",
        "computer name",
        "computer",
        "node name",
        "system name",
        "asset name",
        "device name",
    }
)

# Signal columns for content-based classification of asset-list workbooks.
# The goal: tag a HW/SW inventory xlsx to CM-8 (and family) automatically
# even when nothing in the filename hints at it. The user shouldn't have to
# rename "Server List.xlsx" to "HW_Inventory.xlsx" to get auto-tagging — the
# columns themselves are the strongest signal. Detection runs on the first
# sheet's header row only (sheet 0, row 0); whatever evidence_type the
# columns produce there wins.
_HW_SIGNAL_HEADERS = frozenset(
    {
        "serial number",
        "serial",
        "serial #",
        "serial no",
        "asset tag",
        "asset id",
        "mac address",
        "mac",
        "ip address",
        "ip",
        "manufacturer",
        "make",
        "model",
        "model number",
        "os",
        "operating system",
        "os version",
        "cpu",
        "processor",
        "memory",
        "ram",
        "form factor",
        "device type",
        "hardware",
        "chassis",
        "firmware version",
        "bios version",
    }
)

_SW_SIGNAL_HEADERS = frozenset(
    {
        "version",
        "sw version",
        "software version",
        "publisher",
        "vendor",
        "license",
        "license key",
        "install date",
        "installed on",
        "installed date",
        "software",
        "software name",
        "application",
        "application name",
        "product",
        "product name",
        "app name",
        "package",
        "package name",
    }
)

# Signal columns for an account / roles / user-access matrix. These workbooks
# (e.g. "O&I Roles & User Accounts.xlsx" — an AD/O365 user export or a
# program-authored access matrix) carry no doc number, no CCI token, and no
# control ID in their cells, so Tiers 1-3 produce zero tags and the artifact
# is invisible to AC-2 / AC-6 / IA-2 (BUG B, 2026-06-10). The *columns* are the
# signal: User/Username, Role, Account Type, Privilege, Group, Permission,
# Access Level, etc. Matched the same way as HW/SW — 2+ distinct account
# signals on the first-sheet header row → "account_matrix". Terms are kept
# identity-specific (no bare "title"/"name") so an HW or SW list never
# mis-classifies as an account matrix.
_ACCOUNT_SIGNAL_HEADERS = frozenset(
    {
        "user",
        "username",
        "user name",
        "user id",
        "userid",
        "account",
        "account name",
        "account type",
        "account status",
        "login",
        "logon",
        "last login",
        "last logon",
        "role",
        "roles",
        "role name",
        "privilege",
        "privileges",
        "privilege level",
        "permission",
        "permissions",
        "access level",
        "group",
        "groups",
        "group membership",
        "security group",
        "manager",
        "supervisor",
        "mfa",
        "mfa enabled",
        "two-factor",
        "2fa",
        "sid",
        "sam account name",
        "samaccountname",
        "display name",
        "department",
        "job title",
    }
)


# Lever C (2026-06-11): a Plan of Action & Milestones workbook. These are the
# program's own remediation tracker — the literal artifact CA-5 requires. The
# column shape is highly distinctive: weakness rows, scheduled-completion dates,
# residual-risk levels, milestones. Detection is CORE-anchored: a generic
# project plan can carry "milestone" + "point of contact" but never "weakness" /
# "residual risk" / "poa&m id", so we require >=1 POA&M-defining core header AND
# >=2 total POA&M signals. That bijects to real POA&Ms and routes them to CA-5
# (evidence.tagger EVIDENCE_TYPE_TO_CONTROLS["poam"]).
_POAM_CORE_HEADERS = frozenset(
    {
        "poa&m id",
        "poam id",
        "poa&m",
        "poam",
        "weakness",
        "weaknesses",
        "weakness name",
        "weakness description",
        "source identifying weakness",
        "source of weakness",
        "raw severity",
        "raw severity value",
        "residual risk",
        "residual risk level",
    }
)
_POAM_SUPPORT_HEADERS = frozenset(
    {
        "scheduled completion date",
        "scheduled completion",
        "estimated completion date",
        "milestone",
        "milestones",
        "milestone with completion dates",
        "milestones with completion dates",
        "milestone changes",
        "resources required",
        "resource required",
        "point of contact",
        "poc",
        "remediation plan",
        "planned mitigations",
        "mitigations",
        "status",
    }
)

# Lever C (2026-06-11): a training / security-awareness completion roster. The
# literal evidence for AT-4 (training records); corroborates AT-2 (awareness)
# and AT-3 (role-based). CORE-anchored so a generic personnel roster ("Name,
# Date, Status" — no training/course token) is NOT swept in: require >=1
# training-defining core header AND >=2 total training signals. Identity columns
# are deliberately NOT training signals, so an access matrix never mis-routes
# here (and account_matrix is classified first anyway).
_TRAINING_CORE_HEADERS = frozenset(
    {
        "training",
        "training name",
        "training title",
        "training type",
        "course",
        "course name",
        "course title",
        "awareness training",
        "security awareness training",
        "role-based training",
        "role based training",
        "cbt",
    }
)
_TRAINING_SUPPORT_HEADERS = frozenset(
    {
        "completion date",
        "training completion date",
        "date completed",
        "completed on",
        "training status",
        "completion status",
        "due date",
        "assigned date",
        "score",
        "pass/fail",
    }
)


def _classify_asset_workbook(
    header_row: tuple,
) -> tuple[str | None, list[str]]:
    """Inspect a first-sheet header row and classify it by content shape.

    Returns ``(evidence_type, matched_signals)`` where ``evidence_type`` is
    one of ``"hw_inventory"``, ``"sw_inventory"``, ``"account_matrix"``,
    ``"poam"``, ``"training_record"``, ``"asset_inventory"``, or ``None`` if
    nothing matched. ``matched_signals`` is the list of header strings that
    fired — surfaced into the tagger rationale so the assessor can see *why* a
    workbook was auto-mapped.

    Rules (intentionally simple, easy to tune):
      * 2+ HW signal columns → ``hw_inventory`` (→ CM-8)
      * 2+ SW signal columns → ``sw_inventory`` (→ CM-8 family; HW wins ties —
        most asset lists in this program lead with HW columns and add an "OS
        Version" SW column that we don't want to mis-classify)
      * 2+ account signal columns → ``account_matrix`` (→ AC-2 / AC-6 / IA-2).
        Checked after HW/SW because an inventory never carries 2+ account
        columns, so HW/SW precedence is safe.
      * (Lever C) >=1 POA&M-core + >=2 total POA&M signals → ``poam`` (→ CA-5).
      * (Lever C) >=1 training-core + >=2 total training signals →
        ``training_record`` (→ AT-2 / AT-3 / AT-4).
      * Hostname column + zero HW/SW/account signals → ``asset_inventory``
        (generic — could be HW or SW; CM-8 still applies)
      * Otherwise → ``None`` (don't auto-tag; not enough signal)

    The two Lever-C shapes are evaluated AFTER every pre-existing branch so they
    can only ever catch a workbook that currently classifies as ``None`` —
    pure recall addition, never a re-classification of an artifact the
    inventory/account/asset branches already place (precision-over-recall: no
    existing verdict moves). Both are CORE-anchored: a generic project plan or
    personnel roster lacks the defining "weakness"/"residual risk" or
    "training"/"course" header, so it stays ``None``.
    """
    hw_hits: list[str] = []
    sw_hits: list[str] = []
    account_hits: list[str] = []
    poam_core: list[str] = []
    poam_support: list[str] = []
    training_core: list[str] = []
    training_support: list[str] = []
    has_hostname = False
    for cell in header_row:
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if not key:
            continue
        if key in _HOSTNAME_HEADERS:
            has_hostname = True
        if key in _HW_SIGNAL_HEADERS:
            hw_hits.append(key)
        elif key in _SW_SIGNAL_HEADERS:
            sw_hits.append(key)
        elif key in _ACCOUNT_SIGNAL_HEADERS:
            account_hits.append(key)
        # POA&M / training signals are tracked independently (not part of the
        # HW/SW/account elif chain) so a shared generic header like "status"
        # can still corroborate them without disturbing the inventory tallies.
        if key in _POAM_CORE_HEADERS:
            poam_core.append(key)
        elif key in _POAM_SUPPORT_HEADERS:
            poam_support.append(key)
        if key in _TRAINING_CORE_HEADERS:
            training_core.append(key)
        elif key in _TRAINING_SUPPORT_HEADERS:
            training_support.append(key)
    if len(hw_hits) >= 2:
        return "hw_inventory", hw_hits + (["hostname"] if has_hostname else [])
    if len(sw_hits) >= 2:
        return "sw_inventory", sw_hits + (["hostname"] if has_hostname else [])
    if len(account_hits) >= 2:
        return "account_matrix", account_hits + (["hostname"] if has_hostname else [])
    # Lever C — core-anchored, only reached when no inventory/account shape won.
    poam_hits = poam_core + poam_support
    if poam_core and len(poam_hits) >= 2:
        return "poam", poam_hits
    training_hits = training_core + training_support
    if training_core and len(training_hits) >= 2:
        return "training_record", training_hits
    if has_hostname:
        return "asset_inventory", ["hostname"]
    return None, []


@register(".xlsx", ".xlsm")
def extract_xlsx(stream: BinaryIO, name: str) -> ExtractedDoc:
    """Extract cell text from every sheet of a workbook."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise ExtractorError(
            "openpyxl is not installed — add it to backend/pyproject.toml."
        ) from exc

    stem = PurePosixPath(name).stem

    try:
        wb = load_workbook(
            filename=stream,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:  # pragma: no cover - openpyxl errors vary
        raise ExtractorError(f"openpyxl failed on {name}: {exc}") from exc

    # Refuse to ingest a CCIS workbook as evidence — it's the target.
    # ExtractorSkip (not ExtractorError) so the orchestrator drops it
    # quietly: no Evidence row, no red "failed" tile in the UI. Users
    # who drop a folder containing the workbook itself shouldn't see it
    # surface as a broken artifact.
    if any(sheet_name in _CCIS_MARKER_SHEETS for sheet_name in wb.sheetnames):
        wb.close()
        raise ExtractorSkip(
            f"{name} is a CCIS workbook (has WORKING SHEET tab); "
            "open it via Workbooks, not Evidence."
        )

    sheet_chunks: list[str] = []
    hostnames: list[str] = []
    # Distinct CCI tokens collected from a dedicated CCI column (Lever B). A
    # generic evidence workbook carrying an explicit CCI column is a structured
    # signal on par with a STIG checklist's CCI_REF, so its validated values
    # feed the tagger's ungated 0.95 Tier-2 branch via metadata["cci_refs"].
    # Insertion-ordered (dict) for deterministic output.
    cci_refs: dict[str, None] = {}
    # Best classification seen ANYWHERE in the workbook. The header row with the
    # most matched signals wins, regardless of which sheet/row it sits on — so a
    # Cover/TOC first sheet no longer suppresses a real asset/account header on a
    # later sheet (BUG B re-fix, 2026-06-10). Hostname detection is likewise per
    # sheet: the first hostname column found within the scan window drives host
    # capture for the rest of that sheet, feeding the asset cross-check.
    evidence_type: str | None = None
    evidence_type_signals: list[str] = []
    best_signal_count = 0
    for sheet in wb.worksheets:
        rows: list[str] = []
        sheet_host_col: int | None = None
        sheet_cci_col: int | None = None
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx < _HEADER_SCAN_ROWS:
                # Treat this row as a candidate header. Keep the strongest
                # classification found across the whole workbook so a banner or
                # blank preamble row (which classifies as None) can't lock us in.
                cand_type, cand_signals = _classify_asset_workbook(row)
                if cand_type is not None and len(cand_signals) > best_signal_count:
                    evidence_type = cand_type
                    evidence_type_signals = cand_signals
                    best_signal_count = len(cand_signals)
                # Sniff for a hostname column on this sheet (first one wins).
                if sheet_host_col is None:
                    for i, cell in enumerate(row):
                        if cell is None:
                            continue
                        if str(cell).strip().lower() in _HOSTNAME_HEADERS:
                            sheet_host_col = i
                            break
                # Sniff for a dedicated CCI column on this sheet (first one
                # wins). Detected only from the header allow-list; values are
                # CCI-####-validated below so a false header match is inert.
                if sheet_cci_col is None:
                    for i, cell in enumerate(row):
                        if cell is None:
                            continue
                        if str(cell).strip().lower() in _CCI_HEADER_HEADERS:
                            sheet_cci_col = i
                            break
            if sheet_host_col is not None and sheet_host_col < len(row):
                cell = row[sheet_host_col]
                if cell is not None:
                    s = str(cell).strip()
                    # Skip the header cell itself; capture real host values only.
                    if s and s.lower() not in _HOSTNAME_HEADERS:
                        hostnames.append(s)
            if sheet_cci_col is not None and sheet_cci_col < len(row):
                cell = row[sheet_cci_col]
                if cell is not None:
                    # A CCI cell may hold one token ("CCI-000015") or several
                    # ("CCI-000015, CCI-000018"); collect every validated match.
                    for m in _CCI_RE.finditer(str(cell)):
                        cci_refs.setdefault(m.group(0).upper(), None)
            cells = [str(c) for c in row if c not in (None, "")]
            if cells:
                rows.append(",".join(cells))
        if rows:
            sheet_chunks.append(f"## {sheet.title}\n" + "\n".join(rows))
    wb.close()

    text = "\n\n".join(sheet_chunks)
    metadata: dict = {"sheet_count": len(sheet_chunks)}
    if hostnames:
        # Surface raw values; ingest._capture_host_inventory normalizes.
        metadata["hostnames"] = hostnames
    if evidence_type is not None:
        metadata["evidence_type"] = evidence_type
        metadata["evidence_type_signals"] = evidence_type_signals
    if cci_refs:
        # Consumed by ingest.py → tag_evidence(cci_refs=...) → ungated 0.95
        # Tier-2. Insertion-ordered list (dict keys) for deterministic output.
        metadata["cci_refs"] = list(cci_refs)
    return ExtractedDoc(
        text=text,
        title=stem,
        doc_number=resolve_doc_number(name, stem, text),
        kind=EvidenceKind.XLSX,
        metadata=metadata,
    )
