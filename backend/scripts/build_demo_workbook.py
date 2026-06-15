"""Generate a demo eMASS CCIS workbook fixture.

The shipped product reads workbooks produced by eMASS's Export. To let a
new user try the app without having a real CUI workbook on hand, this
script reproduces the eMASS layout with a handful of fabricated CCIs
spanning a few control families.

The layout MUST match what ``cybersecurity_assessor/excel/ccis_reader.py``
expects (which in turn mirrors the nist-assessor plugin's
``ccis-workbook-guide.md``):

    Sheet name:     "WORKING SHEET"
    Rows 1-5:       system metadata / banner
    Row 6:          column headers
    Rows 7+:        one row per CCI
    Col A:          "YES" or blank (in-scope flag)
    Col B:          Control Acronym (e.g. "AC-2(1)")
    Col G:          AP Acronym (e.g. "AC-2.1") -- objective key
    Col H:          CCI ("CCI-000015" or bare "000015")
    Cols I/J/K:     CCI definition / implementation guidance / procedures
    Cols N-Q:       Assessor fields (status / date / tester / results) --
                    intentionally blank so the user can fill them
    Cols R-U:       PREVIOUS cycle status/date/tester/results -- a few rows
                    pre-populated so the find-evidence flow has something
                    to bite on

Run from ``backend/`` with the project venv::

    cd backend
    uv run python scripts/build_demo_workbook.py

Output: ``backend/tests/fixtures/demo_ccis.xlsx``.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

SHEET_NAME = "WORKING SHEET"

# Header row sits at row 6 in real eMASS exports; rows 1-5 carry the
# program-banner metadata that the reader skips. Reproducing them keeps
# the visual feel of a real workbook even though the reader ignores them.
BANNER_ROW = 1
SYSTEM_ROW = 2
PACKAGE_ROW = 3
ASSESSOR_ROW = 4
NOTE_ROW = 5
HEADER_ROW = 6
DATA_START_ROW = 7

HEADERS: tuple[str, ...] = (
    "Required for assessment?",       # A
    "Control Acronym",                # B
    "Control Information",            # C
    "Control Implementation Status",  # D
    "Security Control Designation",   # E
    "Implementation Narrative",       # F
    "AP Acronym",                     # G
    "CCI",                            # H
    "CCI Definition",                 # I
    "Implementation Guidance",        # J
    "Assessment Procedures",          # K
    "Inherited",                      # L
    "Remote Inheritance Instance",    # M
    "Compliance Status",              # N
    "Date Tested",                    # O
    "Tested By",                      # P
    "Test Results",                   # Q
    "PREVIOUS Compliance Status",     # R
    "PREVIOUS Date Tested",           # S
    "PREVIOUS Tested By",             # T
    "PREVIOUS Test Results",          # U
)

# Per-column display widths (in Excel character units). Tuned so column F
# (narrative) and Q (results) read comfortably without dwarfing the rest.
COLUMN_WIDTHS: dict[str, int] = {
    "A": 10,
    "B": 14,
    "C": 40,
    "D": 16,
    "E": 18,
    "F": 50,
    "G": 12,
    "H": 14,
    "I": 45,
    "J": 45,
    "K": 45,
    "L": 14,
    "M": 22,
    "N": 18,
    "O": 14,
    "P": 18,
    "Q": 50,
    "R": 18,
    "S": 14,
    "T": 18,
    "U": 50,
}


@dataclass
class DemoCci:
    """One demo CCI row.

    Fields map 1:1 to columns A-U; ``cci`` stays as the bare 6-digit form
    eMASS itself uses in column H. ``required`` becomes "YES"/blank.
    """

    required: bool
    control_id: str                # B
    control_information: str       # C
    impl_status: str               # D
    designation: str               # E
    narrative: str                 # F
    ap_acronym: str                # G
    cci: str                       # H -- bare 6-digit
    definition: str                # I
    guidance: str                  # J
    procedures: str                # K
    inherited: str                 # L
    remote_inheritance: str = ""   # M

    # Optional prior-cycle carry-forward (cols R-U). When supplied, the
    # find-evidence flow surfaces these as historical references the way
    # it would for a real returning workbook.
    previous_status: str = ""
    previous_date: str = ""        # ISO yyyy-mm-dd
    previous_tester: str = ""
    previous_results: str = ""


# A small, representative slice across families. Numbers / IDs are
# pulled from the actual DISA CCI list so anyone cross-referencing
# against NIST or eMASS sees recognisable values, but the narratives,
# guidance, and prior-results text are all fabricated for the demo.
DEMO_ROWS: list[DemoCci] = [
    DemoCci(
        required=True,
        control_id="AC-2",
        control_information=(
            "ACCOUNT MANAGEMENT -- The organization manages information system "
            "accounts including establishing, activating, modifying, reviewing, "
            "disabling, and removing accounts."
        ),
        impl_status="Implemented",
        designation="Hybrid",
        narrative=(
            "Account management is handled jointly: the DoW Enterprise Active "
            "Directory tenant provisions standard user accounts via the "
            "automated joiner/mover/leaver workflow; system-local privileged "
            "accounts on the SN application servers are managed by the "
            "platform team per the System Security Plan section 4.2."
        ),
        ap_acronym="AC-2.1",
        cci="000015",
        definition=(
            "The organization identifies and selects which types of "
            "information system accounts support organizational missions/"
            "business functions."
        ),
        guidance=(
            "Documented account types live in the SSP appendix A; refresh "
            "quarterly. Evidence: SSP excerpt, AD group memberships export."
        ),
        procedures=(
            "Examine the SSP for the list of account types; interview the "
            "system owner to confirm the list is current; sample 5 accounts "
            "and verify each maps to a documented type."
        ),
        inherited="Hybrid",
        remote_inheritance="DoW Enterprise AD",
        previous_status="Compliant",
        previous_date="2025-11-12",
        previous_tester="J. Prior-Assessor",
        previous_results=(
            "Examined SSP v3.1 section 4.2; account types documented; "
            "sampled 5 accounts (3 standard, 2 privileged) -- all mapped."
        ),
    ),
    DemoCci(
        required=True,
        control_id="AC-2",
        control_information="ACCOUNT MANAGEMENT (see AC-2 above).",
        impl_status="Implemented",
        designation="Hybrid",
        narrative=(
            "Account assignment is governed by the SDA-issued account "
            "management SOP; managers approve creation in ServiceNow ticket "
            "form RITM00xxxxx prior to account provisioning."
        ),
        ap_acronym="AC-2.2",
        cci="000016",
        definition=(
            "The organization assigns account managers for information "
            "system accounts."
        ),
        guidance=(
            "ServiceNow ticket queue 'Example-SN-AcctMgmt' is the system of "
            "record. Evidence: assignment SOP, ticket sample for last 90 days."
        ),
        procedures=(
            "Examine the account-management SOP; query ServiceNow for the "
            "last 90 days of account-creation tickets; verify each names "
            "the approving manager."
        ),
        inherited="Local",
        previous_status="Compliant",
        previous_date="2025-11-12",
        previous_tester="J. Prior-Assessor",
        previous_results=(
            "Reviewed Account Management SOP (USD00123456); pulled "
            "ServiceNow report 'AcctMgmt-90d-Nov2025' -- all 47 tickets "
            "named an approving manager."
        ),
    ),
    DemoCci(
        required=True,
        control_id="AC-7",
        control_information=(
            "UNSUCCESSFUL LOGON ATTEMPTS -- enforces a limit on consecutive "
            "invalid logon attempts and automatically locks the account."
        ),
        impl_status="Implemented",
        designation="System-Specific",
        narrative=(
            "Account lockout is enforced via GPO 'Example-SN-LockoutPolicy' "
            "applied to the SN OU: 3 invalid attempts within 15 minutes "
            "locks the account for 30 minutes (lockout duration matches "
            "STIG WN10-AC-000005)."
        ),
        ap_acronym="AC-7.1",
        cci="000043",
        definition=(
            "The organization defines the maximum number of consecutive "
            "invalid logon attempts to the information system by a user."
        ),
        guidance=(
            "Lockout threshold and duration set via Group Policy and "
            "documented in the system STIG checklist. Evidence: GPO export, "
            "Windows STIG .cklb showing AC-7 finding closed."
        ),
        procedures=(
            "Inspect the lockout GPO; review STIG checklist findings for "
            "WN10-AC-000005; verify the value matches the ODP in the SSP "
            "(3 attempts)."
        ),
        inherited="Local",
        previous_status="Compliant",
        previous_date="2025-11-13",
        previous_tester="J. Prior-Assessor",
        previous_results=(
            "Reviewed GPO export Example-SN-LockoutPolicy.html dated 2025-10-30; "
            "threshold = 3, duration = 30 min; Windows 10 STIG .cklb shows "
            "WN10-AC-000005 NOT_A_FINDING."
        ),
    ),
    DemoCci(
        required=True,
        control_id="AU-2",
        control_information=(
            "EVENT LOGGING -- identifies the types of events that are "
            "logged within the information system."
        ),
        impl_status="Implemented",
        designation="Hybrid",
        narrative=(
            "Audit event selection is centrally defined by the DoW Enterprise "
            "SIEM team and forwarded by the SN application servers via "
            "Windows Event Forwarding to the Splunk indexer 'sn-prod-idx01'."
        ),
        ap_acronym="AU-2.1",
        cci="000130",
        definition=(
            "The organization determines that the information system is "
            "capable of auditing the organization-defined auditable events."
        ),
        guidance=(
            "Audit policy is enforced via STIG-aligned GPO; verify event "
            "IDs 4624, 4625, 4634, 4720, 4732, 4756 generate. Evidence: "
            "GPO export, Splunk search showing event volume."
        ),
        procedures=(
            "Inspect the auditpol /get /category:* output on a sample SN "
            "server; query Splunk for the last 7 days of each required "
            "event ID; verify non-zero counts."
        ),
        inherited="Hybrid",
        remote_inheritance="DoW Enterprise SIEM",
    ),
    DemoCci(
        required=True,
        control_id="AU-12",
        control_information="AUDIT RECORD GENERATION on system components.",
        impl_status="Implemented",
        designation="System-Specific",
        narrative=(
            "All SN application servers run the Splunk Universal Forwarder "
            "8.2.x; configuration is managed via Ansible playbook "
            "'sn-splunk-uf.yml' and verified weekly by the platform team."
        ),
        ap_acronym="AU-12.1",
        cci="000169",
        definition=(
            "The organization defines the information system components "
            "that provide audit record generation capability."
        ),
        guidance=(
            "List of audit-generating components lives in the SSP appendix B. "
            "Evidence: Ansible inventory, Splunk deployment-server export."
        ),
        procedures=(
            "Examine SSP appendix B; cross-check against the Splunk "
            "deployment-server forwarder list; sample 3 hosts and verify "
            "the UF service is running."
        ),
        inherited="Local",
    ),
    DemoCci(
        required=True,
        control_id="CM-6",
        control_information=(
            "CONFIGURATION SETTINGS -- establishes and documents "
            "configuration settings using security configuration checklists."
        ),
        impl_status="Implemented",
        designation="System-Specific",
        narrative=(
            "Windows Server 2022 and RHEL 9 STIGs are applied via DISA "
            "SCAP content and tracked in Tenable.sc. Findings are "
            "remediated per CCB-approved POAMs."
        ),
        ap_acronym="CM-6.1",
        cci="000363",
        definition=(
            "The organization establishes and documents configuration "
            "settings for information technology products employed within "
            "the information system that reflect the most restrictive mode "
            "consistent with operational requirements."
        ),
        guidance=(
            "STIG checklists (.cklb) live in \\\\sn-fs01\\Compliance\\STIGs. "
            "Evidence: latest .cklb for each platform, Tenable.sc dashboard."
        ),
        procedures=(
            "Examine the most recent .cklb for Windows Server 2022 and "
            "RHEL 9; verify findings counts match the Tenable.sc compliance "
            "report; sample 5 OPEN findings and confirm each has a POAM."
        ),
        inherited="Local",
        previous_status="Non-Compliant",
        previous_date="2025-11-14",
        previous_tester="J. Prior-Assessor",
        previous_results=(
            "Reviewed Windows Server 2022 STIG .cklb dated 2025-10-28: "
            "12 CAT II findings open without POAMs. RHEL 9 STIG clean. "
            "Recommend opening POAM for the 12 open WN findings."
        ),
    ),
    DemoCci(
        required=True,
        control_id="CM-7",
        control_information="LEAST FUNCTIONALITY -- restricts functions and services.",
        impl_status="Implemented",
        designation="System-Specific",
        narrative=(
            "Application allow-listing is enforced via Microsoft AppLocker "
            "on user-facing systems and Carbon Black on servers. Service "
            "baseline is defined per the SN platform hardening guide."
        ),
        ap_acronym="CM-7.1",
        cci="000380",
        definition=(
            "The organization configures the information system to provide "
            "only essential capabilities."
        ),
        guidance=(
            "AppLocker policy is GPO 'Example-SN-AppLocker'; Carbon Black "
            "policy is 'SN-Servers-Restrict'. Evidence: GPO export, CB "
            "policy export, services.msc snapshot."
        ),
        procedures=(
            "Inspect the AppLocker GPO; verify it is in Enforce mode; "
            "examine the Carbon Black policy export; sample a server "
            "and run 'Get-Service' -- verify only baseline services run."
        ),
        inherited="Local",
    ),
    DemoCci(
        required=True,
        control_id="SI-2",
        control_information=(
            "FLAW REMEDIATION -- identifies, reports, and corrects "
            "information system flaws."
        ),
        impl_status="Implemented",
        designation="Hybrid",
        narrative=(
            "Patching is performed monthly via SCCM for Windows and Satellite "
            "for RHEL. Tenable.sc validates remediation within 30 days for "
            "CAT I/II vulnerabilities; CAT III tracked via POAM."
        ),
        ap_acronym="SI-2.1",
        cci="001225",
        definition=(
            "The organization identifies, reports, and corrects information "
            "system flaws."
        ),
        guidance=(
            "Patch cadence and severity SLAs documented in the Vulnerability "
            "Management Plan (USD00345678). Evidence: SCCM compliance report, "
            "Tenable.sc trending dashboard, last 3 monthly patch summaries."
        ),
        procedures=(
            "Examine the Vuln Mgmt Plan; pull SCCM compliance for the "
            "last 90 days; verify CAT I/II patched within SLA; pull a "
            "Tenable.sc report and confirm no overdue CAT I findings."
        ),
        inherited="Hybrid",
        remote_inheritance="DoW Enterprise SCCM",
        previous_status="Compliant",
        previous_date="2025-11-15",
        previous_tester="J. Prior-Assessor",
        previous_results=(
            "Reviewed Vuln Mgmt Plan USD00345678 v2.4; SCCM compliance at "
            "98.2% for Nov 2025 patch cycle; Tenable.sc shows 0 overdue "
            "CAT I, 3 overdue CAT II (POAM-2025-014 open)."
        ),
    ),
    DemoCci(
        required=True,
        control_id="SC-7",
        control_information=(
            "BOUNDARY PROTECTION -- monitors and controls communications "
            "at the external boundary and key internal boundaries."
        ),
        impl_status="Implemented",
        designation="Hybrid",
        narrative=(
            "External boundary is protected by the DoW Enterprise CAP stack "
            "(Palo Alto firewalls + Zscaler proxy). Internal segmentation "
            "between the SN management VLAN and user VLAN is enforced by "
            "the on-prem Cisco Firepower 2140 cluster 'sn-fw-01/02'."
        ),
        ap_acronym="SC-7.1",
        cci="001097",
        definition=(
            "The organization monitors and controls communications at the "
            "external boundary of the information system and at key "
            "internal boundaries within the system."
        ),
        guidance=(
            "Firewall ruleset baseline lives in GitLab project "
            "sda-oi/networking/sn-firewall-config. Evidence: latest rule "
            "export, change tickets for the last 90 days."
        ),
        procedures=(
            "Examine the firewall ruleset export; sample 5 rules and verify "
            "each maps to a documented business need; pull the change-ticket "
            "log and verify each rule modification has CCB approval."
        ),
        inherited="Hybrid",
        remote_inheritance="DoW Enterprise CAP",
    ),
    DemoCci(
        required=True,
        control_id="SC-13",
        control_information=(
            "CRYPTOGRAPHIC PROTECTION -- implements FIPS-validated or NSA-"
            "approved cryptography."
        ),
        impl_status="Implemented",
        designation="System-Specific",
        narrative=(
            "TLS 1.2+ with FIPS-validated cipher suites is enforced on all "
            "SN application endpoints; the Windows servers run in FIPS mode "
            "(GPO 'Example-SN-FIPS') and the RHEL servers have crypto-policies "
            "set to FIPS."
        ),
        ap_acronym="SC-13.1",
        cci="002450",
        definition=(
            "The organization defines the cryptographic uses, and type of "
            "cryptography required for each use."
        ),
        guidance=(
            "Approved crypto uses listed in SSP appendix C. Evidence: "
            "FIPS mode GPO export, output of 'update-crypto-policies "
            "--show' on a sample RHEL host."
        ),
        procedures=(
            "Examine SSP appendix C; verify FIPS GPO is applied; on a "
            "sample RHEL host run 'update-crypto-policies --show' and "
            "confirm output is 'FIPS'."
        ),
        inherited="Local",
    ),
    # Intentionally out-of-scope -- col A blank. Exercises the baseline-load
    # path that has to drop non-required rows.
    DemoCci(
        required=False,
        control_id="SC-28",
        control_information="PROTECTION OF INFORMATION AT REST.",
        impl_status="Not Applicable",
        designation="System-Specific",
        narrative=(
            "Out of scope for this assessment cycle: SN application servers "
            "process no data at rest above the boundary's classification "
            "baseline. Re-scope if/when classified storage is added."
        ),
        ap_acronym="SC-28.1",
        cci="002472",
        definition=(
            "The organization defines information at rest that requires "
            "protection."
        ),
        guidance="N/A -- control de-scoped via tailoring decision documented in SSP.",
        procedures="N/A.",
        inherited="Local",
    ),
    DemoCci(
        required=True,
        control_id="AC-17",
        control_information="REMOTE ACCESS.",
        impl_status="Implemented",
        designation="Common",
        narrative=(
            "Inherited from the DoW Enterprise remote-access service "
            "(GFE laptop + AppGate SDP). No system-local remote access "
            "is permitted to the SN management VLAN."
        ),
        ap_acronym="AC-17.1",
        cci="000067",
        definition=(
            "The organization establishes and documents usage restrictions, "
            "configuration/connection requirements, and implementation "
            "guidance for each type of remote access allowed."
        ),
        guidance=(
            "Inheritance package from DoW Enterprise covers AC-17.1; CRM "
            "row 'AC-17.1 -- Provider' marked Provider Responsibility. "
            "Evidence: CRM excerpt, AppGate SDP policy export."
        ),
        procedures=(
            "Examine the Enterprise CRM AC-17 section; verify the system "
            "consumes the inherited control; confirm no local exceptions."
        ),
        inherited="DoW Enterprise",
        remote_inheritance="DoW Enterprise AppGate",
        previous_status="Compliant",
        previous_date="2025-11-15",
        previous_tester="J. Prior-Assessor",
        previous_results=(
            "Reviewed Enterprise CRM dated 2025-09-01; AC-17.1 marked "
            "Provider Responsibility; no local exceptions documented. "
            "Inherited Compliant."
        ),
    ),
]


def _write_banner(ws: Worksheet) -> None:
    """Populate rows 1-5 with system-banner metadata.

    The reader skips these rows entirely -- they exist purely so the file
    looks like a real eMASS export when the user opens it in Excel.
    """
    bold = Font(bold=True, color="FFFFFF")
    banner_fill = PatternFill("solid", fgColor="0E2C4A")  # nuon navy-deep

    ws.cell(row=BANNER_ROW, column=1, value="DEMO -- CCIS Working Sheet (synthetic data)")
    ws.cell(row=BANNER_ROW, column=1).font = bold
    ws.cell(row=BANNER_ROW, column=1).fill = banner_fill
    ws.merge_cells(start_row=BANNER_ROW, start_column=1, end_row=BANNER_ROW, end_column=21)

    ws.cell(row=SYSTEM_ROW, column=1, value="System:")
    ws.cell(row=SYSTEM_ROW, column=1).font = Font(bold=True)
    ws.cell(row=SYSTEM_ROW, column=2, value="Demo System (Sample Network -- not a real authorization)")

    ws.cell(row=PACKAGE_ROW, column=1, value="Package:")
    ws.cell(row=PACKAGE_ROW, column=1).font = Font(bold=True)
    ws.cell(row=PACKAGE_ROW, column=2, value="DEMO-PKG-001 / IATT")

    ws.cell(row=ASSESSOR_ROW, column=1, value="Assessor:")
    ws.cell(row=ASSESSOR_ROW, column=1).font = Font(bold=True)
    ws.cell(row=ASSESSOR_ROW, column=2, value="(your name) -- fill in column P per row")

    ws.cell(
        row=NOTE_ROW,
        column=1,
        value=(
            "This is a synthetic demo workbook -- no real CUI. Use it to "
            "exercise the assessor app's open / baseline / find-evidence / "
            "assess / generate-POAM flows end to end."
        ),
    )


def _write_headers(ws: Worksheet) -> None:
    """Row 6: column headers, styled to read as a real eMASS export."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")  # darker navy
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for col_idx, label in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[HEADER_ROW].height = 36


def _write_data(ws: Worksheet, rows: list[DemoCci]) -> None:
    wrap = Alignment(wrap_text=True, vertical="top")
    for offset, demo in enumerate(rows):
        r = DATA_START_ROW + offset
        ws.cell(row=r, column=1, value="YES" if demo.required else None)
        ws.cell(row=r, column=2, value=demo.control_id)
        ws.cell(row=r, column=3, value=demo.control_information)
        ws.cell(row=r, column=4, value=demo.impl_status)
        ws.cell(row=r, column=5, value=demo.designation)
        ws.cell(row=r, column=6, value=demo.narrative)
        ws.cell(row=r, column=7, value=demo.ap_acronym)
        ws.cell(row=r, column=8, value=demo.cci)
        ws.cell(row=r, column=9, value=demo.definition)
        ws.cell(row=r, column=10, value=demo.guidance)
        ws.cell(row=r, column=11, value=demo.procedures)
        ws.cell(row=r, column=12, value=demo.inherited)
        ws.cell(row=r, column=13, value=demo.remote_inheritance or None)
        # Columns N-Q (14-17) -- assessor-writable, intentionally blank.
        # Columns R-U (18-21) -- prior-cycle carry-forward, optional.
        if demo.previous_status:
            ws.cell(row=r, column=18, value=demo.previous_status)
        if demo.previous_date:
            ws.cell(row=r, column=19, value=demo.previous_date)
        if demo.previous_tester:
            ws.cell(row=r, column=20, value=demo.previous_tester)
        if demo.previous_results:
            ws.cell(row=r, column=21, value=demo.previous_results)

        # Wrap the long narrative / definition / guidance / procedure / results
        # columns so the row renders readable instead of single-line truncated.
        for col in (3, 6, 9, 10, 11, 17, 21):
            ws.cell(row=r, column=col).alignment = wrap

        # Generous height so wrapped text actually shows; openpyxl doesn't
        # auto-fit on save.
        ws.row_dimensions[r].height = 110


def _apply_widths(ws: Worksheet) -> None:
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width


def build() -> Path:
    wb = Workbook()
    # openpyxl seeds a default "Sheet" -- rename rather than create-and-delete
    # so the workbook has exactly one sheet and the WORKING SHEET resolver
    # finds it without falling back to its case-insensitive search.
    ws = wb.active
    assert ws is not None  # for the type checker
    ws.title = SHEET_NAME

    # Freeze panes so the header stays visible while scrolling through CCIs.
    # G7 freezes rows 1-6 (banner + header) and cols A-F (the most-referenced
    # leading columns).
    ws.freeze_panes = "G7"

    _write_banner(ws)
    _write_headers(ws)
    _write_data(ws, DEMO_ROWS)
    _apply_widths(ws)

    # Repo layout: this file lives at backend/scripts/, fixture lives at
    # backend/tests/fixtures/. Resolve relative to this file so the script
    # works regardless of caller cwd.
    fixtures_dir = Path(__file__).resolve().parents[1] / "tests" / "fixtures"
    fixtures_dir.mkdir(parents=True, exist_ok=True)
    out = fixtures_dir / "demo_ccis.xlsx"
    wb.save(out)
    return out


def main() -> None:
    out = build()
    cols_used = len(HEADERS)
    last_col_letter = get_column_letter(cols_used)
    print(f"wrote {out}")
    print(
        f"  sheet={SHEET_NAME!r}  headers=row {HEADER_ROW} (A..{last_col_letter})  "
        f"rows={len(DEMO_ROWS)}  in_scope={sum(1 for r in DEMO_ROWS if r.required)}"
    )


if __name__ == "__main__":
    main()
