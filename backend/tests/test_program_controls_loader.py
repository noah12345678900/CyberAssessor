"""Unit tests for catalogs/program_controls_loader.py.

Focus: the synthetic Baseline materialization added so program overlays
(SDA Enterprise Services Controls, T1TL Ground Security Controls, etc.)
appear in the WorkbookOverlay surface. Without these, an overlay load
succeeds (RequirementSource + RequirementMap rows written) but the
Workbooks page "Overlays" column never renders the chip because it
joins WorkbookOverlay → Baseline.

Covered scenarios:
  * First-load creates a Baseline + BaselineControl rows + auto-attaches
    to existing workbooks on the same framework.
  * Reload (same overlay, same source_name) is idempotent — no duplicate
    Baseline, BaselineControl rows are wiped+rewritten, and prior
    detach choices stick (no re-attach).
  * Workbooks bound to a different framework do NOT get auto-attached.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook as XlsxWorkbook
from openpyxl.styles import Border, Side
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

_BACKEND = Path(__file__).resolve().parents[1]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.catalogs.program_controls_loader import (  # noqa: E402
    LOADER_VERSION,
    load_program_controls,
)
from cybersecurity_assessor.models import (  # noqa: E402
    Baseline,
    BaselineControl,
    BaselineSourceType,
    Control,
    Framework,
    IngestReport,
    Objective,
    RequirementMap,
    RequirementSource,
    Workbook,
    WorkbookOverlay,
)


def _build_overlay_xlsx(path: Path) -> None:
    """Write a minimal CCI-grain overlay sheet at ``path``.

    Two shall rows referencing two CCIs under AC-2. Header at row 1.
    Matches the SDA Enterprise Services Controls shape (Control No /
    Threshold / CCI columns), which is the easiest path to exercise.
    """
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Overlay"
    ws.append(["Control No", "Threshold", "CCI"])
    ws.append(["SDA-001", "The system shall do X.", "CCI-000015"])
    ws.append(["SDA-002", "The system shall do Y.", "CCI-000007"])
    wb.save(path)


@pytest.fixture
def env(tmp_path: Path):
    """In-memory SQLite seeded with two frameworks (r5, r4), AC-2 + two
    CCI Objectives on r5, plus a workbook on each framework. Returns the
    session + the path to a freshly written overlay xlsx.
    """
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    s = Session(engine)

    fw_r5 = Framework(name="NIST SP 800-53", version="Rev 5")
    fw_r4 = Framework(name="NIST SP 800-53", version="Rev 4")
    s.add_all([fw_r5, fw_r4])
    s.commit()
    s.refresh(fw_r5)
    s.refresh(fw_r4)

    ctrl = Control(
        framework_id=fw_r5.id,
        control_id="AC-2",
        title="Account Management",
        family="AC",
    )
    s.add(ctrl)
    s.commit()
    s.refresh(ctrl)

    s.add_all(
        [
            Objective(
                control_id_fk=ctrl.id,
                objective_id="CCI-000015",
                source="CCI",
                text="Account types defined.",
            ),
            Objective(
                control_id_fk=ctrl.id,
                objective_id="CCI-000007",
                source="CCI",
                text="Account approval recorded.",
            ),
        ]
    )
    s.commit()

    wb_on_r5 = Workbook(
        path="C:/fake/wb_r5.xlsx", filename="wb_r5.xlsx", framework_id=fw_r5.id
    )
    wb_on_r4 = Workbook(
        path="C:/fake/wb_r4.xlsx", filename="wb_r4.xlsx", framework_id=fw_r4.id
    )
    s.add_all([wb_on_r5, wb_on_r4])
    s.commit()
    s.refresh(wb_on_r5)
    s.refresh(wb_on_r4)

    overlay_path = tmp_path / "overlay.xlsx"
    _build_overlay_xlsx(overlay_path)

    return {
        "session": s,
        "fw_r5_id": fw_r5.id,
        "fw_r4_id": fw_r4.id,
        "ctrl_id": ctrl.id,
        "wb_on_r5_id": wb_on_r5.id,
        "wb_on_r4_id": wb_on_r4.id,
        "overlay_path": overlay_path,
    }


def test_first_load_materializes_baseline_without_auto_attach(env):
    """A brand-new program overlay → synthetic Baseline + BaselineControl
    rows in the global catalog, but NO WorkbookOverlay rows. Attach is the
    caller's responsibility (Manage Overlays dialog) so loader semantics
    are predictable on reload — every attach is explicit, never gated on
    "first creation" branching."""
    s: Session = env["session"]
    source = load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    # RequirementMaps wrote both rows.
    assert source.__dict__["_rows_seen"] == 2
    assert source.__dict__["_maps_written"] == 2

    # Synthetic Baseline exists, tagged PROGRAM_CONTROLS.
    bl = s.exec(
        select(Baseline).where(
            Baseline.framework_id == env["fw_r5_id"],
            Baseline.name == "SDA Enterprise Services Controls",
        )
    ).first()
    assert bl is not None
    assert bl.source_type == BaselineSourceType.PROGRAM_CONTROLS
    assert source.__dict__["_baseline_was_created"] is True
    assert source.__dict__["_baseline_id"] == bl.id

    # Both CCIs roll up to AC-2 → exactly one BaselineControl row.
    bcs = s.exec(
        select(BaselineControl).where(BaselineControl.baseline_id == bl.id)
    ).all()
    assert len(bcs) == 1
    assert bcs[0].control_id == env["ctrl_id"]
    assert bcs[0].in_scope is True

    # Pure catalog op — no WorkbookOverlay rows on either same-framework
    # or cross-framework workbooks. Caller must explicitly attach.
    overlays = s.exec(
        select(WorkbookOverlay).where(WorkbookOverlay.baseline_id == bl.id)
    ).all()
    assert overlays == [], "Loader must not auto-attach to any workbook"
    assert source.__dict__["_auto_attached_workbook_ids"] == []


def test_reload_is_idempotent_and_respects_detach(env):
    """Re-running the loader against the same overlay must not duplicate
    Baseline rows AND must not write WorkbookOverlay rows on its own —
    if a user explicitly attached then detached, a reload must not
    silently re-attach.
    """
    s: Session = env["session"]
    load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    bl = s.exec(
        select(Baseline).where(
            Baseline.name == "SDA Enterprise Services Controls"
        )
    ).first()
    assert bl is not None

    # Simulate the user explicitly attaching the overlay to the r5
    # workbook via Manage Overlays — the loader itself never writes this
    # row under the pure-catalog contract. Then simulate detach.
    s.add(
        WorkbookOverlay(workbook_id=env["wb_on_r5_id"], baseline_id=bl.id)
    )
    s.commit()
    detach_target = s.exec(
        select(WorkbookOverlay).where(
            WorkbookOverlay.workbook_id == env["wb_on_r5_id"],
            WorkbookOverlay.baseline_id == bl.id,
        )
    ).first()
    assert detach_target is not None
    s.delete(detach_target)
    s.commit()

    # Re-run the loader.
    reloaded = load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    # No duplicate Baseline.
    baselines = s.exec(
        select(Baseline).where(
            Baseline.name == "SDA Enterprise Services Controls"
        )
    ).all()
    assert len(baselines) == 1
    assert reloaded.__dict__["_baseline_was_created"] is False
    assert reloaded.__dict__["_baseline_id"] == bl.id

    # No duplicate BaselineControl rows (idempotent rewrite).
    bcs = s.exec(
        select(BaselineControl).where(BaselineControl.baseline_id == bl.id)
    ).all()
    assert len(bcs) == 1

    # User detach choice stuck — no re-attach.
    overlays = s.exec(
        select(WorkbookOverlay).where(WorkbookOverlay.baseline_id == bl.id)
    ).all()
    assert overlays == [], "Reload must not re-attach detached overlays"
    assert reloaded.__dict__["_auto_attached_workbook_ids"] == []

    # RequirementMaps wiped + rewritten — no duplicates.
    maps = s.exec(
        select(RequirementMap).where(
            RequirementMap.requirement_source_id == reloaded.id
        )
    ).all()
    assert len(maps) == 2


def test_two_files_same_source_name_produce_two_rows(env, tmp_path):
    """Regression: loading two distinct files that happen to share the same
    ``source_name`` must produce two ``RequirementSource`` rows (and two
    synthetic ``Baseline`` rows), not collapse into one.

    Previously the loader keyed the upsert on ``(framework_id, name)``,
    which silently overwrote the first load when a second file with the
    same label arrived — visible to the user as "I attached two overlays
    but only one shows in Settings → Catalogs." The fix re-keys on
    ``(framework_id, path)`` for both RequirementSource and Baseline.
    """
    s: Session = env["session"]

    # Same label, two distinct files. Second file built to a different path
    # but with the same content shape so both produce mappable rows.
    second_path = tmp_path / "overlay_second.xlsx"
    _build_overlay_xlsx(second_path)

    shared_label = "SDA Enterprise Services Controls"

    src_a = load_program_controls(
        s,
        source_name=shared_label,
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    src_b = load_program_controls(
        s,
        source_name=shared_label,
        workbook_path=second_path,
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )

    # Two distinct RequirementSource rows.
    assert src_a.id != src_b.id
    assert src_a.path != src_b.path
    sources = s.exec(
        select(RequirementSource).where(
            RequirementSource.framework_id == env["fw_r5_id"],
            RequirementSource.name == shared_label,
        )
    ).all()
    assert len(sources) == 2
    assert {src.path for src in sources} == {
        str(env["overlay_path"]),
        str(second_path),
    }

    # And two distinct synthetic Baseline rows — keyed on source_ref/path
    # the same way, so the Settings → Catalogs view sees both.
    baselines = s.exec(
        select(Baseline).where(
            Baseline.framework_id == env["fw_r5_id"],
            Baseline.name == shared_label,
            Baseline.source_type == BaselineSourceType.PROGRAM_CONTROLS,
        )
    ).all()
    assert len(baselines) == 2
    assert {bl.source_ref for bl in baselines} == {
        str(env["overlay_path"]),
        str(second_path),
    }


def test_reupload_same_path_upserts_in_place(env):
    """Reloading the same file path must NOT create a duplicate row — that's
    the 'reload to refresh' UX the loader has always supported."""
    s: Session = env["session"]
    src_a = load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    src_b = load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls (renamed)",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    assert src_a.id == src_b.id
    # Name was updated to reflect the latest load.
    s.refresh(src_b)
    assert src_b.name == "SDA Enterprise Services Controls (renamed)"
    # And exactly one Baseline row exists for this path.
    baselines = s.exec(
        select(Baseline).where(
            Baseline.framework_id == env["fw_r5_id"],
            Baseline.source_ref == str(env["overlay_path"]),
            Baseline.source_type == BaselineSourceType.PROGRAM_CONTROLS,
        )
    ).all()
    assert len(baselines) == 1


def test_other_framework_workbooks_not_attached(env):
    """The r4 workbook must never receive an r5 overlay regardless of
    load order — this is the bug that originally hid SDA from the
    Workbooks page chip list."""
    s: Session = env["session"]
    load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    r4_overlays = s.exec(
        select(WorkbookOverlay).where(
            WorkbookOverlay.workbook_id == env["wb_on_r4_id"]
        )
    ).all()
    assert r4_overlays == []


def _build_overlay_with_unmerged_tall_block(
    path: Path, *, continuation_top_border: bool
) -> None:
    """Two-row overlay where row 3 has a blank col-A cell.

    Row 2 holds "SDA-001" as the program req number for the first CCI.
    Row 3 col A is ``None`` — the question the loader has to answer is
    "did Excel encode this as a tall-cell continuation (no top border on
    the blank cell) or as a real workbook gap (top border drawn)?"

    ``continuation_top_border=False`` mimics the T1TL Ground Security
    Controls AU-2 block where col A spans sub-bullets a-l visually but
    is unmerged in the XML — Excel renders the visual continuity by
    drawing NO top border on the continuation rows. The loader must
    forward-fill "SDA-001" into row 3.

    ``continuation_top_border=True`` mimics a genuine workbook gap (an
    orphan shall-statement) where Excel draws the normal cell border.
    The loader must keep "(unnumbered)" so the operator sees the row in
    the ingest report and decides.
    """
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Overlay"
    ws.append(["Control No", "Threshold", "CCI"])
    ws.append(["SDA-001", "The system shall do X.", "CCI-000015"])
    ws.append([None, "The system shall do Y.", "CCI-000007"])
    if continuation_top_border:
        ws.cell(row=3, column=1).border = Border(top=Side(style="thin"))
    wb.save(path)


def test_forward_fills_continuation_row_without_top_border(env, tmp_path):
    """Unmerged tall col-A cell continuation rows (no top border on the
    blank cell) must inherit the parent req number — this is the T1TL
    AU-2 / SC-7(9) / SI-3 / CP-9 / CP-10 / AC-17 bug surface. Before the
    border-gated fix, openpyxl saw ``None`` in col A on continuation rows
    and the loader silently stamped them as "(unnumbered)", which
    surfaced in the catalog UI as orphan rows even though the workbook
    visually shows them under their parent control.
    """
    s: Session = env["session"]
    overlay = tmp_path / "tall_block_no_border.xlsx"
    _build_overlay_with_unmerged_tall_block(overlay, continuation_top_border=False)

    source = load_program_controls(
        s,
        source_name="T1TL Ground Security Controls",
        workbook_path=overlay,
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )

    # Both rows produced a RequirementMap — the continuation row didn't
    # vanish, it just inherited the parent's req number.
    maps = s.exec(
        select(RequirementMap)
        .where(RequirementMap.requirement_source_id == source.id)
        .order_by(RequirementMap.id)  # type: ignore[arg-type]
    ).all()
    assert len(maps) == 2
    assert maps[0].requirement_number == "SDA-001"
    assert maps[1].requirement_number == "SDA-001", (
        "Row 3 must inherit SDA-001 from row 2 because col-A has no "
        "top border — that's how Excel encodes unmerged tall cells."
    )
    # No "(unnumbered)" sentinels written.
    assert not any(m.requirement_number == "(unnumbered)" for m in maps)

    # Counters reflect the decision.
    assert source.__dict__["_rows_seen"] == 2
    assert source.__dict__["_rows_forward_filled"] == 1
    assert source.__dict__["_rows_unnumbered"] == 0

    # Action log captured the forward-fill with the source value so a
    # 3PAO can reconstruct the decision from the IngestReport JSON.
    actions = source.__dict__["_actions"]
    forward_fills = [a for a in actions if a["action"] == "forward_fill"]
    assert len(forward_fills) == 1
    assert forward_fills[0]["from_value"] == "SDA-001"
    assert forward_fills[0]["row"] == 3


def test_preserves_unnumbered_when_top_border_present(env, tmp_path):
    """A blank col-A cell WITH a visible top border is a genuine workbook
    gap, not a continuation row. The loader must NOT forward-fill — the
    "(unnumbered)" sentinel is what surfaces the row to the operator so
    they can decide if it needs a manual mapping."""
    s: Session = env["session"]
    overlay = tmp_path / "real_gap.xlsx"
    _build_overlay_with_unmerged_tall_block(overlay, continuation_top_border=True)

    source = load_program_controls(
        s,
        source_name="T1TL Ground Security Controls",
        workbook_path=overlay,
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )

    maps = s.exec(
        select(RequirementMap)
        .where(RequirementMap.requirement_source_id == source.id)
        .order_by(RequirementMap.id)  # type: ignore[arg-type]
    ).all()
    assert len(maps) == 2
    assert maps[0].requirement_number == "SDA-001"
    assert maps[1].requirement_number == "(unnumbered)", (
        "Top border on the blank col-A cell signals a real gap — the "
        "row must not silently inherit SDA-001."
    )

    assert source.__dict__["_rows_forward_filled"] == 0
    assert source.__dict__["_rows_unnumbered"] == 1

    actions = source.__dict__["_actions"]
    unnumbered = [a for a in actions if a["action"] == "unnumbered_block_start"]
    assert len(unnumbered) == 1
    assert unnumbered[0]["reason"] == "top_border_present"
    assert unnumbered[0]["row"] == 3


def test_ingest_report_persists_with_loader_version_and_counters(env):
    """Every loader run writes exactly one IngestReport row — the audit
    trail that 3PAO reviewers join from RequirementSource to answer
    "which loader version produced these maps, and what structural
    decisions did it make?" without re-running the loader.
    """
    s: Session = env["session"]
    source = load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )

    reports = s.exec(
        select(IngestReport).where(
            IngestReport.requirement_source_id == source.id
        )
    ).all()
    assert len(reports) == 1
    report = reports[0]
    assert report.loader_version == LOADER_VERSION
    assert report.framework_id == env["fw_r5_id"]
    assert report.source_path == str(env["overlay_path"])
    assert report.sheet_name == "Overlay"
    assert report.rows_seen == 2
    assert report.maps_written == 2
    # The bare-shape SDA fixture has no unmerged tall blocks, so the
    # forward-fill / unnumbered counters stay at zero — that's exactly
    # the audit signal a 3PAO wants on a clean load.
    assert report.rows_forward_filled == 0
    assert report.rows_unnumbered == 0
    assert report.unmapped_ccis == []
    assert report.unmapped_control_ids == []
    assert report.actions == []
    # Transient attr exposes the new row's id for the HTTP response.
    assert source.__dict__["_ingest_report_id"] == report.id
    assert source.__dict__["_loader_version"] == LOADER_VERSION


def test_ingest_report_rewritten_on_reload(env):
    """Reload semantics: a second load against the same source path adds
    a SECOND IngestReport row rather than overwriting the first. The
    history is the point — a 3PAO needs to see "this RequirementSource
    has been re-ingested three times; here's what changed between v1
    and v3" — so we never collapse audit rows.
    """
    s: Session = env["session"]
    load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    reloaded = load_program_controls(
        s,
        source_name="SDA Enterprise Services Controls",
        workbook_path=env["overlay_path"],
        framework_id=env["fw_r5_id"],
        sheet_name="Overlay",
    )
    reports = s.exec(
        select(IngestReport)
        .where(IngestReport.requirement_source_id == reloaded.id)
        .order_by(IngestReport.id)  # type: ignore[arg-type]
    ).all()
    assert len(reports) == 2
    # Both rows record the same source path + loader version — the
    # delta lives in created_at and (eventually) in actions when the
    # underlying workbook content drifts.
    assert {r.loader_version for r in reports} == {LOADER_VERSION}
    assert {r.source_path for r in reports} == {str(env["overlay_path"])}
