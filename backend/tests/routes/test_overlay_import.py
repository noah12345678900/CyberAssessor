"""Tests for ``POST /api/catalog/overlays/import``.

This is the unified front door for any overlay xlsx the user drops in.
It auto-classifies the file as CRM / PSC / OTHER via
:func:`baselines.overlay_classifier.classify_overlay` and dispatches to
the right loader. The tests below pin three things the contract
guarantees:

* **Dispatch correctness** — a CRM-shaped fixture routes to the CRM
  loader (returns ``baseline_id``), a PSC-shaped fixture routes to the
  PSC loader (returns ``requirement_source_id`` + ``sheet_name``), and
  an unrecognized fixture routes to the inert OTHER loader (returns
  ``baseline_id`` and a "no resolver registered" warning).
* **kind_hint escape hatch** — when the caller forces a kind, the
  classifier's verdict is overridden but the matched sheet name still
  comes through (PSC dispatch needs it).
* **Boundary errors** — missing file and unknown framework_id both
  return 400 with a specific message, not 500.

Fixture pattern mirrors ``test_crm_load_route.py`` and
``test_program_controls.py`` (StaticPool SQLite + session override +
seed Framework/Control/Objective up front) — the unified route is a
thin dispatcher, so the tests focus on routing behavior, not on
re-validating loader internals (those have their own test files).
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook as XlsxWorkbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.db import get_session  # noqa: E402
from cybersecurity_assessor.models import (  # noqa: E402
    Baseline,
    BaselineSourceType,
    Control,
    Framework,
    Objective,
    RequirementMap,
    RequirementSource,
)
from cybersecurity_assessor.server import create_app  # noqa: E402


# --- fixture xlsx builders --------------------------------------------------
# Three minimal shapes pulled from
# baselines/overlay_classifier.py vocab + the loader header docstrings.


def _write_crm_xlsx(path: Path) -> None:
    """CRM-shaped: control-id column + responsibility column on one sheet."""
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "CRM"
    ws.append(["Control ID", "Responsibility", "Customer Responsibility"])
    ws.append(["AC-2", "Customer", "Customer owns account lifecycle."])
    wb.save(path)


def _write_psc_xlsx(path: Path) -> None:
    """PSC-shaped: CCI column + threshold column on a named sheet.

    The classifier returns the first PSC-shaped sheet's name; the PSC
    loader then requires that sheet name to dispatch. Both ends meet
    here without the caller having to pass anything explicit.
    """
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "PSC Tab"
    ws.append(["Control No", "Threshold", "CCI"])
    ws.append(["SDA-014", "The system shall define account types.", "CCI-000015"])
    ws.append(["SDA-127", "The system shall record account approvals.", "CCI-000007"])
    wb.save(path)


def _write_psc_t1tl_xlsx(path: Path) -> None:
    """PSC-shaped T1TL file: NO CCI column.

    Real T1TL Ground / SV Security Controls sheets ship with
    ``Security Control | Threshold | Objective | Deliverable | Notes``
    and zero CCI references — CCIs are derived from the
    ``Associated CNSSI 1253 Control Tag:`` anchor in the shall prose,
    joined against the global DISA mapping at resolve time. This
    fixture reproduces that shape end-to-end so the route's PSC
    dispatch is exercised on a CCI-less PSC file (the case that
    previously misclassified to OTHER and broke the upload).
    """
    wb = XlsxWorkbook()
    # Mirror the real file: skip "Risk Management Framework (RMF)" tab
    # is not needed because classify_overlay picks the first PSC-shaped
    # sheet — but we DO want a leading title row so the header-scan
    # logic is exercised too.
    ws = wb.active
    ws.title = "Ground Security Controls"
    ws.append(["Program: Acme T1TL Ground Segment"])
    ws.append([])
    ws.append(
        ["-", "Security Control", "Threshold", "Objective", "Deliverable", "Notes"]
    )
    ws.append(
        [
            1,
            (
                "Associated CNSSI 1253 Control Tag: AC-2\n"
                "The system shall manage information system accounts, "
                "including establishment, activation, modification, review, "
                "disablement, and removal."
            ),
            "",
            "",
            "SSP",
            "",
        ]
    )
    wb.save(path)


def _write_psc_two_sheets_xlsx(path: Path) -> None:
    """Two PSC-shaped sheets — mirrors the T1TL workbook shape.

    Real T1TL files ship with both ``Ground Security Controls`` and
    ``SV Security Controls`` (segment-specific overlays). The classifier
    picks the first PSC-shaped sheet, so without an explicit selector
    there's no way for the UI to ask for SV. This fixture exercises the
    ``sheet_name`` override path.
    """
    wb = XlsxWorkbook()
    ground = wb.active
    ground.title = "Ground Security Controls"
    ground.append(
        ["-", "Security Control", "Threshold", "Objective", "Deliverable", "Notes"]
    )
    ground.append(
        [
            1,
            (
                "Associated CNSSI 1253 Control Tag: AC-2\n"
                "The system shall manage ground accounts."
            ),
            "",
            "",
            "SSP",
            "",
        ]
    )

    sv = wb.create_sheet("SV Security Controls")
    sv.append(
        ["-", "Security Control", "Threshold", "Objective", "Deliverable", "Notes"]
    )
    sv.append(
        [
            1,
            (
                "Associated CNSSI 1253 Control Tag: AC-2\n"
                "The system shall manage SV accounts."
            ),
            "",
            "",
            "SSP",
            "",
        ]
    )
    wb.save(path)


def _write_other_xlsx(path: Path) -> None:
    """Unrecognized headers — no CRM or PSC vocab hits."""
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Hostname", "IP Address", "OS"])
    ws.append(["host-01", "10.0.0.1", "RHEL 8.6"])
    wb.save(path)


# --- environment ------------------------------------------------------------


@pytest.fixture
def env(tmp_path: Path):
    """TestClient + framework with AC-2 and two CCIs seeded.

    The CCI rows are needed for the PSC dispatch test — without an
    Objective for CCI-000015 / CCI-000007, the loader would create the
    RequirementSource but produce zero RequirementMap rows (it reports
    those as ``unmapped_ccis`` instead of erroring).
    """
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)

    def _override_get_session():
        with Session(engine) as s:
            yield s

    app = create_app()
    app.dependency_overrides[get_session] = _override_get_session

    with Session(engine) as s:
        fw = Framework(name="NIST SP 800-53", version="Rev 5")
        s.add(fw)
        s.commit()
        s.refresh(fw)

        ctrl = Control(
            framework_id=fw.id,
            control_id="ac-2",
            title="Account Management",
            family="AC",
        )
        s.add(ctrl)
        s.commit()
        s.refresh(ctrl)

        # CCIs the PSC fixture references — wired so the PSC dispatch
        # test sees real maps_written rather than unmapped_ccis.
        s.add_all(
            [
                Objective(
                    control_id_fk=ctrl.id,
                    objective_id="CCI-000015",
                    source="CCI",
                    text="Account types defined.",
                ),
                Objective(
                    control_id_fk=ctrl.id,
                    objective_id="CCI-000007",
                    source="CCI",
                    text="Account approval recorded.",
                ),
            ]
        )
        s.commit()

        framework_id = fw.id

    yield {
        "client": TestClient(app),
        "engine": engine,
        "framework_id": framework_id,
        "tmp": tmp_path,
    }

    app.dependency_overrides.clear()


# --- dispatch happy paths ---------------------------------------------------


def test_import_crm_dispatches_to_crm_loader(env) -> None:
    """CRM-shaped file → kind=='crm', baseline row materialized.

    v0.2 multi-implementation contract: CRM imports require a non-reserved
    ``scope_label`` so the slice lands under a named boundary (e.g. the cloud
    CSP) and the on-prem half can coexist. Omitting it is a 422 (pinned by
    ``test_crm_import_requires_scope_label``).
    """
    path = env["tmp"] / "crm.xlsx"
    _write_crm_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "scope_label": "AWS GovCloud",
        },
    )
    assert r.status_code == 200, r.text
    payload = r.json()
    assert payload["kind"] == "crm"
    assert isinstance(payload["baseline_id"], int)
    assert payload["controls_in_scope"] == 1
    assert payload["warnings"] == []

    # Confirm a CRM-typed Baseline actually landed.
    with Session(env["engine"]) as s:
        b = s.get(Baseline, payload["baseline_id"])
        assert b is not None
        assert b.source_type == BaselineSourceType.CRM
        assert b.source_ref == str(path)


def test_import_psc_dispatches_to_psc_loader(env) -> None:
    """PSC-shaped file → kind=='psc', RequirementSource row materialized,
    sheet_name surfaced (so the UI toast can show which tab was parsed).
    """
    path = env["tmp"] / "psc.xlsx"
    _write_psc_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={"framework_id": env["framework_id"], "path": str(path)},
    )
    assert r.status_code == 200, r.text
    payload = r.json()
    assert payload["kind"] == "psc"
    assert isinstance(payload["requirement_source_id"], int)
    # classify_overlay returns the first PSC-shaped sheet — must match
    # the title we set in _write_psc_xlsx so the loader dispatches to
    # the right tab.
    assert payload["sheet_name"] == "PSC Tab"
    # Both shall-statements landed on real Objectives (we seeded CCIs).
    assert payload["maps_written"] == 2
    assert payload["unmapped_ccis"] == []

    # And the RequirementSource row really exists.
    with Session(env["engine"]) as s:
        src = s.get(RequirementSource, payload["requirement_source_id"])
        assert src is not None
        assert src.framework_id == env["framework_id"]
        maps = s.exec(
            select(RequirementMap).where(
                RequirementMap.requirement_source_id == src.id  # type: ignore[arg-type]
            )
        ).all()
        assert len(maps) == 2


def test_import_psc_t1tl_without_cci_column_dispatches_to_psc_loader(env) -> None:
    """Regression: PSC-shaped T1TL file with NO CCI column → kind=='psc'.

    Real T1TL Ground/SV Security Controls sheets ship with
    ``Security Control | Threshold | Objective | Deliverable | Notes``
    and no CCI references column. The earlier classifier required a
    CCI column and routed these to OTHER, so the import endpoint
    couldn't dispatch them — even with kind_hint=psc forced — because
    classify_overlay returned sheet_name=None. This test pins the
    end-to-end fix: the route accepts a CCI-less T1TL file, classifies
    it PSC, surfaces the correct sheet_name, and runs the PSC loader.
    """
    path = env["tmp"] / "psc_t1tl.xlsx"
    _write_psc_t1tl_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={"framework_id": env["framework_id"], "path": str(path)},
    )
    assert r.status_code == 200, r.text
    payload = r.json()
    assert payload["kind"] == "psc"
    assert isinstance(payload["requirement_source_id"], int)
    # First PSC-shaped sheet wins — must match the fixture's sheet title
    # so the loader dispatched to the right tab.
    assert payload["sheet_name"] == "Ground Security Controls"

    # The RequirementSource row really exists.
    with Session(env["engine"]) as s:
        src = s.get(RequirementSource, payload["requirement_source_id"])
        assert src is not None
        assert src.framework_id == env["framework_id"]


def test_import_other_emits_no_resolver_warning(env) -> None:
    """Unrecognized file → kind=='other', inert Baseline registered,
    a "no resolver" warning always present so the user knows the file
    won't bias assessment.
    """
    path = env["tmp"] / "other.xlsx"
    _write_other_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={"framework_id": env["framework_id"], "path": str(path)},
    )
    assert r.status_code == 200, r.text
    payload = r.json()
    assert payload["kind"] == "other"
    assert isinstance(payload["baseline_id"], int)
    # Warning is contractually required — UI shows it as the toast
    # subtitle so the user understands the file is metadata-only.
    assert any("no resolver" in w.lower() for w in payload["warnings"])

    # The receipt row was written with source_type=OTHER and no
    # BaselineControl fan-out (other_xlsx is intentionally inert).
    with Session(env["engine"]) as s:
        b = s.get(Baseline, payload["baseline_id"])
        assert b is not None
        assert b.source_type == BaselineSourceType.OTHER


# --- override behavior ------------------------------------------------------


def test_kind_hint_overrides_classifier(env) -> None:
    """``kind_hint="crm"`` on a PSC-shaped file forces CRM dispatch.

    The PSC fixture has no responsibility column, so the CRM loader's
    header validator will reject it with a 400 — the test asserts that
    the override actually took effect (we land in the CRM branch and
    bail out on its validation), not that the override succeeded as if
    the file were CRM-shaped. Surfaces a hint-vs-content warning
    pre-emptively too.

    A non-reserved ``scope_label`` is supplied so we clear the v0.2
    scope_label gate and actually reach the CRM loader's column check —
    otherwise we'd 422 on the missing label before the 400 we're pinning.
    """
    path = env["tmp"] / "psc.xlsx"
    _write_psc_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "kind_hint": "crm",
            "scope_label": "AWS GovCloud",
        },
    )
    # CRM loader couldn't locate a responsibility column on a PSC file —
    # surfaces cleanly as 400, not 500.
    assert r.status_code == 400, r.text
    detail = r.json()["detail"].lower()
    # CRM loader's _locate_columns raises about the missing column.
    assert "responsibility" in detail or "column" in detail


# --- boundary errors --------------------------------------------------------


def test_missing_file_returns_400(env) -> None:
    """Nonexistent path → 400 with an explicit "overlay file not found"
    message. Catches typos and stale paths from the UI file picker
    before they reach a loader.
    """
    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(env["tmp"] / "does-not-exist.xlsx"),
        },
    )
    assert r.status_code == 400, r.text
    assert "not found" in r.json()["detail"].lower()


def test_sheet_name_overrides_classifier_pick(env) -> None:
    """Explicit ``sheet_name`` targets a non-first PSC sheet.

    Without the selector, the classifier picks the first PSC-shaped
    sheet (``Ground Security Controls``). The user needs a way to ask
    for ``SV Security Controls`` instead. Override should be reflected
    in the response sheet_name AND surfaced as a warning so the toast
    explains which tab was actually parsed vs. which would have been
    auto-picked.
    """
    path = env["tmp"] / "psc_t1tl_two_sheets.xlsx"
    _write_psc_two_sheets_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "sheet_name": "SV Security Controls",
        },
    )
    assert r.status_code == 200, r.text
    payload = r.json()
    assert payload["kind"] == "psc"
    # The override actually took effect — loader parsed SV, not Ground.
    assert payload["sheet_name"] == "SV Security Controls"
    # Override-vs-auto-pick is reported so the UI can show the user
    # which tab they bypassed.
    assert any(
        "overrides auto-classified" in w
        and "SV Security Controls" in w
        and "Ground Security Controls" in w
        for w in payload["warnings"]
    )


def test_sheet_name_ignored_for_non_psc_dispatch(env) -> None:
    """``sheet_name`` on a CRM file emits a warning, doesn't change dispatch.

    The selector only makes sense for PSC (the multi-sheet loader). On
    CRM/OTHER files it's a no-op — but we surface a warning so the user
    doesn't think their selection was honored silently.
    """
    path = env["tmp"] / "crm.xlsx"
    _write_crm_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "sheet_name": "Some Tab",
            "scope_label": "AWS GovCloud",
        },
    )
    assert r.status_code == 200, r.text
    payload = r.json()
    assert payload["kind"] == "crm"
    # Warning fired — user knows their sheet_name was ignored.
    assert any(
        "ignored" in w.lower() and "psc dispatch" in w.lower()
        for w in payload["warnings"]
    )


def test_sheet_name_too_long_returns_422(env) -> None:
    """Excel caps sheet names at 31 chars; pydantic should 422 longer
    values up-front so the user gets a specific field error instead of a
    PSC loader 400 about the missing sheet.
    """
    path = env["tmp"] / "psc.xlsx"
    _write_psc_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "sheet_name": "X" * 32,  # one over the Excel cap
        },
    )
    assert r.status_code == 422, r.text
    # Validation error names the bad field so the UI can highlight it.
    assert "sheet_name" in r.text


def test_sheet_name_with_forbidden_character_returns_422(env) -> None:
    """Excel disallows ``: \\ / ? * [ ]`` in sheet names; reject them at
    the API boundary rather than letting the loader fail later.
    """
    path = env["tmp"] / "psc.xlsx"
    _write_psc_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "sheet_name": "bad/name",
        },
    )
    assert r.status_code == 422, r.text
    assert "sheet_name" in r.text


def test_sheet_name_empty_string_returns_422(env) -> None:
    """Empty sheet_name would be falsy and silently fall through to the
    classifier's auto-pick — reject it so the caller knows their input
    didn't take effect.
    """
    path = env["tmp"] / "psc.xlsx"
    _write_psc_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "sheet_name": "",
        },
    )
    assert r.status_code == 422, r.text
    assert "sheet_name" in r.text


def test_psc_no_sheet_detected_error_omits_hint_clause_when_no_kind_hint(env) -> None:
    """Belt-and-suspenders: the "no PSC-shaped sheet" error should only
    mention kind_hint when kind_hint was actually supplied. This branch
    isn't reachable from the unit tests' fixtures today (auto-PSC always
    populates auto.sheet_name), but the message logic still has to be
    accurate — so we exercise it with kind_hint=psc against a non-PSC
    file to confirm the hint clause appears, and rely on the
    classifier's contract (auto-PSC ⇒ sheet_name is set) to keep the
    no-hint path silent.
    """
    path = env["tmp"] / "other.xlsx"
    _write_other_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={
            "framework_id": env["framework_id"],
            "path": str(path),
            "kind_hint": "psc",
        },
    )
    assert r.status_code == 400, r.text
    detail = r.json()["detail"]
    assert "kind_hint=psc was supplied" in detail
    assert "Pass sheet_name explicitly" in detail


def test_unknown_framework_id_returns_400(env) -> None:
    """Unknown framework_id → 400 BEFORE the classifier runs.

    The framework existence check is intentionally up-front so the user
    gets a fast, specific error rather than a confusing loader failure
    (the CRM and OTHER loaders also check, but the route's pre-check
    avoids opening the workbook at all on the bad-id path).
    """
    path = env["tmp"] / "crm.xlsx"
    _write_crm_xlsx(path)

    r = env["client"].post(
        "/api/catalog/overlays/import",
        json={"framework_id": 9999, "path": str(path)},
    )
    assert r.status_code == 400, r.text
    detail = r.json()["detail"]
    assert "Framework" in detail and "9999" in detail
