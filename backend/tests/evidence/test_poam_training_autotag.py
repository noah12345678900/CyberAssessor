"""Lever C: bounded family-pure content-shape auto-tagging (2026-06-11).

Two workbook shapes that biject to a tiny, family-pure control set get a
Tier-4 0.6 corroboration tag straight from their column layout — no doc
number, no CCI, no control ID in any cell required:

  * A Plan of Action & Milestones workbook (weakness rows, scheduled
    completion dates, residual-risk, milestones) → CA-5. A POA&M IS the
    literal artifact CA-5 requires the org to maintain.
  * A training / security-awareness completion roster (course +
    completion-date columns) → AT-2 / AT-3 / AT-4. The roster IS the
    literal AT-4 training record and corroborates AT-2 / AT-3.

Both classifiers are CORE-anchored (>=1 defining header AND >=2 total
signals) and are evaluated AFTER every pre-existing inventory/account/asset
branch, so they can only ever catch a workbook that currently classifies as
``None``: pure recall, zero verdict flips (precision over recall).

Precision guards pinned here:
  * A generic project plan ("Milestone, Point of Contact, Status, Due Date"
    — support headers only, no "weakness"/"residual risk" core) stays
    untagged: no CA-5 leak.
  * A generic personnel roster ("Name, Date, Status" — no training/course
    core) stays untagged: no AT leak.
  * The two shapes do not cross-leak (POA&M ≠ AT, training ≠ CA-5) and
    neither leaks into the CM/AC inventory/account families.
  * No ``StigFinding`` ORM row is fabricated — neither artifact is a STIG.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.evidence.ingest import ingest_folder  # noqa: E402
from cybersecurity_assessor.models import (  # noqa: E402
    Control,
    EvidenceTag,
    Framework,
    Objective,
    StigFinding,
)
from cybersecurity_assessor.models import Workbook as WorkbookModel  # noqa: E402


@pytest.fixture
def session():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def wb_id(session) -> int:
    wb = WorkbookModel(path="/tmp/leverc.xlsx", filename="leverc.xlsx")
    session.add(wb)
    session.commit()
    session.refresh(wb)
    return wb.id


@pytest.fixture
def catalog(session) -> dict[str, list[int]]:
    """Seed CA-5, AT-2/3/4 plus CM-8 + AC-2 negative controls.

    Returns ``{control_id: [objective_id, ...]}`` so a test can assert that
    exactly the right family's child objectives were tagged (Tier-4 fans out
    to EVERY child objective of each mapped control). Each control carries 2
    children so a single-control match (CA-5) still clears the Tier-5 low-tag
    gate and the LLM backstop never fires in these deterministic tests.
    """
    fw = Framework(name="NIST SP 800-53", version="Rev 5")
    session.add(fw)
    session.commit()
    session.refresh(fw)

    by_control: dict[str, list[int]] = {}
    for ctl_id, family, objs in [
        ("ca-5", "CA", ["ca-5.1", "ca-5.2"]),
        ("at-2", "AT", ["at-2.1", "at-2.2"]),
        ("at-3", "AT", ["at-3.1", "at-3.2"]),
        ("at-4", "AT", ["at-4.1", "at-4.2"]),
        ("cm-8", "CM", ["cm-8.1", "cm-8.2"]),  # inventory negative control
        ("ac-2", "AC", ["ac-2.1", "ac-2.2"]),  # account negative control
    ]:
        ctrl = Control(
            framework_id=fw.id,
            control_id=ctl_id,
            title=f"{ctl_id.upper()} title",
            family=family,
        )
        session.add(ctrl)
        session.commit()
        session.refresh(ctrl)
        ids: list[int] = []
        for oid in objs:
            obj = Objective(
                control_id_fk=ctrl.id,
                objective_id=oid,
                source="AP",
                text=f"objective text for {oid}",
            )
            session.add(obj)
            session.commit()
            session.refresh(obj)
            ids.append(obj.id)
        by_control[ctl_id] = ids
    return by_control


# ---------------------------------------------------------------------------
# Workbook writers
# ---------------------------------------------------------------------------


def _write_poam(path: Path) -> None:
    """A POA&M tracker: 'Weakness' core + 4 support columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "POA&M"
    ws.append(
        [
            "Weakness",
            "Scheduled Completion Date",
            "Milestones",
            "Point of Contact",
            "Status",
        ]
    )
    ws.append(
        [
            "Unpatched TLS library",
            "2026-09-30",
            "Patch staged 2026-08; deploy 2026-09",
            "J. Smith",
            "Ongoing",
        ]
    )
    ws.append(
        [
            "Missing audit forwarding",
            "2026-10-15",
            "Configure syslog 2026-09",
            "A. Lee",
            "Ongoing",
        ]
    )
    wb.save(path)


def _write_training_roster(path: Path) -> None:
    """A training completion roster: 'Course Name' core + 3 support columns.

    Uses identity column "Name" (NOT an account signal) so the account_matrix
    branch never pre-empts the training classifier.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Training"
    ws.append(["Name", "Course Name", "Completion Date", "Score", "Training Status"])
    ws.append(["Jane Roe", "Annual Security Awareness", "2026-01-15", "95", "Complete"])
    ws.append(["John Doe", "Privileged User Training", "2026-02-01", "88", "Complete"])
    wb.save(path)


def _write_generic_project_plan(path: Path) -> None:
    """Support headers only ('Milestone', 'Point of Contact', 'Status') — no
    POA&M core ('weakness'/'residual risk'). Must stay None → no CA-5 tag."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    ws.append(["Task", "Milestone", "Point of Contact", "Status", "Due Date"])
    ws.append(["Kickoff", "Phase 1", "PM", "Done", "2026-03-01"])
    ws.append(["Build", "Phase 2", "Eng", "Active", "2026-06-01"])
    wb.save(path)


def _write_generic_roster(path: Path) -> None:
    """Plain personnel roster ('Name, Date, Status') — no training/course
    core. Must stay None → no AT tag."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"
    ws.append(["Name", "Date", "Status"])
    ws.append(["Jane Roe", "2026-01-15", "Active"])
    ws.append(["John Doe", "2026-02-01", "Active"])
    wb.save(path)


# ---------------------------------------------------------------------------
# Positive: POA&M → CA-5
# ---------------------------------------------------------------------------


def test_poam_workbook_auto_tags_ca5_at_06(session, catalog, wb_id, tmp_path):
    """A POA&M column shape tags every CA-5 child at Tier-4 0.6."""
    _write_poam(tmp_path / "Program_POAM.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)
    assert summary.ingested == 1
    assert summary.errors == []

    tags = session.exec(select(EvidenceTag)).all()
    tagged = {t.objective_id for t in tags}

    # Every CA-5 child objective must be tagged.
    for oid in catalog["ca-5"]:
        assert oid in tagged, f"CA-5 child {oid} not tagged from POA&M shape"

    ca5_tags = [t for t in tags if t.objective_id in catalog["ca-5"]]
    assert all(t.confidence == 0.6 for t in ca5_tags), "POA&M tags must be Tier-4 (0.6)"
    assert all("poam" in t.rationale.lower() for t in ca5_tags)
    # The rationale must surface a detected POA&M column so the assessor sees why.
    assert any("weakness" in t.rationale.lower() for t in ca5_tags)

    # No cross-leak into AT, CM, or AC families.
    for family in ("at-2", "at-3", "at-4", "cm-8", "ac-2"):
        for oid in catalog[family]:
            assert oid not in tagged, f"POA&M leaked into {family} ({oid})"


# ---------------------------------------------------------------------------
# Positive: training roster → AT-2 / AT-3 / AT-4
# ---------------------------------------------------------------------------


def test_training_roster_auto_tags_at_family_at_06(session, catalog, wb_id, tmp_path):
    """A training-roster column shape tags every AT-2/AT-3/AT-4 child at 0.6."""
    _write_training_roster(tmp_path / "Training_Completion.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)
    assert summary.ingested == 1
    assert summary.errors == []

    tags = session.exec(select(EvidenceTag)).all()
    tagged = {t.objective_id for t in tags}

    for family in ("at-2", "at-3", "at-4"):
        for oid in catalog[family]:
            assert oid in tagged, f"{family} child {oid} not tagged from training shape"

    at_obj_ids = {oid for f in ("at-2", "at-3", "at-4") for oid in catalog[f]}
    at_tags = [t for t in tags if t.objective_id in at_obj_ids]
    assert all(t.confidence == 0.6 for t in at_tags), "training tags must be Tier-4 (0.6)"
    assert all("training record" in t.rationale.lower() for t in at_tags)

    # No cross-leak into CA-5, CM, or AC families.
    for family in ("ca-5", "cm-8", "ac-2"):
        for oid in catalog[family]:
            assert oid not in tagged, f"training leaked into {family} ({oid})"


# ---------------------------------------------------------------------------
# Negative: generic plan / roster stay untagged
# ---------------------------------------------------------------------------


def test_generic_project_plan_does_not_tag_ca5(session, catalog, wb_id, tmp_path):
    """Support headers without a POA&M core header → no CA-5 tag."""
    _write_generic_project_plan(tmp_path / "Project_Plan.xlsx")

    ingest_folder(session, tmp_path, workbook_id=wb_id)

    tags = session.exec(select(EvidenceTag)).all()
    tagged = {t.objective_id for t in tags}
    for oid in catalog["ca-5"]:
        assert oid not in tagged, "generic project plan mis-classified as POA&M"
    # And it certainly produced no Tier-4 0.6 tag anywhere.
    assert not [t for t in tags if t.confidence == 0.6], (
        "generic project plan produced a Tier-4 content-shape tag"
    )


def test_generic_roster_does_not_tag_at(session, catalog, wb_id, tmp_path):
    """A 'Name, Date, Status' roster without a training core → no AT tag."""
    _write_generic_roster(tmp_path / "Personnel.xlsx")

    ingest_folder(session, tmp_path, workbook_id=wb_id)

    tags = session.exec(select(EvidenceTag)).all()
    tagged = {t.objective_id for t in tags}
    at_obj_ids = {oid for f in ("at-2", "at-3", "at-4") for oid in catalog[f]}
    assert not (tagged & at_obj_ids), "generic roster mis-classified as training record"
    assert not [t for t in tags if t.confidence == 0.6], (
        "generic roster produced a Tier-4 content-shape tag"
    )


# ---------------------------------------------------------------------------
# Precision: no StigFinding fabricated
# ---------------------------------------------------------------------------


def test_leverc_shapes_fabricate_no_stig_finding(session, catalog, wb_id, tmp_path):
    """Neither Lever-C shape is a STIG — the StigFinding table stays empty."""
    _write_poam(tmp_path / "Program_POAM.xlsx")
    _write_training_roster(tmp_path / "Training_Completion.xlsx")

    ingest_folder(session, tmp_path, workbook_id=wb_id)

    findings = session.exec(select(StigFinding)).all()
    assert findings == [], "a Lever-C content-shape xlsx fabricated StigFinding rows"
