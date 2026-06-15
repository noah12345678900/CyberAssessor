"""Ingest orchestrator quiet-skip semantics.

Covers the ``ExtractorSkip`` path: when an extractor intentionally
refuses a file (today only the xlsx extractor sniffing a CCIS workbook
by its WORKING SHEET tab), the orchestrator must drop it on the floor
quietly — no Evidence row, no entry in ``summary.errors``, just a bump
to ``skipped_unsupported`` so the totals strip in the UI still
accounts for it.

The bug this pins: before ``ExtractorSkip`` existed the same refusal
raised ``ExtractorError``, which the orchestrator caught and treated
as a recoverable failure — it created an Evidence row with empty
text AND added an error tile to the ingest summary, so users who
dropped a folder containing their own CCIS workbook saw it surface
as a red "failed" artifact.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.evidence.ingest import ingest_folder  # noqa: E402
from cybersecurity_assessor.models import Evidence  # noqa: E402
from cybersecurity_assessor.models import Workbook as WorkbookModel  # noqa: E402


@pytest.fixture
def session():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def wb_id(session) -> int:
    """A persisted Workbook id — ingest_folder requires it (PR 2 scoping)."""
    wb = WorkbookModel(path="/tmp/ingest_skip.xlsx", filename="ingest_skip.xlsx")
    session.add(wb)
    session.commit()
    session.refresh(wb)
    return wb.id


def _write_ccis_workbook(path: Path) -> None:
    """Build a minimal xlsx that the extractor will recognize as CCIS."""
    wb = Workbook()
    ws = wb.active
    ws.title = "WORKING SHEET"
    ws["A1"] = "marker"
    wb.save(path)


def _write_plain_workbook(path: Path) -> None:
    """A normal xlsx with no CCIS marker — should ingest normally."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hosts"
    ws["A1"] = "hostname"
    ws["A2"] = "server01"
    wb.save(path)


def test_ccis_workbook_is_quietly_skipped(session, wb_id, tmp_path):
    """Dropping a CCIS workbook into an evidence folder must not surface it."""
    _write_ccis_workbook(tmp_path / "CCIS_Example System_Demo_System_2026May.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)

    assert summary.scanned == 1
    assert summary.ingested == 0
    assert summary.skipped_unsupported == 1
    assert summary.errors == []

    # The Evidence table should be empty — no row, not even with empty text.
    rows = session.exec(select(Evidence)).all()
    assert rows == []


def test_plain_xlsx_alongside_ccis_still_ingests(session, wb_id, tmp_path):
    """The skip path must not poison the rest of the run."""
    _write_ccis_workbook(tmp_path / "ccis.xlsx")
    _write_plain_workbook(tmp_path / "asset_list.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)

    assert summary.scanned == 2
    assert summary.ingested == 1
    assert summary.skipped_unsupported == 1
    assert summary.errors == []

    rows = session.exec(select(Evidence)).all()
    assert len(rows) == 1
    assert rows[0].title == "asset_list"
