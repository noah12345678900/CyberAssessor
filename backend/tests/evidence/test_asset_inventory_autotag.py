"""Content-based asset-list auto-tagging (xlsx → CM-8).

Pins the user's explicit ask: a HW asset list xlsx must auto-tag to the
CM-8 family WITHOUT needing a filename hint or manual tagging step. The
xlsx extractor classifies the workbook by its first-sheet header columns
(``hostname``, ``serial number``, ``manufacturer``, ...) and stamps
``metadata["evidence_type"] = "hw_inventory"``. The tagger's Tier 4 path
then routes the evidence to every Control in
``EVIDENCE_TYPE_TO_CONTROLS[evidence_type]`` and tags each Control's
direct child Objectives at confidence 0.6.

The bug this pins (regression risk): if either piece silently breaks
— extractor stops emitting the metadata, OR the tagger stops reading
the kwarg — an asset list ingests with zero auto-tags and the user
has to tag it by hand. That's the workflow the user explicitly
rejected (2026-06-04).
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.evidence.ingest import ingest_folder  # noqa: E402
from cybersecurity_assessor.models import (  # noqa: E402
    Control,
    Evidence,
    EvidenceTag,
    Framework,
    Objective,
)
from cybersecurity_assessor.models import Workbook as WorkbookModel  # noqa: E402


@pytest.fixture
def session():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def wb_id(session) -> int:
    """A persisted Workbook so ``ingest_folder`` has a scope to attach to.

    ``ingest_folder`` requires ``workbook_id`` (PR 2 per-workbook hard
    scoping) — Evidence is physically scoped to one workbook, no global
    pool. The auto-tag assertions don't care which workbook; they just
    need a valid id for the ingest to run.
    """
    wb = WorkbookModel(path="/tmp/asset_autotag.xlsx", filename="asset_autotag.xlsx")
    session.add(wb)
    session.commit()
    session.refresh(wb)
    return wb.id


@pytest.fixture
def cm_catalog(session):
    """Seed enough of the CM family for tier-4 routing to land.

    Controls are stored in the OSCAL lowercase / dot-notation form
    (``cm-8``, ``cm-7.5``) because the tagger's ``_normalize_control_id``
    canonicalizes lookup keys to that shape. Each Control gets two
    child Objectives so we can prove "all children" instead of "one".
    An AC-2 control is also seeded as a NOT-tagged negative control:
    a HW inventory must not bleed into AC family tags.
    """
    fw = Framework(name="NIST SP 800-53", version="Rev 5")
    session.add(fw)
    session.commit()
    session.refresh(fw)

    seeded: dict[str, list[Objective]] = {}
    for ctl_id, family, ccis in [
        ("cm-8", "CM", ["CCI-000074", "CCI-000077"]),
        ("cm-7.5", "CM", ["CCI-001774", "CCI-001775"]),
        ("cm-10", "CM", ["CCI-001730"]),
        ("cm-11", "CM", ["CCI-001805"]),
        ("ac-2", "AC", ["CCI-000015", "CCI-000017"]),  # negative control
    ]:
        ctrl = Control(
            framework_id=fw.id,
            control_id=ctl_id,
            title=f"{ctl_id.upper()} title",
            family=family,
        )
        session.add(ctrl)
        session.commit()
        session.refresh(ctrl)

        objs: list[Objective] = []
        for cci in ccis:
            obj = Objective(
                control_id_fk=ctrl.id,
                objective_id=cci,
                source="CCI",
                text=f"objective text for {cci}",
            )
            session.add(obj)
            session.commit()
            session.refresh(obj)
            objs.append(obj)
        seeded[ctl_id] = objs

    return seeded


def _write_hw_inventory(path: Path) -> None:
    """A HW asset list: hostname + serial + manufacturer columns, no hints."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    ws.append(["Hostname", "Serial Number", "Manufacturer", "Model", "MAC Address"])
    ws.append(["server01", "ABC123", "Dell", "R740", "00:11:22:33:44:55"])
    ws.append(["server02", "DEF456", "Dell", "R740", "00:11:22:33:44:66"])
    wb.save(path)


def _write_sw_inventory(path: Path) -> None:
    """A SW inventory: hostname + software/version/publisher columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Software"
    ws.append(["Hostname", "Software", "Version", "Publisher"])
    ws.append(["server01", "OpenSSL", "3.0.7", "OpenSSL Foundation"])
    wb.save(path)


def _write_generic_asset_list(path: Path) -> None:
    """Only a hostname column — should still classify as asset_inventory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hosts"
    ws.append(["Hostname"])
    ws.append(["server01"])
    ws.append(["server02"])
    wb.save(path)


def _write_random_xlsx(path: Path) -> None:
    """Nothing inventory-shaped — must NOT be auto-tagged to CM."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    ws.append(["Topic", "Notes"])
    ws.append(["Coffee", "Decaf preferred"])
    wb.save(path)


def test_hw_inventory_xlsx_auto_tags_cm8(session, cm_catalog, wb_id, tmp_path):
    """Generic filename + HW columns → tagged to all CM-8 child objectives."""
    # Filename intentionally vague — no "asset", "hw", "inventory", control ID.
    _write_hw_inventory(tmp_path / "ServerList.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)
    assert summary.ingested == 1
    assert summary.errors == []

    # Every CM-8 child objective should be tagged.
    cm8_obj_ids = {o.id for o in cm_catalog["cm-8"]}
    tags = session.exec(select(EvidenceTag)).all()
    tagged_obj_ids = {t.objective_id for t in tags}
    assert cm8_obj_ids.issubset(tagged_obj_ids), (
        f"expected CM-8 children {cm8_obj_ids} in tags, got {tagged_obj_ids}"
    )

    # Confidence must be 0.6 (tier 4, not tier 3) and rationale must surface
    # the detected columns so the assessor can audit the auto-classification.
    cm8_tags = [t for t in tags if t.objective_id in cm8_obj_ids]
    assert all(t.confidence == 0.6 for t in cm8_tags)
    assert all("serial number" in t.rationale.lower() for t in cm8_tags)

    # AC-2 objectives must NOT be tagged — HW inventory doesn't bleed into AC.
    ac2_obj_ids = {o.id for o in cm_catalog["ac-2"]}
    assert tagged_obj_ids.isdisjoint(ac2_obj_ids), (
        "HW inventory leaked into AC-2 — Tier 4 mapping is unbounded"
    )


def test_sw_inventory_xlsx_auto_tags_cm8_and_cm10(session, cm_catalog, wb_id, tmp_path):
    """SW columns trigger the multi-control SW mapping (CM-8, CM-7.5, CM-10, CM-11)."""
    _write_sw_inventory(tmp_path / "InstalledApps.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)
    assert summary.ingested == 1

    tags = session.exec(select(EvidenceTag)).all()
    tagged_obj_ids = {t.objective_id for t in tags}

    # SW mapping = ["cm-8", "cm-7.5", "cm-10", "cm-11"].
    for ctl_id in ("cm-8", "cm-7.5", "cm-10", "cm-11"):
        for obj in cm_catalog[ctl_id]:
            assert obj.id in tagged_obj_ids, (
                f"SW inventory missed {ctl_id} objective {obj.objective_id}"
            )

    # AC-2 still off-limits.
    ac2_obj_ids = {o.id for o in cm_catalog["ac-2"]}
    assert tagged_obj_ids.isdisjoint(ac2_obj_ids)


def test_generic_hostname_only_xlsx_classifies_as_asset_inventory(
    session, cm_catalog, wb_id, tmp_path
):
    """A bare hostname column is still enough to route to CM-8."""
    _write_generic_asset_list(tmp_path / "hosts.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)
    assert summary.ingested == 1

    cm8_obj_ids = {o.id for o in cm_catalog["cm-8"]}
    tags = session.exec(select(EvidenceTag)).all()
    tagged_obj_ids = {t.objective_id for t in tags}
    assert cm8_obj_ids.issubset(tagged_obj_ids)


def test_random_xlsx_is_not_auto_tagged_to_cm(session, cm_catalog, wb_id, tmp_path):
    """Workbooks without inventory shape must not be force-mapped to CM-8."""
    _write_random_xlsx(tmp_path / "meeting_notes.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)
    assert summary.ingested == 1

    # No CM tags should exist — auto-classification must not fire on a random
    # two-column notes sheet. (Doc-number / CCI / control-ID tiers will still
    # fire if those tokens appear, but nothing in the fixture triggers them.)
    tags = session.exec(select(EvidenceTag)).all()
    all_cm_obj_ids = {
        o.id
        for ctl_id in ("cm-8", "cm-7.5", "cm-10", "cm-11")
        for o in cm_catalog[ctl_id]
    }
    tagged_obj_ids = {t.objective_id for t in tags}
    assert tagged_obj_ids.isdisjoint(all_cm_obj_ids), (
        "random xlsx was auto-tagged to CM — classifier is too eager"
    )


def test_evidence_metadata_round_trips_evidence_type(
    session, cm_catalog, wb_id, tmp_path
):
    """The extractor's evidence_type metadata is what the tagger consumes.

    Regression guard: if the metadata key is renamed in one place but not
    the other, the tagger silently no-ops on every asset list. We can't
    inspect ExtractedDoc directly through ingest_folder, but a successful
    tag attachment on the HW fixture proves the key flowed end-to-end.
    """
    _write_hw_inventory(tmp_path / "anything.xlsx")
    ingest_folder(session, tmp_path, workbook_id=wb_id)

    ev = session.exec(select(Evidence)).first()
    assert ev is not None
    tags = session.exec(
        select(EvidenceTag).where(EvidenceTag.evidence_id == ev.id)
    ).all()
    # At least one Tier-4 tag (confidence 0.6, source=auto) must exist.
    tier4 = [t for t in tags if t.confidence == 0.6 and t.source == "auto"]
    assert tier4, "no Tier-4 tags created — evidence_type didn't flow through ingest"
