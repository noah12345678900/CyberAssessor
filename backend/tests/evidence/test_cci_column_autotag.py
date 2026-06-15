"""Lever B: a generic evidence xlsx with a dedicated CCI column auto-tags
its named CCIs at the ungated 0.95 Tier-2 confidence (2026-06-11).

A compliance matrix / RTM / SCAP-results export that is NOT a STIG
checklist often still names the CCI it satisfies in a column ("CCI",
"CCI ID", "CCI Ref", ...). Before Lever B that column was invisible to
the tagger: only the STIG/Nessus extractors populated structured
``cci_refs``, so such a workbook fell through to the LLM Tier-5 judge (or
zero-tagged). The xlsx extractor now sniffs a header-allow-listed CCI
column, value-validates every cell with the canonical ``CCI-\\d{6}``
regex, and emits ``metadata["cci_refs"]``; ingest threads it into
``tag_evidence(cci_refs=...)`` which merges it UNGATED into the same 0.95
Tier-2 ``cci_set`` the STIG path feeds.

Precision guards pinned here (the whole point of Lever B — recall up,
verdicts unchanged):
  * A CCI token in a NON-CCI column (free-text body cell) does NOT earn a
    0.95 tag — collection is scoped to the detected column only.
  * A workbook with no CCI column emits no ``cci_refs`` — no Tier-2 tag,
    no behavior change vs. pre-Lever-B.
  * A multi-token cell ("CCI-000015, CCI-000018") tags both.
  * No ``StigFinding`` ORM row is fabricated — the artifact is not a STIG.
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.evidence.ingest import ingest_folder  # noqa: E402
from cybersecurity_assessor.models import (  # noqa: E402
    Control,
    Evidence,
    EvidenceTag,
    Framework,
    Objective,
    StigFinding,
)
from cybersecurity_assessor.models import Workbook as WorkbookModel  # noqa: E402


@pytest.fixture
def session():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def wb_id(session) -> int:
    wb = WorkbookModel(path="/tmp/cci_col.xlsx", filename="cci_col.xlsx")
    session.add(wb)
    session.commit()
    session.refresh(wb)
    return wb.id


@pytest.fixture
def cci_catalog(session) -> dict[str, Objective]:
    """Seed objectives keyed by CCI id so Tier-2 ``_objectives_by_cci`` lands.

    Objectives carry ``objective_id == "CCI-######"`` (the production shape
    for DISA CCIs). A handful across two families so we can prove the column
    drives attribution to exactly the named CCIs and nothing else.
    """
    fw = Framework(name="NIST SP 800-53", version="Rev 5")
    session.add(fw)
    session.commit()
    session.refresh(fw)

    by_cci: dict[str, Objective] = {}
    for ctl_id, family, ccis in [
        ("ac-2", "AC", ["CCI-000015", "CCI-000018"]),
        ("au-12", "AU", ["CCI-000169"]),
        ("ia-2", "IA", ["CCI-000764"]),  # never named in fixtures — stays untagged
    ]:
        ctrl = Control(
            framework_id=fw.id,
            control_id=ctl_id,
            title=f"{ctl_id.upper()} title",
            family=family,
        )
        session.add(ctrl)
        session.commit()
        session.refresh(ctrl)
        for cci in ccis:
            obj = Objective(
                control_id_fk=ctrl.id,
                objective_id=cci,
                source="CCI",
                text=f"objective text for {cci}",
            )
            session.add(obj)
            session.commit()
            session.refresh(obj)
            by_cci[cci] = obj
    return by_cci


def _write_cci_matrix(path: Path) -> None:
    """A compliance matrix: dedicated 'CCI' column + a multi-token cell."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Matrix"
    ws.append(["Requirement", "CCI", "Status", "Notes"])
    ws.append(["Account mgmt", "CCI-000015", "Met", "see policy"])
    ws.append(["Account mgmt II", "CCI-000015, CCI-000018", "Met", "combined"])
    ws.append(["Audit gen", "CCI-000169", "Met", "syslog"])
    wb.save(path)


def _write_cci_in_freetext_only(path: Path) -> None:
    """A CCI token appears, but ONLY in a free-text 'Notes' cell — no CCI col."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["Date", "Notes"])
    ws.append(["2026-06-11", "Reviewed CCI-000015 during the audit walkthrough."])
    wb.save(path)


def _write_no_cci(path: Path) -> None:
    """No CCI column, no CCI tokens anywhere."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Misc"
    ws.append(["Topic", "Owner"])
    ws.append(["Backups", "ops"])
    wb.save(path)


def test_cci_column_auto_tags_named_ccis_at_095(
    session, cci_catalog, wb_id, tmp_path
):
    """Dedicated CCI column → every named CCI tagged at Tier-2 0.95."""
    _write_cci_matrix(tmp_path / "ComplianceMatrix.xlsx")

    summary = ingest_folder(session, tmp_path, workbook_id=wb_id)
    assert summary.ingested == 1
    assert summary.errors == []

    tags = session.exec(select(EvidenceTag)).all()
    tagged_obj_ids = {t.objective_id for t in tags}

    # CCI-000015, CCI-000018 (multi-token cell), CCI-000169 must all be tagged.
    for cci in ("CCI-000015", "CCI-000018", "CCI-000169"):
        assert cci_catalog[cci].id in tagged_obj_ids, f"{cci} not tagged from CCI column"

    # Those tags must be Tier-2: confidence 0.95, "Direct CCI reference" rationale.
    cci_obj_ids = {cci_catalog[c].id for c in ("CCI-000015", "CCI-000018", "CCI-000169")}
    cci_tags = [t for t in tags if t.objective_id in cci_obj_ids]
    assert all(t.confidence == 0.95 for t in cci_tags), (
        "CCI-column tags must be Tier-2 (0.95), not a weaker tier"
    )
    assert all("direct cci reference" in t.rationale.lower() for t in cci_tags)

    # IA-2 (CCI-000764) was never named — must stay untagged.
    assert cci_catalog["CCI-000764"].id not in tagged_obj_ids, (
        "an unnamed CCI was tagged — column scan over-collected"
    )


def test_cci_in_freetext_only_does_not_tier2_tag(
    session, cci_catalog, wb_id, tmp_path
):
    """A CCI token in a non-CCI 'Notes' cell must NOT earn a 0.95 tag.

    This is the precision guard: an xlsx is NOT in
    ``_STRUCTURED_FINDING_KINDS``, so the inline text-scrape Tier-2 branch
    is gated off for it, and the Lever-B column path only collects from the
    detected CCI column. A casual mention in free text therefore yields no
    Tier-2 attribution — exactly the pre-Lever-B behavior, preserved.
    """
    _write_cci_in_freetext_only(tmp_path / "AuditLog.xlsx")

    ingest_folder(session, tmp_path, workbook_id=wb_id)

    tags = session.exec(select(EvidenceTag)).all()
    tier2 = [t for t in tags if t.confidence == 0.95]
    assert not tier2, (
        "a free-text CCI mention produced a 0.95 Tier-2 tag — Lever B leaked "
        "beyond the dedicated CCI column"
    )


def test_no_cci_column_no_tier2_tag(session, cci_catalog, wb_id, tmp_path):
    """A workbook with no CCI column emits no cci_refs → no Tier-2 tag."""
    _write_no_cci(tmp_path / "Plain.xlsx")

    ingest_folder(session, tmp_path, workbook_id=wb_id)

    tags = session.exec(select(EvidenceTag)).all()
    tier2 = [t for t in tags if t.confidence == 0.95]
    assert not tier2, "no-CCI-column workbook produced a Tier-2 tag"


def test_cci_column_fabricates_no_stig_finding(
    session, cci_catalog, wb_id, tmp_path
):
    """Lever B must not pollute the StigFinding table for a non-STIG artifact."""
    _write_cci_matrix(tmp_path / "ComplianceMatrix.xlsx")

    ingest_folder(session, tmp_path, workbook_id=wb_id)

    findings = session.exec(select(StigFinding)).all()
    assert findings == [], (
        "a generic CCI-column xlsx created StigFinding rows — Lever B should "
        "use the dedicated cci_refs path, not fabricate findings"
    )
