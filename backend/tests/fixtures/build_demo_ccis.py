"""Build the demo CCIS workbook used by the test suite.

Regenerates ``demo_ccis.xlsx`` from scratch using openpyxl. Both the
script AND the produced .xlsx are committed — the script is the source
of truth, the .xlsx is the artifact tests load.

Layout mirrors the real eMASS CCIS WORKING SHEET schema as parsed by
``backend/cybersecurity_assessor/excel/ccis_reader.py``:
    - Sheet name: "WORKING SHEET"
    - Rows 1-5: system metadata (left mostly blank for the fixture)
    - Row 6: column headers
    - Row 7+: CCI data rows
    - Column B: Control Acronym, Column F: Implementation Narrative,
      Column H: CCI, Column N: Compliance Status,
      Column U: PREVIOUS Test Results (where supersession-detection scans
      for legacy doc references)

The fixture is intentionally tiny (7 rows) and covers:
    - A mix of control families (AC, AU, SI)
    - A SUPERSEDED USD reference in column U (USD0050010 / "SDA T1 O&I
      Account Management User Guide" — superseded per nist-assessor's
      ``document-supersession.md`` to USD00050010 Example System Account Management
      Plan)
    - A row with valid-looking current evidence references
    - A row with empty column F (no narrative — the "gap" case)
    - One Common-inherited row and one System-Specific row
    - One row with no CCI in column H (should be skipped by the upsert
      flow but still parsed)

Run from anywhere:
    python backend/tests/fixtures/build_demo_ccis.py
or:
    cd backend && uv run python tests/fixtures/build_demo_ccis.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SHEET_NAME = "WORKING SHEET"
OUTPUT = Path(__file__).resolve().parent / "demo_ccis.xlsx"

# Header row -- order MUST match the ccis_reader column constants.
HEADERS = [
    "Required",                       # A
    "Control Acronym",                # B
    "Control Information",            # C
    "Implementation Status",          # D
    "Security Control Designation",   # E
    "Implementation Narrative",       # F
    "AP Acronym",                     # G
    "CCI",                            # H
    "CCI Definition",                 # I
    "Implementation Guidance",        # J
    "Assessment Procedures",          # K
    "Inherited",                      # L
    "Remote Inheritance Instance",    # M
    "Compliance Status",              # N
    "Date Tested",                    # O
    "Tested By",                      # P
    "Test Results",                   # Q
    "PREVIOUS Compliance Status",     # R
    "PREVIOUS Date Tested",           # S
    "PREVIOUS Tested By",             # T
    "PREVIOUS Test Results",          # U
]

# 7 fixture rows. Tuples line up with HEADERS column-for-column.
# Keep narrative / evidence text short but realistic.
ROWS: list[tuple] = [
    # 1. AC-2 / CCI-000015 -- has a SUPERSEDED USD doc ref in col U.
    #    The legacy "SDA T1 O&I Account Management User Guide" is
    #    superseded by USD00050010 Example System Account Management Plan per
    #    nist-assessor's document-supersession.md table.
    (
        "YES",                                                    # A
        "AC-2",                                                   # B
        "The organization manages information system accounts.",  # C
        "Implemented",                                            # D
        "Hybrid",                                                 # E
        "Account management is implemented per current Account "  # F
        "Management Plan; sampled 5 accounts in last 90 days.",
        "AC-2.1",                                                 # G
        "000015",                                                 # H (bare, no prefix)
        "The organization assigns account managers.",             # I
        "Document designated account managers in the AMP.",       # J
        "Examine the Account Management Plan and interview the "  # K
        "account manager.",
        "Local",                                                  # L
        None,                                                     # M
        None,                                                     # N (unassessed)
        None,                                                     # O
        None,                                                     # P
        None,                                                     # Q
        "Compliant",                                              # R
        "2024-08-15",                                             # S
        "Prior Assessor",                                         # T
        # Col U cites the LEGACY doc -- supersession logic should flag
        # this and steer the assessor to USD00050010 instead.
        "Reviewed SDA T1 O&I Account Management User Guide "
        "Section 5.9 -- account creation procedure observed.",
    ),
    # 2. AC-3 / CCI-000213 -- valid-looking current evidence in col U.
    (
        "YES",
        "AC-3",
        "The information system enforces approved authorizations.",
        "Implemented",
        "System-Specific",
        "Access enforcement implemented via Active Directory "
        "group policy; verified GPO settings on domain controllers.",
        "AC-3.1",
        "000213",
        "The system enforces approved authorizations for logical "
        "access to information and system resources.",
        "Configure RBAC per the SSP Section 5.3.",
        "Examine GPO export and AD group memberships.",
        "Local",
        None,
        None,
        None,
        None,
        None,
        "Compliant",
        "2024-08-15",
        "Prior Assessor",
        "Reviewed USD00050015 Rev D SSP Section 5.3 and verified "
        "AD GPO export against documented role matrix.",
    ),
    # 3. AC-2(1) / CCI-001361 -- EMPTY narrative (col F) -- gap case.
    (
        "YES",
        "AC-2(1)",
        "The organization employs automated mechanisms to "
        "support the management of information system accounts.",
        "Planned",
        "System-Specific",
        None,  # F -- intentionally empty (no narrative => gap)
        "AC-2(1).1",
        "001361",
        "The organization employs automated mechanisms to "
        "support the management of accounts.",
        "Deploy an IDM tool integrated with AD.",
        "Examine IDM configuration and account provisioning logs.",
        "Local",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ),
    # 4. AU-2 / CCI-000130 -- inherited from Enterprise.
    (
        "YES",
        "AU-2",
        "The organization determines that the information system "
        "is capable of auditing defined events.",
        "Implemented",
        "Common",
        "Audit events inherited from DoW Enterprise logging "
        "service; supplemental events configured locally per "
        "the Audit Plan.",
        "AU-2.1",
        "000130",
        "The organization identifies the types of events the "
        "system is capable of logging.",
        "Adopt the enterprise audit-event catalog.",
        "Examine the enterprise audit-event catalog and the "
        "local audit configuration.",
        "DoW Enterprise",
        "Enterprise Logging Service",
        None,
        None,
        None,
        None,
        "Compliant",
        "2024-08-20",
        "Prior Assessor",
        "Inherited from DoW Enterprise Logging Service per "
        "Enterprise Services Controls (AU-2 Req #001).",
    ),
    # 5. AU-3 / CCI-000133 -- already-assessed current entry (col N filled).
    (
        "YES",
        "AU-3",
        "The information system generates audit records with the "
        "required content.",
        "Implemented",
        "System-Specific",
        "Audit records include timestamp, source, event type, "
        "outcome, and user identity per Syslog config.",
        "AU-3.1",
        "000133",
        "The system generates audit records containing required "
        "content fields.",
        "Configure syslog format per Audit Plan Section 4.",
        "Examine a sample of audit records.",
        "Local",
        None,
        "Compliant",
        "2026-06-03",
        "Noah Jaskolski",
        "Examined 50 audit records from 2026-05; all contained "
        "required fields per Audit Plan Section 4.",
        "Compliant",
        "2024-08-20",
        "Prior Assessor",
        "Sampled audit records -- format matched the plan.",
    ),
    # 6. SI-2 / CCI-002605 -- non-compliant current finding.
    (
        "YES",
        "SI-2",
        "The organization identifies, reports, and corrects "
        "information system flaws.",
        "Implemented",
        "Hybrid",
        "Patch management performed monthly per Vulnerability "
        "Management Plan; some critical patches > 30 days old.",
        "SI-2.1",
        "002605",
        "The organization installs security-relevant software "
        "updates within the organization-defined time period.",
        "Apply critical patches within 30 days.",
        "Examine the patch management report.",
        "Local",
        None,
        "Non-Compliant",
        "2026-06-03",
        "Noah Jaskolski",
        "Examined patch report 2026-05-30: 3 critical CVEs > 30 "
        "days outstanding on file servers.",
        "Compliant",
        "2024-08-22",
        "Prior Assessor",
        "Reviewed Vulnerability Management Plan and prior month's "
        "patch report -- no outstanding criticals at that time.",
    ),
    # 7. SI-4 / no CCI in col H -- exercises the missing_cci skip path.
    (
        "YES",
        "SI-4",
        "The organization monitors the information system to "
        "detect attacks and indicators of potential attacks.",
        "Implemented",
        "Common",
        "IDS coverage provided by enterprise SOC.",
        "SI-4.1",
        None,  # H -- intentionally empty
        "The organization monitors the information system to "
        "detect attacks.",
        "Subscribe to enterprise SOC monitoring.",
        "Examine the SOC MOA.",
        "DoW Enterprise",
        "Enterprise SOC",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ),
]


def build() -> Path:
    """(Re)generate demo_ccis.xlsx and return its path."""
    wb = Workbook()
    # The default sheet starts as "Sheet" -- rename it.
    sheet = wb.active
    sheet.title = SHEET_NAME

    # Row 1: a single metadata line so it looks like the real export.
    sheet.cell(row=1, column=1, value="SYSTEM: Demo System (fixture)")
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(
        row=2,
        column=1,
        value="Generated by backend/tests/fixtures/build_demo_ccis.py",
    )
    # Rows 3-5 stay blank to mirror real exports' metadata padding.

    # Row 6: headers (the reader ignores the labels and uses absolute
    # column positions, but a real human opening the file expects them).
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, label in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=6, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Row 7+: data
    for row_offset, row_values in enumerate(ROWS):
        excel_row = 7 + row_offset
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row=excel_row, column=col_idx, value=value)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path} ({len(ROWS)} data rows)")
