"""Tests for ``engine.workbook_sync``.

Builds tiny CCIS-shaped workbooks on disk (matching the schema the
re-read tests use), persists a ``Workbook`` row, and walks the diff
through the four event types the spec calls out: added, removed,
moved, edited.

In-memory SQLite with StaticPool is used so a single shared connection
backs every session call within a test.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

# Ensure the backend package is importable when pytest is launched from any cwd.
_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402  -- registers tables
from cybersecurity_assessor.engine.workbook_sync import (  # noqa: E402
    SyncSummary,
    sync_workbook,
)
from cybersecurity_assessor.excel.ccis_reader import (  # noqa: E402
    COL_CCI,
    COL_CONTROL,
    COL_NARRATIVE,
    COL_REQUIRED,
)
from cybersecurity_assessor.models import (  # noqa: E402
    Workbook as WorkbookRow,
)
from cybersecurity_assessor.models import (  # noqa: E402
    WorkbookSyncEvent,
)

SHEET = "WORKING SHEET"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def session():
    """In-memory SQLite, single shared connection per test."""
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def workbook_path(tmp_path):
    return tmp_path / "test_ccis_sync.xlsx"


@pytest.fixture
def workbook_row(session, workbook_path):
    """Persist a minimal Workbook row so events can FK to it."""
    wb = WorkbookRow(path=str(workbook_path), filename=workbook_path.name)
    session.add(wb)
    session.commit()
    session.refresh(wb)
    return wb


# ---------------------------------------------------------------------------
# Workbook fixture helper — mirrors test_reader_reread._make_workbook
# ---------------------------------------------------------------------------


_FIELD_TO_COL = {
    "required": COL_REQUIRED,
    "control_id": COL_CONTROL,
    "cci_id": COL_CCI,
    "narrative": COL_NARRATIVE,
}


def _make_workbook(path: Path, rows: list[dict]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(row=6, column=COL_CONTROL, value="Control Acronym")
    ws.cell(row=6, column=COL_CCI, value="CCI")

    for offset, row in enumerate(rows):
        excel_row = 7 + offset
        for key, val in row.items():
            col = _FIELD_TO_COL.get(key)
            if col is None:
                raise KeyError(f"Unknown fixture key: {key}")
            ws.cell(row=excel_row, column=col, value=val)

    wb.save(path)
    wb.close()
    return path


def _events_for(session, workbook_id: int) -> list[WorkbookSyncEvent]:
    return list(
        session.exec(
            select(WorkbookSyncEvent).where(
                WorkbookSyncEvent.workbook_id == workbook_id
            )
        )
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_first_sync_no_prior_snapshot_emits_added_for_each_keyable_row(
    session, workbook_row, workbook_path
):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )

    summary = sync_workbook(session, workbook_row.id, workbook_path)

    assert isinstance(summary, SyncSummary)
    assert summary.had_prior_snapshot is False
    assert summary.added_count == 2
    assert summary.removed_count == 0
    assert summary.moved_count == 0
    assert summary.edited_count == 0

    persisted = _events_for(session, workbook_row.id)
    assert len(persisted) == 2
    types = {e.event_type for e in persisted}
    assert types == {"added"}
    for ev in persisted:
        assert ev.old_value_json is None
        assert ev.new_value_json is not None
        assert ev.source == "reread"


def test_unchanged_second_sync_emits_zero_events(
    session, workbook_row, workbook_path
):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )

    sync_workbook(session, workbook_row.id, workbook_path)  # establishes snapshot
    summary = sync_workbook(session, workbook_row.id, workbook_path)

    assert summary.had_prior_snapshot is True
    assert summary.added_count == 0
    assert summary.removed_count == 0
    assert summary.moved_count == 0
    assert summary.edited_count == 0
    # Only the first sync's events should be persisted (2 rows added).
    persisted = _events_for(session, workbook_row.id)
    assert len(persisted) == 2


def test_added_row_emits_one_added_event_with_null_old(
    session, workbook_row, workbook_path
):
    _make_workbook(
        workbook_path,
        [{"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"}],
    )
    sync_workbook(session, workbook_row.id, workbook_path)

    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "new"},
        ],
    )
    summary = sync_workbook(session, workbook_row.id, workbook_path)

    assert summary.added_count == 1
    assert summary.removed_count == 0
    new_events = [e for e in summary.events if e.event_type == "added"]
    assert len(new_events) == 1
    assert new_events[0].old_value_json is None
    assert new_events[0].control_id == "AC-3"
    assert new_events[0].cci_id == "CCI-000213"


def test_removed_row_emits_one_removed_event_with_null_new(
    session, workbook_row, workbook_path
):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )
    sync_workbook(session, workbook_row.id, workbook_path)

    _make_workbook(
        workbook_path,
        [{"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"}],
    )
    summary = sync_workbook(session, workbook_row.id, workbook_path)

    assert summary.removed_count == 1
    assert summary.added_count == 0
    removed_events = [e for e in summary.events if e.event_type == "removed"]
    assert len(removed_events) == 1
    assert removed_events[0].new_value_json is None
    assert removed_events[0].old_value_json is not None
    assert removed_events[0].control_id == "AC-3"
    assert removed_events[0].cci_id == "CCI-000213"


def test_edited_narrative_emits_one_edited_event_with_new_narrative(
    session, workbook_row, workbook_path
):
    _make_workbook(
        workbook_path,
        [
            {
                "control_id": "AC-2",
                "cci_id": "000015",
                "narrative": "original narrative",
            }
        ],
    )
    sync_workbook(session, workbook_row.id, workbook_path)

    _make_workbook(
        workbook_path,
        [
            {
                "control_id": "AC-2",
                "cci_id": "000015",
                "narrative": "EDITED narrative",
            }
        ],
    )
    summary = sync_workbook(session, workbook_row.id, workbook_path)

    assert summary.edited_count == 1
    edited_events = [e for e in summary.events if e.event_type == "edited"]
    assert len(edited_events) == 1

    ev = edited_events[0]
    assert ev.control_id == "AC-2"
    assert ev.cci_id == "CCI-000015"
    # Full snapshot dict round-trips through JSON for both old and new.
    new_payload = json.loads(ev.new_value_json)
    old_payload = json.loads(ev.old_value_json)
    assert new_payload["narrative"] == "EDITED narrative"
    assert old_payload["narrative"] == "original narrative"


def test_moved_row_emits_one_moved_event(session, workbook_row, workbook_path):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )
    sync_workbook(session, workbook_row.id, workbook_path)

    # Swap the order — both rows technically "moved".
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
        ],
    )
    summary = sync_workbook(session, workbook_row.id, workbook_path)

    assert summary.moved_count >= 1
    assert summary.edited_count == 0
    moved_events = [e for e in summary.events if e.event_type == "moved"]
    assert len(moved_events) == summary.moved_count
    # Each move event has both old and new payloads (full snapshot dicts).
    for ev in moved_events:
        assert ev.old_value_json is not None
        assert ev.new_value_json is not None
        old = json.loads(ev.old_value_json)
        new = json.loads(ev.new_value_json)
        assert old.get("excel_row") != new.get("excel_row")


def test_events_in_same_call_share_occurred_at(
    session, workbook_row, workbook_path
):
    """All events from one sync_workbook call land in the same ~1-second window."""
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
            {"control_id": "AC-4", "cci_id": "000217", "narrative": "n3"},
        ],
    )
    summary = sync_workbook(session, workbook_row.id, workbook_path)

    assert len(summary.events) == 3
    # SQLite via SQLModel can strip tzinfo on round-trip; normalize before diffing.
    timestamps: list[datetime] = []
    for ev in summary.events:
        ts = ev.occurred_at
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        timestamps.append(ts)
    span = max(timestamps) - min(timestamps)
    assert span <= timedelta(seconds=1)


# ---------------------------------------------------------------------------
# Internal helper + commit-failure paths (the branches the e2e flow never hits)
# ---------------------------------------------------------------------------


def test_current_snapshot_for_returns_none_when_key_not_in_index(
    workbook_path,
):
    """``_current_snapshot_for(... key not present ...)`` → None (workbook_sync.py:83).

    Pins the defensive "row not found" return in the helper that resolves
    moved/edited diff entries back to their current snapshot dict. The
    public sync flow can't normally reach this branch — diff entries are
    constructed FROM the same index the helper scans — but the helper
    is also called speculatively by future tooling (e.g. a UI that
    re-reads the index and asks "show me the current state of <key>")
    that may carry a stale key. Drop the None branch and any such caller
    would see an unbound-local KeyError instead of a clean miss; pin the
    contract: missing key → None, no raise.
    """
    from cybersecurity_assessor.engine.workbook_sync import (
        _current_snapshot_for,
    )
    from cybersecurity_assessor.excel.ccis_reader import reread_workbook

    _make_workbook(
        workbook_path,
        [{"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"}],
    )
    result = reread_workbook(workbook_path, update_snapshot=False)

    # Key that is NOT in the index — helper must return None, not raise.
    missing = _current_snapshot_for(result, "ZZ-99", "CCI-999999")
    assert missing is None
    # Sanity: a key that IS present returns a dict (proves the loop runs).
    present = _current_snapshot_for(result, "AC-2", "CCI-000015")
    assert present is not None
    assert present["control_id"] == "AC-2"


def test_sync_workbook_rolls_back_and_reraises_on_commit_failure(
    session, workbook_row, workbook_path, monkeypatch
):
    """``session.commit()`` raises → ``session.rollback()`` runs, exception propagates.

    Pins workbook_sync.py:184-186 — the try/except wrapper around event
    construction + commit. The audit log MUST stay consistent with the
    snapshot sidecar; a partial commit (some events written, others
    not) would leave the UI showing phantom changes on the next sync.
    The rollback is the guarantee.

    Patch ``session.commit`` to raise on the FIRST call inside
    ``sync_workbook`` and assert (a) the exception bubbles up unmodified,
    (b) ``session.rollback`` was called, (c) zero events end up persisted
    in a fresh query session — proving the rollback actually undid the
    ``session.add`` calls.
    """
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )

    rollback_called: list[bool] = []
    original_rollback = session.rollback

    def _tracking_rollback():
        rollback_called.append(True)
        original_rollback()

    def _exploding_commit():
        raise RuntimeError("simulated db disconnect mid-sync")

    monkeypatch.setattr(session, "commit", _exploding_commit)
    monkeypatch.setattr(session, "rollback", _tracking_rollback)

    with pytest.raises(RuntimeError, match="simulated db disconnect"):
        sync_workbook(session, workbook_row.id, workbook_path)

    # rollback fired on the way out — the except branch did its job.
    assert rollback_called == [True]
    # And nothing for this workbook landed in the audit log.
    monkeypatch.undo()
    persisted = _events_for(session, workbook_row.id)
    assert persisted == []
