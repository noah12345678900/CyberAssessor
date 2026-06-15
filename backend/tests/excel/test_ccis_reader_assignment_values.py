"""Tests for ``read_assignment_values`` in ``ccis_reader``.

The Assignment Values tab is the eMASS-exported list of ODP values for
a program. The reader has three jobs the rest of the ingest pipeline
relies on:

1. Parse value-bearing rows into :class:`AssignmentValueRow` with the
   ``odp_id`` normalized to the placeholder string the renderer looks
   up (``37`` → ``{$37$}``; Rev 5 tokens pass through).
2. Dedup within the parsed list on
   ``(control_id, odp_id, assigned_from)`` so an unchanged workbook
   re-import does not fire the OdpAuditLog diff path on every row.
3. Capture the authoritative slot order per control from the
   parameterized statement column — including slots with no assigned
   value — so the OSCAL positional bridge in
   :mod:`baselines.ccis_workbook` can align sparse workbooks.

Each test builds a tiny openpyxl workbook so the assertions stay
hermetic and don't depend on a real eMASS file.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from cybersecurity_assessor.excel.ccis_reader import (
    AssignmentValueRow,
    _ASSIGNMENT_VALUES_SHEET_NAMES,
    read_assignment_values,
)


def _make_av_workbook(
    path: Path,
    rows: list[dict],
    *,
    include_parameterized: bool = True,
    sheet_name: str | None = None,
) -> Path:
    """Write a minimal Assignment Values workbook to ``path``.

    Header row is row 1. Columns: A=Control Acronym, B=Assignment Value
    ID, C=Assignment Value, D=Assigned From, and (optionally) E=
    Parameterized Control. Each dict in ``rows`` keys against the four
    short names ``control_id`` / ``odp_id`` / ``value`` /
    ``assigned_from`` / ``parameterized``.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or _ASSIGNMENT_VALUES_SHEET_NAMES[0]
    ws.cell(row=1, column=1, value="Control Acronym")
    ws.cell(row=1, column=2, value="Assignment Value ID")
    ws.cell(row=1, column=3, value="Assignment Value")
    ws.cell(row=1, column=4, value="Assigned From")
    if include_parameterized:
        ws.cell(row=1, column=5, value="Parameterized Control")
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row.get("control_id"))
        ws.cell(row=i, column=2, value=row.get("odp_id"))
        ws.cell(row=i, column=3, value=row.get("value"))
        ws.cell(row=i, column=4, value=row.get("assigned_from"))
        if include_parameterized:
            ws.cell(row=i, column=5, value=row.get("parameterized"))
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Sparse workbook: AC-2 declares four slots, only two are filled
# ---------------------------------------------------------------------------


def test_sparse_control_parses_filled_and_captures_full_slot_order(tmp_path):
    """AC-2 in real workbooks declares four ODP slots but only fills
    two of them. The reader must:

      * emit rows ONLY for the value-bearing entries
      * still report all four slot ids in ``slot_orders`` (drawn from
        the parameterized statement column, not the value-bearing rows)

    Without both properties the OSCAL positional bridge would mis-align
    AC-2 every ingest.
    """
    parameterized = (
        "The organization: Identifies and selects {$36$}; "
        "Requires approvals by {$37$}; "
        "Reviews accounts {$38$}; "
        "Notifies account managers within {$39$}."
    )
    path = _make_av_workbook(
        tmp_path / "av.xlsx",
        [
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
                "parameterized": parameterized,
            },
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "24 hours",
                "assigned_from": "DoW Enterprise",
                # parameterized intentionally blank on subsequent rows
            },
        ],
    )

    rows, slot_orders = read_assignment_values(path)

    assert [r.odp_id for r in rows] == ["{$37$}", "{$39$}"]
    assert [r.value for r in rows] == ["ISSM or ISSO", "24 hours"]
    # Slot order MUST include all four declared slots, in template
    # appearance order — even the two with no value-bearing row.
    assert slot_orders["AC-2"] == ["{$36$}", "{$37$}", "{$38$}", "{$39$}"]


# ---------------------------------------------------------------------------
# Rev 4 integer → "{$N$}" normalization
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        (37, "{$37$}"),
        ("37", "{$37$}"),
        ("37.0", "{$37$}"),  # openpyxl sometimes returns "37.0" for numeric cells
        ("{$37$}", "{$37$}"),
        ("ac-02_odp.03", "ac-02_odp.03"),  # Rev 5 passes through
    ],
)
def test_odp_id_normalization(tmp_path, raw, expected):
    path = _make_av_workbook(
        tmp_path / "av.xlsx",
        [
            {
                "control_id": "AC-2",
                "odp_id": raw,
                "value": "v",
                "assigned_from": "DoW Enterprise",
                "parameterized": f"x {expected} y",
            }
        ],
    )
    rows, _ = read_assignment_values(path)
    assert len(rows) == 1
    assert rows[0].odp_id == expected


# ---------------------------------------------------------------------------
# Dedup on (control_id, odp_id, assigned_from)
# ---------------------------------------------------------------------------


def test_duplicate_rows_collapse_within_same_source(tmp_path):
    """Same (control, odp, source) triple appearing twice in the
    workbook must produce ONE row. The OdpAuditLog diff path triggers
    on every divergent row, so silent dedup keeps unchanged re-imports
    silent."""
    path = _make_av_workbook(
        tmp_path / "av.xlsx",
        [
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
                "parameterized": "x {$37$} y",
            },
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
            },
        ],
    )
    rows, _ = read_assignment_values(path)
    assert len(rows) == 1


def test_same_odp_different_source_coexists(tmp_path):
    """The (control, odp, source) PK is THREE-part. The same ODP id
    with two different overlay sources (e.g. DoW Enterprise vs FedRAMP
    HBL) must produce two rows — the render layer's most-recent-wins
    chooses between them.
    """
    path = _make_av_workbook(
        tmp_path / "av.xlsx",
        [
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "24 hours",
                "assigned_from": "DoW Enterprise",
                "parameterized": "x {$39$} y",
            },
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "1 hour",
                "assigned_from": "FedRAMP HBL",
            },
        ],
    )
    rows, _ = read_assignment_values(path)
    assert len(rows) == 2
    sources = sorted(r.assigned_from for r in rows)
    assert sources == ["DoW Enterprise", "FedRAMP HBL"]


# ---------------------------------------------------------------------------
# Sheet absence / structural drift
# ---------------------------------------------------------------------------


def test_no_assignment_values_tab_returns_empty(tmp_path):
    """Older eMASS exports omit the tab entirely. The reader returns
    ``([], {})`` rather than raising so workbook ingest can proceed."""
    wb = Workbook()
    wb.active.title = "WORKING SHEET"
    p = tmp_path / "no_av.xlsx"
    wb.save(p)

    rows, slot_orders = read_assignment_values(p)
    assert rows == []
    assert slot_orders == {}


def test_unrecognized_header_raises(tmp_path):
    """If the tab exists but headers don't map to required fields,
    raise rather than silently skip — silent skip would mask real
    data loss on a drifted template."""
    wb = Workbook()
    ws = wb.active
    ws.title = _ASSIGNMENT_VALUES_SHEET_NAMES[0]
    ws.cell(row=1, column=1, value="totally unexpected header")
    ws.cell(row=1, column=2, value="another unexpected header")
    p = tmp_path / "bad_headers.xlsx"
    wb.save(p)

    with pytest.raises(ValueError, match="unrecognized header layout"):
        read_assignment_values(p)


# ---------------------------------------------------------------------------
# Fallback slot order when parameterized column is absent
# ---------------------------------------------------------------------------


def test_slot_order_falls_back_to_value_rows_when_no_parameterized_col(tmp_path):
    """Older workbooks may not have the parameterized statement column.
    In that case slot order is derived from value-bearing rows in
    first-occurrence order. Fine for non-sparse controls (degrades
    gracefully to pre-Option-A behavior)."""
    path = _make_av_workbook(
        tmp_path / "av.xlsx",
        [
            {
                "control_id": "AU-2",
                "odp_id": 50,
                "value": "monthly",
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AU-2",
                "odp_id": 51,
                "value": "ISSO",
                "assigned_from": "DoW Enterprise",
            },
        ],
        include_parameterized=False,
    )

    rows, slot_orders = read_assignment_values(path)
    assert [r.odp_id for r in rows] == ["{$50$}", "{$51$}"]
    # No parameterized column → fallback to observed slot order.
    assert slot_orders["AU-2"] == ["{$50$}", "{$51$}"]


# ---------------------------------------------------------------------------
# Empty-value rows preserve slot identity
# ---------------------------------------------------------------------------


def test_empty_value_row_is_skipped_in_rows_but_slot_kept_via_parameterized(tmp_path):
    """An empty-value row in the workbook still encodes a slot. The
    reader skips it from ``rows`` (no value to store) but the
    parameterized column carries the slot identity so the bridge stays
    aligned. Verifies the AssignmentValueRow rows list reflects only
    value-bearing entries while slot_orders is full."""
    parameterized = "x {$10$} y {$11$} z {$12$}"
    path = _make_av_workbook(
        tmp_path / "av.xlsx",
        [
            {
                "control_id": "AC-7",
                "odp_id": 10,
                "value": "3 attempts",
                "assigned_from": "DoW Enterprise",
                "parameterized": parameterized,
            },
            {
                "control_id": "AC-7",
                "odp_id": 11,
                "value": "",  # empty — encodes slot identity only
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AC-7",
                "odp_id": 12,
                "value": "lock account",
                "assigned_from": "DoW Enterprise",
            },
        ],
    )
    rows, slot_orders = read_assignment_values(path)
    # Empty-value row is preserved (value coerced to "") so the bridge
    # can stamp it; the renderer surfaces empty as unresolved.
    assert [r.odp_id for r in rows] == ["{$10$}", "{$11$}", "{$12$}"]
    assert [r.value for r in rows] == ["3 attempts", "", "lock account"]
    assert slot_orders["AC-7"] == ["{$10$}", "{$11$}", "{$12$}"]


# ---------------------------------------------------------------------------
# AssignmentValueRow shape contract
# ---------------------------------------------------------------------------


def test_assignment_value_row_carries_excel_row_for_traceability(tmp_path):
    """``excel_row`` is the 1-based row number in the workbook. Used
    by audit-log messages and by manual workbook re-inspection.
    Header is row 1 so the first data row is row 2."""
    path = _make_av_workbook(
        tmp_path / "av.xlsx",
        [
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM",
                "assigned_from": "DoW Enterprise",
                "parameterized": "x {$37$}",
            }
        ],
    )
    rows, _ = read_assignment_values(path)
    assert isinstance(rows[0], AssignmentValueRow)
    assert rows[0].excel_row == 2
