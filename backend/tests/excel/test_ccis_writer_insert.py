"""Tests for ``ccis_writer.insert_cci_row``.

The function closes a workbook completeness gap: when the catalog (DB)
knows a CCI belongs to a control but the workbook's WORKING SHEET
doesn't list that CCI, the assess pipeline 422s into a dead end. The
sidecar's auto-insert path calls this helper, then re-reads the
workbook so the existing assess flow can proceed.

These tests reuse the openpyxl-backed xlwings fake from
``test_writer_safety`` (extended with ``range()`` bulk-read and
``.api.Rows(N).Insert()``).
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from cybersecurity_assessor.excel import ccis_writer
from cybersecurity_assessor.excel.ccis_reader import read_workbook_index
from cybersecurity_assessor.excel.ccis_writer import insert_cci_row

SHEET = "WORKING SHEET"


# ---------------------------------------------------------------------------
# Fake xlwings — extended with range() and api.Rows().Insert()
# ---------------------------------------------------------------------------


class _FakeCell:
    def __init__(self, sheet, row, col):
        self._sheet = sheet
        self._row = row
        self._col = col

    @property
    def value(self):
        return self._sheet._ws.cell(row=self._row, column=self._col).value

    @value.setter
    def value(self, v):
        self._sheet._ws.cell(row=self._row, column=self._col, value=v)


class _FakeRange:
    def __init__(self, sheet, top_left, bottom_right):
        self._sheet = sheet
        self._tl = top_left
        self._br = bottom_right

    @property
    def value(self):
        r1, c1 = self._tl
        r2, c2 = self._br
        ws = self._sheet._ws
        if r1 == r2 and c1 == c2:
            return ws.cell(row=r1, column=c1).value
        if c1 == c2:
            return [ws.cell(row=r, column=c1).value for r in range(r1, r2 + 1)]
        if r1 == r2:
            return [ws.cell(row=r1, column=c).value for c in range(c1, c2 + 1)]
        return [
            [ws.cell(row=r, column=c).value for c in range(c1, c2 + 1)]
            for r in range(r1, r2 + 1)
        ]


class _FakeApiRows:
    def __init__(self, sheet, row_num):
        self._sheet = sheet
        self._row = row_num

    def Insert(self):
        # openpyxl's insert_rows shifts all rows at and below ``self._row``
        # down by one and leaves the new row blank — matches what
        # Excel.Rows(N).Insert() does for cell values (formatting drift is
        # not modeled here since the fake doesn't carry CF/DV).
        self._sheet._ws.insert_rows(self._row)


class _FakeApi:
    def __init__(self, sheet):
        self._sheet = sheet

    def Rows(self, row_num):
        return _FakeApiRows(self._sheet, row_num)


class _FakeSheet:
    def __init__(self, book, name):
        self._book = book
        self.name = name
        self._ws = book._wb[name]
        self.api = _FakeApi(self)

    def cells(self, row, col):
        return _FakeCell(self, row, col)

    def range(self, top_left, bottom_right):
        return _FakeRange(self, top_left, bottom_right)


class _Sheets:
    def __init__(self, book):
        self._book = book
        self._cache: dict[str, _FakeSheet] = {}

    def __iter__(self):
        for name in self._book._wb.sheetnames:
            yield self[name]

    def __getitem__(self, name):
        if name not in self._book._wb.sheetnames:
            raise KeyError(name)
        if name not in self._cache:
            self._cache[name] = _FakeSheet(self._book, name)
        return self._cache[name]


class _FakeBook:
    def __init__(self, path: Path):
        self.fullname = str(path)
        self.name = path.name
        self._path = path
        self._wb = load_workbook(path)
        self.sheets = _Sheets(self)

    def save(self):
        self._wb.save(self._path)

    def close(self):
        self._wb.close()


class _FakeApps(list):
    pass


class _FakeBooks(list):
    def __init__(self, parent_apps):
        super().__init__()
        self._parent_apps = parent_apps

    def open(self, abs_path: str) -> _FakeBook:
        book = _FakeBook(Path(abs_path))
        self.append(book)
        return book


class _FakeApp:
    def __init__(self, owner_apps):
        self.books = _FakeBooks(owner_apps)


class _FakeXw:
    def __init__(self):
        self.apps = _FakeApps()
        self.books = _FakeBooks(self.apps)

    def App(self, visible=False, add_book=False):
        app = _FakeApp(self.apps)
        self.apps.append(app)
        return app


# ---------------------------------------------------------------------------
# Workbook fixtures
# ---------------------------------------------------------------------------


def _seed_row(ws, row: int, *, control: str, cci: str) -> None:
    ws.cell(row=row, column=1, value="YES")  # A
    ws.cell(row=row, column=2, value=control)  # B
    ws.cell(row=row, column=8, value=cci)  # H


def _make_workbook_with_control(path: Path, rows: list[tuple[str, str]]) -> None:
    """Build a stub WORKING SHEET seeded with the given (control, cci) rows
    starting at the first data row (7)."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for col in range(1, 22):
        ws.cell(row=6, column=col, value=f"H{col}")
    for offset, (control, cci) in enumerate(rows):
        _seed_row(ws, 7 + offset, control=control, cci=cci)
    wb.save(path)


@pytest.fixture
def fake_xw(monkeypatch):
    """No-op shim — ``ccis_writer`` no longer uses xlwings; row insertion
    is now driven by :func:`xlsx_surgery.patch_cells`. The fixture is
    retained so existing tests still resolve their ``fake_xw`` parameter,
    but it doesn't patch anything. The fake-xlwings classes above are
    dead-code carry-over kept only to keep the diff small; safe to delete
    in a follow-up cleanup.
    """
    return None


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_insert_after_last_row_of_same_control(tmp_path, fake_xw):
    """Workbook has AC-1 rows at 7-9 and AC-2 rows at 10-11. Inserting a
    new CCI for AC-1 should land at row 10 (after the last AC-1 row)."""
    path = tmp_path / "ccis.xlsx"
    _make_workbook_with_control(
        path,
        [
            ("AC-1", "CCI-000001"),
            ("AC-1", "CCI-000002"),
            ("AC-1", "CCI-000003"),
            ("AC-2", "CCI-000004"),
            ("AC-2", "CCI-000005"),
        ],
    )

    new_row = insert_cci_row(
        path,
        control_id="AC-1",
        cci_id="CCI-000007",
        definition="def text",
        guidance="guidance text",
        procedures="procedure text",
    )

    assert new_row == 10

    wb = load_workbook(path)
    ws = wb[SHEET]
    assert ws.cell(row=10, column=1).value == "YES"
    assert ws.cell(row=10, column=2).value == "AC-1"
    assert ws.cell(row=10, column=8).value == "CCI-000007"
    assert ws.cell(row=10, column=9).value == "def text"
    assert ws.cell(row=10, column=10).value == "guidance text"
    assert ws.cell(row=10, column=11).value == "procedure text"
    # AC-2 rows shifted down.
    assert ws.cell(row=11, column=2).value == "AC-2"
    assert ws.cell(row=11, column=8).value == "CCI-000004"
    assert ws.cell(row=12, column=8).value == "CCI-000005"
    wb.close()

    # And the parser sees the new row via its by_cci index.
    index = read_workbook_index(path)
    assert "CCI-000007" in index.by_cci()
    assert index.by_cci()["CCI-000007"].excel_row == 10


def test_insert_when_control_not_yet_in_workbook(tmp_path, fake_xw):
    """If the control has no rows, the new row is appended after the last
    non-blank row in col B."""
    path = tmp_path / "ccis.xlsx"
    _make_workbook_with_control(
        path,
        [
            ("AC-1", "CCI-000001"),
            ("AC-1", "CCI-000002"),
        ],
    )

    new_row = insert_cci_row(
        path,
        control_id="AC-2",
        cci_id="CCI-000099",
        definition="brand-new control",
    )

    assert new_row == 9

    wb = load_workbook(path)
    ws = wb[SHEET]
    assert ws.cell(row=9, column=2).value == "AC-2"
    assert ws.cell(row=9, column=8).value == "CCI-000099"
    wb.close()


def test_insert_canonicalises_cci_id(tmp_path, fake_xw):
    """Caller may pass 'CCI-7' or '7' or 'CCI-000007' — col H must always
    be canonical 6-digit form so the parser's by_cci index matches."""
    path = tmp_path / "ccis.xlsx"
    _make_workbook_with_control(path, [("AC-1", "CCI-000001")])

    insert_cci_row(path, control_id="AC-1", cci_id="7")

    wb = load_workbook(path)
    assert wb[SHEET].cell(row=8, column=8).value == "CCI-000007"
    wb.close()


def test_insert_creates_backup_and_passes_verification(tmp_path, fake_xw):
    """The safe_write harness should snapshot before the insert and the
    post-insert read-back should match — leaving exactly one .bak-*."""
    path = tmp_path / "ccis.xlsx"
    _make_workbook_with_control(path, [("AC-1", "CCI-000001")])

    insert_cci_row(
        path,
        control_id="AC-1",
        cci_id="CCI-000007",
        definition="def",
    )

    backups = sorted(path.parent.glob(path.name + ".bak-*"))
    assert len(backups) == 1


def test_insert_rejects_empty_control_id(tmp_path, fake_xw):
    path = tmp_path / "ccis.xlsx"
    _make_workbook_with_control(path, [("AC-1", "CCI-000001")])

    with pytest.raises(ValueError, match="control_id"):
        insert_cci_row(path, control_id="   ", cci_id="CCI-000007")


def test_insert_rejects_unparseable_cci(tmp_path, fake_xw):
    path = tmp_path / "ccis.xlsx"
    _make_workbook_with_control(path, [("AC-1", "CCI-000001")])

    with pytest.raises(ValueError, match="Cannot parse CCI"):
        insert_cci_row(path, control_id="AC-1", cci_id="not-a-cci")


def test_insert_missing_file_raises(tmp_path, fake_xw):
    with pytest.raises(FileNotFoundError):
        insert_cci_row(
            tmp_path / "nope.xlsx", control_id="AC-1", cci_id="CCI-000007"
        )
