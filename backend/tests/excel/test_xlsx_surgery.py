"""Tests for the headless .xlsx cell-patcher.

The patcher replaces xlwings/COM for the assessor-write path. These tests
verify it actually does what xlwings was chosen to do — preserve every
feature openpyxl would otherwise round-trip-strip:

    - Comments
    - Named ranges
    - Data validation
    - Conditional formatting
    - Merged cells
    - Formulas
    - Cell styles
    - Existing shared strings

Each test builds a real openpyxl workbook with the relevant feature, then
patches a target cell and asserts both the new value AND that the
preserved feature survives the round-trip.
"""

from __future__ import annotations

import zipfile
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from cybersecurity_assessor.excel import xlsx_surgery
from cybersecurity_assessor.excel.xlsx_surgery import (
    find_sheet_xml_path,
    patch_cells,
)

SHEET = "WORKING SHEET"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_workbook(path: Path) -> None:
    """Build a workbook with every preservation-critical feature populated."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    # Headers + a few data rows.
    for col_idx in range(1, 22):
        ws.cell(row=6, column=col_idx, value=f"H{col_idx}")
    for row_idx in range(7, 12):
        ws.cell(row=row_idx, column=1, value=f"REQ-{row_idx}")
        ws.cell(row=row_idx, column=2, value=f"AC-{row_idx}")
        ws.cell(row=row_idx, column=8, value=f"CCI-00000{row_idx}")
        # Pre-populate N (status) so we can verify replacement preserves style.
        cell_n = ws.cell(row=row_idx, column=14, value="Compliant")
        cell_n.font = Font(bold=True, color="FF0000")
        cell_n.fill = PatternFill("solid", fgColor="FFFF00")

    # Data validation (status dropdown) on col N rows 7-500.
    dv = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant,Not Applicable"',
        allow_blank=True,
    )
    dv.add("N7:N500")
    ws.add_data_validation(dv)

    # Conditional formatting on col N rows 7-500 — red fill for Non-Compliant.
    red = PatternFill("solid", fgColor="FFC7CE")
    ws.conditional_formatting.add(
        "N7:N500",
        CellIsRule(operator="equal", formula=['"Non-Compliant"'], fill=red),
    )

    # Merged cells in the metadata header band.
    ws.merge_cells("A1:Q1")
    ws.merge_cells("A2:Q2")

    # A comment on a header cell.
    ws["A6"].comment = Comment("Required flag", "Test")

    # A formula somewhere harmless.
    ws.cell(row=14, column=18, value="=SUM(N7:N12)")

    # A named range scoping the assessor write area.
    wb.defined_names["AssessorCells"] = DefinedName(
        name="AssessorCells",
        attr_text=f"'{SHEET}'!$N$7:$Q$500",
    )

    wb.save(path)


@pytest.fixture
def workbook(tmp_path: Path) -> Path:
    path = tmp_path / "ccis.xlsx"
    _make_workbook(path)
    return path


# ---------------------------------------------------------------------------
# Sheet path resolution
# ---------------------------------------------------------------------------


def test_find_sheet_xml_path_exact(workbook: Path) -> None:
    assert find_sheet_xml_path(workbook, SHEET) == "xl/worksheets/sheet1.xml"


def test_find_sheet_xml_path_fallback_substring(workbook: Path) -> None:
    # Even with a totally different cased name, the "working" substring
    # fallback should resolve to the same sheet.
    assert find_sheet_xml_path(workbook, "working") == "xl/worksheets/sheet1.xml"


def test_find_sheet_xml_path_missing_raises(workbook: Path) -> None:
    with pytest.raises(ValueError, match="Sheet not found"):
        find_sheet_xml_path(workbook, "DoesNotExist")


# ---------------------------------------------------------------------------
# Cell value writes
# ---------------------------------------------------------------------------


def test_patch_string_cell_writes_value(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    wb = load_workbook(workbook)
    assert wb[SHEET]["N7"].value == "Non-Compliant"
    wb.close()


def test_patch_multiple_cells_in_one_call(workbook: Path) -> None:
    patch_cells(
        workbook,
        SHEET,
        {
            "N7": "Non-Compliant",
            "O7": "2026-06-05",
            "P7": "Noah Jaskolski",
            "Q7": "Examined evidence; gap found.",
        },
    )
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    assert ws["N7"].value == "Non-Compliant"
    assert ws["O7"].value == "2026-06-05"
    assert ws["P7"].value == "Noah Jaskolski"
    assert ws["Q7"].value == "Examined evidence; gap found."
    wb.close()


def test_patch_date_value(workbook: Path) -> None:
    # Native date → Excel serial. openpyxl recognizes it as a date when read
    # back, but only if the cell carries a date number format. Without a date
    # style the value comes back as an int — assert the serial directly.
    patch_cells(workbook, SHEET, {"O7": date(2026, 6, 5)})
    wb = load_workbook(workbook)
    val = wb[SHEET]["O7"].value
    # 2026-06-05 → serial 46178 (days since 1899-12-30).
    assert val == 46178
    wb.close()


def test_patch_numeric_value(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"R7": 42})
    wb = load_workbook(workbook)
    assert wb[SHEET]["R7"].value == 42
    wb.close()


def test_patch_none_clears_cell(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": None})
    wb = load_workbook(workbook)
    assert wb[SHEET]["N7"].value is None
    wb.close()


def test_patch_new_string_appends_to_shared_strings(workbook: Path) -> None:
    """A novel value gets added to sharedStrings and resolved on read-back."""
    novel = "This exact string is not present in the workbook anywhere."
    patch_cells(workbook, SHEET, {"Q8": novel})
    wb = load_workbook(workbook)
    assert wb[SHEET]["Q8"].value == novel
    wb.close()


def test_patch_reuses_existing_shared_string(workbook: Path) -> None:
    """An already-present value should reuse its sst index, not duplicate.

    openpyxl may emit inline strings on initial save (no sharedStrings.xml
    entry yet), so seed the sst first with one explicit patch, then verify
    the second patch with the same value does not grow it.
    """
    patch_cells(workbook, SHEET, {"N12": "SharedSentinel"})
    with zipfile.ZipFile(workbook, "r") as zf:
        before = zf.read("xl/sharedStrings.xml").decode("utf-8")
    sst_count_before = before.count("<si")

    patch_cells(workbook, SHEET, {"N11": "SharedSentinel"})
    with zipfile.ZipFile(workbook, "r") as zf:
        after = zf.read("xl/sharedStrings.xml").decode("utf-8")
    sst_count_after = after.count("<si")
    assert sst_count_after == sst_count_before


# ---------------------------------------------------------------------------
# Preservation: every feature xlwings was chosen to protect
# ---------------------------------------------------------------------------


def test_comments_survive_write(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    wb = load_workbook(workbook)
    cmt = wb[SHEET]["A6"].comment
    assert cmt is not None
    assert cmt.text == "Required flag"
    wb.close()


def test_named_ranges_survive_write(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    wb = load_workbook(workbook)
    assert "AssessorCells" in wb.defined_names
    # Reference still points at $N$7:$Q$500 (sheet name may be quoted).
    name = wb.defined_names["AssessorCells"]
    assert "$N$7:$Q$500" in name.attr_text
    wb.close()


def test_data_validation_survives_write(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    dvs = list(ws.data_validations.dataValidation)
    assert any("N7:N500" in str(dv.sqref) for dv in dvs)
    wb.close()


def test_conditional_formatting_survives_write(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    cf_ranges = [str(r) for r in ws.conditional_formatting._cf_rules]
    assert any("N7:N500" in r for r in cf_ranges)
    wb.close()


def test_merged_cells_survive_write(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    wb = load_workbook(workbook)
    ranges = {str(r) for r in wb[SHEET].merged_cells.ranges}
    assert "A1:Q1" in ranges
    assert "A2:Q2" in ranges
    wb.close()


def test_formulas_survive_write(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    wb = load_workbook(workbook)
    # Formula cell at R14.
    assert wb[SHEET]["R14"].value == "=SUM(N7:N12)"
    wb.close()


def test_cell_style_preserved_on_replace(workbook: Path) -> None:
    """Replacing an existing styled cell keeps its `s=` attribute."""
    # Read original style index for N7 directly from the XML.
    with zipfile.ZipFile(workbook, "r") as zf:
        before = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    import re

    m = re.search(r'<c\b[^>]*\br="N7"[^>]*?\bs="(\d+)"', before)
    assert m is not None, "N7 should have a style attribute in the fixture"
    style_before = m.group(1)

    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})

    with zipfile.ZipFile(workbook, "r") as zf:
        after = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    m2 = re.search(r'<c\b[^>]*\br="N7"[^>]*?\bs="(\d+)"', after)
    assert m2 is not None, "N7 must still carry its style after patch"
    assert m2.group(1) == style_before


# ---------------------------------------------------------------------------
# Row insertion
# ---------------------------------------------------------------------------


def test_insert_row_bumps_existing_rows(workbook: Path) -> None:
    """Inserting at row 8 shifts old row 8's content down to row 9."""
    patch_cells(workbook, SHEET, {}, insert_row_before=8)
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    # Row 7 untouched.
    assert ws.cell(row=7, column=1).value == "REQ-7"
    assert ws.cell(row=7, column=14).value == "Compliant"
    # Old row 8 content now at row 9.
    assert ws.cell(row=9, column=1).value == "REQ-8"
    assert ws.cell(row=9, column=14).value == "Compliant"
    # Row 8 is the new empty row.
    assert ws.cell(row=8, column=1).value is None
    wb.close()


def test_insert_row_extends_data_validation_sqref(workbook: Path) -> None:
    """A dataValidation that spanned N7:N500 should now span N7:N501."""
    patch_cells(workbook, SHEET, {}, insert_row_before=8)
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    dv_sqrefs = [str(dv.sqref) for dv in ws.data_validations.dataValidation]
    assert any("N7:N501" in s for s in dv_sqrefs), (
        f"Expected N7:N501 in {dv_sqrefs!r}"
    )
    wb.close()


def test_insert_row_extends_conditional_formatting_sqref(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {}, insert_row_before=8)
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    cf_ranges = [str(r) for r in ws.conditional_formatting._cf_rules]
    assert any("N7:N501" in r for r in cf_ranges)
    wb.close()


def test_insert_row_then_populate_in_same_call(workbook: Path) -> None:
    """patch_cells with both insert_row_before AND cells should populate the
    new row in one round trip — addresses reference the post-insert row."""
    patch_cells(
        workbook,
        SHEET,
        {"A8": "REQ-new", "B8": "AC-new", "H8": "CCI-099999"},
        insert_row_before=8,
    )
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    assert ws.cell(row=8, column=1).value == "REQ-new"
    assert ws.cell(row=8, column=2).value == "AC-new"
    assert ws.cell(row=8, column=8).value == "CCI-099999"
    # The old row 8 still has its content, now at row 9.
    assert ws.cell(row=9, column=1).value == "REQ-8"
    wb.close()


def test_insert_row_preserves_formula_reference_after_bump(workbook: Path) -> None:
    """The formula at R14 references rows 7-12. After inserting at row 8,
    the formula's cell ref itself moves to R15 — its body is left intact
    (we don't rewrite formula bodies)."""
    patch_cells(workbook, SHEET, {}, insert_row_before=8)
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    # Formula bumped one row down.
    assert ws.cell(row=15, column=18).value == "=SUM(N7:N12)"
    wb.close()


# ---------------------------------------------------------------------------
# Atomicity / error paths
# ---------------------------------------------------------------------------


def test_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        patch_cells(tmp_path / "nope.xlsx", SHEET, {"N7": "x"})


def test_no_temp_files_left_behind_on_success(workbook: Path) -> None:
    patch_cells(workbook, SHEET, {"N7": "Non-Compliant"})
    stragglers = list(workbook.parent.glob(workbook.name + ".*.tmp"))
    assert stragglers == []
