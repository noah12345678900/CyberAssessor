"""Tests for the safety harness in ``ccis_writer``.

The writer drives a headless ``.xlsx`` zip-surgery path in production
(no xlwings, no live Excel). These tests exercise the backup-and-verify
behavior of ``safe_write`` directly against real ``.xlsx`` files on disk.

Coverage:
    - Successful write: backup is created, verification passes, backup
      is retained (up to the rolling 5).
    - Verification failure: writer claims it wrote a value but the
      on-disk file doesn't reflect it → ``WorkbookWriteVerificationError``
      raised and the workbook is restored from the snapshot.
    - Retention: more than 5 writes prunes oldest backups.
    - Exception during write: backup restores original.
"""

from __future__ import annotations

import time
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from cybersecurity_assessor.excel import ccis_writer
from cybersecurity_assessor.excel.ccis_writer import (
    CcisWrite,
    WorkbookWriteVerificationError,
    write_assessment,
)

SHEET = "WORKING SHEET"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    # Headers row at row 6 to mirror real layout (writer only addresses
    # cells by absolute index so labels don't matter).
    for col in range(1, 22):
        ws.cell(row=6, column=col, value=f"H{col}")
    # A pre-existing data row at row 7 with stub content in writable cols.
    ws.cell(row=7, column=14, value="Compliant")
    ws.cell(row=7, column=15, value="2024-01-01")
    ws.cell(row=7, column=16, value="Prior Assessor")
    ws.cell(row=7, column=17, value="Prior results")
    wb.save(path)


@pytest.fixture
def workbook(tmp_path: Path) -> Path:
    path = tmp_path / "ccis.xlsx"
    _make_workbook(path)
    return path


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_successful_write_creates_backup_and_verifies(workbook):
    """A normal write produces one .bak-* snapshot next to the workbook
    and the post-write read-back matches what was written."""
    result = write_assessment(
        workbook,
        [
            CcisWrite(
                excel_row=7,
                status="Non-Compliant",
                date_tested=datetime(2026, 6, 3),
                tester="Noah Jaskolski",
                results="Examined evidence; gap found.",
            )
        ],
    )

    # Sanity: the writer reports the write.
    assert result["rows_written"] == 1
    assert result["cells_changed"] == 4

    # On-disk file actually reflects the write.
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    assert ws.cell(row=7, column=14).value == "Non-Compliant"
    assert ws.cell(row=7, column=16).value == "Noah Jaskolski"
    wb.close()

    # Exactly one backup exists.
    backups = sorted(workbook.parent.glob(workbook.name + ".bak-*"))
    assert len(backups) == 1

    # The backup contains the PRE-write state. openpyxl rejects the
    # .bak-* suffix as an unknown format, so copy to a .xlsx temp and
    # read that — verifies the bytes themselves are still a valid xlsx.
    import shutil as _sh
    bak_xlsx = workbook.parent / "verify_backup.xlsx"
    _sh.copy2(backups[0], bak_xlsx)
    wb_bak = load_workbook(bak_xlsx)
    ws_bak = wb_bak[SHEET]
    assert ws_bak.cell(row=7, column=14).value == "Compliant"  # original
    assert ws_bak.cell(row=7, column=16).value == "Prior Assessor"
    wb_bak.close()


def test_verification_failure_restores_backup(workbook, monkeypatch):
    """If the patcher silently drops a write, the harness detects the
    mismatch on read-back, restores the snapshot, and raises.

    Patches ``ccis_writer._patch_cells`` to strip the N7 entry before
    delegating to the real patcher — simulating a writer that *claims*
    it wrote N7 but didn't.
    """
    real_patch = ccis_writer._patch_cells

    def patch_dropping_n7(path, sheet_name, cells, *, insert_row_before=None):
        filtered = {k: v for k, v in cells.items() if k != "N7"}
        return real_patch(
            path,
            sheet_name,
            filtered,
            insert_row_before=insert_row_before,
        )

    monkeypatch.setattr(ccis_writer, "_patch_cells", patch_dropping_n7)

    with pytest.raises(WorkbookWriteVerificationError) as exc_info:
        write_assessment(
            workbook,
            [CcisWrite(excel_row=7, status="Non-Compliant")],
        )

    assert "N7" in str(exc_info.value)

    # Workbook on disk was restored to its original Compliant value.
    wb = load_workbook(workbook)
    ws = wb[SHEET]
    assert ws.cell(row=7, column=14).value == "Compliant"
    wb.close()

    # Backup was cleaned up on rollback (it would be redundant with the
    # restored original).
    backups = list(workbook.parent.glob(workbook.name + ".bak-*"))
    assert backups == []


def test_exception_during_write_restores_backup(workbook):
    """If the safe_write block raises, the original is restored."""
    from cybersecurity_assessor.excel.ccis_writer import safe_write

    # Mutate the file inside the context, then raise.
    with pytest.raises(RuntimeError):
        with safe_write(workbook) as ctx:
            wb = load_workbook(workbook)
            wb[SHEET].cell(row=7, column=14, value="Tampered")
            wb.save(workbook)
            wb.close()
            ctx["expected"]["N7"] = "Tampered"
            raise RuntimeError("boom")

    # Original value restored.
    wb = load_workbook(workbook)
    assert wb[SHEET].cell(row=7, column=14).value == "Compliant"
    wb.close()

    # Backup pruned on rollback.
    backups = list(workbook.parent.glob(workbook.name + ".bak-*"))
    assert backups == []


def test_backup_retention_keeps_only_five(workbook):
    """Seven successive writes should leave only the 5 newest backups."""
    for i in range(7):
        write_assessment(
            workbook,
            [CcisWrite(excel_row=7, tester=f"Tester {i}")],
        )
        # Force distinct mtimes — the timestamp suffix has 1-second
        # resolution and tests can complete sub-second on fast disks.
        time.sleep(1.1)

    backups = sorted(workbook.parent.glob(workbook.name + ".bak-*"))
    assert len(backups) == 5
