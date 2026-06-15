"""Tests for the re-read mode in ``ccis_reader``.

Builds tiny in-memory openpyxl workbooks (saved to a tmp path so the
sidecar snapshot has somewhere to live) and walks the diff through
the four states the spec calls out: added, removed, moved, edited.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from cybersecurity_assessor.excel.ccis_reader import (
    COL_CCI,
    COL_CONTROL,
    COL_DATE_TESTED,
    COL_DEFINITION,
    COL_GUIDANCE,
    COL_NARRATIVE,
    COL_PREV_RESULTS,
    COL_PROCEDURES,
    COL_REQUIRED,
    COL_RESULTS,
    COL_STATUS,
    COL_TESTER,
    RereadDiff,
    RereadResult,
    reread_workbook,
)

SHEET = "WORKING SHEET"


def _make_workbook(path: Path, rows: list[dict]) -> Path:
    """Write a minimal CCIS-shaped workbook to ``path``.

    ``rows`` is a list of dicts keyed by short field names; only fields
    set will be written, leaving other columns blank — which matches how
    the reader handles real workbooks.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    # Reader expects header on row 6 (rows 1-5 are system metadata).
    ws.cell(row=6, column=COL_CONTROL, value="Control Acronym")
    ws.cell(row=6, column=COL_CCI, value="CCI")

    field_to_col = {
        "required": COL_REQUIRED,
        "control_id": COL_CONTROL,
        "cci_id": COL_CCI,
        "narrative": COL_NARRATIVE,
        "definition": COL_DEFINITION,
        "guidance": COL_GUIDANCE,
        "procedures": COL_PROCEDURES,
        "status": COL_STATUS,
        "date_tested": COL_DATE_TESTED,
        "tester": COL_TESTER,
        "results": COL_RESULTS,
        "previous_results": COL_PREV_RESULTS,
    }

    for offset, row in enumerate(rows):
        excel_row = 7 + offset
        for key, val in row.items():
            col = field_to_col.get(key)
            if col is None:
                raise KeyError(f"Unknown fixture key: {key}")
            ws.cell(row=excel_row, column=col, value=val)

    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def workbook_path(tmp_path):
    return tmp_path / "test_ccis.xlsx"


def test_first_read_has_no_prior_snapshot_and_lists_all_as_added(workbook_path):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )
    result = reread_workbook(workbook_path)

    assert isinstance(result, RereadResult)
    assert result.had_prior_snapshot is False
    assert len(result.index.rows) == 2
    assert len(result.diff.added) == 2
    assert result.diff.removed == []
    assert result.diff.moved == []
    assert result.diff.edited == []
    # Snapshot sidecar must now exist.
    assert (workbook_path.parent / (workbook_path.name + ".snapshot.json")).exists()


def test_unchanged_reread_produces_empty_diff(workbook_path):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )
    reread_workbook(workbook_path)  # establishes snapshot
    result = reread_workbook(workbook_path)

    assert result.had_prior_snapshot is True
    assert result.diff.is_empty, result.diff


def test_added_row_detected(workbook_path):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
        ],
    )
    reread_workbook(workbook_path)

    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "new"},
        ],
    )
    result = reread_workbook(workbook_path)

    assert len(result.diff.added) == 1
    assert result.diff.added[0]["key"] == ["AC-3", "CCI-000213"]
    assert result.diff.removed == []
    assert result.diff.edited == []


def test_removed_row_detected(workbook_path):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )
    reread_workbook(workbook_path)

    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
        ],
    )
    result = reread_workbook(workbook_path)

    assert len(result.diff.removed) == 1
    assert result.diff.removed[0]["key"] == ["AC-3", "CCI-000213"]
    assert result.diff.added == []


def test_moved_row_detected(workbook_path):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
        ],
    )
    reread_workbook(workbook_path)

    # Swap the order — same content, different excel_row positions.
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-3", "cci_id": "000213", "narrative": "n2"},
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
        ],
    )
    result = reread_workbook(workbook_path)

    assert len(result.diff.moved) == 2
    move_keys = {tuple(m["key"]) for m in result.diff.moved}
    assert move_keys == {("AC-2", "CCI-000015"), ("AC-3", "CCI-000213")}
    # No edits because the tracked columns didn't change.
    assert result.diff.edited == []


def test_edited_columns_detected_for_each_tracked_column(workbook_path):
    _make_workbook(
        workbook_path,
        [
            {
                "control_id": "AC-2",
                "cci_id": "000015",
                "narrative": "orig narrative",
                "definition": "orig definition",
                "guidance": "orig guidance",
                "procedures": "orig procedures",
                "status": "Compliant",
                "tester": "Noah Jaskolski",
                "results": "orig results",
                "previous_results": "orig prev",
            },
        ],
    )
    reread_workbook(workbook_path)

    _make_workbook(
        workbook_path,
        [
            {
                "control_id": "AC-2",
                "cci_id": "000015",
                "narrative": "EDITED narrative",
                "definition": "EDITED definition",
                "guidance": "EDITED guidance",
                "procedures": "EDITED procedures",
                "status": "Non-Compliant",
                "tester": "Someone Else",
                "results": "EDITED results",
                "previous_results": "EDITED prev",
            },
        ],
    )
    result = reread_workbook(workbook_path)

    assert len(result.diff.edited) == 1
    edited = result.diff.edited[0]
    assert edited["key"] == ["AC-2", "CCI-000015"]
    # F, I, J, K, N, P, Q, U all touched (O / date_tested not changed).
    assert set(edited["changed_cols"]) == {"F", "I", "J", "K", "N", "P", "Q", "U"}


def test_whitespace_only_change_is_not_an_edit(workbook_path):
    _make_workbook(
        workbook_path,
        [{"control_id": "AC-2", "cci_id": "000015", "narrative": "narrative text"}],
    )
    reread_workbook(workbook_path)

    _make_workbook(
        workbook_path,
        [{"control_id": "AC-2", "cci_id": "000015", "narrative": "  narrative text  "}],
    )
    result = reread_workbook(workbook_path)

    assert result.diff.is_empty, result.diff


def test_rows_without_cci_are_excluded_from_diff(workbook_path):
    _make_workbook(
        workbook_path,
        [
            {"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"},
            # No CCI — not keyable.
            {"control_id": "AC-3", "narrative": "no cci"},
        ],
    )
    result = reread_workbook(workbook_path)
    # Only the keyable row shows up in "added" (the other has no key).
    assert len(result.diff.added) == 1
    assert result.diff.added[0]["key"] == ["AC-2", "CCI-000015"]
    # But the index itself parsed both rows.
    assert len(result.index.rows) == 2


def test_update_snapshot_false_does_not_persist(workbook_path):
    _make_workbook(
        workbook_path,
        [{"control_id": "AC-2", "cci_id": "000015", "narrative": "n1"}],
    )
    reread_workbook(workbook_path, update_snapshot=False)
    sidecar = workbook_path.parent / (workbook_path.name + ".snapshot.json")
    assert not sidecar.exists()


def test_reread_diff_is_empty_property(workbook_path):
    empty = RereadDiff()
    assert empty.is_empty
    populated = RereadDiff(added=[{"key": ["X", "Y"]}])
    assert not populated.is_empty
