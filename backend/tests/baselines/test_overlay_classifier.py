"""Unit tests for baselines/overlay_classifier.sniff_overlay_kind.

Synthesizes xlsx fixtures in tmp_path so the tests don't depend on
real CRM / PSC files shipping with the repo. The header tokens used
here are pulled directly from the loader vocabularies the classifier
sniffs against:

  * CRM:  baselines/crm_xlsx.py _CONTROL_ID_HEADERS + _RESPONSIBILITY_HEADERS_*
  * PSC:  catalogs/program_controls_loader.py _HEADER_ALIASES
"""

from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook as XlsxWorkbook

_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor.baselines.overlay_classifier import (  # noqa: E402
    OverlayKind,
    sniff_overlay_kind,
)


def _write_crm_fixture(path: Path) -> None:
    """A CRM-shaped sheet: control-id column + responsibility column."""
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "CRM"
    ws.append(["Control ID", "Responsibility", "Customer Responsibility"])
    ws.append(["AC-2", "Customer", "Customer configures user accounts."])
    ws.append(["AC-2(1)", "Hybrid", "Provider supplies IdP; customer manages users."])
    ws.append(["AU-12", "Provider", "Provider runs audit pipeline."])
    wb.save(path)


def _write_psc_fixture(path: Path) -> None:
    """A PSC-shaped sheet: threshold/shall column.

    Matches the SDA Enterprise Services Controls shape (Control No /
    Threshold / CCI columns) used by the program_controls_loader tests.
    The CCI column is incidental — PSC classification keys off the
    requirement-text column alone.
    """
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Overlay"
    ws.append(["Control No", "Threshold", "CCI"])
    ws.append(["SDA-001", "The system shall do X.", "CCI-000015"])
    ws.append(["SDA-002", "The system shall do Y.", "CCI-000007"])
    wb.save(path)


def _write_psc_t1tl_fixture(path: Path) -> None:
    """A control-grain PSC sheet with NO CCI column — T1TL shape.

    Real-world T1TL Ground / SV Security Controls sheets ship with
    ``Security Control | Threshold | Objective | Deliverable | Notes``
    and no CCI references column. CCIs are derived later from the
    ``Associated CNSSI 1253 Control Tag:`` anchor in the shall prose,
    joined against the global DISA mapping. The classifier must accept
    these — requiring a CCI column would block the unified-import
    dispatch on the T1TL files the assessor actually loads.
    """
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Ground Security Controls"
    ws.append(["Program: Acme T1TL"])
    ws.append([])
    ws.append(
        ["-", "Security Control", "Threshold", "Objective", "Deliverable", "Notes"]
    )
    ws.append(
        [
            1,
            "Associated CNSSI 1253 Control Tag: AC-2\nThe system shall manage accounts.",
            "",
            "",
            "SSP",
            "",
        ]
    )
    wb.save(path)


def _write_other_fixture(path: Path) -> None:
    """Arbitrary xlsx with no recognizable CRM or PSC vocab.

    Resembles a config baseline / inventory export the user might drop
    in — has headers, just none that match the loader vocabularies.
    """
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Hostname", "IP Address", "OS", "Last Patched"])
    ws.append(["host-01", "10.0.0.1", "RHEL 8.6", "2026-05-30"])
    ws.append(["host-02", "10.0.0.2", "Win Server 2022", "2026-05-28"])
    wb.save(path)


def _write_mixed_fixture(path: Path) -> None:
    """Two sheets — first CRM-shaped, second PSC-shaped. First wins."""
    wb = XlsxWorkbook()
    crm_sheet = wb.active
    crm_sheet.title = "CRM Sheet"
    crm_sheet.append(["Control ID", "Responsibility"])
    crm_sheet.append(["AC-2", "Customer"])

    psc_sheet = wb.create_sheet("PSC Sheet")
    psc_sheet.append(["Control No", "Threshold", "CCI"])
    psc_sheet.append(["SDA-001", "The system shall enforce X.", "CCI-000015"])

    wb.save(path)


def _write_empty_headers_fixture(path: Path) -> None:
    """Blank sheet — no header row at all. Must classify OTHER, not error."""
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Empty"
    wb.save(path)


def test_crm_fixture_classifies_as_crm(tmp_path: Path) -> None:
    path = tmp_path / "crm.xlsx"
    _write_crm_fixture(path)
    assert sniff_overlay_kind(path) is OverlayKind.CRM


def test_psc_fixture_classifies_as_psc(tmp_path: Path) -> None:
    path = tmp_path / "psc.xlsx"
    _write_psc_fixture(path)
    assert sniff_overlay_kind(path) is OverlayKind.PSC


def test_psc_t1tl_without_cci_column_classifies_as_psc(tmp_path: Path) -> None:
    """Regression: T1TL-style PSC sheets ship with no CCI column. The
    classifier must still call them PSC — earlier versions required a
    CCI column and misrouted these to OTHER, blocking the unified
    import dispatch even when the user forced kind_hint=psc."""
    path = tmp_path / "psc_t1tl.xlsx"
    _write_psc_t1tl_fixture(path)
    assert sniff_overlay_kind(path) is OverlayKind.PSC


def test_other_fixture_classifies_as_other(tmp_path: Path) -> None:
    path = tmp_path / "other.xlsx"
    _write_other_fixture(path)
    assert sniff_overlay_kind(path) is OverlayKind.OTHER


def test_mixed_fixture_first_sheet_wins(tmp_path: Path) -> None:
    """When a file has both a CRM-shaped sheet AND a PSC-shaped sheet,
    the first matched sheet wins. The mixed fixture orders CRM first,
    so the result is CRM. Documented in the classify_overlay module
    docstring."""
    path = tmp_path / "mixed.xlsx"
    _write_mixed_fixture(path)
    assert sniff_overlay_kind(path) is OverlayKind.CRM


def test_empty_workbook_classifies_as_other(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    _write_empty_headers_fixture(path)
    assert sniff_overlay_kind(path) is OverlayKind.OTHER


def test_missing_file_raises_file_not_found(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        sniff_overlay_kind(tmp_path / "does-not-exist.xlsx")


def test_header_normalization_tolerates_punctuation_and_case(tmp_path: Path) -> None:
    """``control_id`` / ``CONTROL #`` / ``Control-ID`` must all match the
    same normalized token as ``Control ID``. Catches loaders that hand-
    rolled their own case-folding and missed a separator."""
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "Quirky"
    ws.append(["CONTROL #", "RESPONSIBILITY"])
    ws.append(["AC-2", "Customer"])
    path = tmp_path / "quirky.xlsx"
    wb.save(path)
    assert sniff_overlay_kind(path) is OverlayKind.CRM


def test_header_within_first_ten_rows_is_found(tmp_path: Path) -> None:
    """PSC files often have a title block above the actual header row —
    the classifier scans the top _HEADER_SCAN_DEPTH rows so it sees the
    same row the PSC loader will parse."""
    wb = XlsxWorkbook()
    ws = wb.active
    ws.title = "TitleBlock"
    ws.append(["Program: Acme T1TL"])
    ws.append(["Revision: 2026-06-01"])
    ws.append([])
    ws.append(["Control No", "Threshold", "CCI"])
    ws.append(["SDA-001", "The system shall do X.", "CCI-000015"])
    path = tmp_path / "title_block.xlsx"
    wb.save(path)
    assert sniff_overlay_kind(path) is OverlayKind.PSC
