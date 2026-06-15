"""End-to-end tests for the Step 6/6b ODP ingest path in
``CcisWorkbookBaselineSource.apply()``.

The two prior test files cover the building blocks in isolation
(:mod:`test_ccis_reader_assignment_values` for the parse,
:mod:`test_odp_render` for the render). This file covers the wiring
that turns a workbook file on disk into rows in ``odp_assignment`` plus
diff rows in ``odp_audit_log``. Three properties matter:

1. **Round-trip.** A first ingest of a fresh workbook lands every
   value-bearing row as an ``OdpAssignment`` with the right
   ``framework_version`` / ``control_id`` / ``odp_id`` / ``value`` /
   ``source_ingest='CCIS-workbook'``. The Step 6b positional bridge
   stamps ``oscal_param_id`` on rows whose slot matches an OSCAL
   ``{{ insert: param, X }}`` reference in the seeded
   ``Control.statement``.
2. **Idempotent re-apply.** Re-importing the same workbook unchanged
   must produce ZERO new ``OdpAuditLog`` rows — the diff path is
   value-gated, and a no-op re-import is a routine operation
   (re-opening the same workbook in the UI). Spurious audit rows would
   pollute the SAR history.
3. **Diff fires on a real change.** Edit one cell in the Assignment
   Values tab, re-ingest, and exactly one ``OdpAuditLog`` row appears
   with ``prev_value`` / ``new_value`` / ``who='CCIS-workbook-ingest:<file>'``.

A fourth check covers Step 6b's orphan-tracking: when a value-bearing
row's ``odp_id`` isn't in the parameterized statement column's slot
list, the row stays ``oscal_param_id=NULL`` and the count surfaces in
``BaselineApplyResult.notes['odp_assignments']['value_rows_without_slot']``.
Observed in the wild on the May 2026 Example System workbook (AC-7, SI-5,
CM-3, SA-19, SI-3) — the bridge correctly abstaining instead of
guessing is what keeps the renderer from substituting the wrong value.
"""

from __future__ import annotations

import sys
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine, select

_BACKEND = Path(__file__).resolve().parents[2]
if str(_BACKEND) not in sys.path:
    sys.path.insert(0, str(_BACKEND))

from cybersecurity_assessor import models  # noqa: F401,E402 -- registers tables
from cybersecurity_assessor.baselines.ccis_workbook import (  # noqa: E402
    CcisWorkbookBaselineSource,
)
from cybersecurity_assessor.excel.ccis_reader import _INDEX_CACHE  # noqa: E402
from cybersecurity_assessor.models import (  # noqa: E402
    Control,
    Framework,
    Objective,
    OdpAssignment,
    OdpAuditLog,
)

FW_ID = "NIST-800-53r4"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def session():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def seeded_framework(session: Session) -> Framework:
    """Framework + AC-2 Control + one Objective (CCI-000015).

    The Control.statement carries two OSCAL ``{{ insert: param, X }}``
    references so the Step 6b positional bridge has something to align
    against. Two params + two slots in the workbook = clean
    positional match.
    """
    fw = Framework(
        name="NIST SP 800-53",
        version="Rev 4",
        framework_id=FW_ID,
    )
    session.add(fw)
    session.commit()
    session.refresh(fw)

    ctl = Control(
        framework_id=fw.id,  # type: ignore[arg-type]
        control_id="ac-2",
        title="Account Management",
        family="AC",
        statement=(
            "The organization: "
            "Requires approvals by {{ insert: param, ac-2_prm_1 }} for requests; "
            "Notifies account managers within {{ insert: param, ac-2_prm_2 }}."
        ),
    )
    session.add(ctl)
    session.commit()
    session.refresh(ctl)

    obj = Objective(
        control_id_fk=ctl.id,  # type: ignore[arg-type]
        objective_id="CCI-000015",
        source="CCI",
        text="The organization defines personnel or roles for account management approvals.",
    )
    session.add(obj)
    session.commit()
    return fw


def _build_workbook(
    path: Path,
    *,
    assignment_rows: list[dict],
    parameterized: str,
    include_assignment_tab: bool = True,
) -> Path:
    """Build a workbook with a minimal Working Sheet + Assignment Values tab.

    The Working Sheet only needs one CCI row referencing AC-2 / CCI-15
    so that ``populate_objectives`` finds the seeded catalog row and
    ``apply()`` reaches Step 6 with a Baseline row to attach to. Header
    row is row 6; data starts row 7 per the eMASS template.

    ``assignment_rows`` is a list of dicts shaped like the Assignment
    Values tab: ``control_id`` / ``odp_id`` / ``value`` /
    ``assigned_from`` / optional ``parameterized``. Only the FIRST row
    needs the parameterized statement filled — the reader pulls slot
    order off any non-blank cell for the control.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "WORKING SHEET"
    # Rows 1-5: workbook metadata (irrelevant, just need to exist).
    for i in range(1, 6):
        ws.cell(row=i, column=1, value=f"meta row {i}")
    # Row 6: headers (the reader doesn't actually check them — it
    # streams from row 7 by column index — but real workbooks have them).
    ws.cell(row=6, column=1, value="Required for assessment?")
    ws.cell(row=6, column=2, value="Control Acronym")
    ws.cell(row=6, column=7, value="AP Acronym")
    ws.cell(row=6, column=8, value="CCI")
    # Row 7: one CCI tying the workbook to the seeded catalog.
    ws.cell(row=7, column=1, value="YES")
    ws.cell(row=7, column=2, value="AC-2")
    ws.cell(row=7, column=7, value="AC-2.1")
    ws.cell(row=7, column=8, value="CCI-000015")

    if include_assignment_tab:
        av = wb.create_sheet(title="Assignment Values")
        av.cell(row=1, column=1, value="Control Acronym")
        av.cell(row=1, column=2, value="Assignment Value ID")
        av.cell(row=1, column=3, value="Assignment Value")
        av.cell(row=1, column=4, value="Assigned From")
        av.cell(row=1, column=5, value="Parameterized Control")
        for i, row in enumerate(assignment_rows, start=2):
            av.cell(row=i, column=1, value=row.get("control_id"))
            av.cell(row=i, column=2, value=row.get("odp_id"))
            av.cell(row=i, column=3, value=row.get("value"))
            av.cell(row=i, column=4, value=row.get("assigned_from"))
            # Only the first row carries the parameterized statement —
            # mirrors what real eMASS exports do.
            if i == 2:
                av.cell(row=i, column=5, value=parameterized)

    wb.save(path)
    return path


def _bust_workbook_cache() -> None:
    """The ccis_reader index cache is keyed by (path, mtime, size).
    Re-saving the SAME path in a single test would normally bump mtime
    but openpyxl preserves enough byte-identity that the cache can
    occasionally hit. Clearing per-test keeps the read fresh.
    """
    _INDEX_CACHE.clear()


# ---------------------------------------------------------------------------
# Round-trip: first ingest lands rows + stamps the bridge
# ---------------------------------------------------------------------------


def test_first_apply_creates_odp_assignment_rows(
    tmp_path, session: Session, seeded_framework: Framework
):
    """Two value-bearing rows in the workbook → two OdpAssignment rows
    with framework_version, OSCAL-canonical control_id, verbatim
    odp_id, value, assigned_from, and source_ingest stamped correctly.
    """
    _bust_workbook_cache()
    parameterized = (
        "Requires approvals by {$37$} for requests; "
        "Notifies account managers within {$39$}."
    )
    path = _build_workbook(
        tmp_path / "wb.xlsx",
        assignment_rows=[
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "24 hours",
                "assigned_from": "DoW Enterprise",
            },
        ],
        parameterized=parameterized,
    )

    result = CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )

    rows = session.exec(select(OdpAssignment)).all()
    assert len(rows) == 2

    by_odp = {r.odp_id: r for r in rows}
    assert set(by_odp.keys()) == {"{$37$}", "{$39$}"}
    for r in rows:
        assert r.framework_version == FW_ID
        # Stored canonical OSCAL form, not the workbook's "AC-2".
        assert r.control_id == "ac-2"
        assert r.assigned_from == "DoW Enterprise"
        assert r.source_ingest == "CCIS-workbook"
    assert by_odp["{$37$}"].value == "ISSM or ISSO"
    assert by_odp["{$39$}"].value == "24 hours"

    # Step 6 counts surface in the result notes.
    notes = result.notes["odp_assignments"]
    assert notes["inserted"] == 2
    assert notes["updated"] == 0
    assert notes["rows_parsed"] == 2


def test_first_apply_stamps_oscal_param_bridge(
    tmp_path, session: Session, seeded_framework: Framework
):
    """The seeded Control.statement declares two OSCAL params in order
    (``ac-2_prm_1``, ``ac-2_prm_2``). The workbook's parameterized
    column lists two slots in the same order (``{$37$}``, ``{$39$}``).
    Counts match → Step 6b zips positionally and stamps each row's
    ``oscal_param_id`` with the matching OSCAL id.
    """
    _bust_workbook_cache()
    parameterized = (
        "Requires approvals by {$37$} for requests; "
        "Notifies account managers within {$39$}."
    )
    path = _build_workbook(
        tmp_path / "wb.xlsx",
        assignment_rows=[
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "24 hours",
                "assigned_from": "DoW Enterprise",
            },
        ],
        parameterized=parameterized,
    )

    result = CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )

    rows = {
        r.odp_id: r for r in session.exec(select(OdpAssignment)).all()
    }
    # Positional alignment: first OSCAL param → first workbook slot.
    assert rows["{$37$}"].oscal_param_id == "ac-2_prm_1"
    assert rows["{$39$}"].oscal_param_id == "ac-2_prm_2"

    notes = result.notes["odp_assignments"]
    assert notes["oscal_mapped"] == 2
    assert notes["oscal_mapping_abstained"] == 0


# ---------------------------------------------------------------------------
# Idempotent re-apply: no audit-log noise
# ---------------------------------------------------------------------------


def test_reapply_unchanged_workbook_emits_no_audit_rows(
    tmp_path, session: Session, seeded_framework: Framework
):
    """Re-importing the same file twice produces zero OdpAuditLog rows
    AND zero new OdpAssignment rows. The diff path is value-gated so a
    no-op re-open in the UI must not pollute the audit trail."""
    _bust_workbook_cache()
    parameterized = (
        "Requires approvals by {$37$} for requests; "
        "Notifies account managers within {$39$}."
    )
    path = _build_workbook(
        tmp_path / "wb.xlsx",
        assignment_rows=[
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "24 hours",
                "assigned_from": "DoW Enterprise",
            },
        ],
        parameterized=parameterized,
    )

    CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )
    first_count = len(session.exec(select(OdpAssignment)).all())
    assert first_count == 2
    assert session.exec(select(OdpAuditLog)).all() == []

    # Second apply — same file, same byte contents.
    _bust_workbook_cache()
    result2 = CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )

    assert len(session.exec(select(OdpAssignment)).all()) == first_count
    assert session.exec(select(OdpAuditLog)).all() == []
    notes = result2.notes["odp_assignments"]
    assert notes["inserted"] == 0
    assert notes["updated"] == 0


# ---------------------------------------------------------------------------
# Diff fires on a real value change
# ---------------------------------------------------------------------------


def test_value_change_emits_single_audit_row(
    tmp_path, session: Session, seeded_framework: Framework
):
    """Edit one cell, re-ingest. Exactly one OdpAuditLog row appears,
    with prev_value=old, new_value=new, who tagging the workbook
    filename. The OdpAssignment row is mutated in place (still 2 rows
    total) and its ingested_at timestamp advances."""
    _bust_workbook_cache()
    parameterized = (
        "Requires approvals by {$37$} for requests; "
        "Notifies account managers within {$39$}."
    )
    path = _build_workbook(
        tmp_path / "wb.xlsx",
        assignment_rows=[
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "24 hours",
                "assigned_from": "DoW Enterprise",
            },
        ],
        parameterized=parameterized,
    )

    CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )
    pre_ingested_at = {
        r.odp_id: r.ingested_at
        for r in session.exec(select(OdpAssignment)).all()
    }

    # Mutate {$39$} in place — load existing file, edit cell, re-save.
    wb = load_workbook(path)
    av = wb["Assignment Values"]
    # Row 3 (header=1, first data=2, second data=3) is the {$39$} row.
    assert av.cell(row=3, column=2).value == 39
    av.cell(row=3, column=3, value="1 hour")
    wb.save(path)

    _bust_workbook_cache()
    result2 = CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )

    # Still two rows total — no duplicate insert.
    rows = {r.odp_id: r for r in session.exec(select(OdpAssignment)).all()}
    assert len(rows) == 2
    assert rows["{$39$}"].value == "1 hour"
    assert rows["{$37$}"].value == "ISSM or ISSO"  # untouched
    # Mutated row's ingested_at advances; untouched row's stays put.
    assert rows["{$39$}"].ingested_at >= pre_ingested_at["{$39$}"]
    assert rows["{$37$}"].ingested_at == pre_ingested_at["{$37$}"]

    # Exactly one audit row.
    audit = session.exec(select(OdpAuditLog)).all()
    assert len(audit) == 1
    a = audit[0]
    assert a.framework_version == FW_ID
    assert a.control_id == "ac-2"
    assert a.odp_id == "{$39$}"
    assert a.assigned_from == "DoW Enterprise"
    assert a.prev_value == "24 hours"
    assert a.new_value == "1 hour"
    assert a.who == f"CCIS-workbook-ingest:{path.name}"

    notes = result2.notes["odp_assignments"]
    assert notes["inserted"] == 0
    assert notes["updated"] == 1


# ---------------------------------------------------------------------------
# Orphan tracking — value row whose odp_id is absent from the slot list
# ---------------------------------------------------------------------------


def test_orphan_value_row_surfaces_in_notes_without_bridge(
    tmp_path, session: Session, seeded_framework: Framework
):
    """The parameterized statement column declares slots {$37$} and
    {$39$}, but the workbook also carries an extra value-bearing row
    for {$99$} that's NOT in the slot list. The orphan must:

      * still land as an OdpAssignment row (precision over recall —
        we store what the workbook said, we just can't bridge it)
      * have oscal_param_id=None (no positional target)
      * be counted in ``value_rows_without_slot`` so the result notes
        warn the assessor about workbook drift

    Matches the Example System behavior observed on AC-7, SI-5, CM-3,
    SA-19, SI-3 in the May 2026 workbook.
    """
    _bust_workbook_cache()
    # Parameterized only lists two slots; {$99$} is intentionally absent.
    parameterized = (
        "Requires approvals by {$37$} for requests; "
        "Notifies account managers within {$39$}."
    )
    path = _build_workbook(
        tmp_path / "wb.xlsx",
        assignment_rows=[
            {
                "control_id": "AC-2",
                "odp_id": 37,
                "value": "ISSM or ISSO",
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AC-2",
                "odp_id": 39,
                "value": "24 hours",
                "assigned_from": "DoW Enterprise",
            },
            {
                "control_id": "AC-2",
                "odp_id": 99,
                "value": "drift-value",
                "assigned_from": "DoW Enterprise",
            },
        ],
        parameterized=parameterized,
    )

    result = CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )

    rows = {r.odp_id: r for r in session.exec(select(OdpAssignment)).all()}
    assert set(rows.keys()) == {"{$37$}", "{$39$}", "{$99$}"}
    # In-slot rows get the bridge stamped.
    assert rows["{$37$}"].oscal_param_id == "ac-2_prm_1"
    assert rows["{$39$}"].oscal_param_id == "ac-2_prm_2"
    # Orphan row stays NULL — no slot to align to.
    assert rows["{$99$}"].oscal_param_id is None

    notes = result.notes["odp_assignments"]
    assert notes["value_rows_without_slot"] == 1
    assert "ac-2" in notes["controls_with_orphan_values"]


# ---------------------------------------------------------------------------
# No Assignment Values tab → Step 6 is a no-op, apply still succeeds
# ---------------------------------------------------------------------------


def test_apply_without_assignment_values_tab_is_a_noop(
    tmp_path, session: Session, seeded_framework: Framework
):
    """Older eMASS exports omit the tab. The reader returns empty, Step
    6/6b short-circuits, and the rest of apply() proceeds normally —
    no exception, no rows, no notes pollution."""
    _bust_workbook_cache()
    path = _build_workbook(
        tmp_path / "wb.xlsx",
        assignment_rows=[],
        parameterized="",
        include_assignment_tab=False,
    )

    result = CcisWorkbookBaselineSource(path).apply(
        session, framework_id=seeded_framework.id  # type: ignore[arg-type]
    )

    assert session.exec(select(OdpAssignment)).all() == []
    assert session.exec(select(OdpAuditLog)).all() == []
    notes = result.notes["odp_assignments"]
    assert notes["inserted"] == 0
    assert notes["updated"] == 0
    assert notes["rows_parsed"] == 0
