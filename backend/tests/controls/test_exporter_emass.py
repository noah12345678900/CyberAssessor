"""eMASS exporter tests — HEADLESS (openpyxl, no Excel/COM).

The exporter was rewritten from xlwings/COM to openpyxl after the COM path
(a) FROZE on the second export (sync route on FastAPI's threadpool, no COM
init, 2nd xw.App() hangs in the long-lived sidecar) and (b) wrote controls
top-down from row 2, OVERWRITING the template's pre-populated per-acronym
formula rows. These tests pin the new contract:

  - Header detection works on the canonical 1-row header template.
  - Each control is matched to its EXISTING row by acronym (the template
    ships one pre-populated row per control); rows are never repositioned.
  - NO SILENT SKIP: every in-scope control with a template row is written;
    a needs_review control gets the "Needs Review" status, not dropped.
  - Re-export is STABLE — same control → same row, and a missing Status
    column is appended ONCE (never duplicated on the second run).
  - A control with no matching template row is reported in ``skipped``.
  - Workbook.exported_at is stamped on success.

No ``requires_excel`` marker — the path is pure openpyxl now and runs in CI.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook as PyXlWorkbook
from openpyxl import load_workbook

from cybersecurity_assessor.controls.exporter import export_controls_to_emass
from cybersecurity_assessor.models import ComplianceStatus, Workbook


def _make_template(path, sheet="Controls", acronyms=("AC-2", "AC-3", "AC-4", "AC-5")):
    """Template with a header row + one PRE-POPULATED row per control.

    The real enterprise-services template ships pre-populated with one row
    per control (each carrying its own formulas); the exporter matches into
    those rows by acronym. We replicate that shape so the row-match path is
    exercised. A Status column is included by default.
    """
    pwb = PyXlWorkbook()
    ws = pwb.active
    ws.title = sheet
    ws.cell(1, 1, "Control Acronym")
    ws.cell(1, 2, "Status")
    ws.cell(1, 3, "Implementation Narrative")
    for i, acr in enumerate(acronyms, start=2):
        ws.cell(i, 1, acr)
    pwb.save(str(path))


def _seed_assessments(assess, ctx):
    wb_id = ctx["workbook"].id
    objs = ctx["objectives"]
    # AC-2: single-bucket Compliant.
    assess(wb_id, objs["CCI-000015"].id, ComplianceStatus.COMPLIANT,
           narrative="Compliant per system inventory.")
    assess(wb_id, objs["CCI-000016"].id, ComplianceStatus.COMPLIANT,
           narrative="Compliant per system inventory.")
    # AC-3: mixed → multi-line rollup.
    assess(wb_id, objs["CCI-000213"].id, ComplianceStatus.COMPLIANT,
           narrative="Reviewed quarterly per SOP.")
    assess(wb_id, objs["CCI-000214"].id, ComplianceStatus.NON_COMPLIANT,
           narrative="No documented quarterly review cadence.")
    # AC-5: needs_review → written as "Needs Review" (NOT skipped).
    assess(wb_id, objs["CCI-000038"].id, ComplianceStatus.COMPLIANT,
           narrative="Dual-pass disagreed.",
           needs_review=True)


def _status_col(ws):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and "status" in str(v).lower():
            return c
    return None


def _row_for(ws, acronym):
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == acronym:
            return r
    return None


class TestEmassExport:
    def test_writes_every_control_no_skip(
        self, session, controls_catalog, assess, tmp_path
    ):
        _seed_assessments(assess, controls_catalog)
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        _make_template(tpl)

        result = export_controls_to_emass(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            template_path=str(tpl),
            output_path=str(out),
        )

        # NO SKIP: all four in-scope controls written, including AC-5 (needs_review).
        assert result.rows_written == 4
        assert result.skipped == []

        pwb = load_workbook(str(out))
        ws = pwb["Controls"]
        sc = _status_col(ws)

        # AC-3 multi-line rollup (Compliant + Non-Compliant).
        ac3_status = ws.cell(_row_for(ws, "AC-3"), sc).value or ""
        assert "Compliant:" in ac3_status
        assert "Non-Compliant:" in ac3_status

        # AC-5 written as Needs Review — present, not dropped.
        ac5_status = ws.cell(_row_for(ws, "AC-5"), sc).value or ""
        assert "Needs Review" in ac5_status

        # AC-4 inherited-via-CRM rolls up to Compliant.
        ac4_status = ws.cell(_row_for(ws, "AC-4"), sc).value or ""
        assert "Compliant" in ac4_status

    def test_matches_existing_rows_does_not_reposition(
        self, session, controls_catalog, assess, tmp_path
    ):
        """Controls land in their PRE-EXISTING template rows (matched by
        acronym), not top-down from row 2."""
        _seed_assessments(assess, controls_catalog)
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        # Deliberately non-alphabetical template order to prove we match by
        # acronym, not by position.
        _make_template(tpl, acronyms=("AC-5", "AC-4", "AC-3", "AC-2"))

        export_controls_to_emass(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            template_path=str(tpl),
            output_path=str(out),
        )
        ws = load_workbook(str(out))["Controls"]
        # AC-2 stays on its template row (row 5), AC-5 on row 2.
        assert ws.cell(2, 1).value == "AC-5"
        assert ws.cell(5, 1).value == "AC-2"
        sc = _status_col(ws)
        assert (ws.cell(5, sc).value or "") == "Compliant"  # AC-2 single bucket

    def test_reexport_is_stable_no_duplicate_columns(
        self, session, controls_catalog, assess, tmp_path
    ):
        """Re-export onto the previous output keeps rows stable and does not
        append a second Status column (the duplicate-column regression)."""
        _seed_assessments(assess, controls_catalog)
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        # Template WITHOUT a Status column → exporter appends one; re-export
        # must reuse it, not append a second.
        pwb = PyXlWorkbook()
        ws0 = pwb.active
        ws0.title = "Controls"
        ws0.cell(1, 1, "Control Acronym")
        for i, acr in enumerate(("AC-2", "AC-3", "AC-4", "AC-5"), start=2):
            ws0.cell(i, 1, acr)
        pwb.save(str(tpl))

        export_controls_to_emass(
            session=session, workbook_id=controls_catalog["workbook"].id,
            template_path=str(tpl), output_path=str(out),
        )
        export_controls_to_emass(
            session=session, workbook_id=controls_catalog["workbook"].id,
            template_path=str(out), output_path=str(out),
        )
        ws = load_workbook(str(out))["Controls"]
        status_cols = sum(
            1 for c in range(1, ws.max_column + 1)
            if (ws.cell(1, c).value or "").strip().lower() == "implementation status"
        )
        assert status_cols == 1

    def test_control_missing_from_template_is_reported(
        self, session, controls_catalog, assess, tmp_path
    ):
        """An in-scope control with no template row is reported in skipped,
        not silently dropped and not appended as a formula-less row."""
        _seed_assessments(assess, controls_catalog)
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        # Template omits AC-5's row.
        _make_template(tpl, acronyms=("AC-2", "AC-3", "AC-4"))

        result = export_controls_to_emass(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            template_path=str(tpl),
            output_path=str(out),
        )
        assert result.rows_written == 3
        assert [a for a, _ in result.skipped] == ["AC-5"]

    def test_appends_status_column_when_template_lacks_one(
        self, session, controls_catalog, assess, tmp_path
    ):
        _seed_assessments(assess, controls_catalog)
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        pwb = PyXlWorkbook()
        ws0 = pwb.active
        ws0.title = "Controls"
        ws0.cell(1, 1, "Control Acronym")
        for i, acr in enumerate(("AC-2", "AC-3", "AC-4", "AC-5"), start=2):
            ws0.cell(i, 1, acr)
        pwb.save(str(tpl))

        result = export_controls_to_emass(
            session=session, workbook_id=controls_catalog["workbook"].id,
            template_path=str(tpl), output_path=str(out),
        )
        assert any("Status column" in w for w in result.template_warnings)
        ws = load_workbook(str(out))["Controls"]
        assert _status_col(ws) is not None

    def test_stamps_workbook_exported_at(
        self, session, controls_catalog, assess, tmp_path
    ):
        _seed_assessments(assess, controls_catalog)
        wb_id = controls_catalog["workbook"].id
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        _make_template(tpl)

        assert session.get(Workbook, wb_id).exported_at is None
        export_controls_to_emass(
            session=session, workbook_id=wb_id,
            template_path=str(tpl), output_path=str(out),
        )
        session.expire_all()
        assert session.get(Workbook, wb_id).exported_at is not None

    def test_missing_template_raises(self, session, controls_catalog, tmp_path):
        with pytest.raises(FileNotFoundError):
            export_controls_to_emass(
                session=session,
                workbook_id=controls_catalog["workbook"].id,
                template_path=str(tmp_path / "nope.xlsx"),
                output_path=str(tmp_path / "out.xlsx"),
            )

    def test_missing_sheet_raises_value_error(
        self, session, controls_catalog, tmp_path
    ):
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        pwb = PyXlWorkbook()
        pwb.active.title = "WrongSheet"
        pwb.save(str(tpl))

        with pytest.raises(ValueError, match="Controls"):
            export_controls_to_emass(
                session=session,
                workbook_id=controls_catalog["workbook"].id,
                template_path=str(tpl),
                output_path=str(out),
            )

    def test_missing_acronym_header_raises_value_error(
        self, session, controls_catalog, tmp_path
    ):
        tpl = tmp_path / "tpl.xlsx"
        out = tmp_path / "out.xlsx"
        pwb = PyXlWorkbook()
        ws = pwb.active
        ws.title = "Controls"
        ws.cell(1, 1, "Unrelated")
        ws.cell(1, 2, "Other")
        pwb.save(str(tpl))

        with pytest.raises(ValueError, match="Control Acronym"):
            export_controls_to_emass(
                session=session,
                workbook_id=controls_catalog["workbook"].id,
                template_path=str(tpl),
                output_path=str(out),
            )
