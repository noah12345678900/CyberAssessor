"""Endpoint tests for /api/controls/export/{emass,working}.

Both endpoints are headless (openpyxl) — the eMASS path was migrated off
xlwings/COM, so the happy-path tests run in CI without Excel.
"""

from __future__ import annotations

import pytest
from fastapi.testclient import TestClient

from cybersecurity_assessor.db import get_session
from cybersecurity_assessor.models import ComplianceStatus
from cybersecurity_assessor.server import create_app


@pytest.fixture
def client(session):
    """Production FastAPI app with our in-memory session swapped in via
    ``dependency_overrides`` — same pattern as the CRM suspicion tests
    so the URL prefix + route decorators reflect what ships."""
    app = create_app()
    app.dependency_overrides[get_session] = lambda: session
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


# ---------------------------------------------------------------------------
# /export/working — openpyxl, no Excel COM required
# ---------------------------------------------------------------------------


class TestWorkingEndpoint:
    def test_happy_path_returns_dto(
        self, client, controls_catalog, assess, tmp_path
    ):
        wb_id = controls_catalog["workbook"].id
        objs = controls_catalog["objectives"]
        assess(wb_id, objs["CCI-000015"].id, ComplianceStatus.COMPLIANT)
        assess(wb_id, objs["CCI-000213"].id, ComplianceStatus.NON_COMPLIANT)

        out_path = tmp_path / "working_endpoint.xlsx"
        resp = client.post(
            "/api/controls/export/working",
            json={
                "workbook_id": wb_id,
                "output_path": str(out_path),
            },
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["output_path"] == str(out_path)
        assert body["rows_written"] == 6
        assert body["controls_with_psc"] == 2
        assert body["skipped"] == []
        assert out_path.exists()

    def test_family_filter_passed_through(
        self, client, controls_catalog, tmp_path
    ):
        wb_id = controls_catalog["workbook"].id
        out_path = tmp_path / "fam.xlsx"

        # All seeded controls are AC; "AU" filter must zero them out.
        resp = client.post(
            "/api/controls/export/working",
            json={
                "workbook_id": wb_id,
                "output_path": str(out_path),
                "family": "AU",
            },
        )
        assert resp.status_code == 200
        assert resp.json()["rows_written"] == 0

    def test_missing_workbook_404(self, client, tmp_path):
        resp = client.post(
            "/api/controls/export/working",
            json={
                "workbook_id": 999_999,
                "output_path": str(tmp_path / "x.xlsx"),
            },
        )
        # ValueError with 'not found' → 404 per routes/controls.py mapper.
        assert resp.status_code == 404

    def test_no_baseline_422(self, client, session, tmp_path):
        """Workbook exists but has no baseline → ValueError without
        'not found' → 422 per the mapper."""
        from cybersecurity_assessor.models import Workbook

        p = tmp_path / "no-bl.xlsx"
        p.write_bytes(b"x")
        wb = Workbook(path=str(p), filename=p.name)
        session.add(wb)
        session.commit()
        session.refresh(wb)

        resp = client.post(
            "/api/controls/export/working",
            json={
                "workbook_id": wb.id,
                "output_path": str(tmp_path / "out.xlsx"),
            },
        )
        assert resp.status_code == 422


# ---------------------------------------------------------------------------
# /export/emass — validation paths (no Excel needed)
# ---------------------------------------------------------------------------


class TestEmassEndpointValidation:
    def test_missing_template_410(self, client, controls_catalog, tmp_path):
        """A bogus template_path bails with FileNotFoundError —
        mapped to 410 (Gone) per the routes layer."""
        resp = client.post(
            "/api/controls/export/emass",
            json={
                "workbook_id": controls_catalog["workbook"].id,
                "template_path": str(tmp_path / "does-not-exist.xlsx"),
                "output_path": str(tmp_path / "out.xlsx"),
            },
        )
        assert resp.status_code == 410

    def test_missing_workbook_404(self, client, tmp_path):
        # Even with a missing workbook we have to point template_path at
        # *something*; the workbook lookup runs first.
        tpl = tmp_path / "tpl.xlsx"
        tpl.write_bytes(b"x")
        resp = client.post(
            "/api/controls/export/emass",
            json={
                "workbook_id": 999_999,
                "template_path": str(tpl),
                "output_path": str(tmp_path / "out.xlsx"),
            },
        )
        assert resp.status_code == 404

    def test_missing_body_field_422(self, client, tmp_path):
        """FastAPI rejects body missing required fields with 422 before
        the route handler runs."""
        resp = client.post(
            "/api/controls/export/emass",
            json={"output_path": str(tmp_path / "out.xlsx")},
        )
        assert resp.status_code == 422


# ---------------------------------------------------------------------------
# /export/emass — full path (headless openpyxl, runs in CI)
# ---------------------------------------------------------------------------


class TestEmassEndpointHappyPath:
    def test_returns_dto_and_stamps_exported_at(
        self, client, session, controls_catalog, assess, tmp_path
    ):
        """The eMASS endpoint writes the xlsx headlessly and stamps
        Workbook.exported_at so the UI's "Last exported" badge refreshes."""
        from cybersecurity_assessor.models import Workbook
        from openpyxl import Workbook as PyXlWorkbook

        wb_id = controls_catalog["workbook"].id
        objs = controls_catalog["objectives"]
        assess(wb_id, objs["CCI-000015"].id, ComplianceStatus.COMPLIANT)
        assess(wb_id, objs["CCI-000016"].id, ComplianceStatus.COMPLIANT)

        # Template pre-populated with one row per control (the shape the
        # row-match exporter expects), header at row 1.
        tpl = tmp_path / "tpl.xlsx"
        pwb = PyXlWorkbook()
        ws = pwb.active
        ws.title = "Controls"
        ws.cell(1, 1, "Control Acronym")
        ws.cell(1, 2, "Status")
        ws.cell(1, 3, "Narrative")
        for i, acr in enumerate(("AC-2", "AC-3", "AC-4", "AC-5"), start=2):
            ws.cell(i, 1, acr)
        pwb.save(str(tpl))

        out_path = tmp_path / "emass_out.xlsx"
        resp = client.post(
            "/api/controls/export/emass",
            json={
                "workbook_id": wb_id,
                "template_path": str(tpl),
                "output_path": str(out_path),
            },
        )
        assert resp.status_code == 200, resp.text
        body = resp.json()
        # All 4 in-scope controls land (no silent skip), matched to their rows.
        assert body["rows_written"] == 4
        assert body["skipped"] == []
        assert out_path.exists()

        # exported_at stamped.
        session.expire_all()
        fresh = session.get(Workbook, wb_id)
        assert fresh.exported_at is not None
