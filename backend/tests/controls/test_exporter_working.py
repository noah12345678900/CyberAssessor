"""Working-view export tests — openpyxl path, no Excel COM required.

The working-view export is the assessor's personal triage workbook:
one row per objective (not per control), needs_review surfaced as a
column, PSC mappings rendered the same way as the eMASS export. These
tests pin:

  - One row per objective (not per control).
  - Family / status / search filters honored.
  - needs_review rows INCLUDED (unlike eMASS export, which excludes them).
  - PSC column rendered per control with source-prefixed lines.
"""

from __future__ import annotations

from openpyxl import load_workbook

from cybersecurity_assessor.controls.exporter import (
    ControlsFilterState,
    export_controls_working_view,
)
from cybersecurity_assessor.models import ComplianceStatus


def _seed_assessments(assess, ctx):
    """Populate AC-2 (compliant), AC-3 (mixed), AC-5 (needs_review).

    AC-4 is left without assessments so the CRM short-circuit drives its
    status — that's the realistic "the LLM never had to look at this"
    inherited-control case.
    """
    wb_id = ctx["workbook"].id
    objs = ctx["objectives"]

    # AC-2: both CCIs compliant.
    assess(wb_id, objs["CCI-000015"].id, ComplianceStatus.COMPLIANT,
           narrative="AC-2 CCI-000015 compliant narrative.")
    assess(wb_id, objs["CCI-000016"].id, ComplianceStatus.COMPLIANT,
           narrative="AC-2 CCI-000016 compliant narrative.")

    # AC-3: one compliant + one NC → mixed rollup.
    assess(wb_id, objs["CCI-000213"].id, ComplianceStatus.COMPLIANT,
           narrative="AC-3 CCI-000213 compliant.")
    assess(wb_id, objs["CCI-000214"].id, ComplianceStatus.NON_COMPLIANT,
           narrative="AC-3 CCI-000214 gap on quarterly review cadence.")

    # AC-5: needs_review — surfaced in working view, excluded from eMASS.
    assess(wb_id, objs["CCI-000038"].id, ComplianceStatus.COMPLIANT,
           narrative="Dual-pass disagreed, awaiting tester.",
           needs_review=True,
           review_reason="Status confidence below threshold.")


class TestRowExpansion:
    def test_one_row_per_objective_not_per_control(
        self, session, controls_catalog, assess, tmp_path
    ):
        _seed_assessments(assess, controls_catalog)
        out_path = tmp_path / "working.xlsx"

        result = export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
        )

        # 4 controls, total 6 objectives (AC-2:2, AC-3:2, AC-4:1, AC-5:1).
        assert result.rows_written == 6
        assert out_path.exists()

        py_wb = load_workbook(str(out_path))
        ws = py_wb.active
        # Header row + 6 data rows.
        assert ws.max_row == 7

    def test_needs_review_rows_included_with_flag(
        self, session, controls_catalog, assess, tmp_path
    ):
        """Unlike the eMASS export (precision-over-recall gate), the
        working view surfaces needs_review rows so the assessor can
        triage them in Excel."""
        _seed_assessments(assess, controls_catalog)
        out_path = tmp_path / "working_nr.xlsx"

        export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
        )

        py_wb = load_workbook(str(out_path))
        ws = py_wb.active
        headers = [c.value for c in ws[1]]
        nr_idx = headers.index("Needs Review")
        cci_idx = headers.index("CCI")

        # Locate the AC-5 / CCI-000038 row.
        nr_rows = [
            row for row in ws.iter_rows(min_row=2, values_only=True)
            if row[cci_idx] == "CCI-000038"
        ]
        assert len(nr_rows) == 1
        assert nr_rows[0][nr_idx] == "Yes"


class TestFilters:
    def test_family_filter_only_emits_matching_controls(
        self, session, controls_catalog, assess, tmp_path
    ):
        """All seeded controls are family=AC so a non-AC filter must
        drop everything — empty data set, header row still present."""
        _seed_assessments(assess, controls_catalog)
        out_path = tmp_path / "fam_none.xlsx"

        result = export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
            filter_state=ControlsFilterState(family="AU"),
        )

        assert result.rows_written == 0
        py_wb = load_workbook(str(out_path))
        ws = py_wb.active
        assert ws.max_row == 1  # header only

    def test_family_filter_ac_emits_all(
        self, session, controls_catalog, assess, tmp_path
    ):
        _seed_assessments(assess, controls_catalog)
        out_path = tmp_path / "fam_ac.xlsx"

        result = export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
            filter_state=ControlsFilterState(family="AC"),
        )

        # Family filter is case-insensitive in the loader (upper()).
        assert result.rows_written == 6

    def test_search_filter_substring_on_control_id(
        self, session, controls_catalog, assess, tmp_path
    ):
        """search='AC-2' must match only AC-2 control, leaving 2 rows."""
        _seed_assessments(assess, controls_catalog)
        out_path = tmp_path / "search.xlsx"

        result = export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
            filter_state=ControlsFilterState(search="AC-2"),
        )

        # AC-2 has two objectives.
        assert result.rows_written == 2

    def test_status_filter_drops_non_matching_objectives(
        self, session, controls_catalog, assess, tmp_path
    ):
        """status='Non-Compliant' keeps only the AC-3 CCI-000214 row.

        Note: the AC-4 inherited row has no Assessment, so it carries
        the loader's default NOT_APPLICABLE — it doesn't match the NC
        filter either. The status filter compares raw ComplianceStatus
        value, not the CRM-classified bucket from _rollup_status."""
        _seed_assessments(assess, controls_catalog)
        out_path = tmp_path / "status.xlsx"

        result = export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
            filter_state=ControlsFilterState(status="Non-Compliant"),
        )

        assert result.rows_written == 1
        py_wb = load_workbook(str(out_path))
        ws = py_wb.active
        headers = [c.value for c in ws[1]]
        cci_idx = headers.index("CCI")
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert data_rows[0][cci_idx] == "CCI-000214"


class TestPscColumn:
    def test_psc_text_appears_for_controls_with_overlays(
        self, session, controls_catalog, assess, tmp_path
    ):
        """AC-2 has two PSC mappings (SDA-127, T1TL-031); AC-3 has one
        (SDA-201). AC-4 and AC-5 have none."""
        _seed_assessments(assess, controls_catalog)
        out_path = tmp_path / "psc.xlsx"

        result = export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
        )

        # controls_with_psc counts unique controls, not rows.
        assert result.controls_with_psc == 2

        py_wb = load_workbook(str(out_path))
        ws = py_wb.active
        headers = [c.value for c in ws[1]]
        ctl_idx = headers.index("Control")
        psc_idx = headers.index("Program-Specific Controls")

        psc_by_control: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            psc_by_control[row[ctl_idx]] = row[psc_idx]

        # AC-2: source-prefixed lines from both SDA and T1TL, sorted by
        # source name then requirement_number.
        ac2 = psc_by_control["AC-2"] or ""
        assert "SDA-127:" in ac2
        assert "T1TL-031:" in ac2
        # SDA sorts before T1TL alphabetically.
        assert ac2.index("SDA-127:") < ac2.index("T1TL-031:")

        # AC-3: just SDA-201.
        ac3 = psc_by_control["AC-3"] or ""
        assert "SDA-201:" in ac3

        # AC-4, AC-5: no PSC overlays.
        assert not psc_by_control["AC-4"]
        assert not psc_by_control["AC-5"]


class TestHeaderContract:
    def test_header_row_columns(
        self, session, controls_catalog, tmp_path
    ):
        """Header row pinned — UI/CSV consumers index by column name."""
        out_path = tmp_path / "headers.xlsx"

        export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
        )

        py_wb = load_workbook(str(out_path))
        ws = py_wb.active
        headers = [c.value for c in ws[1]]
        assert headers == [
            "Control",
            "Title",
            "Family",
            "Program-Specific Controls",
            "CCI",
            "Status",
            "Needs Review",
            "Narrative",
            "Narrative (On-Prem)",
            "Narrative (Cloud)",
            "Inheritance Rule",
            "Confidence",
            "CRM Responsibility (Cloud)",
            "CRM Responsibility (On-Prem)",
        ]

    def test_sheet_title(self, session, controls_catalog, tmp_path):
        out_path = tmp_path / "title.xlsx"
        export_controls_working_view(
            session=session,
            workbook_id=controls_catalog["workbook"].id,
            output_path=str(out_path),
        )
        py_wb = load_workbook(str(out_path))
        assert py_wb.active.title == "Controls (Working View)"


class TestErrorPaths:
    def test_missing_workbook_raises_value_error(self, session, tmp_path):
        import pytest
        with pytest.raises(ValueError, match="not found"):
            export_controls_working_view(
                session=session,
                workbook_id=999_999,
                output_path=str(tmp_path / "x.xlsx"),
            )

    def test_no_baseline_raises_value_error(self, session, tmp_path):
        """A workbook without a baseline_id can't materialize an in-scope
        set — the loader bails before touching xlsx output."""
        import pytest
        from cybersecurity_assessor.models import Workbook

        p = tmp_path / "no-baseline.xlsx"
        p.write_bytes(b"x")
        wb = Workbook(path=str(p), filename=p.name)
        session.add(wb)
        session.commit()
        session.refresh(wb)

        with pytest.raises(ValueError, match="Baseline"):
            export_controls_working_view(
                session=session,
                workbook_id=wb.id,
                output_path=str(tmp_path / "out.xlsx"),
            )
