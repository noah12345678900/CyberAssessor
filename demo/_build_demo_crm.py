"""Generate a realistic AWS GovCloud FedRAMP High CRM xlsx for smoke testing.

Mimics the real AWS CRM column layout (Control ID, Control Name,
Responsibility, AWS Implementation, Customer Responsibility, Implementation
Status). The loader (baselines/crm_xlsx.py) only reads Control ID +
Responsibility + Customer Responsibility -- the other columns are cosmetic
realism so the file looks like something AWS would actually publish.

Coverage targets every code path the CRM short-circuit / hybrid block
exercises:
  - inherited      -> COMPLIANT short-circuit       (PE -- AWS-owned physical)
  - shared         -> prompt enrichment             (AC/AU/CM/IA/SC/SI/MA/...)
                      (loader maps "shared" -> hybrid internally)
  - customer       -> full LLM path                 (AT/PL/PM/RA -- pure policy)

TEST-DATA DIVERGENCES from a real AWS CRM (intentional, for path coverage --
remove when the real Monday CRM arrives):
  - provider       -> NOT_APPLICABLE short-circuit  (PE-4, PE-5 marked
                      "Service Provider"). Real AWS calls these "Inherited";
                      we keep "Service Provider" so the assessor's
                      provider->NOT_APPLICABLE branch gets exercised.
  - not_applicable -> NOT_APPLICABLE short-circuit  (PE-10, PE-11 marked
                      "Not Applicable"). Real AWS marks these "Inherited"
                      because FedRAMP High mandates them; we keep "Not
                      Applicable" so the assessor's not_applicable branch
                      gets exercised.

Gemini-3-Pro review applied (2026-06-04):
  - "Hybrid" -> "Shared" (real FedRAMP terminology; loader still routes it
    to the hybrid prompt-enrichment path)
  - MA-2/3/4/5 moved Inherited -> Shared (customer owns guest-OS maintenance)
  - Added -1 controls (AC-1, AU-1, SC-1) and IR-8, CM-9 as Shared
  - Densified narratives: explicit boundary demarcation, mandate citations
  - SI-2 narrative now cites CISA KEV BOD 22-01 15-day SLA
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(__file__).parent / "crm" / "AWS_GovCloud_FedRAMP_High_CRM.xlsx"

# (control_id, control_name, responsibility, aws_impl, customer_resp, impl_status)
ROWS = [
    # === INHERITED (fully AWS-owned, customer just records inheritance) =====
    ("PE-1",  "Physical and Environmental Protection Policy and Procedures",
     "Inherited",
     "AWS develops, documents, disseminates, reviews, and updates physical and "
     "environmental protection policy and procedures for all AWS GovCloud (US) "
     "datacenter facilities. Policy reviewed annually and following any "
     "significant change to the physical environment. Coverage extends to all "
     "AWS-owned, AWS-leased, and AWS-colocated facilities supporting GovCloud.",
     "Customer fully inherits the AWS physical and environmental protection "
     "policy and procedures for all customer workloads hosted in AWS GovCloud "
     "(US). No customer action required for the underlying infrastructure "
     "layer. Customer SSP must reference the AWS FedRAMP High SSP (available "
     "via AWS Artifact) as the authoritative source for the inherited control.",
     "Implemented"),
    ("PE-2",  "Physical Access Authorizations",
     "Inherited",
     "AWS develops, approves, and maintains a list of personnel with authorized "
     "access to AWS datacenter facilities. Access is reviewed and re-approved "
     "every 90 days by AWS Security personnel. Access is removed within 24 "
     "hours of personnel termination or change in role.",
     "Customer fully inherits AWS physical access authorization controls for "
     "datacenter facilities. Customer personnel are not granted physical access "
     "to AWS GovCloud datacenters. No customer-side implementation required.",
     "Implemented"),
    ("PE-3",  "Physical Access Control",
     "Inherited",
     "AWS enforces physical access controls at AWS datacenter facilities via "
     "(1) multi-factor authentication at all perimeter ingress points, (2) "
     "biometric authentication at server-room boundaries, (3) mantrap "
     "vestibules, (4) 24x7 on-site professional security personnel, and (5) "
     "visitor escort requirements. Access logs retained per AU-11.",
     "Customer fully inherits all AWS physical access controls. Customer "
     "personnel are not granted physical access to AWS datacenter facilities "
     "and have no compensating-control responsibility.",
     "Implemented"),
    ("PE-6",  "Monitoring Physical Access",
     "Inherited",
     "AWS monitors physical access to AWS datacenter facilities via (1) CCTV "
     "coverage of all access points with minimum 90-day retention, (2) "
     "intrusion detection sensors at all perimeter and internal access "
     "boundaries, (3) continuous review by AWS Security Operations Center, and "
     "(4) automated alerting on anomalous access patterns.",
     "Customer fully inherits AWS physical access monitoring. Customer may "
     "review the AWS SOC 2 Type II and FedRAMP High audit reports via AWS "
     "Artifact for assurance evidence.",
     "Implemented"),
    ("PE-12", "Emergency Lighting",
     "Inherited",
     "AWS datacenters provide and maintain emergency lighting throughout "
     "facility egress paths and operational areas. Emergency lighting is "
     "powered by UPS-backed circuits with monthly inspection and testing per "
     "NFPA 101 Life Safety Code.",
     "Customer fully inherits AWS datacenter emergency lighting controls. No "
     "customer-side implementation responsibility.",
     "Implemented"),
    ("PE-13", "Fire Protection",
     "Inherited",
     "AWS datacenters deploy (1) VESDA (Very Early Smoke Detection Apparatus) "
     "for early-warning detection, (2) pre-action dry-pipe fire suppression "
     "systems, (3) handheld extinguishers per NFPA 10, and (4) quarterly "
     "inspection and annual recertification of all suppression equipment.",
     "Customer fully inherits AWS datacenter fire protection. Customer-managed "
     "AMIs, EBS volumes, and S3 data are protected against facility-level fire "
     "events through customer-implemented multi-AZ / multi-region replication "
     "per CP-9.",
     "Implemented"),
    ("PE-14", "Environmental Controls",
     "Inherited",
     "AWS maintains datacenter temperature and humidity within ASHRAE TC 9.9 "
     "recommended ranges via N+1 redundant HVAC systems with continuous "
     "monitoring and automated failover. Environmental parameters are logged "
     "and reviewed by AWS facility operations.",
     "Customer fully inherits AWS datacenter environmental controls. No "
     "customer-side implementation responsibility.",
     "Implemented"),

    # === SERVICE PROVIDER (test scaffolding -- exercises provider->NA branch) ===
    # NOTE: A real AWS CRM marks these "Inherited". We keep "Service Provider"
    # only so the assessor's provider->NOT_APPLICABLE short-circuit gets
    # exercised end-to-end. Replace with "Inherited" when the real CRM lands.
    ("PE-4",  "Access Control for Transmission",
     "Service Provider",
     "AWS controls physical access to all information system distribution and "
     "transmission lines within AWS datacenter facilities, including all "
     "intra-facility fiber, copper, and power distribution.",
     "Provider-owned. Customer has no role in physical transmission line access "
     "control within AWS datacenter facilities and no compensating-control "
     "responsibility.",
     "Implemented"),
    ("PE-5",  "Access Control for Output Devices",
     "Service Provider",
     "AWS controls physical access to output devices (printers, displays, "
     "console monitors) within AWS datacenter facilities. AWS does not provide "
     "physical output devices to customers; all customer interaction with the "
     "service occurs via authenticated remote APIs.",
     "Provider-owned. Customer does not maintain physical output devices in "
     "AWS facilities. Not applicable to customer.",
     "Implemented"),

    # === SHARED (loader routes "shared" -> hybrid prompt-enrichment path) ===
    ("AC-1",  "Access Control Policy and Procedures",
     "Shared",
     "AWS develops, documents, and reviews AWS-internal access control policy "
     "and procedures applicable to AWS personnel and AWS-managed infrastructure "
     "annually.",
     "Customer is responsible for developing, documenting, disseminating, and "
     "annually reviewing the customer-side access control policy and "
     "supporting procedures governing customer use of AWS services and "
     "customer-deployed workloads. Customer SSP must document the customer's "
     "AC-1 policy and its alignment with the customer organization's mission.",
     "Implemented"),
    ("AC-2",  "Account Management",
     "Shared",
     "AWS manages the AWS root master account and AWS-internal personnel "
     "accounts per AWS Personnel Security policy. AWS provides IAM, IAM "
     "Identity Center (formerly SSO), Organizations, and Cognito as account "
     "management primitives for customer use.",
     "Customer is responsible for the full lifecycle management of customer-"
     "side IAM users, IAM roles, federated identities, and root account "
     "credentials, including: (a) provisioning per documented request/approval "
     "workflow, (b) deprovisioning within 24 hours of role change or "
     "separation, (c) quarterly access reviews via IAM Access Analyzer and "
     "Access Advisor, (d) automated disabling of IAM users inactive >=90 days, "
     "and (e) MFA enforcement on the root account.",
     "Implemented"),
    ("AC-3",  "Access Enforcement",
     "Shared",
     "AWS enforces access at the AWS service API layer via IAM policy "
     "evaluation; AWS denies all requests not explicitly authorized by an "
     "applicable IAM policy, resource policy, SCP, or session policy.",
     "Customer is responsible for authoring and maintaining IAM identity-based "
     "policies, resource-based policies, Service Control Policies, permission "
     "boundaries, and session policies that enforce least-privilege access to "
     "customer-owned resources. Customer must validate effective permissions "
     "with IAM Policy Simulator and IAM Access Analyzer prior to deployment.",
     "Implemented"),
    ("AC-5",  "Separation of Duties",
     "Shared",
     "AWS enforces separation of duties across AWS personnel roles "
     "(operations, security, audit) per AWS Personnel Security policy. AWS "
     "personnel cannot unilaterally bypass change-management or audit-review "
     "controls.",
     "Customer is responsible for documenting customer-side separation of "
     "duties and enforcing the separation via IAM (distinct roles for "
     "developers, operators, security engineers, auditors), Service Control "
     "Policies that prevent any single principal from both modifying audit "
     "logs and performing the actions being audited, and permission boundaries "
     "that bound delegated IAM administration.",
     "Implemented"),
    ("AC-6",  "Least Privilege",
     "Shared",
     "AWS implements least privilege for AWS personnel and AWS-managed "
     "service-linked roles via narrowly scoped IAM policies.",
     "Customer is responsible for implementing least privilege within customer "
     "AWS accounts: (a) IAM policies scoped to the minimum actions and "
     "resources required, (b) periodic IAM Access Analyzer findings review and "
     "remediation, (c) removal of unused IAM permissions and credentials per "
     "Access Advisor, and (d) just-in-time elevation for privileged operations "
     "via IAM Identity Center.",
     "Implemented"),
    ("AU-1",  "Audit and Accountability Policy and Procedures",
     "Shared",
     "AWS develops and maintains AWS-internal audit and accountability policy "
     "for AWS-managed infrastructure, reviewed annually.",
     "Customer is responsible for developing and maintaining the customer-side "
     "audit and accountability policy and procedures, including the "
     "specification of auditable events (per AU-2), retention duration (per "
     "AU-11), and review cadence (per AU-6) for customer-deployed workloads.",
     "Implemented"),
    ("AU-2",  "Event Logging",
     "Shared",
     "AWS logs all API calls to AWS-internal infrastructure via AWS-internal "
     "audit systems. AWS provides CloudTrail (management events, data events, "
     "Insights events), VPC Flow Logs, ELB access logs, S3 server access "
     "logs, and CloudWatch Logs for customer event capture.",
     "Customer is responsible for (a) enabling CloudTrail in all regions and "
     "accounts including organization-wide multi-region trails, (b) enabling "
     "data-event logging on customer-sensitive S3 buckets and Lambda "
     "functions, (c) configuring CloudWatch Logs for application-level events, "
     "and (d) documenting the customer-defined list of auditable event types "
     "in the customer SSP per the customer AU-1 policy.",
     "Implemented"),
    ("AU-6",  "Audit Record Review, Analysis, and Reporting",
     "Shared",
     "AWS reviews AWS-internal audit logs continuously via AWS Security "
     "Operations Center tooling and reports relevant findings to customers per "
     "the AWS shared responsibility model and IR-6 communications.",
     "Customer is responsible for reviewing customer CloudTrail events, VPC "
     "Flow Logs, GuardDuty findings, Security Hub findings, Macie alerts, and "
     "application-layer audit logs on the customer-defined cadence (minimum "
     "weekly for non-critical, continuous for critical). Findings reportable "
     "per IR-6 must be escalated within customer-defined timeframes.",
     "Implemented"),
    ("AU-12", "Audit Record Generation",
     "Shared",
     "AWS generates audit records for all AWS API calls, management-plane "
     "events, and AWS-internal service interactions per AWS audit policy.",
     "Customer is responsible for enabling audit record generation on all "
     "customer-deployed systems including (a) EC2 guest OS auditd / Windows "
     "Event Log, (b) RDS database audit logs, (c) container runtime logs, and "
     "(d) application-layer logs, ensuring records cover the AU-2 event types "
     "and are forwarded to a tamper-resistant store (CloudWatch Logs / S3 "
     "with Object Lock).",
     "Implemented"),
    ("CM-2",  "Baseline Configuration",
     "Shared",
     "AWS maintains baseline configurations for all AWS-managed infrastructure "
     "and AWS-managed services per AWS configuration management policy.",
     "Customer is responsible for developing, documenting, and maintaining "
     "baseline configurations for customer-managed AMIs, container images, "
     "RDS parameter groups, and Infrastructure-as-Code templates "
     "(CloudFormation, Terraform, CDK). Baselines reviewed and re-approved at "
     "minimum annually and after any significant change. Drift detection via "
     "AWS Config or equivalent.",
     "Implemented"),
    ("CM-6",  "Configuration Settings",
     "Shared",
     "AWS configures AWS-managed services per AWS security baselines aligned "
     "to applicable benchmarks (CIS AWS Foundations, AWS Well-Architected "
     "Security Pillar).",
     "Customer is responsible for (a) hardening customer-managed OS images per "
     "DISA STIG or CIS Benchmarks at the highest applicable level, (b) "
     "enforcing configuration settings continuously via AWS Config managed "
     "rules and custom rules, (c) remediating configuration drift within "
     "customer-defined SLAs, and (d) documenting any approved deviations from "
     "the baseline with compensating controls.",
     "Implemented"),
    ("CM-8",  "System Component Inventory",
     "Shared",
     "AWS maintains a complete inventory of all AWS-managed physical and "
     "virtual infrastructure supporting the GovCloud service offering.",
     "Customer is responsible for maintaining an accurate, up-to-date "
     "inventory of customer-deployed resources via AWS Config, AWS Resource "
     "Groups, Systems Manager Inventory, and tag-based resource organization. "
     "Inventory accuracy reviewed and reconciled at minimum monthly; new "
     "resource types incorporated into inventory within 30 days of first use.",
     "Implemented"),
    ("CM-9",  "Configuration Management Plan",
     "Shared",
     "AWS maintains an AWS-internal configuration management plan governing "
     "changes to AWS-managed infrastructure.",
     "Customer is responsible for developing and maintaining the customer-side "
     "configuration management plan covering customer-deployed resources, "
     "including (a) roles and responsibilities, (b) configuration items under "
     "management, (c) change request and approval workflow, (d) integration "
     "with customer change advisory board, and (e) coordination with AWS-"
     "published change notifications.",
     "Implemented"),
    ("CP-9",  "System Backup",
     "Shared",
     "AWS replicates and backs up all AWS-managed control-plane and service "
     "data per AWS contingency planning policy. AWS guarantees the durability "
     "of S3 standard storage at 99.999999999% (11 nines).",
     "Customer is responsible for backing up customer-owned data with backup "
     "frequency aligned to customer-defined RPO: AMI snapshots, EBS snapshots, "
     "RDS automated backups + cross-region snapshot copies, S3 cross-region "
     "replication, DynamoDB point-in-time recovery, and AWS Backup plans for "
     "centralized policy-based backup. Backup integrity tested per CP-4.",
     "Implemented"),
    ("CP-10", "System Recovery and Reconstitution",
     "Shared",
     "AWS provides multi-AZ and multi-region service availability and service-"
     "level reconstitution capabilities. AWS publishes service-level RTO/RPO "
     "via the AWS Service Health Dashboard and Personal Health Dashboard.",
     "Customer is responsible for (a) designing customer workloads for multi-"
     "AZ deployment by default and multi-region deployment for tier-1 systems, "
     "(b) documenting customer-defined RTO and RPO per workload, (c) "
     "implementing automated failover where supported, and (d) exercising "
     "recovery procedures at minimum annually with documented results.",
     "Implemented"),
    ("IA-2",  "Identification and Authentication (Organizational Users)",
     "Shared",
     "AWS authenticates AWS personnel via AWS-internal MFA and PIV-equivalent "
     "credentials. AWS provides IAM (long-term credentials), IAM Identity "
     "Center (federated SSO), and hardware MFA token support for customer use.",
     "Customer is responsible for (a) enforcing MFA on all customer IAM users, "
     "with hardware MFA mandatory for the root account and for principals "
     "authorized to perform privileged operations, (b) configuring SAML 2.0 or "
     "OIDC federation with the customer enterprise identity provider, (c) "
     "rotating IAM access keys at maximum 90-day intervals per IA-5, and (d) "
     "enforcing FIPS 140-2 validated authenticators for all government-user "
     "access.",
     "Implemented"),
    ("IA-5",  "Authenticator Management",
     "Shared",
     "AWS manages authenticators for AWS personnel per AWS Personnel Security "
     "policy including initial issuance, rotation, and revocation.",
     "Customer is responsible for managing customer-side authenticators: (a) "
     "IAM access key rotation at maximum 90-day intervals, (b) IAM password "
     "policy enforcement (minimum 14 characters, complexity, reuse-prevention "
     "history, expiration), (c) MFA device lifecycle including loss reporting "
     "and re-issuance, and (d) AWS Secrets Manager or Parameter Store rotation "
     "schedules for application credentials with automatic rotation enabled.",
     "Implemented"),
    ("IA-8",  "Identification and Authentication (Non-Organizational Users)",
     "Shared",
     "AWS provides Amazon Cognito and IAM identity federation services "
     "supporting SAML 2.0, OIDC, and social identity providers for non-"
     "organizational user authentication to customer applications.",
     "Customer is responsible for configuring Cognito user pools and identity "
     "pools, integrating with approved external identity providers, enforcing "
     "MFA on non-organizational users where required by customer policy, and "
     "documenting the trust relationships in the customer SSP.",
     "Implemented"),
    ("IR-4",  "Incident Handling",
     "Shared",
     "AWS handles incidents affecting AWS-managed infrastructure per the AWS "
     "Incident Response plan and notifies affected customers per the AWS "
     "shared responsibility model and the AWS Customer Agreement.",
     "Customer is responsible for incident handling within the customer "
     "environment: (a) detection via GuardDuty, Security Hub, Macie, and "
     "customer SIEM, (b) triage and investigation per customer IR playbooks, "
     "(c) containment via IAM revocation, Security Group quarantine, and "
     "instance isolation, (d) eradication and recovery per customer IR plan, "
     "and (e) breach notification per customer regulatory obligations.",
     "Implemented"),
    ("IR-6",  "Incident Reporting",
     "Shared",
     "AWS reports incidents affecting AWS-managed infrastructure to affected "
     "customers and to FedRAMP PMO per AWS incident reporting procedures and "
     "the AWS FedRAMP High SSP.",
     "Customer is responsible for reporting customer-environment incidents to "
     "(a) customer-organization authorities per customer policy, (b) US-CERT "
     "per FISMA reporting timeframes, (c) the FedRAMP PMO within 1 hour of "
     "incident determination (per FedRAMP Incident Communications Procedure) "
     "if the system is FedRAMP-authorized, and (d) any sector-specific "
     "regulator (CISA, sector ISAC) as required.",
     "Implemented"),
    ("IR-8",  "Incident Response Plan",
     "Shared",
     "AWS maintains an AWS-internal incident response plan for AWS-managed "
     "infrastructure, reviewed annually and after lessons-learned activities.",
     "Customer is responsible for developing, documenting, distributing, and "
     "annually reviewing the customer Incident Response Plan covering the "
     "customer-deployed environment. Plan must address coordination with AWS "
     "for incidents spanning the shared responsibility boundary and must "
     "designate the customer point-of-contact for AWS incident notifications.",
     "Implemented"),
    ("MA-2",  "Controlled Maintenance",
     "Shared",
     "AWS performs all controlled maintenance on AWS-managed physical "
     "infrastructure per AWS change management procedures with documented "
     "schedules, approvals, and post-maintenance verification.",
     "Customer is responsible for controlled maintenance of customer-managed "
     "logical components: guest OS patching, application maintenance, RDS "
     "engine version upgrades, container image updates, and Lambda runtime "
     "updates. Maintenance scheduled per customer change management with "
     "documented approvals and rollback procedures.",
     "Implemented"),
    ("MA-3",  "Maintenance Tools",
     "Shared",
     "AWS controls, inspects, and sanitizes all maintenance tools entering "
     "AWS datacenter facilities per AWS physical security procedures.",
     "Customer is responsible for approving and controlling customer-side "
     "maintenance tools used against customer-managed resources, including "
     "Systems Manager Run Command documents, Session Manager preferences, "
     "third-party EDR maintenance agents, and any customer scripts with "
     "elevated privileges.",
     "Implemented"),
    ("MA-4",  "Nonlocal Maintenance",
     "Shared",
     "AWS authorizes, monitors, and audits all nonlocal maintenance on AWS-"
     "managed infrastructure via approved AWS-internal access paths with MFA "
     "and full session recording.",
     "Customer is responsible for nonlocal maintenance of customer-managed "
     "resources: (a) all administrative access via Systems Manager Session "
     "Manager (no direct SSH/RDP), (b) MFA enforced on all administrative "
     "sessions, (c) session recording enabled and stored in tamper-resistant "
     "S3 with Object Lock, and (d) FIPS-validated transport for all "
     "administrative connections.",
     "Implemented"),
    ("MA-5",  "Maintenance Personnel",
     "Shared",
     "AWS maintenance personnel undergo AWS Personnel Security screening per "
     "PS-3 (inherited) prior to access authorization.",
     "Customer is responsible for vetting customer personnel authorized to "
     "perform maintenance on customer-managed AWS resources: background "
     "investigation per customer PS-3, designated maintenance role in IAM, "
     "documented authorization, and revocation upon role change or separation.",
     "Implemented"),
    ("SA-9",  "External System Services",
     "Shared",
     "AWS provides FedRAMP High authorized services within GovCloud; AWS "
     "publishes the FedRAMP authorization package, SOC reports, and ISO "
     "certifications via AWS Artifact for customer due-diligence review.",
     "Customer is responsible for (a) assessing external (non-AWS) services "
     "used by customer workloads against customer security requirements, (b) "
     "documenting all customer-managed third-party connections and the data "
     "flows traversing them, (c) ensuring contractual security commitments "
     "from third-party providers, and (d) periodic re-assessment of external "
     "service authorizations.",
     "Implemented"),
    ("SC-1",  "System and Communications Protection Policy and Procedures",
     "Shared",
     "AWS maintains AWS-internal system and communications protection policy "
     "and procedures applicable to AWS-managed infrastructure, reviewed "
     "annually.",
     "Customer is responsible for developing, documenting, and annually "
     "reviewing the customer-side system and communications protection policy "
     "and procedures, addressing customer-deployed network architecture, "
     "encryption requirements, and boundary protection design.",
     "Implemented"),
    ("SC-7",  "Boundary Protection",
     "Shared",
     "AWS provides VPC, Security Groups, Network ACLs, AWS WAF, AWS Shield, "
     "Network Firewall, Transit Gateway, and PrivateLink as boundary "
     "protection primitives. AWS protects the AWS service boundary at the "
     "service API layer.",
     "Customer is responsible for designing the customer VPC boundary "
     "architecture including (a) public/private subnet segmentation, (b) "
     "Security Group rules enforcing least-privilege traffic flows, (c) NACL "
     "rules at the subnet boundary, (d) WAF rules on internet-facing ALBs and "
     "API Gateway, (e) Network Firewall deployment for VPC-level egress "
     "filtering, and (f) enforcement of the customer-defined boundary between "
     "trust zones per the customer architecture.",
     "Implemented"),
    ("SC-8",  "Transmission Confidentiality and Integrity",
     "Shared",
     "AWS encrypts data in transit between AWS services and across AWS "
     "regions using TLS 1.2 or higher with FIPS-validated cipher suites. AWS "
     "physically protects inter-datacenter dark fiber.",
     "Customer is responsible for (a) enforcing TLS 1.2+ on all customer-"
     "facing endpoints (ALB/NLB listeners, API Gateway, CloudFront "
     "distributions), (b) configuring only approved FIPS-validated cipher "
     "suites, (c) enforcing HTTPS-only access via redirect or policy, (d) "
     "ensuring application-layer TLS for service-to-service communication, "
     "and (e) using FIPS endpoints (-fips. suffix) where the customer system "
     "is FedRAMP- or FISMA-regulated.",
     "Implemented"),
    ("SC-12", "Cryptographic Key Establishment and Management",
     "Shared",
     "AWS manages keys for AWS-managed encryption (AWS-owned and AWS-managed "
     "KMS keys). AWS provides AWS KMS (FIPS 140-2 Level 2 validated HSMs) "
     "and AWS CloudHSM (FIPS 140-2 Level 3 validated single-tenant HSMs) for "
     "customer key management.",
     "Customer is responsible for managing customer-managed KMS CMKs and "
     "CloudHSM clusters where applicable: (a) key creation with appropriate "
     "key spec and usage, (b) automatic annual rotation enabled for "
     "symmetric CMKs, (c) IAM and key policies enforcing least-privilege key "
     "access, (d) CloudTrail logging of all key usage, and (e) scheduled "
     "deletion with appropriate waiting period prior to permanent removal.",
     "Implemented"),
    ("SC-13", "Cryptographic Protection",
     "Shared",
     "AWS provides FIPS 140-2 validated cryptographic modules within GovCloud "
     "via AWS KMS (Level 2), AWS CloudHSM (Level 3), AWS Certificate Manager, "
     "and FIPS-validated service endpoints. NIAP-approved configurations "
     "available per service.",
     "Customer is responsible for using only FIPS 140-2 validated "
     "cryptographic implementations for all customer workloads: (a) enable "
     "FIPS service endpoints (-fips. suffix) for all customer service calls, "
     "(b) configure FIPS mode in customer EC2 guest operating systems where "
     "supported, (c) avoid non-FIPS algorithms in customer applications, and "
     "(d) document customer cryptographic implementations in the customer SSP.",
     "Implemented"),
    ("SI-2",  "Flaw Remediation",
     "Shared",
     "AWS identifies and remediates flaws in AWS-managed services per AWS "
     "Vulnerability Management policy. AWS publishes service-level security "
     "bulletins via the AWS Security Bulletins page.",
     "Customer is responsible for remediating flaws in customer-managed "
     "components: (a) guest OS, container, and application patching via AWS "
     "Systems Manager Patch Manager, (b) Inspector findings remediation per "
     "customer-defined SLAs, (c) Critical CVEs remediated within 30 days of "
     "vendor disclosure, (d) High CVEs within 90 days, (e) CISA Known "
     "Exploited Vulnerabilities Catalog entries remediated within the BOD "
     "22-01 mandated timeframe (typically 15 days for federal systems), and "
     "(f) documented exception process for any deviation.",
     "Implemented"),
    ("SI-3",  "Malicious Code Protection",
     "Shared",
     "AWS protects AWS-managed infrastructure against malicious code via AWS-"
     "internal endpoint protection and AWS-internal SOC monitoring.",
     "Customer is responsible for deploying malicious code protection on "
     "customer workloads: (a) anti-malware on customer EC2 / EKS / ECS "
     "workloads via vendor EDR or ClamAV, (b) Amazon GuardDuty Malware "
     "Protection enabled across all accounts, (c) signature and engine "
     "updates within 24 hours of vendor release, and (d) remediation of "
     "detections per customer IR-4.",
     "Implemented"),
    ("SI-4",  "System Monitoring",
     "Shared",
     "AWS monitors AWS-managed infrastructure via AWS-internal SOC tooling "
     "with 24x7 coverage and automated alerting.",
     "Customer is responsible for monitoring customer-deployed workloads via "
     "(a) CloudWatch metrics and alarms, (b) Amazon GuardDuty across all "
     "accounts and regions, (c) AWS Security Hub for cross-service finding "
     "aggregation, (d) AWS Inspector for vulnerability assessment, (e) Amazon "
     "Macie for sensitive-data discovery in S3, and (f) customer SIEM "
     "integration via Kinesis Data Firehose or partner connector. Alerts "
     "triaged per IR-4.",
     "Implemented"),
    ("MP-4",  "Media Storage",
     "Shared",
     "AWS physically secures and controls access to storage media within AWS "
     "datacenter facilities per AWS Physical Security policy.",
     "Customer is responsible for protecting customer logical media: (a) data "
     "classification labeling via S3 object tags, (b) S3 bucket policies "
     "denying public access with S3 Block Public Access at account and bucket "
     "level, (c) encryption-at-rest for all customer media (EBS, S3, RDS, "
     "DynamoDB) using customer-managed KMS keys, and (d) MFA Delete on "
     "critical S3 buckets.",
     "Implemented"),
    ("MP-6",  "Media Sanitization",
     "Shared",
     "AWS sanitizes decommissioned physical storage media per NIST SP 800-88 "
     "Rev. 1 guidelines including degaussing, shredding, or pulverization "
     "with documented certificates of destruction.",
     "Customer is responsible for sanitizing customer-managed data: (a) EBS "
     "volume deletion (KMS key deletion = NIST 800-88 cryptographic erasure "
     "when customer-managed CMK is used), (b) emptying versioned S3 buckets "
     "including all noncurrent versions and delete markers, (c) RDS snapshot "
     "deletion, and (d) scheduled deletion of customer-managed KMS CMKs that "
     "encrypted retired data.",
     "Implemented"),

    # === CUSTOMER (full LLM assessment -- pure customer policy) =============
    ("AT-1",  "Awareness and Training Policy and Procedures",
     "Customer",
     "Not applicable to the AWS service boundary. Customer maintains the "
     "customer-side awareness and training policy independently.",
     "Customer is responsible for developing, documenting, disseminating, "
     "reviewing, and updating the customer security awareness and training "
     "policy and supporting procedures applicable to all customer personnel "
     "with access to customer information systems.",
     "Implemented"),
    ("AT-2",  "Literacy Training and Awareness",
     "Customer",
     "Not applicable to the AWS service boundary.",
     "Customer is responsible for providing security literacy training and "
     "awareness to all customer information system users on initial access, "
     "annually thereafter, and following any significant change to the system "
     "environment, including role-based content per the customer training "
     "program.",
     "Implemented"),
    ("AT-3",  "Role-Based Training",
     "Customer",
     "Not applicable to the AWS service boundary.",
     "Customer is responsible for providing role-based security training to "
     "customer personnel with significant security responsibilities (cloud "
     "administrators, security engineers, developers with production access) "
     "prior to authorizing access and annually thereafter.",
     "Implemented"),
    ("PL-2",  "System Security and Privacy Plans",
     "Customer",
     "Not applicable to the AWS service boundary. Customer maintains the "
     "customer-side System Security Plan independently.",
     "Customer is responsible for developing, documenting, and maintaining "
     "the customer System Security Plan covering customer-deployed workloads, "
     "customer responsibilities, and the explicit inheritance and shared "
     "responsibility relationships with AWS per this CRM.",
     "Implemented"),
    ("PL-4",  "Rules of Behavior",
     "Customer",
     "Not applicable to the AWS service boundary.",
     "Customer is responsible for establishing and enforcing rules of "
     "behavior for customer users with access to customer-managed AWS "
     "resources, including acknowledgment prior to access authorization.",
     "Implemented"),
    ("PM-1",  "Information Security Program Plan",
     "Customer",
     "Not applicable to the AWS service boundary.",
     "Customer is responsible for establishing and maintaining the customer "
     "Information Security Program Plan, including customer governance, risk, "
     "compliance functions, and senior agency information security officer "
     "designation.",
     "Implemented"),
    ("PM-9",  "Risk Management Strategy",
     "Customer",
     "Not applicable to the AWS service boundary.",
     "Customer is responsible for developing and implementing the customer "
     "Risk Management Strategy, including risk tolerance levels for customer-"
     "managed workloads in AWS GovCloud and risk framing for the inherited "
     "and shared portions of the AWS shared responsibility model.",
     "Implemented"),
    ("RA-3",  "Risk Assessment",
     "Customer",
     "Not applicable to the AWS service boundary.",
     "Customer is responsible for conducting risk assessments of customer-"
     "deployed information systems at minimum annually and following any "
     "significant change, incorporating the AWS shared responsibility model "
     "and the customer-specific threat landscape.",
     "Implemented"),
    ("RA-5",  "Vulnerability Monitoring and Scanning",
     "Customer",
     "Not applicable to the AWS service boundary. AWS scans AWS-managed "
     "infrastructure independently per the AWS Vulnerability Management "
     "program.",
     "Customer is responsible for vulnerability scanning of customer-deployed "
     "workloads via Amazon Inspector (EC2, ECR, Lambda), customer third-party "
     "scanners (Tenable Nessus, Qualys, Rapid7), and remediating findings per "
     "customer-defined SLAs aligned to SI-2.",
     "Implemented"),

    # === NOT APPLICABLE (test scaffolding -- exercises NA short-circuit) ====
    # NOTE: A real AWS CRM marks these "Inherited" because FedRAMP High
    # mandates them. We keep "Not Applicable" only so the assessor's
    # not_applicable->NOT_APPLICABLE short-circuit gets exercised end-to-end.
    # Replace with "Inherited" when the real CRM lands.
    ("PE-10", "Emergency Shutoff",
     "Not Applicable",
     "Not applicable to AWS GovCloud (US) customer service offering. AWS "
     "datacenter emergency power shutoff is an AWS-internal facility control "
     "not exposed to customer interaction.",
     "Not applicable to customer. Customer has no role in datacenter "
     "emergency shutoff procedures and no requirement to implement an "
     "equivalent compensating control.",
     "Not Implemented"),
    ("PE-11", "Emergency Power",
     "Not Applicable",
     "Not applicable to AWS GovCloud (US) customer service offering. AWS "
     "provides datacenter emergency power (UPS, generators, fuel reserves) "
     "as an internal infrastructure control.",
     "Not applicable to customer. AWS GovCloud datacenter emergency power "
     "implementation is fully internal to AWS facilities; customer has no "
     "implementation or compensating-control responsibility.",
     "Not Implemented"),
]

wb = Workbook()
ws = wb.active
ws.title = "Customer Responsibility Matrix"

# Header row -- realistic AWS CRM column layout. Loader matches on
# "Control ID" / "Responsibility" / "Customer Responsibility".
headers = [
    "Control ID",
    "Control Name",
    "Responsibility",
    "AWS Implementation",
    "Customer Responsibility",
    "Implementation Status",
]
ws.append(headers)

# Header styling
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E78")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Data rows
wrap = Alignment(wrap_text=True, vertical="top")
for row_data in ROWS:
    ws.append(row_data)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).alignment = wrap

# Column widths roughly matching AWS published CRMs
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 65
ws.column_dimensions["E"].width = 65
ws.column_dimensions["F"].width = 20

ws.freeze_panes = "A2"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)

# Summary
if __name__ == "__main__":
    from collections import Counter
    counts = Counter(r[2] for r in ROWS)
    print(f"Wrote {OUT}")
    print(f"  {len(ROWS)} rows")
    for resp, n in sorted(counts.items()):
        print(f"  {resp:20s} {n}")
