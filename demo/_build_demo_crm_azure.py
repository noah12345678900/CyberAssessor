"""Generate a realistic Azure Government FedRAMP High CRM xlsx for smoke testing.

Second example CRM (alongside _build_demo_crm.py's AWS GovCloud CRM) so the
assessor can be exercised against more than one cloud-service-provider format.
Mimics the Microsoft Azure Government Customer Responsibility Matrix column
layout (Control ID, Control Name, Responsibility, Azure Implementation,
Customer Responsibility, Implementation Status). The loader
(baselines/crm_xlsx.py) only reads Control ID + Responsibility + Customer
Responsibility -- the other columns are cosmetic realism so the file looks like
something Microsoft would actually publish.

Coverage targets every code path the CRM short-circuit / hybrid block
exercises:
  - inherited      -> COMPLIANT short-circuit       (PE -- Microsoft-owned physical)
  - shared         -> prompt enrichment             (AC/AU/CM/IA/SC/SI/...)
                      (loader maps "shared" -> hybrid internally)
  - customer       -> full LLM path                 (AT/PL/PM/RA -- pure policy)
  - provider       -> NOT_APPLICABLE short-circuit  (PE-4, PE-5 "Microsoft")
  - not_applicable -> NOT_APPLICABLE short-circuit  (PE-10, PE-11)

Intentional Azure-vs-AWS divergences so the two demo CRMs are not carbon
copies: Microsoft terminology ("Microsoft" responsibility label instead of
"Service Provider"; Entra ID, Defender for Cloud, Sentinel, Key Vault,
managed HSM, NSGs, Azure Policy, Update Manager) and a slightly different
control selection.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = Path(__file__).parent / "crm" / "Azure_Government_FedRAMP_High_CRM.xlsx"

# (control_id, control_name, responsibility, azure_impl, customer_resp, impl_status)
ROWS = [
    # === INHERITED (fully Microsoft-owned datacenter controls) ==============
    ("PE-1",  "Physical and Environmental Protection Policy and Procedures",
     "Inherited",
     "Microsoft develops, documents, and annually reviews physical and "
     "environmental protection policy and procedures for all Azure Government "
     "datacenter regions. Policy applies to all Microsoft-owned, -leased, and "
     "-colocated facilities supporting the Azure Government cloud.",
     "Customer fully inherits the Microsoft physical and environmental "
     "protection policy for all customer workloads hosted in Azure Government. "
     "No customer action required for the underlying infrastructure layer. "
     "Customer SSP must reference the Azure Government FedRAMP High SSP "
     "(available via the Service Trust Portal) as the authoritative source.",
     "Implemented"),
    ("PE-2",  "Physical Access Authorizations",
     "Inherited",
     "Microsoft maintains and reviews the list of personnel authorized to "
     "access Azure datacenter facilities. Access is re-approved at least every "
     "90 days and revoked within 24 hours of separation or role change.",
     "Customer fully inherits Azure physical access authorization controls. "
     "Customer personnel are not granted physical access to Azure Government "
     "datacenters. No customer-side implementation required.",
     "Implemented"),
    ("PE-3",  "Physical Access Control",
     "Inherited",
     "Microsoft enforces datacenter physical access via perimeter fencing, "
     "badge plus biometric authentication at building and cage boundaries, "
     "mantrap vestibules, 24x7 security staff, and full visitor escort. "
     "Access events are logged and retained per AU-11.",
     "Customer fully inherits Microsoft datacenter physical access controls "
     "and has no compensating-control responsibility.",
     "Implemented"),
    ("PE-6",  "Monitoring Physical Access",
     "Inherited",
     "Microsoft monitors datacenter physical access via CCTV with 90-day "
     "minimum retention, intrusion-detection sensors, and continuous review by "
     "the Microsoft datacenter security operations team with automated "
     "alerting on anomalous access.",
     "Customer fully inherits Azure physical access monitoring. Customer may "
     "review the Azure SOC 2 Type II and FedRAMP High audit reports via the "
     "Service Trust Portal for assurance evidence.",
     "Implemented"),
    ("PE-13", "Fire Protection",
     "Inherited",
     "Azure datacenters deploy very-early smoke detection, pre-action dry-pipe "
     "suppression, handheld extinguishers per NFPA 10, and conduct quarterly "
     "inspection with annual recertification of suppression equipment.",
     "Customer fully inherits Azure datacenter fire protection. Customer data "
     "is protected against facility-level events through customer-implemented "
     "availability-zone / geo-redundant replication per CP-9.",
     "Implemented"),
    ("PE-14", "Environmental Controls",
     "Inherited",
     "Microsoft maintains datacenter temperature and humidity within ASHRAE "
     "ranges via N+1 redundant cooling with continuous monitoring and "
     "automated failover.",
     "Customer fully inherits Azure datacenter environmental controls. No "
     "customer-side implementation responsibility.",
     "Implemented"),

    # === MICROSOFT / PROVIDER (exercises provider -> NA short-circuit) ======
    ("PE-4",  "Access Control for Transmission",
     "Microsoft",
     "Microsoft controls physical access to all information system "
     "transmission and distribution lines within Azure datacenter facilities, "
     "including intra-facility fiber, copper, and power distribution.",
     "Provider-owned. Customer has no role in physical transmission-line "
     "access control within Azure datacenters and no compensating-control "
     "responsibility.",
     "Implemented"),
    ("PE-5",  "Access Control for Output Devices",
     "Microsoft",
     "Microsoft controls physical access to output devices within Azure "
     "datacenter facilities. No physical output devices are provided to "
     "customers; all customer interaction occurs via authenticated remote APIs "
     "and the Azure portal.",
     "Provider-owned. Customer does not maintain physical output devices in "
     "Azure facilities. Not applicable to customer.",
     "Implemented"),

    # === SHARED (loader routes "shared" -> hybrid prompt-enrichment path) ===
    ("AC-1",  "Access Control Policy and Procedures",
     "Shared",
     "Microsoft maintains internal access control policy for Microsoft "
     "personnel and Azure-managed infrastructure, reviewed annually.",
     "Customer is responsible for developing, documenting, disseminating, and "
     "annually reviewing the customer-side access control policy governing "
     "customer use of Azure services and customer-deployed workloads. Customer "
     "SSP must document the AC-1 policy and its mission alignment.",
     "Implemented"),
    ("AC-2",  "Account Management",
     "Shared",
     "Microsoft manages Microsoft-internal personnel accounts per Microsoft "
     "Personnel Security policy and provides Microsoft Entra ID, Privileged "
     "Identity Management (PIM), and managed identities as account-management "
     "primitives for customer use.",
     "Customer is responsible for the full lifecycle of customer Entra ID "
     "users, groups, service principals, and managed identities, including "
     "(a) provisioning per documented request/approval workflow, (b) "
     "deprovisioning within 24 hours of role change or separation, (c) "
     "quarterly access reviews via Entra ID Access Reviews, (d) automated "
     "disabling of accounts inactive >=90 days, and (e) MFA enforcement on all "
     "Global Administrator accounts.",
     "Implemented"),
    ("AC-3",  "Access Enforcement",
     "Shared",
     "Microsoft enforces access at the Azure Resource Manager control-plane "
     "layer; requests not authorized by an applicable Azure RBAC role "
     "assignment or Azure Policy are denied.",
     "Customer is responsible for authoring and maintaining Azure RBAC role "
     "assignments, custom roles, Azure Policy assignments, and Conditional "
     "Access policies that enforce least-privilege access to customer-owned "
     "resources, and for validating effective access prior to deployment.",
     "Implemented"),
    ("AC-6",  "Least Privilege",
     "Shared",
     "Microsoft implements least privilege for Microsoft personnel and Azure "
     "service principals via narrowly scoped built-in roles.",
     "Customer is responsible for implementing least privilege within customer "
     "subscriptions: (a) custom RBAC roles scoped to minimum actions, (b) "
     "just-in-time privileged elevation via Entra PIM, (c) periodic review and "
     "removal of unused role assignments, and (d) management-group-level Azure "
     "Policy to constrain delegated administration.",
     "Implemented"),
    ("AU-2",  "Event Logging",
     "Shared",
     "Microsoft logs all control-plane operations via Azure-internal audit "
     "systems and provides Azure Activity Log, Azure Monitor, diagnostic "
     "settings, and Microsoft Entra audit logs for customer event capture.",
     "Customer is responsible for (a) routing Activity Log and resource "
     "diagnostic logs to a Log Analytics workspace and immutable storage, (b) "
     "enabling data-plane logging on sensitive storage accounts, Key Vaults, "
     "and databases, (c) configuring Entra ID sign-in and audit log export, "
     "and (d) documenting the customer-defined auditable event list per AU-1.",
     "Implemented"),
    ("AU-6",  "Audit Record Review, Analysis, and Reporting",
     "Shared",
     "Microsoft reviews Azure-internal audit logs continuously via the "
     "Microsoft security operations team and reports relevant findings to "
     "customers per IR-6.",
     "Customer is responsible for reviewing customer Activity Logs, Defender "
     "for Cloud alerts, Microsoft Sentinel incidents, and application logs on "
     "the customer-defined cadence (minimum weekly for non-critical, "
     "continuous for critical). Reportable findings escalated per IR-6.",
     "Implemented"),
    ("AU-12", "Audit Record Generation",
     "Shared",
     "Microsoft generates audit records for all Azure control-plane and "
     "Azure-internal service operations per Microsoft audit policy.",
     "Customer is responsible for enabling audit record generation on all "
     "customer-deployed systems: VM guest OS auditd / Windows Event Log via "
     "Azure Monitor Agent, Azure SQL audit, container runtime logs, and "
     "application logs, forwarded to a tamper-resistant Log Analytics "
     "workspace or immutable storage.",
     "Implemented"),
    ("CM-2",  "Baseline Configuration",
     "Shared",
     "Microsoft maintains baseline configurations for all Azure-managed "
     "infrastructure and services per Microsoft configuration management "
     "policy.",
     "Customer is responsible for developing and maintaining baseline "
     "configurations for customer-managed VM images, container images, and "
     "Infrastructure-as-Code templates (ARM, Bicep, Terraform). Baselines "
     "reviewed at least annually and after significant change; drift detected "
     "via Azure Policy guest configuration / Machine Configuration.",
     "Implemented"),
    ("CM-6",  "Configuration Settings",
     "Shared",
     "Microsoft configures Azure-managed services per Microsoft security "
     "baselines aligned to the Microsoft cloud security benchmark.",
     "Customer is responsible for (a) hardening customer-managed OS images per "
     "DISA STIG or CIS Benchmarks, (b) continuously enforcing configuration "
     "via Azure Policy and Machine Configuration, (c) remediating drift within "
     "customer-defined SLAs, and (d) documenting approved deviations with "
     "compensating controls.",
     "Implemented"),
    ("CM-8",  "System Component Inventory",
     "Shared",
     "Microsoft maintains a complete inventory of Azure-managed physical and "
     "virtual infrastructure supporting the Azure Government offering.",
     "Customer is responsible for maintaining an accurate inventory of "
     "customer-deployed resources via Azure Resource Graph, Azure Resource "
     "tags, and Defender for Cloud asset inventory, reconciled at least "
     "monthly.",
     "Implemented"),
    ("CP-9",  "System Backup",
     "Shared",
     "Microsoft replicates and backs up Azure-managed control-plane and "
     "service data per Microsoft contingency planning policy, with locally and "
     "geo-redundant storage durability guarantees.",
     "Customer is responsible for backing up customer-owned data aligned to "
     "customer-defined RPO via Azure Backup vaults, VM and disk snapshots, "
     "Azure SQL automated backups with geo-restore, and geo-redundant storage "
     "replication. Backup integrity tested per CP-4.",
     "Implemented"),
    ("IA-2",  "Identification and Authentication (Organizational Users)",
     "Shared",
     "Microsoft authenticates Microsoft personnel via internal MFA and "
     "PIV-equivalent credentials and provides Microsoft Entra ID with "
     "Conditional Access and FIDO2 / certificate-based authentication for "
     "customer use.",
     "Customer is responsible for (a) enforcing MFA on all customer Entra ID "
     "users via Conditional Access, with phishing-resistant MFA for privileged "
     "roles, (b) federating with the customer enterprise IdP where applicable, "
     "(c) enforcing FIPS 140-validated authenticators for government-user "
     "access, and (d) blocking legacy authentication protocols.",
     "Implemented"),
    ("IA-5",  "Authenticator Management",
     "Shared",
     "Microsoft manages authenticators for Microsoft personnel per Microsoft "
     "Personnel Security policy including issuance, rotation, and revocation.",
     "Customer is responsible for managing customer-side authenticators: (a) "
     "Entra ID password protection and ban-list policy, (b) service-principal "
     "secret and certificate rotation, (c) MFA device lifecycle including loss "
     "reporting and re-issuance, and (d) Azure Key Vault rotation policies for "
     "application secrets and certificates.",
     "Implemented"),
    ("IR-4",  "Incident Handling",
     "Shared",
     "Microsoft handles incidents affecting Azure-managed infrastructure per "
     "the Microsoft Security Response process and notifies affected customers "
     "per the Microsoft Online Services Terms.",
     "Customer is responsible for incident handling within the customer "
     "environment: detection via Defender for Cloud and Microsoft Sentinel, "
     "triage per customer playbooks, containment via Conditional Access "
     "revocation and NSG isolation, eradication and recovery per the customer "
     "IR plan, and breach notification per regulatory obligations.",
     "Implemented"),
    ("IR-6",  "Incident Reporting",
     "Shared",
     "Microsoft reports incidents affecting Azure-managed infrastructure to "
     "affected customers and to the FedRAMP PMO per Microsoft incident "
     "reporting procedures.",
     "Customer is responsible for reporting customer-environment incidents to "
     "customer authorities, to US-CERT per FISMA timeframes, and to the "
     "FedRAMP PMO within 1 hour of incident determination for FedRAMP-"
     "authorized systems.",
     "Implemented"),
    ("MA-4",  "Nonlocal Maintenance",
     "Shared",
     "Microsoft authorizes, monitors, and audits all nonlocal maintenance on "
     "Azure-managed infrastructure via approved internal access paths with MFA "
     "and full session recording.",
     "Customer is responsible for nonlocal maintenance of customer-managed "
     "resources: administrative access via Azure Bastion (no direct public "
     "SSH/RDP), MFA on all administrative sessions, session logging to "
     "immutable storage, and FIPS-validated transport for administrative "
     "connections.",
     "Implemented"),
    ("SC-7",  "Boundary Protection",
     "Shared",
     "Microsoft provides Virtual Network, Network Security Groups, Azure "
     "Firewall, DDoS Protection, Application Gateway WAF, and Private Link as "
     "boundary-protection primitives, and protects the Azure service boundary "
     "at the control-plane layer.",
     "Customer is responsible for designing the customer virtual-network "
     "boundary: (a) subnet segmentation, (b) NSG rules enforcing "
     "least-privilege flows, (c) Azure Firewall for centralized egress "
     "filtering, (d) WAF on internet-facing Application Gateways and Front "
     "Door, (e) Private Endpoints for PaaS services, and (f) enforcement of "
     "the customer-defined trust-zone boundary.",
     "Implemented"),
    ("SC-8",  "Transmission Confidentiality and Integrity",
     "Shared",
     "Microsoft encrypts data in transit between Azure services and across "
     "regions using TLS 1.2+ with FIPS-validated cipher suites and physically "
     "protects inter-datacenter fiber.",
     "Customer is responsible for (a) enforcing TLS 1.2+ on all customer-"
     "facing endpoints, (b) configuring only approved FIPS-validated cipher "
     "suites, (c) enforcing HTTPS-only access, (d) application-layer TLS for "
     "service-to-service traffic, and (e) using Azure Government FIPS endpoints "
     "for FISMA-regulated systems.",
     "Implemented"),
    ("SC-12", "Cryptographic Key Establishment and Management",
     "Shared",
     "Microsoft manages keys for Microsoft-managed encryption and provides "
     "Azure Key Vault (FIPS 140-2 Level 2 validated HSMs) and Azure Key Vault "
     "Managed HSM (FIPS 140-2 Level 3) for customer key management.",
     "Customer is responsible for managing customer-managed keys: (a) creation "
     "with appropriate key type and size, (b) rotation policy enabled, (c) "
     "Key Vault access policies / RBAC enforcing least privilege, (d) "
     "diagnostic logging of all key operations, and (e) soft-delete and purge "
     "protection prior to permanent removal.",
     "Implemented"),
    ("SC-13", "Cryptographic Protection",
     "Shared",
     "Microsoft provides FIPS 140-2 validated cryptographic modules within "
     "Azure Government via Key Vault, Managed HSM, and FIPS-validated service "
     "endpoints.",
     "Customer is responsible for using only FIPS 140-2 validated "
     "cryptographic implementations: enable FIPS-validated endpoints for "
     "customer service calls, configure FIPS mode in customer guest operating "
     "systems where supported, avoid non-FIPS algorithms in applications, and "
     "document customer cryptographic implementations in the SSP.",
     "Implemented"),
    ("SI-2",  "Flaw Remediation",
     "Shared",
     "Microsoft identifies and remediates flaws in Azure-managed services per "
     "the Microsoft Vulnerability Management program and publishes service "
     "security advisories.",
     "Customer is responsible for remediating flaws in customer-managed "
     "components: (a) guest OS and application patching via Azure Update "
     "Manager, (b) Defender for Cloud recommendations remediation per SLA, (c) "
     "Critical CVEs within 30 days and High within 90 days of disclosure, (d) "
     "CISA Known Exploited Vulnerabilities entries within the BOD 22-01 "
     "mandated window, and (e) a documented exception process.",
     "Implemented"),
    ("SI-3",  "Malicious Code Protection",
     "Shared",
     "Microsoft protects Azure-managed infrastructure against malicious code "
     "via internal endpoint protection and continuous SOC monitoring.",
     "Customer is responsible for deploying malicious code protection on "
     "customer workloads: Microsoft Defender for Servers or vendor EDR on VMs, "
     "Defender for Containers on AKS, signature/engine updates within 24 hours "
     "of release, and remediation of detections per IR-4.",
     "Implemented"),
    ("SI-4",  "System Monitoring",
     "Shared",
     "Microsoft monitors Azure-managed infrastructure via internal SOC tooling "
     "with 24x7 coverage and automated alerting.",
     "Customer is responsible for monitoring customer-deployed workloads via "
     "(a) Azure Monitor metrics and alerts, (b) Microsoft Defender for Cloud "
     "across all subscriptions, (c) Microsoft Sentinel for SIEM/SOAR "
     "correlation, (d) Defender for Cloud vulnerability assessment, and (e) "
     "customer SIEM integration where applicable. Alerts triaged per IR-4.",
     "Implemented"),
    ("MP-6",  "Media Sanitization",
     "Shared",
     "Microsoft sanitizes decommissioned physical storage media per NIST SP "
     "800-88 Rev. 1 with documented certificates of destruction.",
     "Customer is responsible for sanitizing customer-managed data: disk "
     "deletion (customer-managed key deletion = cryptographic erasure), "
     "purging versioned storage including soft-deleted blobs, database "
     "deletion, and scheduled purge of customer-managed Key Vault keys that "
     "encrypted retired data.",
     "Implemented"),

    # === CUSTOMER (full LLM assessment -- pure customer policy) =============
    ("AT-1",  "Awareness and Training Policy and Procedures",
     "Customer",
     "Not applicable to the Azure service boundary. Customer maintains the "
     "awareness and training policy independently.",
     "Customer is responsible for developing, documenting, disseminating, and "
     "annually reviewing the customer security awareness and training policy "
     "applicable to all personnel with access to customer information systems.",
     "Implemented"),
    ("AT-2",  "Literacy Training and Awareness",
     "Customer",
     "Not applicable to the Azure service boundary.",
     "Customer is responsible for providing security literacy training and "
     "awareness to all customer information system users on initial access, "
     "annually thereafter, and following significant environmental change, "
     "including role-based content.",
     "Implemented"),
    ("AT-3",  "Role-Based Training",
     "Customer",
     "Not applicable to the Azure service boundary.",
     "Customer is responsible for providing role-based security training to "
     "personnel with significant security responsibilities (cloud "
     "administrators, security engineers, developers with production access) "
     "prior to access authorization and annually thereafter.",
     "Implemented"),
    ("PL-2",  "System Security and Privacy Plans",
     "Customer",
     "Not applicable to the Azure service boundary. Customer maintains the "
     "System Security Plan independently.",
     "Customer is responsible for developing and maintaining the customer "
     "System Security Plan covering customer-deployed workloads, customer "
     "responsibilities, and the explicit inheritance and shared "
     "responsibility relationships with Microsoft per this CRM.",
     "Implemented"),
    ("PM-9",  "Risk Management Strategy",
     "Customer",
     "Not applicable to the Azure service boundary.",
     "Customer is responsible for developing and implementing the customer "
     "Risk Management Strategy, including risk tolerance for customer-managed "
     "workloads in Azure Government and risk framing for the inherited and "
     "shared portions of the shared responsibility model.",
     "Implemented"),
    ("RA-3",  "Risk Assessment",
     "Customer",
     "Not applicable to the Azure service boundary.",
     "Customer is responsible for conducting risk assessments of customer-"
     "deployed systems at least annually and following significant change, "
     "incorporating the Azure shared responsibility model and the customer-"
     "specific threat landscape.",
     "Implemented"),
    ("RA-5",  "Vulnerability Monitoring and Scanning",
     "Customer",
     "Not applicable to the Azure service boundary. Microsoft scans Azure-"
     "managed infrastructure independently.",
     "Customer is responsible for vulnerability scanning of customer-deployed "
     "workloads via Microsoft Defender for Cloud vulnerability assessment, "
     "customer third-party scanners (Tenable Nessus, Qualys), and remediating "
     "findings per customer SLAs aligned to SI-2.",
     "Implemented"),

    # === NOT APPLICABLE (exercises NA short-circuit) ========================
    ("PE-10", "Emergency Shutoff",
     "Not Applicable",
     "Not applicable to the Azure Government customer service offering. "
     "Datacenter emergency power shutoff is a Microsoft-internal facility "
     "control not exposed to customer interaction.",
     "Not applicable to customer. Customer has no role in datacenter emergency "
     "shutoff procedures and no requirement to implement an equivalent "
     "compensating control.",
     "Not Implemented"),
    ("PE-11", "Emergency Power",
     "Not Applicable",
     "Not applicable to the Azure Government customer service offering. "
     "Microsoft provides datacenter emergency power (UPS, generators, fuel "
     "reserves) as an internal infrastructure control.",
     "Not applicable to customer. Azure datacenter emergency power is fully "
     "internal to Microsoft facilities; customer has no implementation or "
     "compensating-control responsibility.",
     "Not Implemented"),
]

wb = Workbook()
ws = wb.active
ws.title = "Customer Responsibility Matrix"

# Header row -- realistic Azure Government CRM column layout. Loader matches on
# "Control ID" / "Responsibility" / "Customer Responsibility".
headers = [
    "Control ID",
    "Control Name",
    "Responsibility",
    "Azure Implementation",
    "Customer Responsibility",
    "Implementation Status",
]
ws.append(headers)

# Header styling -- Azure blue
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="0078D4")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Data rows
wrap = Alignment(wrap_text=True, vertical="top")
for row_data in ROWS:
    ws.append(row_data)
    r = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).alignment = wrap

# Column widths roughly matching Microsoft published CRMs
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 38
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 65
ws.column_dimensions["E"].width = 65
ws.column_dimensions["F"].width = 20

ws.freeze_panes = "A2"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)

# Summary
if __name__ == "__main__":
    from collections import Counter
    counts = Counter(r[2] for r in ROWS)
    print(f"Wrote {OUT}")
    print(f"  {len(ROWS)} rows")
    for resp, n in sorted(counts.items()):
        print(f"  {resp:20s} {n}")
