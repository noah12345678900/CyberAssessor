"""Generate sample anonymized PSC (Program-Specific Control) overlays, one
per supported NIST 800-53 revision.

This mirrors the *control-grain* T1TL "Ground Security Controls" shape that
backend/cybersecurity_assessor/catalogs/program_controls_loader.py auto-detects
and backend/cybersecurity_assessor/baselines/overlay_classifier.py routes to
``OverlayKind.PSC``:

* There is NO CCI column. CCIs are resolved later by joining the parsed NIST
  control IDs against the global DISA CCI <-> 800-53 mapping at resolve time.
* The NIST control ID is carried two ways so the demo exercises BOTH loader
  paths:
    1. a dedicated "Security Control" column with structured IDs ("AC-2(13)"
       or "AC-6(1), AC-6(2)") -> _extract_control_ids_direct (no anchor)
    2. the shall text itself carries an
       "Associated CNSSI 1253 Control Tag: <id>" anchor -> the prose anchor
       extractor, as a fallback for rows where the structured column is blank.
* The sheet header uses PSC vocabulary ("Threshold", "Objective",
  "Security Control") and deliberately has NO "Responsibility" column, so the
  overlay classifier does NOT mistake it for a CRM.

REVISION TARGETING (Rev 5)
--------------------------
PSC -> CCI resolution is NIST-800-53-specific (CCIs are a DISA 800-53
construct) AND revision-aware: the loader keys its objective lookup by 800-53
control id, scoped to the framework the DISA CCIs were loaded against. A
control enhancement that exists in one revision can be absent in, or
renumbered between, others (e.g. AC-2(13) is Rev 4+, not Rev 3; the
application allow-list enhancement is CM-7(2) in Rev 3 but CM-7(5) in BOTH
Rev 4 and Rev 5 -- it did NOT renumber to CM-7(8); CM-7(8) is an unrelated
"Binary or Machine Executable Code" control). A PSC whose IDs are absent at
the loaded revision sees those IDs land in ``unmapped_control_ids`` and
produce the "got nothing" failure.

This demo corpus targets Rev 5 exclusively. Critically, none of the control
IDs it uses renumber between Rev 4 and Rev 5, so the single Rev 5 fixture
resolves 100% clean against the shipped demo DISA source
(demo/cci_list/stig-mapping-to-nist-800-53.xlsx) even though that source
carries Rev 4 references -- the loader falls back to the highest available
revision (Rev 4) and every ID still matches. To exercise the fixture
end-to-end, load the DISA CCI catalog into the framework you resolve the PSC
against; a true Rev 5 DISA source is not required for the demo to map cleanly.

EVERYTHING here is synthetic. The fictional program is "Example Program";
control text is paraphrased generic 800-53 language, not real program data.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_DIR = Path(__file__).parent / "psc"

SHEET = "Ground Security Controls"

# Header layout. Order matters only for readability; the loader resolves
# columns by header alias, not position.
#   Control No.       -> req_number
#   Security Control  -> control_id (structured) + requirement_text fallback
#   Threshold         -> requirement_text (shall statement; wins over Objective)
#   Objective         -> objective-level shall (lower-priority req_text alias)
HEADERS = ["Control No.", "Security Control", "Threshold", "Objective"]
WIDTHS = [16, 20, 78, 78]

# Revisions to emit. The demo corpus targets NIST 800-53 Rev 5 exclusively.
# No PSC control renumbers between Rev 4 and Rev 5 (AC-2(13) and CM-7(5) are
# identical in both), so a single Rev 5 fixture resolves 100% clean against the
# shipped Rev-4-only demo DISA source. Rev 3/Rev 4 fixtures are not emitted; the
# rev-specific machinery below is retained so a future revision can be re-added.
REVISIONS: tuple[str, ...] = ("5",)

# ---------------------------------------------------------------------------
# Rows that are identical across every supported revision (the control IDs
# resolve against both Rev 3 and Rev 4 of the demo DISA source).
# (control_no, security_control, threshold_shall, objective_shall)
# ---------------------------------------------------------------------------
INVARIANT_ROWS: list[tuple[str, str, str, str]] = [
    (
        "EP-AC-002", "AC-6(1), AC-6(2)",
        "The Example Program ground system shall authorize access for "
        "privileged functions only to designated administrators using "
        "separate privileged accounts.",
        "Verify privileged functions are restricted to named administrators "
        "and that privileged actions use non-default accounts.",
    ),
    (
        "EP-AC-003", "AC-17(2)",
        "The Example Program ground system shall protect the confidentiality "
        "and integrity of remote access sessions using FIPS 140-3 validated "
        "cryptographic mechanisms.",
        "Verify remote sessions are tunneled through a validated cryptographic "
        "module and that the module certificate is current.",
    ),
    (
        # * structured col blank on purpose -> exercises the CNSSI prose anchor
        "EP-AU-001", "",
        "The Example Program ground system shall generate audit records for "
        "the program-defined auditable events. Associated CNSSI 1253 Control "
        "Tag: AU-2",
        "Verify the program auditable-event list is implemented and that "
        "records are produced for each event type.",
    ),
    (
        "EP-AU-002", "AU-9(4)",
        "The Example Program ground system shall restrict management of audit "
        "logging functionality to a subset of privileged users separate from "
        "system administrators.",
        "Verify audit-management privileges are segregated from general "
        "system administration.",
    ),
    (
        "EP-CM-001", "CM-2(2)",
        "The Example Program ground system shall maintain a current baseline "
        "configuration using an automated configuration management tool with "
        "drift detection.",
        "Verify an automated tool maintains the baseline and reports "
        "unauthorized configuration changes.",
    ),
    (
        # * prose-anchor path, multi-control list
        "EP-IA-001", "",
        "The Example Program ground system shall implement multifactor "
        "authentication for all interactive accounts. Associated CNSSI 1253 "
        "Control Tag: IA-2(1), IA-2(2)",
        "Verify MFA is enforced for both privileged and non-privileged "
        "interactive logons.",
    ),
    (
        "EP-IA-002", "IA-5(1)",
        "The Example Program ground system shall enforce a minimum password "
        "length of 15 characters and complexity meeting the program standard.",
        "Verify the enforced password policy meets or exceeds the program "
        "threshold for length and complexity.",
    ),
    (
        "EP-SC-001", "SC-7(3)",
        "The Example Program ground system shall limit external interfaces to "
        "the program-approved managed interfaces and deny all other traffic "
        "by default.",
        "Verify only approved managed interfaces are exposed and that the "
        "default boundary posture is deny-all.",
    ),
    (
        "EP-SC-002", "SC-8(1)",
        "The Example Program ground system shall protect the confidentiality "
        "and integrity of transmitted information using validated encryption "
        "between all ground nodes.",
        "Verify node-to-node traffic is encrypted with a validated module and "
        "that plaintext fallbacks are disabled.",
    ),
    (
        "EP-SI-001", "SI-4(2)",
        "The Example Program ground system shall employ automated tools to "
        "support near real-time analysis of monitoring events.",
        "Verify a SIEM or equivalent performs automated near-real-time "
        "correlation of monitored events.",
    ),
    (
        "EP-SI-002", "SI-7(1)",
        "The Example Program ground system shall perform integrity checks of "
        "program-defined software and firmware at startup and on a defined "
        "frequency.",
        "Verify integrity verification runs at startup and on the program "
        "frequency, and that failures are alerted.",
    ),
    (
        "EP-RA-001", "RA-5(2)",
        "The Example Program ground system shall update the vulnerability "
        "scanning tool's plugin set prior to each authenticated scan.",
        "Verify the scanner plugin feed is updated before each scan cycle.",
    ),
]

# ---------------------------------------------------------------------------
# Rows whose control ID (and matching shall text) changes by revision because
# the relevant 800-53 enhancement is numbered/scoped differently per rev.
# Keyed by req_number, then revision. Every revision in REVISIONS must have an
# entry whose control IDs resolve in that revision's DISA CCI set.
# ---------------------------------------------------------------------------
REV_SPECIFIC_ROWS: dict[str, dict[str, tuple[str, str, str, str]]] = {
    "EP-AC-001": {
        # AC-2(13) "Disable Accounts for High-risk Individuals" exists in both
        # Rev 4 and Rev 5; Rev 3 has no equivalent (would use AC-2(3)).
        "5": (
            "EP-AC-001", "AC-2(13)",
            "The Example Program ground system shall disable accounts of "
            "users posing a significant risk within 30 minutes of discovery "
            "of the risk.",
            "Verify that high-risk accounts are disabled within the program "
            "threshold and that the disabling action is logged.",
        ),
        "4": (
            "EP-AC-001", "AC-2(13)",
            "The Example Program ground system shall disable accounts of "
            "users posing a significant risk within 30 minutes of discovery "
            "of the risk.",
            "Verify that high-risk accounts are disabled within the program "
            "threshold and that the disabling action is logged.",
        ),
        "3": (
            "EP-AC-001", "AC-2(3)",
            "The Example Program ground system shall automatically disable "
            "inactive accounts after the program-defined period of inactivity.",
            "Verify that inactive accounts are disabled within the program "
            "threshold and that the disabling action is logged.",
        ),
    },
    "EP-CM-002": {
        # Application allow-list ("allow-by-exception"): CM-7(5) in BOTH Rev 4
        # and Rev 5 -- it did NOT renumber to CM-7(8) (that is a different,
        # unrelated control). CM-7(2) "Prevent Program Execution" was the Rev 3
        # form.
        "5": (
            "EP-CM-002", "CM-7(5)",
            "The Example Program ground system shall prevent execution of "
            "unauthorized software via an application allow-list "
            "(allow-by-exception) enforced in blocking mode.",
            "Verify the allow-list is enforced in blocking (not audit) mode "
            "and that unlisted binaries are denied.",
        ),
        "4": (
            "EP-CM-002", "CM-7(5)",
            "The Example Program ground system shall prevent execution of "
            "unauthorized software via an application allow-list "
            "(allow-by-exception) enforced in blocking mode.",
            "Verify the allow-list is enforced in blocking (not audit) mode "
            "and that unlisted binaries are denied.",
        ),
        "3": (
            "EP-CM-002", "CM-7(2)",
            "The Example Program ground system shall prevent program "
            "execution of unauthorized software (deny-by-default) per the "
            "program software policy.",
            "Verify unauthorized binaries are blocked from executing and that "
            "the policy is enforced, not audit-only.",
        ),
    },
}

# Stable display order by req_number prefix. The two rev-specific rows are
# slotted into their natural family position.
ROW_ORDER: list[str] = [
    "EP-AC-001",  # rev-specific
    "EP-AC-002",
    "EP-AC-003",
    "EP-AU-001",
    "EP-AU-002",
    "EP-CM-001",
    "EP-CM-002",  # rev-specific
    "EP-IA-001",
    "EP-IA-002",
    "EP-SC-001",
    "EP-SC-002",
    "EP-SI-001",
    "EP-SI-002",
    "EP-RA-001",
]


def _rows_for_revision(rev: str) -> list[tuple[str, str, str, str]]:
    """Assemble the full ordered row list for one revision."""
    invariant = {r[0]: r for r in INVARIANT_ROWS}
    by_req: dict[str, tuple[str, str, str, str]] = {}
    for req in ROW_ORDER:
        if req in REV_SPECIFIC_ROWS:
            specific = REV_SPECIFIC_ROWS[req]
            if rev not in specific:
                raise KeyError(
                    f"{req} has no row defined for revision {rev}; add one to "
                    "REV_SPECIFIC_ROWS before building it."
                )
            by_req[req] = specific[rev]
        else:
            by_req[req] = invariant[req]
    return [by_req[req] for req in ROW_ORDER]


def _out_path(rev: str) -> Path:
    return OUT_DIR / f"Example_Program_Ground_Security_Controls_PSC_800-53r{rev}.xlsx"


def build_one(rev: str) -> tuple[Path, int]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = _rows_for_revision(rev)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="548235")  # green = program overlay
    header_font = Font(bold=True, color="FFFFFF")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    # --- Rows 1-4: banner (loader scans first 10 rows for the header) ---
    # IMPORTANT: keep banner prose free of any header-alias substring that
    # gates header detection -- specifically "security control" (a
    # requirement_text/control_id alias) and "cci" (a cci_refs alias). The
    # loader's _find_header_row returns the FIRST row in which a
    # requirement_text or cci_refs alias resolves to a populated column, so a
    # banner cell containing those words is mistaken for the header and the
    # real tabular header at row 6 is never reached. Real T1TL overlays keep
    # those words out of their preamble columns, which is why they load
    # cleanly; the demo must do the same. (The sheet *title* may still say
    # "Ground Security Controls" -- the title is not scanned.)
    ws["A1"] = (
        f"EXAMPLE PROGRAM - GROUND SYSTEM OVERLAY "
        f"(PSC, 800-53 Rev {rev}, DEMO COPY)"
    )
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    ws["A2"] = "Program: Example Program (synthetic)"
    ws["A2"].font = bold
    ws["A3"] = (
        f"Overlay type: Program-Specific Controls (control-grain, T1TL-style); "
        f"control IDs valid for NIST SP 800-53 Revision {rev}"
    )
    ws["A4"] = (
        "Notice: SYNTHETIC demo overlay. Control IDs resolve to DISA "
        "control-correlation tags via the global mapping at load time -- load "
        f"the DISA mapping at Rev {rev} into the framework you resolve this PSC "
        "against. Not real program data."
    )
    ws["A4"].font = Font(italic=True, color="C00000")
    ws.merge_cells("A4:D4")

    # --- Row 6: headers ---
    header_row = 6
    for i, (label, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 30

    # --- Rows 7+: data ---
    for ri, row in enumerate(rows, start=header_row + 1):
        for ci, value in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = wrap_top
            cell.border = border
        ws.row_dimensions[ri].height = 64

    ws.freeze_panes = "A7"

    path = _out_path(rev)
    wb.save(str(path))
    return path, len(rows)


def build() -> list[Path]:
    return [build_one(rev)[0] for rev in REVISIONS]


if __name__ == "__main__":
    for rev in REVISIONS:
        path, n = build_one(rev)
        rows = _rows_for_revision(rev)
        direct = sum(1 for r in rows if r[1])
        anchored = sum(1 for r in rows if not r[1])
        print(
            f"WROTE  {path.relative_to(path.parent.parent)} "
            f"(Rev {rev}: {n} reqs: {direct} structured + {anchored} CNSSI-anchored)"
        )
