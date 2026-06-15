"""Generate a sample CCIS workbook for the Example System Demo system.

Mirrors the eMASS CCIS layout that backend/cybersecurity_assessor/excel/ccis_reader.py
and ccis_writer.py expect: WORKING SHEET, headers at row 6, data from row 7,
columns A-U.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).parent / "ccis" / "CCIS_Example System_Demo_System_2026May.xlsx"

# Column layout (1-based) -- MUST match ccis_reader.py
COLS = [
    ("Required", 10),
    ("Control", 14),
    ("Information", 28),
    ("Implementation Status", 18),
    ("Designation", 18),
    ("Implementation Narrative (Col F)", 60),
    ("AP Acronym", 14),
    ("CCI", 14),
    ("CCI Definition (Col I)", 60),
    ("Implementation Guidance (Col J)", 60),
    ("Assessment Procedures (Col K)", 60),
    ("Inherited (Col L)", 18),
    ("Remote Inheritance (Col M)", 22),
    ("Compliance Status (Col N)", 20),
    ("Date Tested (Col O)", 16),
    ("Tested By (Col P)", 22),
    ("Test Results (Col Q)", 70),
    ("Originating Requirement (R)", 22),
    ("Implementation Type (S)", 18),
    ("Previous Assessor (T)", 22),
    ("Previous Test Results (Col U)", 70),
]

# ---------------------------------------------------------------------------
# Sample CCI rows
# ---------------------------------------------------------------------------
# Tuple order matches COLS. Use "" for blanks.
# N/O/P/Q (status/date/tester/results) are sometimes prefilled (Compliant)
# and sometimes blank so the assessor has something to do.

EXAMPLE_SYSTEM = "Example System Example System Demo IATT"
TESTER = "Noah Jaskolski"

ROWS: list[tuple] = [
    # Required, Control, Info,
    # ImplStatus, Designation,
    # Narrative,
    # AP, CCI,
    # Definition,
    # Guidance,
    # Procedures,
    # Inherited, RemoteInheritance,
    # Status (N), DateTested (O), Tester (P), Results (Q),
    # OriginatingReq (R), ImplType (S), PrevAssessor (T), PrevResults (U)

    # --- AC-1 (policy & procedures) -- rule 8a: DoD-level auto-compliant (col K) ---
    (
        "Yes", "AC-1", "Policy and Procedures",
        "Implemented", "",
        "Access control policy and procedures are established at the DoD "
        "enterprise level and adopted by the Example System Demo system.",
        "AC-1", "CCI-000001",
        "The organization develops, documents, and disseminates an access "
        "control policy.",
        "Provide the governing access control policy. Cite the document number.",
        "This control is covered at the DoD level via DoDI 8500.01; no system-"
        "level test is required for the policy baseline.",
        "", "",
        "Compliant", "2026-05-19", TESTER,
        "Access control policy is inherited from the DoD enterprise policy set "
        "(DoDI 8500.01) and requires no system-level assessment.",
        "NIST 800-53", "Common", "T. Prior",
        "",
    ),

    # --- AC-2.1 (account management) -- Compliant, prefilled ---
    (
        "Yes", "AC-2.1", "Account Management",
        "Implemented", "",
        "The Example System Demo system implements account management procedures per "
        "the Information System Account Management Policy (USD20240315). "
        "Account types are documented (Section 3); approval authority is "
        "ISSM + supervisor; account lifecycle steps are defined in Section 4.",
        "AC-2.1", "CCI-000015",
        "The organization defines the information system account types to "
        "be identified and selected to support organizational missions/"
        "business functions.",
        "Provide the policy or procedure that identifies and selects the "
        "account types. Document numbers should be cited.",
        "Examine the account management policy. Verify account types are "
        "identified. Interview the ISSM. Test by creating a sample account.",
        "Local", "",
        "Compliant", "2026-05-19", TESTER,
        "Examined Information System Account Management Policy "
        "(USD20240315) Section 3, which identifies five account types: "
        "individual user, privileged, service/application, emergency, and "
        "temporary. Verified approval authority and review intervals are "
        "documented for each. Examined Windows account inventory dated "
        "2026-05-19; sampled five accounts and confirmed each maps to one "
        "of the documented types.",
        "NIST 800-53", "Common", "T. Prior",
        "Examined Account Management SOP (USD20210715) and verified the "
        "five account types are still documented.",
    ),

    # --- AC-2.4 (account audit) -- requires audit log evidence ---
    (
        "Yes", "AC-2.4", "Account Management",
        "Implemented", "",
        "Account creation, modification, enabling, disabling, and removal "
        "events are audited via Windows Security Event Log and forwarded "
        "to the Example System SIEM. Weekly review per AU-6 procedures.",
        "AC-2.4", "CCI-000018",
        "The information system automatically audits account creation "
        "actions.",
        "Provide evidence that account-creation events are captured and "
        "reviewed (Event ID 4720 on Windows).",
        "Examine audit log; verify Event ID 4720 is captured. Examine "
        "review records.",
        "Local", "",
        "", "", "", "",
        "NIST 800-53", "System-Specific", "T. Prior",
        "Examined Splunk index audit-win for the period 2025-11 through "
        "2026-01; confirmed Event ID 4720 captured for all new accounts "
        "in that window.",
    ),

    # --- AC-7.a (unsuccessful logon attempts) -- Compliant, prefilled ---
    (
        "Yes", "AC-7.a", "Unsuccessful Logon Attempts",
        "Implemented", "",
        "The system enforces a limit of 5 invalid logon attempts in a 15-"
        "minute window via the Default Domain Policy GPO.",
        "AC-7.a", "CCI-000044",
        "The organization defines the number of consecutive invalid logon "
        "attempts allowed to the information system by a user during an "
        "organization-defined time period.",
        "Provide the GPO export showing the lockout threshold and reset "
        "counter.",
        "Examine the GPO export. Verify threshold is consistent with the "
        "organizational definition.",
        "Local", "",
        "Compliant", "2026-05-19", TESTER,
        "Examined GPO_Password_Policy_Export.xlsx (USD20240218) row 'Account "
        "lockout threshold' value '5 invalid logon attempts' and row 'Reset "
        "account lockout counter after' value '15 minutes'. Threshold "
        "matches the AC-7 organizationally defined value.",
        "NIST 800-53", "Common", "T. Prior",
        "Examined gpresult /h output from 2025-12-04 showing lockout "
        "threshold = 3 attempts. (NOTE: setting was tightened to 5 in 2026.)",
    ),

    # --- AC-11 (session lock) -- blank, ready for assessment ---
    (
        "Yes", "AC-11", "Session Lock",
        "", "",
        "",
        "AC-11", "CCI-000056",
        "The information system prevents further access to the system by "
        "initiating a session lock after an organization-defined time "
        "period of inactivity.",
        "Provide the GPO setting that enforces inactivity lockout.",
        "Examine GPO or local policy. Verify the inactivity timeout.",
        "Local", "",
        "", "", "", "",
        "NIST 800-53", "Common", "T. Prior",
        "",
    ),

    # --- AC-18 (wireless access) -- rule 8b: NA via documented scope exclusion (col Q) ---
    (
        "Yes", "AC-18", "Wireless Access",
        "Not Applicable", "",
        "The Example System Demo authorization boundary contains no wireless "
        "networking components.",
        "AC-18", "CCI-001438",
        "The organization establishes configuration requirements, connection "
        "requirements, and implementation guidance for each type of wireless "
        "access.",
        "If wireless is in the boundary, provide the wireless configuration "
        "policy. Otherwise document the scope exclusion.",
        "Examine the network diagram and asset inventory. Confirm whether any "
        "wireless components are in scope.",
        "", "",
        "Not Applicable", "2026-05-20", TESTER,
        "Per system scoping, this CCI is not applicable -- the authorization "
        "boundary contains no wireless networking components. Examined the "
        "asset inventory and ACAS scan; no wireless interfaces detected.",
        "NIST 800-53", "System-Specific", "T. Prior",
        "",
    ),

    # --- AT-2.1 (security awareness training) -- Compliant ---
    (
        "Yes", "AT-2.1", "Security Awareness Training",
        "Implemented", "",
        "All Example System Demo users complete annual security awareness training "
        "(USD20240518); roster maintained in the LMS.",
        "AT-2.1", "CCI-000100",
        "The organization provides basic security awareness training to "
        "information system users (including managers, senior executives, "
        "and contractors) as part of initial training for new users.",
        "Provide the training brief and the completion roster.",
        "Examine the training material. Sample the roster to verify "
        "completion within 30 days of access.",
        "Local", "",
        "Compliant", "2026-05-19", TESTER,
        "Examined Security_Awareness_Training_Brief_2026Q2 (USD20240518) "
        "covering insider threat (AT-2(2)), phishing, CUI handling, and "
        "privileged user duties. Examined LMS completion roster dated "
        "2026-05-15; sampled 10 of 42 active users, all show completion "
        "within 30 days of access grant.",
        "NIST 800-53", "Common", "T. Prior",
        "Examined prior training brief from 2025; confirmed annual "
        "refresh cadence is documented.",
    ),

    # --- AU-4 (audit storage capacity) -- rule 8a: internal inheritance (col K) ---
    (
        "Yes", "AU-4", "Audit Log Storage Capacity",
        "Inherited", "",
        "Audit log storage is provided by the enterprise logging service; the "
        "Example System Demo system forwards records and retains no local "
        "long-term audit store.",
        "AU-4", "CCI-000140",
        "The organization allocates audit record storage capacity in "
        "accordance with organization-defined audit record storage "
        "requirements.",
        "Cite the providing system if audit storage is inherited.",
        "Audit record storage capacity is inherited from the enterprise "
        "(SDA Enterprise Logging Service).",
        "", "",
        "Compliant", "2026-05-22", TESTER,
        "Storage capacity for audit records is provided by the SDA Enterprise "
        "Logging Service. Local forwarding is configured; no local retention "
        "store is maintained.",
        "NIST 800-53", "Hybrid", "T. Prior",
        "",
    ),

    # --- AU-6.1 (audit review) -- Compliant; cites text evidence ---
    (
        "Yes", "AU-6.1", "Audit Review, Analysis, and Reporting",
        "Implemented", "",
        "The ISSO reviews Security Event Logs weekly and the SIEM produces "
        "weekly correlation reports.",
        "AU-6.1", "CCI-000148",
        "The organization reviews and analyzes information system audit "
        "records weekly for indications of organization-defined "
        "inappropriate or unusual activity.",
        "Provide a sample weekly review record and a SIEM correlation "
        "report.",
        "Examine the review record. Examine the SIEM report. Verify the "
        "review covered the defined period.",
        "Local", "",
        "Compliant", "2026-05-22", TESTER,
        "Examined audit_log_review_2026-05-19.txt covering period 2026-05-"
        "12 to 2026-05-18, signed by ISSO. Examined siem_weekly_"
        "correlation_report_2026-05-22.txt covering period 2026-05-15 to "
        "2026-05-21. Both records show the weekly review was performed and "
        "all flagged events were dispositioned.",
        "NIST 800-53", "System-Specific", "T. Prior",
        "Examined Splunk dashboard 'Example System Demo Weekly Review' showing 6 of 6 "
        "weeks completed in Q4 2025.",
    ),

    # --- IA-5(1)(a) (password complexity) -- Compliant ---
    (
        "Yes", "IA-5 (1) (a)", "Authenticator Management - Password Complexity",
        "Implemented", "",
        "The Default Domain Policy enforces a 15-character minimum and "
        "complexity requirements per Identification and Authentication "
        "Procedures (USD20240212) Section 3.",
        "IA-5 (1) (a)", "CCI-000192",
        "The information system enforces minimum password complexity of "
        "case sensitivity, number of characters, mix of upper-case "
        "letters, lower-case letters, numbers, and special characters, "
        "including minimum requirements for each type.",
        "Provide the GPO export showing complexity = Enabled and minimum "
        "length.",
        "Examine the GPO export. Attempt to set a non-compliant password "
        "and verify rejection.",
        "Local", "",
        "Compliant", "2026-05-19", TESTER,
        "Examined GPO_Password_Policy_Export.xlsx (USD20240218) row "
        "'Minimum password length' value '15 characters' and row "
        "'Password must meet complexity requirements' value 'Enabled'. "
        "Both exceed the 800-53 baseline.",
        "NIST 800-53", "Common", "T. Prior",
        "Examined Default Domain Policy export from 2025-12-04; same "
        "settings observed (length 15, complexity Enabled).",
    ),

    # --- CM-6 (configuration settings) -- Compliant; cites STIG ---
    (
        "Yes", "CM-6.1", "Configuration Settings",
        "Implemented", "",
        "The Example System Demo workstation is hardened to the Windows Server 2022 "
        "STIG. Most recent assessment recorded in "
        "Windows_Server_2022_STIG_Sample.ckl on 2026-05-18.",
        "CM-6.1", "CCI-000363",
        "The organization establishes and documents configuration settings "
        "for information technology products employed within the "
        "information system using organization-defined security "
        "configuration checklists that reflect the most restrictive mode "
        "consistent with operational requirements.",
        "Provide a STIG checklist (.ckl/.cklb) for each major OS / "
        "application in the boundary.",
        "Examine the checklist. Verify open findings are tracked in the "
        "POA&M.",
        "Local", "",
        "Compliant", "2026-05-19", TESTER,
        "Examined Windows_Server_2022_STIG_Sample.ckl dated 2026-05-18, "
        "evaluated against the Windows Server 2022 STIG. Open findings "
        "are tracked. Also examined RHEL_9_STIG_Sample.cklb for the "
        "supporting Linux host. STIG checklists are present for all "
        "major OSes in the boundary.",
        "NIST 800-53", "System-Specific", "T. Prior",
        "Examined STIG checklists from Q4 2025; same OS coverage.",
    ),

    # --- PE-3 (physical access control) -- rule 8a: CSP-provided inheritance (col Q) ---
    (
        "Yes", "PE-3", "Physical Access Control",
        "Inherited", "",
        "Physical access control for the hosting facility is provided by the "
        "cloud service provider; the Example System Demo system has no "
        "on-premises data center footprint.",
        "PE-3", "CCI-000919",
        "The organization enforces physical access authorizations at "
        "organization-defined entry/exit points to the facility where the "
        "information system resides.",
        "If hosting is in a CSP, cite the provider authorization. Otherwise "
        "provide the facility access control records.",
        "Examine the inheritance source. If on-premises, examine the badge "
        "system records.",
        "", "",
        "Compliant", "2026-05-20", TESTER,
        "Physical access control is implemented by AWS (AWS GovCloud) and "
        "inherited via the FedRAMP High authorization. The program operates "
        "no on-premises facility for this boundary; facility entry/exit "
        "enforcement is the provider's responsibility.",
        "NIST 800-53", "Inherited", "T. Prior",
        "",
    ),

    # --- CP-7 (alternate processing site) -- rule 8c: UNCLEAR bare "inherited from" (col K) ---
    (
        "Yes", "CP-7", "Alternate Processing Site",
        "", "",
        "Alternate processing arrangements are referenced but the providing "
        "system is not yet identified in the package.",
        "CP-7", "CCI-000146",
        "The organization establishes an alternate processing site including "
        "necessary agreements to permit the transfer and resumption of "
        "information system operations for essential missions/business "
        "functions.",
        "Cite the alternate site agreement or the providing system if "
        "inherited.",
        "Alternate processing is inherited from a provider; the providing "
        "system is not documented in the current package.",
        "", "",
        "", "", "", "",
        "NIST 800-53", "Hybrid", "T. Prior",
        "",
    ),

    # --- RA-5 (vulnerability scanning) -- Compliant; cites Nessus ---
    (
        "Yes", "RA-5.1", "Vulnerability Monitoring and Scanning",
        "Implemented", "",
        "ACAS / Tenable.sc credentialed scans are run against the Example System Demo "
        "subnet on a defined cadence; results managed in Tenable.sc.",
        "RA-5.1", "CCI-001067",
        "The organization scans for vulnerabilities in the information "
        "system and hosted applications and when new vulnerabilities "
        "potentially affecting the system/applications are identified and "
        "reported.",
        "Provide a recent ACAS/Nessus scan report covering the boundary.",
        "Examine the scan report. Verify the scan was credentialed and "
        "covered the in-scope subnet.",
        "Local", "",
        "Compliant", "2026-05-21", TESTER,
        "Examined acas_nessus_subnet_10_10_5_0.nessus dated 2026-05-20, "
        "policy 'Example System ACAS Credentialed Scan - Demo Subnet'. Plugin output "
        "confirms credentialed scan. ReportHost example-system-demo-ws01.demo.local "
        "(10.10.5.21) in scope; Server 2022 host scanned successfully.",
        "NIST 800-53", "System-Specific", "T. Prior",
        "Examined ACAS scan from 2025-12-15; subnet 10.10.5.0/24 covered.",
    ),

    # --- SC-7 (boundary protection) -- not applicable per inheritance ---
    (
        "Yes", "SC-7.1", "Boundary Protection",
        "Inherited", "Inherited",
        "Boundary protection is inherited from the enclosing SDA Enterprise "
        "Service. Per SDA Controls overlay row 412, SC-7 is inheritable "
        "for this system.",
        "SC-7.1", "CCI-001097",
        "The information system monitors and controls communications at "
        "the external boundary of the system and at key internal "
        "boundaries within the system.",
        "If inheriting, cite the provider system. Otherwise provide the "
        "firewall configuration and ACL.",
        "Examine the inheritance memorandum. If implementing locally, "
        "examine the firewall config.",
        "SDA Enterprise Service", "SDA Enterprise Service",
        "Compliant", "2026-05-19", TESTER,
        "Inheritance from SDA Enterprise Service is documented in the SDA "
        "Controls overlay row 412. The Example System Demo system does not implement "
        "boundary protection locally; rule sets and ingress/egress "
        "filtering are managed by the SDA enterprise edge.",
        "NIST 800-53", "Inherited", "T. Prior",
        "Examined prior SDA Controls overlay; SC-7 was previously marked "
        "inheritable. No change.",
    ),

    # --- SC-13 (cryptographic protection) -- LLM feeder, expected Non-Compliant ---
    (
        "Yes", "SC-13", "Cryptographic Protection",
        "Implemented", "",
        "Most services use FIPS 140-3 validated modules, but a legacy "
        "interface component has not yet been migrated.",
        "SC-13", "CCI-002450",
        "The information system implements organization-defined cryptographic "
        "uses and type of cryptography required for each use in accordance "
        "with applicable federal laws, Executive Orders, directives, "
        "policies, regulations, and standards.",
        "Provide the FIPS 140-3 validation certificates for the cryptographic "
        "modules in use.",
        "Examine the module inventory. Verify each module has a current "
        "FIPS 140-3 (or 140-2 transitional) certificate.",
        "Local", "",
        "", "", "", "",
        "NIST 800-53", "System-Specific", "T. Prior",
        "Examined the FIPS module inventory in Q4 2025; one legacy interface "
        "component (LEGACY-XFER-01) was still using a non-validated "
        "cryptographic library and was open in the POA&M.",
    ),

    # --- SI-4 (system monitoring) -- partial; ready for assessor ---
    (
        "Yes", "SI-4.1", "System Monitoring",
        "Implemented", "",
        "SIEM coverage is verified weekly via the correlation report.",
        "SI-4.1", "CCI-002645",
        "The organization monitors the information system to detect "
        "attacks and indicators of potential attacks in accordance with "
        "organization-defined monitoring objectives.",
        "Provide the SIEM correlation report and a list of monitored log "
        "sources.",
        "Examine the SIEM report. Verify each in-boundary asset has a "
        "healthy log source.",
        "Local", "",
        "", "", "", "",
        "NIST 800-53", "System-Specific", "T. Prior",
        "",
    ),
]


def build() -> Path:
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "WORKING SHEET"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    inheritable_fill = PatternFill("solid", fgColor="FFF2CC")

    # --- Rows 1-5: metadata banner (mimics eMASS export) ---
    ws["A1"] = "Example System DEMO - CCIS WORKBOOK (DEMO COPY)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")

    ws["A2"] = f"System: {EXAMPLE_SYSTEM}"
    ws["A2"].font = bold
    ws["H2"] = "eMASS Export Date:"
    ws["I2"] = "2026-05-15"

    ws["A3"] = "Framework: NIST SP 800-53 Rev. 5"
    ws["A4"] = "Assessment Lead: Noah Jaskolski, ISSO"
    ws["A5"] = (
        "Notice: This is a SYNTHETIC demo workbook used to exercise the "
        "cybersecurity-assessor app. Do NOT use for a real authorization."
    )
    ws["A5"].font = Font(italic=True, color="C00000")
    ws.merge_cells("A5:U5")

    # --- Row 6: headers ---
    for i, (label, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=6, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[6].height = 42

    # --- Rows 7+: data ---
    for ri, row in enumerate(ROWS, start=7):
        for ci, value in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = wrap_top
            cell.border = border
        # Highlight inheritance-tagged rows so they're scannable
        if row[12]:  # Col M Remote Inheritance names a source
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=ri, column=ci).fill = inheritable_fill
        ws.row_dimensions[ri].height = 110

    # Freeze panes below header row, after Control column
    ws.freeze_panes = "C7"

    # Data validation on col N (Compliance Status)
    dv = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant,Not Applicable"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Status must be Compliant, Non-Compliant, or Not Applicable."
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add(f"N7:N{7 + len(ROWS) - 1}")

    # --- Second sheet: SDA Controls overlay stub (program-specific reqs) ---
    ws2 = wb.create_sheet("SDA Controls")
    ws2["A1"] = "SDA Controls Overlay (extract)"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:F1")
    hdr = ["Row", "Control", "CCI", "Requirement Source", "Notes", "Inheritable"]
    for c, h in enumerate(hdr, start=1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    overlay_rows = [
        (101, "AC-2.1", "CCI-000015", "SDA Enterprise Identity Service", "Local account procedures still required.", "No"),
        (243, "AC-7.a", "CCI-000044", "SDA Enterprise Identity Service", "GPO must enforce 5/15.", "No"),
        (304, "AU-6.1", "CCI-000148", "SDA Enterprise Logging Service", "Weekly local review still required.", "No"),
        (412, "SC-7.1", "CCI-001097", "SDA Enterprise Boundary Service", "Fully inherited.", "Yes"),
        (518, "RA-5.1", "CCI-001067", "SDA Enterprise Vulnerability Mgmt", "Local scan execution still required.", "No"),
    ]
    for ri, r in enumerate(overlay_rows, start=4):
        for ci, v in enumerate(r, start=1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.alignment = wrap_top
    for col_letter, width in zip("ABCDEF", (8, 14, 16, 36, 50, 14)):
        ws2.column_dimensions[col_letter].width = width

    wb.save(str(OUT))
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"WROTE  {path.relative_to(path.parent.parent)} ({len(ROWS)} CCI rows)")
