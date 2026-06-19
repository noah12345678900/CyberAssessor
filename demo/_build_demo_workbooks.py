"""Generate one importable demo workbook per framework.

Every file mirrors the eMASS CCIS WORKING SHEET layout that
backend/cybersecurity_assessor/excel/ccis_reader.py expects: WORKING SHEET tab,
metadata banner rows 1-5, headers at row 6, data from row 7, columns A-U, and a
data-validation list on column N (Compliance Status).

The only thing that varies per framework is:
  - the banner label in cell A3,
  - the column-B control acronyms (each framework's native control-ID format),
  - the column-H "CCI" key. Only NIST 800-53 publishes DoD CCIs; every other
    framework's lowest assessable unit IS the requirement itself, so column H
    repeats the column-B id (the reader keys one Objective per row off it).

Import binds rows to the framework the USER PICKS in the ComplianceTargetPicker
(the A3 banner is cosmetic). For the license-aware catalogs (ISO 27001, CIS v8,
PCI DSS, SOC 2) the workbook imports structurally even before the org uploads
its licensed catalog export -- objectives just won't resolve to catalog control
text until that export is loaded.

One output file per framework lands in demo/workbooks/.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = Path(__file__).parent / "workbooks"

EXAMPLE_SYSTEM = "Example System Example System Demo IATT"
TESTER = "Noah Jaskolski"

# Column layout (1-based) -- MUST match ccis_reader.py
COLS = [
    ("Required", 10),
    ("Control", 16),
    ("Information", 28),
    ("Implementation Status", 18),
    ("Designation", 18),
    ("Implementation Narrative (Col F)", 60),
    ("AP Acronym", 16),
    ("CCI", 16),
    ("CCI Definition (Col I)", 60),
    ("Implementation Guidance (Col J)", 60),
    ("Assessment Procedures (Col K)", 60),
    ("Inherited (Col L)", 18),
    ("Remote Inheritance (Col M)", 22),
    ("Compliance Status (Col N)", 20),
    ("Date Tested (Col O)", 16),
    ("Tested By (Col P)", 22),
    ("Test Results (Col Q)", 70),
    ("Originating Requirement (R)", 22),
    ("Implementation Type (S)", 18),
    ("Previous Assessor (T)", 22),
    ("Previous Test Results (Col U)", 70),
]


@dataclass
class Req:
    """One assessable requirement -> one workbook data row.

    control: column-B acronym in the framework's native format.
    cci:     column-H key. "" lets the builder default it to `control`.
    """

    control: str
    title: str
    family: str
    narrative: str
    definition: str
    guidance: str
    procedures: str
    status: str = ""          # col N: "", "Compliant", "Non-Compliant", "Not Applicable"
    results: str = ""         # col Q
    cci: str = ""             # col H ("" -> defaults to control)
    inherited: str = "No"     # col L
    remote: str = ""          # col M
    framework_tag: str = ""   # col R originating requirement


@dataclass
class FrameworkDemo:
    framework_label: str      # cell A3 banner
    filename: str
    reqs: list[Req]


# ---------------------------------------------------------------------------
# NIST SP 800-53 Rev. 5 -- DoD CCIs in column H
# ---------------------------------------------------------------------------
NIST_800_53 = FrameworkDemo(
    framework_label="NIST SP 800-53 Rev. 5",
    filename="DEMO_NIST_800-53r5_Example System.xlsx",
    reqs=[
        Req(
            "AC-2 (1)", "Account Management | Automated System Account Management",
            "AC", "The Example System Demo system uses automated account management via "
            "Active Directory; account lifecycle actions are scripted and logged.",
            "The organization employs automated mechanisms to support the "
            "management of information system accounts.",
            "Provide the AD/IdAM configuration showing automated provisioning.",
            "Examine the IdAM config. Verify provisioning/deprovisioning is "
            "automated. Sample an onboarding ticket.",
            status="Compliant",
            results="Examined Information System Account Management Policy "
            "(USD20240315), which documents the automated AD create/disable/remove "
            "workflow; confirmed account lifecycle actions are logged in the AD "
            "audit channel. Sampled three onboarding tickets; all show automated "
            "account creation.",
            cci="CCI-000015",
        ),
        Req(
            "AC-7 a", "Unsuccessful Logon Attempts",
            "AC", "The Default Domain Policy enforces a lockout after 5 invalid "
            "attempts in a 15-minute window.",
            "The organization defines the number of consecutive invalid logon "
            "attempts allowed during an organization-defined time period.",
            "Provide the GPO export showing lockout threshold and reset counter.",
            "Examine the GPO export. Verify threshold matches the ODV.",
            status="Compliant",
            results="Examined GPO_Password_Policy_Export.xlsx (USD20240218): "
            "'Account lockout threshold' = '5 invalid logon attempts'; 'Reset "
            "account lockout counter after' = '15 minutes'.",
            cci="CCI-000044",
        ),
        Req(
            "AC-11", "Session Lock",
            "AC", "The Default Domain Policy enforces a session lock after 15 "
            "minutes of inactivity, requiring re-authentication to resume.",
            "The information system initiates a session lock after an "
            "organization-defined period of inactivity.",
            "Provide the GPO setting that enforces inactivity lockout.",
            "Examine GPO or local policy. Verify the inactivity timeout.",
            status="Compliant",
            results="Examined GPO_Password_Policy_Export.xlsx (USD20240218): "
            "'Interactive logon: Machine inactivity limit' = '900 seconds' "
            "(15 minutes), enforcing an automatic session lock.",
            cci="CCI-000057",
        ),
        Req(
            "IA-5 (1) (a)", "Authenticator Management | Password-Based Authentication",
            "IA", "The Default Domain Policy enforces a 15-character minimum "
            "password length per I&A Procedures (USD20240212) Section 3.",
            "The information system enforces minimum password length.",
            "Provide the GPO export showing complexity = Enabled and minimum "
            "length.",
            "Examine the GPO export. Attempt a non-compliant password and verify "
            "rejection.",
            status="Compliant",
            results="Examined GPO_Password_Policy_Export.xlsx (USD20240218): "
            "'Minimum password length' = '15 characters'; 'Password must meet "
            "complexity requirements' = 'Enabled'.",
            cci="CCI-000205",
        ),
        Req(
            "CM-6 a", "Configuration Settings",
            "CM", "The Example System Demo hosts are hardened to their DISA STIGs; "
            "the RHEL 9 and Windows Server 2022 checklists are the most recent "
            "configuration baselines on record.",
            "The organization implements the security configuration settings.",
            "Provide a STIG checklist for each major OS/application in scope.",
            "Examine the checklist. Verify open findings are tracked in the POA&M.",
            status="Compliant",
            results="Examined RHEL_9_STIG_Sample.cklb and "
            "Windows_Server_2022_STIG_Sample.ckl. Open findings are tracked in "
            "the POA&M.",
            cci="CCI-000366",
        ),
        Req(
            "RA-5 a", "Vulnerability Monitoring and Scanning",
            "RA", "ACAS/Tenable.sc credentialed scans run on a defined cadence "
            "against the Example System Demo subnet.",
            "The organization scans for vulnerabilities in the information system "
            "and hosted applications.",
            "Provide a recent ACAS/Nessus scan report covering the boundary.",
            "Examine the scan report. Verify credentialed and in-scope coverage.",
            status="Compliant",
            results="Examined acas_nessus_subnet_10_10_5_0.nessus dated "
            "2026-05-20; credentialed scan confirmed; host example-system-demo-ws01 in scope.",
            cci="CCI-001054",
        ),
        Req(
            "SC-7", "Boundary Protection",
            "SC", "Boundary protection is inherited from the enclosing SDA "
            "Enterprise Service per SDA Controls overlay row 412.",
            "The information system monitors and controls communications at the "
            "external boundary and key internal boundaries.",
            "If inheriting, cite the provider system; otherwise provide the "
            "firewall config and ACL.",
            "Examine the inheritance memorandum or the firewall config.",
            status="Compliant",
            results="Boundary protection is fully inherited from the SDA "
            "Enterprise Service per SDA Controls overlay row 412. An inherited "
            "control whose provider satisfies the requirement is Compliant "
            "(applicable but met by the provider) — not Not Applicable.",
            cci="CCI-001097", inherited="Yes", remote="SDA Enterprise Service",
        ),
        # 8th control — the multi-tenant divergence showcase. Left UNASSESSED
        # (blank status) so it flows to the assessor, and deliberately
        # responsibility-split across the two demo CRMs: AWS GovCloud marks
        # AC-17 "Customer" (the customer configures Client VPN + conditional
        # access), while Azure Government marks it "Inherited" (Azure Bastion is
        # a fully-managed remote-access plane). With BOTH CRMs attached, the
        # assessor must write a different per-scope narrative for each tenant —
        # the exact thing the boundary/CRM work exists to demonstrate.
        Req(
            "AC-17", "Remote Access",
            "AC", "Remote access to the Example System Demo workloads differs by "
            "cloud: the AWS GovCloud enclave uses customer-configured Client VPN "
            "with conditional access, while the Azure Government enclave brokers "
            "all administrative remote access through managed Azure Bastion.",
            "The organization authorizes remote access to the information system "
            "prior to connection and enforces approved remote-access methods.",
            "Provide the remote-access configuration for each cloud enclave; cite "
            "the provider-managed plane where access is inherited.",
            "Examine the remote-access method per scope. For the customer-owned "
            "scope verify VPN/conditional-access enforcement; for the inherited "
            "scope confirm the managed plane covers the requirement.",
            status="",
            results="",
            cci="CCI-000063", inherited="No",
        ),
        # 9th control — a genuine NOT APPLICABLE. AC-18 (Wireless Access) is
        # documented as out of scope: the Example System Demo boundary has no
        # wireless networking, so the control is Not Applicable (distinct from
        # inherited-Compliant on SC-7). Exercises the deterministic 8b
        # scope-exclusion -> NOT_APPLICABLE lane.
        Req(
            "AC-18", "Wireless Access",
            "AC", "The Example System Demo authorization boundary contains no "
            "wireless networking. All connectivity is wired within the data "
            "center enclave; wireless is not deployed and is prohibited by the "
            "system configuration baseline.",
            "The organization establishes configuration requirements and "
            "authorizes wireless access to the information system.",
            "If wireless is present, provide the wireless configuration; "
            "otherwise document the scope exclusion.",
            "Confirm the boundary has no wireless components. Verify the "
            "scope-exclusion statement in the SSP.",
            status="Not Applicable",
            results="Not Applicable — the Example System Demo boundary deploys no "
            "wireless networking; all access is wired within the data-center "
            "enclave. Wireless is prohibited by the configuration baseline, so "
            "AC-18 does not apply to this system.",
            cci="CCI-001438", inherited="No",
        ),
        # N/A-SLICE showcase #1 — ALL-SLICES-N/A rollup. PE-10 (Emergency
        # Shutoff) is "Not Applicable" on BOTH cloud CRMs (datacenter emergency
        # shutoff is a provider-internal facility control) AND column N is
        # prefilled "Not Applicable" so the flex (On-Premises/workbook) slice is
        # N/A too. Every slice is N/A → the control rolls up NOT APPLICABLE (the
        # "N/A only if ALL slices N/A" rule). The flex chip reads "N/A". Left
        # with a deterministic col-N N/A so it short-circuits via rule 8b.
        Req(
            "PE-10", "Emergency Shutoff",
            "PE", "Emergency power shutoff for the Example System Demo is a "
            "provider-internal datacenter facility control on every cloud "
            "enclave; the system has no customer-operated facility requiring an "
            "emergency power-shutoff capability.",
            "The organization provides the capability to shut off power to the "
            "information system or individual system components in emergencies.",
            "If a customer-operated facility exists, provide the emergency "
            "power-shutoff procedure; otherwise document the scope exclusion.",
            "Confirm emergency shutoff is provider-internal on every cloud and "
            "that no customer facility requires a local capability.",
            status="Not Applicable",
            results="Not Applicable — datacenter emergency power shutoff is a "
            "provider-internal facility control on every cloud enclave (AWS "
            "GovCloud, Azure Government) and the system operates no customer "
            "facility requiring an equivalent capability. PE-10 does not apply "
            "to any slice of this system.",
            cci="CCI-000813", inherited="No",
        ),
        # N/A-SLICE showcase #2 — PER-CLOUD N/A (mixed rollup). MP-6 (Media
        # Sanitization) is "Not Applicable" on the AWS GovCloud CRM (AWS performs
        # cryptographic erasure of customer-managed media on its own) but
        # "Inherited" on the Azure Government CRM. Column L = "No" so the flex
        # slice is customer-assessed. With both CRMs attached the grid shows an
        # AWS "N/A" chip beside an Azure "Inherited" chip — exercising a single
        # N/A slice inside a mixed worst-of rollup (one N/A + one Compliant +
        # the flex slice). Left UNASSESSED so it flows to the assessor.
        Req(
            "MP-6", "Media Sanitization",
            "MP", "Media sanitization for the Example System Demo differs by "
            "cloud: the AWS GovCloud enclave treats it as not applicable "
            "(provider performs cryptographic erasure of customer-managed media), "
            "while the Azure Government enclave inherits the control from the "
            "provider's media-sanitization plane.",
            "The organization sanitizes information system media prior to "
            "disposal, release, or reuse.",
            "Provide the media-sanitization procedure or the provider "
            "attestation per cloud enclave; document any scope exclusion.",
            "Examine media sanitization per scope: confirm the AWS N/A exclusion "
            "and the Azure inherited coverage; assess the flex scope locally.",
            status="",
            results="",
            cci="CCI-001028", inherited="No",
        ),
        # 10th control — the PIE-SLICE FULLY-INHERITED showcase. PE-3 (Physical
        # Access Control) is marked "Inherited" by BOTH demo CRMs (AWS GovCloud +
        # Azure Government own datacenter physical security), AND the workbook's
        # Column L names "DoW Enterprise" as the inheritance source for the flex
        # (On-Premises/workbook) slice — the facility itself is a DoW enterprise
        # data center whose physical-access plane the system inherits. Left
        # UNASSESSED (blank status) so it flows to the assessor: the two cloud
        # slices short-circuit Compliant-by-inheritance (crm_inherited) AND the
        # flex slice resolves Compliant from Column L (resolve_col_l_flex_status
        # -> INHERITED) WITHOUT an LLM call. The whole control rolls up Compliant.
        # This demonstrates the pie-slice authority split: Column L (eMASS) governs
        # the flex slice's status, the CRMs govern the cloud slices. (Contrast the
        # AU-9 customer-owned abstain and the AC-18 scope-exclusion NA.)
        Req(
            "PE-3", "Physical Access Control",
            "PE", "Physical access control for the Example System Demo is inherited "
            "on every scope: the AWS GovCloud and Azure Government enclaves inherit "
            "datacenter physical security from the cloud service providers, and the "
            "on-premises/workbook footprint inherits physical access from the DoW "
            "Enterprise data center (Column L) — multi-factor perimeter access, "
            "biometric server-room boundaries, mantrap vestibules, 24x7 guards.",
            "The organization enforces physical access authorizations at entry "
            "and exit points to the facility where the information system "
            "resides.",
            "If inheriting, cite the provider CRM (cloud) and the workbook "
            "inheritance source (Column L); otherwise provide the facility "
            "physical-access procedures and access logs.",
            "Examine the provider CRM inheritance for each cloud enclave and the "
            "Column L inheritance source for the flex scope. Confirm every scope's "
            "physical-access plane is inherited.",
            status="",
            results="",
            cci="CCI-000919", inherited="Remote", remote="DoW Enterprise",
        ),
        # 11th control — the ABSTAIN showcase. AU-4 (Audit Log Storage Capacity)
        # is in-scope and left UNASSESSED. Its ONLY evidence
        # (Audit_Log_Storage_Capacity_Memo_USD20240623.docx) internally
        # contradicts itself — Section 2 asserts the 500 GB allocation meets the
        # 1-year retention requirement, Section 4's utilization analysis shows
        # the partition fills in ~40 days and the fix is only PLANNED. A control
        # whose sole artifact both affirms and refutes the requirement is the
        # textbook abstain case: the assessor cannot reach a confident verdict,
        # so it writes a needs_review row (no status) for human adjudication
        # rather than fabricating Compliant or Non-Compliant.
        Req(
            "AU-4", "Audit Log Storage Capacity",
            "AU", "Audit log storage capacity for the Example System Demo SIEM "
            "tier is documented in the Audit Log Storage Capacity Memo "
            "(USD20240623). The memo is a draft whose capacity sections are not "
            "yet reconciled.",
            "The organization allocates audit log storage capacity to "
            "accommodate the organization-defined audit log retention "
            "requirements.",
            "Provide the audit storage capacity allocation and a utilization "
            "analysis demonstrating the retention period is met.",
            "Examine the allocated capacity against the 1-year retention "
            "requirement and the measured ingest/utilization rate.",
            status="",
            results="",
            cci="CCI-000137", inherited="No",
        ),
        # 12th control — the HYBRID + CUSTOMER per-tenant split showcase. AU-6
        # (Audit Record Review) is marked "Shared" (→ hybrid) by the AWS
        # GovCloud CRM but "Customer" by the Azure Government CRM: on AWS the
        # provider does some SOC-side review, while on Azure the customer runs
        # its own Sentinel workspace and owns review end-to-end. Left UNASSESSED
        # so it flows to the assessor; with BOTH CRMs attached the grid shows an
        # AWS GovCloud "Hybrid" chip alongside an Azure Government "Customer"
        # chip — the mixed-responsibility combo this control exists to exercise.
        Req(
            "AU-6", "Audit Record Review, Analysis, and Reporting",
            "AU", "Audit record review for the Example System Demo differs by "
            "cloud: the AWS GovCloud enclave shares review between the provider "
            "SOC and the customer, while the Azure Government enclave is reviewed "
            "end-to-end by the customer's Microsoft Sentinel workspace.",
            "The organization reviews and analyzes information system audit "
            "records for indications of inappropriate or unusual activity.",
            "Provide the audit-review cadence and sample findings for each cloud "
            "enclave; cite the provider-shared review where applicable.",
            "Examine the audit-review process per scope. For the customer-owned "
            "scope verify the SIEM review cadence; for the shared scope confirm "
            "the provider/customer review split.",
            status="",
            results="",
            cci="CCI-000148", inherited="No",
        ),
        # 13th control — the HYBRID + INHERITED per-tenant split showcase. SC-13
        # (Cryptographic Protection) is marked "Shared" (→ hybrid) by the AWS
        # GovCloud CRM but "Inherited" by the Azure Government CRM: on AWS the
        # customer must select FIPS endpoints / configure guest-OS FIPS mode,
        # while on Azure the workload consumes only Microsoft-provided validated
        # modules and fully inherits the control. Left UNASSESSED so it flows to
        # the assessor; with BOTH CRMs attached the grid shows an AWS GovCloud
        # "Hybrid" chip alongside an Azure Government "Inherited" chip.
        Req(
            "SC-13", "Cryptographic Protection",
            "SC", "Cryptographic protection for the Example System Demo differs by "
            "cloud: the AWS GovCloud enclave shares responsibility (customer "
            "selects FIPS endpoints and guest-OS FIPS mode), while the Azure "
            "Government enclave consumes only Microsoft-provided FIPS 140-2 "
            "validated modules and fully inherits the control.",
            "The information system implements organization-defined cryptographic "
            "uses and type of cryptography in accordance with applicable laws.",
            "Provide the FIPS-validated module configuration per cloud enclave; "
            "cite the provider CRM where the control is inherited.",
            "Examine the cryptographic implementation per scope. For the shared "
            "scope verify FIPS endpoint/guest-OS configuration; for the inherited "
            "scope confirm the provider supplies the validated modules.",
            status="",
            results="",
            cci="CCI-002450", inherited="No",
        ),
        # 14th control — the ABSTAIN-WITH-BOTH-CRMS showcase. AU-9 (Protection
        # of Audit Information) is CUSTOMER-owned on BOTH the AWS GovCloud and
        # Azure Government CRMs (the provider supplies the storage primitives;
        # the customer must configure immutability/access controls on its own
        # log stores). Its ONLY evidence — the Audit Information Protection
        # Memo (USD20240624) — INTERNALLY CONTRADICTS itself: Section 2 asserts
        # S3 Object Lock / immutable blob retention is ENABLED and AU-9 is met,
        # while Section 4's configuration review found Object Lock is NOT
        # enabled on the production log bucket and the fix is only PLANNED.
        # A customer-owned control whose sole artifact both affirms and refutes
        # the protection — with BOTH CRMs attached — is the textbook abstain in
        # the multi-scope context: the assessor cannot reach a confident
        # verdict, so it writes a needs_review row (no status) for human
        # adjudication rather than fabricating Compliant or Non-Compliant.
        # Left UNASSESSED so it flows to the assessor.
        Req(
            "AU-9", "Protection of Audit Information",
            "AU", "Protection of audit information for the Example System Demo "
            "is customer-owned on both cloud enclaves; immutability and access "
            "controls on the customer log stores are documented in the Audit "
            "Information Protection Memo (USD20240624), a draft whose protection "
            "sections are not yet reconciled.",
            "The information system protects audit information and audit tools "
            "from unauthorized access, modification, and deletion.",
            "Provide the immutability / access-control configuration for the "
            "customer audit log stores on each cloud enclave.",
            "Examine the configured object-lock / immutable-retention and access "
            "controls on the customer log stores against the protection "
            "requirement; confirm the as-configured state matches the memo.",
            status="",
            results="",
            cci="CCI-001493", inherited="No",
        ),
    ],
)

# ---------------------------------------------------------------------------
# NIST SP 800-171 Rev. 3 -- requirement number is the assessable unit
# ---------------------------------------------------------------------------
NIST_800_171 = FrameworkDemo(
    framework_label="NIST SP 800-171 Rev. 3",
    filename="DEMO_NIST_800-171r3_Example System.xlsx",
    reqs=[
        Req(
            "03.01.01", "Account Management",
            "Access Control", "Account types, approvals, and lifecycle are "
            "documented in the Account Management Policy (USD20240315).",
            "Define the types of system accounts allowed and prohibited; create, "
            "enable, modify, disable, and remove accounts per policy.",
            "Provide the account management policy and an account inventory.",
            "Examine the policy. Sample accounts and verify each maps to a "
            "documented type.",
            status="Compliant",
            results="Examined Account Management Policy (USD20240315) Section 3 "
            "(five account types) and the 2026-05-19 account inventory; sampled "
            "five accounts, all map to a documented type.",
        ),
        Req(
            "03.01.05", "Least Privilege",
            "Access Control", "Privileged roles are assigned per the least-"
            "privilege matrix; admin rights are role-based.",
            "Allow only authorized system access necessary to accomplish assigned "
            "organizational tasks.",
            "Provide the role/permission matrix and a privileged-account list.",
            "Examine the matrix. Verify privileged accounts match documented "
            "roles.",
            status="Compliant",
            results="Examined least-privilege role matrix (USD20240401) and the "
            "privileged-account list; all four admins map to a documented role.",
        ),
        Req(
            "03.05.07", "Password Management",
            "Identification and Authentication", "",
            "Enforce a minimum password complexity and change of characters when "
            "new passwords are created.",
            "Provide the GPO export showing length and complexity.",
            "Examine the GPO export; attempt a weak password and verify "
            "rejection.",
            cci="03.05.07.a",
        ),
        Req(
            "03.08.03", "Media Sanitization",
            "Media Protection", "Removable media is sanitized per the Media "
            "Protection SOP before reuse or disposal.",
            "Sanitize system media containing CUI before disposal, release, or "
            "reuse.",
            "Provide the sanitization SOP and a sample sanitization record.",
            "Examine the SOP. Examine a sanitization log entry.",
            status="Compliant",
            results="Examined Media Protection SOP (USD20240210) and a 2026-04 "
            "sanitization record showing NIST 800-88 purge for two drives.",
        ),
        Req(
            "03.11.02", "Vulnerability Monitoring and Scanning",
            "Risk Assessment", "ACAS credentialed scans run on a defined cadence "
            "against the CUI enclave.",
            "Scan for vulnerabilities in the system and applications periodically "
            "and when new vulnerabilities are identified.",
            "Provide a recent credentialed scan report.",
            "Examine the scan report; verify credentialed and in-scope.",
            status="Compliant",
            results="Examined acas_nessus_subnet_10_10_5_0.nessus dated "
            "2026-05-20; credentialed scan of the CUI enclave confirmed.",
        ),
        Req(
            "03.14.06", "System Monitoring",
            "System and Information Integrity", "",
            "Monitor the system to detect attacks and indicators of potential "
            "attacks; identify unauthorized use.",
            "Provide the SIEM correlation report and monitored-source list.",
            "Examine the SIEM report. Verify each asset has a healthy log source.",
        ),
    ],
)

# ---------------------------------------------------------------------------
# NIST CSF 2.0 -- subcategory ids in column B
# ---------------------------------------------------------------------------
NIST_CSF = FrameworkDemo(
    framework_label="NIST Cybersecurity Framework 2.0",
    filename="DEMO_NIST_CSF_2.0_Example System.xlsx",
    reqs=[
        Req(
            "GV.OC-01", "Organizational Context | Mission Understanding",
            "GV", "The mission and stakeholder expectations for the Example System Demo "
            "system are documented in the System Security Plan front matter.",
            "The organizational mission is understood and informs cybersecurity "
            "risk management.",
            "Provide the SSP section describing mission and stakeholders.",
            "Examine the SSP front matter. Confirm mission/stakeholders are "
            "stated.",
            status="Compliant",
            results="Examined SSP Section 1 (USD20240101); mission statement and "
            "stakeholder list are documented and current.",
        ),
        Req(
            "ID.AM-01", "Asset Management | Hardware Inventory",
            "ID", "A hardware inventory is auto-derived from ACAS + CKL + the "
            "declared inventory and reconciled monthly.",
            "Inventories of hardware managed by the organization are maintained.",
            "Provide the hardware inventory and the reconciliation record.",
            "Examine the inventory. Cross-check against the latest scan host "
            "list.",
            status="Compliant",
            results="Examined the 2026-05 hardware inventory; reconciled against "
            "acas_nessus host list with zero unexplained hosts.",
        ),
        Req(
            "PR.AA-01", "Identity Management & Access Control | Identities",
            "PR", "",
            "Identities and credentials for authorized users are managed by the "
            "organization.",
            "Provide the IdAM configuration and the account inventory.",
            "Examine the IdAM config. Sample identities for lifecycle controls.",
        ),
        Req(
            "DE.CM-01", "Continuous Monitoring | Network Monitoring",
            "DE", "Networks and network services are monitored by the SIEM with "
            "weekly correlation reporting.",
            "Networks and network services are monitored to find potentially "
            "adverse events.",
            "Provide the SIEM correlation report and monitored-source list.",
            "Examine the SIEM report. Verify network sources are healthy.",
            status="Compliant",
            results="Examined siem_weekly_correlation_report_2026-05-22.txt; "
            "network sensors and flow sources report healthy.",
        ),
        Req(
            "RS.MA-01", "Incident Management | Response Execution",
            "RS", "",
            "The incident response plan is executed in coordination with relevant "
            "third parties once an incident is declared.",
            "Provide the incident response plan and a tabletop exercise record.",
            "Examine the IR plan. Examine the most recent exercise after-action.",
        ),
        Req(
            "RC.RP-01", "Incident Recovery Plan Execution",
            "RC", "Recovery procedures are documented in the Contingency Plan and "
            "tested annually.",
            "The recovery portion of the incident response plan is executed once "
            "initiated from the incident response process.",
            "Provide the contingency/recovery plan and a test record.",
            "Examine the recovery plan. Examine the annual test after-action.",
            status="Compliant",
            results="Examined Contingency Plan (USD20240320) and the 2026-03 "
            "recovery test after-action; RTO/RPO objectives were met.",
        ),
    ],
)

# ---------------------------------------------------------------------------
# ISO/IEC 27001:2022 -- Annex A control numbers (license-aware catalog)
# ---------------------------------------------------------------------------
ISO_27001 = FrameworkDemo(
    framework_label="ISO/IEC 27001:2022 (Annex A)",
    filename="DEMO_ISO_27001_2022_Example System.xlsx",
    reqs=[
        Req(
            "A.5.1", "Policies for information security",
            "Organizational", "An information security policy set is approved by "
            "management and reviewed annually.",
            "Information security policy and topic-specific policies shall be "
            "defined, approved, published, and reviewed.",
            "Provide the approved ISMS policy and the review record.",
            "Examine the policy. Confirm management approval and review date.",
            status="Compliant",
            results="Examined the ISMS Information Security Policy (rev 2026-01); "
            "management approval signature and annual review date present.",
        ),
        Req(
            "A.5.15", "Access control",
            "Organizational", "Access is granted role-based per the access "
            "control policy; reviews occur quarterly.",
            "Rules to control physical and logical access shall be established and "
            "implemented based on business and security requirements.",
            "Provide the access control policy and a quarterly review record.",
            "Examine the policy. Examine a quarterly access review.",
            status="Compliant",
            results="Examined the Access Control Policy (rev 2026-02) and the "
            "Q1-2026 access review; all entitlements were attested.",
        ),
        Req(
            "A.8.7", "Protection against malware",
            "Technological", "",
            "Protection against malware shall be implemented and supported by "
            "appropriate user awareness.",
            "Provide the endpoint anti-malware configuration and a signature-age "
            "report.",
            "Examine the EDR/AV console. Confirm signatures are current.",
        ),
        Req(
            "A.8.16", "Monitoring activities",
            "Technological", "Networks, systems, and applications are monitored "
            "by the SIEM with weekly review.",
            "Networks, systems, and applications shall be monitored for anomalous "
            "behaviour and appropriate actions taken.",
            "Provide the SIEM correlation report and the review record.",
            "Examine the SIEM report. Verify anomalous events were dispositioned.",
            status="Compliant",
            results="Examined siem_weekly_correlation_report_2026-05-22.txt; "
            "flagged events were triaged and closed.",
        ),
        Req(
            "A.8.8", "Management of technical vulnerabilities",
            "Technological", "Technical vulnerabilities are identified via ACAS "
            "scanning and remediated per the vulnerability management SOP.",
            "Information about technical vulnerabilities shall be obtained and "
            "appropriate measures taken to address the associated risk.",
            "Provide a recent scan report and the remediation tracker.",
            "Examine the scan report and the remediation tracker.",
            status="Compliant",
            results="Examined acas_nessus_subnet_10_10_5_0.nessus dated "
            "2026-05-20 and the remediation tracker; criticals are within SLA.",
        ),
    ],
)

# ---------------------------------------------------------------------------
# CIS Controls v8 -- Safeguard ids (license-aware catalog)
# ---------------------------------------------------------------------------
CIS_V8 = FrameworkDemo(
    framework_label="CIS Controls v8",
    filename="DEMO_CIS_v8_Example System.xlsx",
    reqs=[
        Req(
            "1.1", "Establish and Maintain Detailed Enterprise Asset Inventory",
            "1", "A detailed asset inventory is auto-derived and reconciled "
            "monthly.",
            "Establish and maintain an accurate, detailed, and up-to-date "
            "inventory of all enterprise assets.",
            "Provide the asset inventory and the reconciliation record.",
            "Examine the inventory. Cross-check against the latest scan.",
            status="Compliant",
            results="Examined the 2026-05 enterprise asset inventory; reconciled "
            "against the ACAS host list with no unexplained assets.",
        ),
        Req(
            "4.1", "Establish and Maintain a Secure Configuration Process",
            "4", "Secure configuration baselines follow DISA STIGs and are "
            "tracked in the configuration management plan.",
            "Establish and maintain a secure configuration process for enterprise "
            "assets and software.",
            "Provide the configuration management plan and a STIG checklist.",
            "Examine the CM plan. Examine a STIG checklist.",
            status="Compliant",
            results="Examined the CM Plan (USD20240118) and "
            "Windows_Server_2022_STIG_Sample.ckl dated 2026-05-18.",
        ),
        Req(
            "5.2", "Use Unique Passwords",
            "5", "",
            "Use unique passwords for all enterprise assets; enforce minimum "
            "length and complexity.",
            "Provide the GPO export showing length and complexity.",
            "Examine the GPO export.",
        ),
        Req(
            "7.1", "Establish and Maintain a Vulnerability Management Process",
            "7", "A documented vulnerability management process drives monthly "
            "ACAS scanning and remediation.",
            "Establish and maintain a documented vulnerability management process "
            "for enterprise assets.",
            "Provide the vulnerability management SOP and a recent scan report.",
            "Examine the SOP. Examine a recent scan report.",
            status="Compliant",
            results="Examined the Vulnerability Management SOP (USD20240222) and "
            "acas_nessus_subnet_10_10_5_0.nessus dated 2026-05-20.",
        ),
        Req(
            "8.2", "Collect Audit Logs",
            "8", "Audit logs are collected from all in-scope assets and "
            "forwarded to the SIEM.",
            "Collect audit logs; ensure logging has been enabled across "
            "enterprise assets.",
            "Provide the log-source inventory and a SIEM ingest report.",
            "Examine the log-source inventory. Verify each asset forwards logs.",
            status="Compliant",
            results="Examined the SIEM log-source inventory; all in-scope assets "
            "show healthy ingest in the 2026-05-22 report.",
        ),
    ],
)

# ---------------------------------------------------------------------------
# PCI DSS 4.0 -- dotted requirement ids (license-aware catalog)
# ---------------------------------------------------------------------------
PCI_DSS = FrameworkDemo(
    framework_label="PCI DSS 4.0",
    filename="DEMO_PCI_DSS_4.0_Example System.xlsx",
    reqs=[
        Req(
            "1.2.1", "Network Security Controls Configuration",
            "1", "NSC rulesets are defined and reviewed every six months per the "
            "firewall management procedure.",
            "Configuration standards for network security control rulesets are "
            "defined, implemented, and maintained.",
            "Provide the NSC ruleset export and the review record.",
            "Examine the ruleset. Confirm a six-month review occurred.",
            status="Compliant",
            results="Examined the NSC ruleset export and the 2026-04 review "
            "record; ruleset is documented and approved.",
        ),
        Req(
            "8.3.6", "Strong Authentication | Password Length",
            "8", "Passwords meet a 12-character minimum with complexity, enforced "
            "via the domain policy.",
            "Passwords/passphrases meet a minimum length of 12 characters and "
            "contain both numeric and alphabetic characters.",
            "Provide the GPO export showing length and complexity.",
            "Examine the GPO export.",
            status="Compliant",
            results="Examined GPO_Password_Policy_Export.xlsx (USD20240218): "
            "length 15 and complexity Enabled, exceeding the PCI minimum.",
        ),
        Req(
            "10.2.1", "Audit Logs | Capture",
            "10", "",
            "Audit logs are enabled and active for all system components and "
            "capture the required event types.",
            "Provide the audit policy export and a sample log.",
            "Examine the audit policy. Verify required events are captured.",
        ),
        Req(
            "11.3.1", "External and Internal Vulnerability Scans",
            "11", "Internal and external vulnerability scans run quarterly and "
            "after significant change.",
            "Internal vulnerability scans are performed at least once every three "
            "months.",
            "Provide the most recent quarterly scan reports.",
            "Examine the scan reports. Confirm quarterly cadence.",
            status="Compliant",
            results="Examined acas_nessus_subnet_10_10_5_0.nessus dated "
            "2026-05-20; quarterly cadence confirmed against the scan calendar.",
        ),
        Req(
            "12.1.1", "Information Security Policy",
            "12", "An information security policy is established, published, and "
            "reviewed at least annually.",
            "An overall information security policy is established, published, "
            "maintained, and disseminated.",
            "Provide the approved policy and the annual review record.",
            "Examine the policy. Confirm publication and annual review.",
            status="Compliant",
            results="Examined the Information Security Policy (rev 2026-01); "
            "publication and annual management review are documented.",
        ),
    ],
)

# ---------------------------------------------------------------------------
# SOC 2 -- Trust Services Criteria ids (license-aware catalog)
# ---------------------------------------------------------------------------
SOC_2 = FrameworkDemo(
    framework_label="SOC 2 (2017 Trust Services Criteria)",
    filename="DEMO_SOC_2_TSC_Example System.xlsx",
    reqs=[
        Req(
            "CC1.1", "Control Environment | Integrity and Ethical Values",
            "CC", "A code of conduct is published and acknowledged annually by "
            "all personnel.",
            "The entity demonstrates a commitment to integrity and ethical "
            "values.",
            "Provide the code of conduct and acknowledgement roster.",
            "Examine the code of conduct. Sample the acknowledgement roster.",
            status="Compliant",
            results="Examined the Code of Conduct (rev 2026-01) and the "
            "acknowledgement roster; sampled 10 staff, all acknowledged.",
        ),
        Req(
            "CC6.1", "Logical Access | Security Software & Infrastructure",
            "CC", "Logical access is restricted via role-based AD groups and MFA "
            "for privileged access.",
            "The entity implements logical access security software, "
            "infrastructure, and architectures over protected information assets.",
            "Provide the IdAM config and the MFA enrollment report.",
            "Examine the IdAM config. Confirm MFA on privileged accounts.",
            status="Compliant",
            results="Examined the IdAM role mapping and the MFA enrollment "
            "report; all four privileged accounts are MFA-enrolled.",
        ),
        Req(
            "CC6.6", "Logical Access | Boundary Protection",
            "CC", "",
            "The entity implements logical access security measures to protect "
            "against threats from sources outside its system boundaries.",
            "Provide the boundary protection configuration or inheritance memo.",
            "Examine the firewall config or the inheritance memorandum.",
        ),
        Req(
            "CC7.2", "System Operations | Monitoring",
            "CC", "Infrastructure and software are monitored by the SIEM to "
            "detect anomalies and security events.",
            "The entity monitors system components and the operation of controls "
            "to detect anomalies indicative of security events.",
            "Provide the SIEM correlation report and the alerting policy.",
            "Examine the SIEM report. Verify anomalies were dispositioned.",
            status="Compliant",
            results="Examined siem_weekly_correlation_report_2026-05-22.txt; "
            "flagged anomalies were triaged and closed within SLA.",
        ),
        Req(
            "A1.2", "Availability | Environmental Protections & Backup",
            "A", "Backups run nightly and recovery is tested annually per the "
            "contingency plan.",
            "The entity authorizes, designs, develops, implements, operates, "
            "maintains, and monitors environmental protections, software, data "
            "backup processes, and recovery infrastructure.",
            "Provide the backup schedule and the most recent recovery test.",
            "Examine the backup logs. Examine the annual recovery test.",
            status="Compliant",
            results="Examined nightly backup logs for 2026-05 and the 2026-03 "
            "recovery test after-action; RTO/RPO objectives were met.",
        ),
    ],
)


ALL_FRAMEWORKS = [
    NIST_800_53,
    NIST_800_171,
    NIST_CSF,
    ISO_27001,
    CIS_V8,
    PCI_DSS,
    SOC_2,
]


def _build_one(fd: FrameworkDemo) -> Path:
    out = OUT_DIR / fd.filename

    wb = Workbook()
    ws = wb.active
    ws.title = "WORKING SHEET"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF")
    wrap_top = Alignment(wrap_text=True, vertical="top")
    inheritable_fill = PatternFill("solid", fgColor="FFF2CC")

    # Rows 1-5: metadata banner
    ws["A1"] = "Example System DEMO - eMASS CONTROLS WORKBOOK (DEMO COPY)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")

    ws["A2"] = f"System: {EXAMPLE_SYSTEM}"
    ws["A2"].font = bold
    ws["H2"] = "eMASS Export Date:"
    ws["I2"] = "2026-05-15"

    ws["A3"] = f"Framework: {fd.framework_label}"
    ws["A3"].font = bold
    ws["A4"] = "Assessment Lead: Noah Jaskolski, ISSO"
    ws["A5"] = (
        "Notice: This is a SYNTHETIC demo workbook used to exercise the "
        "cybersecurity-assessor app. Do NOT use for a real authorization."
    )
    ws["A5"].font = Font(italic=True, color="C00000")
    ws.merge_cells("A5:U5")

    # Row 6: headers
    for i, (label, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=6, column=i, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(
            wrap_text=True, vertical="center", horizontal="center"
        )
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[6].height = 42

    # Rows 7+: data
    for ri, req in enumerate(fd.reqs, start=7):
        cci = req.cci or req.control
        row = [
            "Yes",                  # A Required
            req.control,            # B Control
            req.title,              # C Information
            "Implemented" if req.status == "Compliant" else (
                "Inherited" if req.inherited == "Yes" else ""
            ),                       # D Implementation Status
            "Inherited" if req.inherited == "Yes" else "",  # E Designation
            req.narrative,          # F Narrative
            req.control,            # G AP Acronym
            cci,                    # H CCI / requirement key
            req.definition,         # I Definition
            req.guidance,           # J Guidance
            req.procedures,         # K Procedures
            req.inherited,          # L Inherited
            req.remote,             # M Remote Inheritance
            req.status,             # N Compliance Status
            "2026-05-19" if req.status else "",  # O Date Tested
            TESTER if req.status else "",        # P Tested By
            req.results,            # Q Test Results
            req.framework_tag or fd.framework_label,  # R Originating Req
            "Inherited" if req.inherited == "Yes" else (
                "Common" if req.status == "Compliant" else "System-Specific"
            ),                       # S Implementation Type
            "T. Prior",             # T Previous Assessor
            "",                     # U Previous Test Results
        ]
        for ci, value in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.alignment = wrap_top
            cell.border = border
        if req.inherited == "Yes":
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=ri, column=ci).fill = inheritable_fill
        ws.row_dimensions[ri].height = 110

    ws.freeze_panes = "C7"

    dv = DataValidation(
        type="list",
        formula1='"Compliant,Non-Compliant,Not Applicable"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Status must be Compliant, Non-Compliant, or Not Applicable."
    dv.errorTitle = "Invalid status"
    ws.add_data_validation(dv)
    dv.add(f"N7:N{7 + len(fd.reqs) - 1}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


def build() -> list[Path]:
    written = []
    for fd in ALL_FRAMEWORKS:
        written.append(_build_one(fd))
    return written


if __name__ == "__main__":
    for path in build():
        print(f"WROTE  workbooks/{path.name}")
